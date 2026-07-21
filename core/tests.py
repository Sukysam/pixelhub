from decimal import Decimal
from datetime import date, timedelta
import hashlib
import hmac
import io
import json
import os
import secrets
import tempfile
import urllib.parse
from unittest.mock import patch
from django.db.utils import DatabaseError

from django.urls import reverse
from django.contrib.auth.models import User, Permission
from django.utils import timezone
from django.core.cache import cache
from django.core import mail
from django.test import override_settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from rest_framework import status
from rest_framework.test import APITestCase

from PIL import Image
from openpyxl import Workbook, load_workbook

from .models import (
    Item,
    Invoice,
    Customer,
    Receipt,
    InvoiceItem,
    AuditLog,
    Expense,
    SourceAccount,
    Currency,
    ExchangeRate,
    GlobalSettings,
    UserSettings,
    UserProfile,
    SocialAuthConnection,
    EmailVerificationToken,
    AccessToken,
    Role,
    UserRole,
    SavedInvoiceView,
    DocumentDelivery,
    SavedDocument,
    PaymentTransaction,
    BusinessAccount,
    BusinessMembership,
    LogoAsset,
    AdminUserInvitation,
    evaluate_invoice_payment_status,
)
from .import_export import (
    process_batch_import,
    import_customers_from_upload,
    import_expenses_from_upload,
    import_invoices_from_upload,
    import_items_from_upload,
)
from .expense_security import decrypt_expense_text, is_encrypted_expense_value
from .finance_services import (
    outstanding_invoice_amount,
    resolve_invoice_discount,
    sync_invoice_status_with_payments,
    invoice_discount_context,
)
from .serializers import build_internal_remarks_preview
from .rendering_service import (
    detect_country,
    effective_company_identity_for_user,
    effective_region_settings_for_user,
    effective_templates_for_user,
    format_money,
)
from .auth_service import role_for_name
from .views import _rate_limit
from .documents import RenderedDocument, create_delivery, render_invoice, render_receipt


def _test_secret() -> str:
    return f"Aa1!{secrets.token_urlsafe(12)}"


def _png_logo_upload_file(name: str = "logo.png") -> SimpleUploadedFile:
    img = Image.new("RGBA", (400, 120), (200, 10, 10, 255))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return SimpleUploadedFile(name, buf.getvalue(), content_type="image/png")


def _webp_logo_upload_file(name: str = "logo.webp") -> SimpleUploadedFile:
    img = Image.new("RGBA", (320, 120), (10, 120, 210, 255))
    buf = io.BytesIO()
    img.save(buf, format="WEBP")
    return SimpleUploadedFile(name, buf.getvalue(), content_type="image/webp")


def _pdf_upload_file(name: str = "receipt.pdf") -> SimpleUploadedFile:
    data = b"%PDF-1.4\n1 0 obj<</Type/Catalog>>endobj\ntrailer<</Root 1 0 R>>\n%%EOF"
    return SimpleUploadedFile(name, data, content_type="application/pdf")


def _csv_upload_file(name: str, body: str) -> SimpleUploadedFile:
    return SimpleUploadedFile(name, body.encode("utf-8"), content_type="text/csv")


def _xlsx_upload_file(name: str, header: list[str], rows: list[list[object]]) -> SimpleUploadedFile:
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    return SimpleUploadedFile(
        name,
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class _MockJsonResponse:
    def __init__(self, payload):
        self.payload = payload

    def read(self):
        return json.dumps(self.payload).encode("utf-8")

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False


GW_SHARED = "test"


class PersistenceTests(APITestCase):
    def setUp(self):
        super().setUp()
        self.cred = _test_secret()
        self.user = User.objects.create_user(username="admin", password=self.cred)
        self.client.force_authenticate(user=self.user)
        perms = Permission.objects.filter(
            codename__in=[
                "change_customer",
                "delete_customer",
                "change_item",
                "delete_item",
                "change_invoice",
                "delete_invoice",
                "change_receipt",
                "delete_receipt",
                "change_invoiceitem",
                "delete_invoiceitem",
            ]
        )
        self.user.user_permissions.add(*perms)

    def test_customers_crud_and_persistence(self):
        create_url = reverse("customer-list")
        res = self.client.post(
            create_url,
            {"name": "Acme", "email": "acme@example.com", "phone": "123", "billing_address": "Addr"},
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        customer_id = res.data["id"]

        detail_url = reverse("customer-detail", args=[customer_id])
        res = self.client.get(detail_url)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data["name"], "Acme")
        updated_at = res.data["updated_at"]

        self.client.force_authenticate(user=self.user)
        res = self.client.patch(detail_url, {"phone": "456", "updated_at": updated_at}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data["phone"], "456")
        updated_at = res.data["updated_at"]

        res = self.client.delete(f"{detail_url}?updated_at={updated_at}")
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)
        self.assertTrue(AuditLog.objects.filter(action="delete", object_id=str(customer_id)).exists())

    def test_customer_internal_remarks_are_permission_gated_and_never_render_on_documents(self):
        create_url = reverse("customer-list")
        denied_create = self.client.post(
            create_url,
            {"name": "Internal Buyer", "email": "internal@example.com", "internal_remarks": "TOPSECRET"},
            format="json",
        )
        self.assertEqual(denied_create.status_code, status.HTTP_403_FORBIDDEN)

        customer = self.client.post(
            create_url,
            {"name": "Buyer", "email": "buyer_notes@example.com"},
            format="json",
        ).data
        detail_url = reverse("customer-detail", args=[customer["id"]])

        denied_patch = self.client.patch(
            detail_url,
            {"internal_remarks": "TOPSECRET", "updated_at": customer["updated_at"]},
            format="json",
        )
        self.assertEqual(denied_patch.status_code, status.HTTP_403_FORBIDDEN)

        read_res = self.client.get(detail_url)
        self.assertEqual(read_res.status_code, status.HTTP_200_OK)
        self.assertNotIn("internal_remarks", read_res.data)

        staff = User.objects.create_user(username="staff_notes", password=_test_secret(), is_staff=True)
        self.client.force_authenticate(user=staff)
        notes = "  TOPSECRET internal note with enough detail to exceed the fifty character preview limit.  "
        staff_patch = self.client.patch(
            detail_url,
            {"internal_remarks": notes, "updated_at": customer["updated_at"]},
            format="json",
        )
        self.assertEqual(staff_patch.status_code, status.HTTP_200_OK)
        self.assertEqual(staff_patch.data["internal_remarks"], notes.strip())
        self.assertEqual(staff_patch.data["internal_remarks_preview"], build_internal_remarks_preview(notes))
        self.assertTrue(staff_patch.data["has_internal_remarks"])

        staff_list = self.client.get(create_url)
        self.assertEqual(staff_list.status_code, status.HTTP_200_OK)
        staff_row = next(row for row in staff_list.data["results"] if row["id"] == customer["id"])
        self.assertEqual(staff_row["internal_remarks_preview"], build_internal_remarks_preview(notes))
        self.assertTrue(staff_row["has_internal_remarks"])
        self.assertNotIn("internal_remarks", staff_row)

        staff_detail = self.client.get(detail_url)
        self.assertEqual(staff_detail.status_code, status.HTTP_200_OK)
        self.assertEqual(staff_detail.data["internal_remarks"], notes.strip())

        item = self.client.post(
            reverse("item-list"),
            {"name": "Widget", "sku": "W-002", "unit_price": "10.00", "tax_rate": "0", "stock_quantity": 5},
            format="json",
        ).data
        invoice_res = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        )
        self.assertEqual(invoice_res.status_code, status.HTTP_201_CREATED)
        invoice_id = invoice_res.data["id"]
        invoice_html = self.client.get(f"/api/invoices/{invoice_id}/print_html/")
        self.assertEqual(invoice_html.status_code, status.HTTP_200_OK)
        self.assertNotIn("TOPSECRET", invoice_html.content.decode("utf-8", errors="ignore"))

        receipt_res = self.client.post(
            reverse("receipt-list"),
            {"invoice": invoice_id, "amount_paid": "10.00", "payment_method": "Cash", "reference_number": "R-2"},
            format="json",
        )
        self.assertEqual(receipt_res.status_code, status.HTTP_201_CREATED)
        receipt_id = receipt_res.data["id"]
        receipt_html = self.client.get(f"/api/receipts/{receipt_id}/print_html/")
        self.assertEqual(receipt_html.status_code, status.HTTP_200_OK)
        self.assertNotIn("TOPSECRET", receipt_html.content.decode("utf-8", errors="ignore"))

        self.client.force_authenticate(user=self.user)
        unreadable_list = self.client.get(create_url)
        self.assertEqual(unreadable_list.status_code, status.HTTP_200_OK)
        unreadable_row = next(row for row in unreadable_list.data["results"] if row["id"] == customer["id"])
        self.assertNotIn("internal_remarks", unreadable_row)
        self.assertNotIn("internal_remarks_preview", unreadable_row)
        self.assertNotIn("has_internal_remarks", unreadable_row)

    def test_internal_remarks_preview_helper_truncates_to_fifty_characters(self):
        original = "First line with extra spacing\nand additional detail for preview rendering."
        preview = build_internal_remarks_preview(original)
        self.assertLessEqual(len(preview), 50)
        self.assertTrue(preview.endswith("..."))
        self.assertEqual(preview, "First line with extra spacing and additional de...")

    def test_inventory_invoice_and_receipt_flow(self):
        customer = self.client.post(
            reverse("customer-list"),
            {"name": "Buyer", "email": "buyer@example.com"},
            format="json",
        ).data

        item = self.client.post(
            reverse("item-list"),
            {"name": "Widget", "sku": "W-001", "unit_price": "10.00", "tax_rate": "10.0", "stock_quantity": 5},
            format="json",
        ).data

        invoice_res = self.client.post(
            reverse("invoice-list"),
            {
                "customer": customer["id"],
                "status": "Draft",
                "items": [{"item": item["id"], "quantity": 2}],
            },
            format="json",
        )
        self.assertEqual(invoice_res.status_code, status.HTTP_201_CREATED)
        invoice_id = invoice_res.data["id"]
        self.assertTrue(invoice_res.data["invoice_number"].startswith("INV-"))
        self.assertEqual(Decimal(invoice_res.data["subtotal"]), Decimal("20.00"))
        self.assertEqual(Decimal(invoice_res.data["tax_total"]), Decimal("2.00"))
        self.assertEqual(Decimal(invoice_res.data["total_amount"]), Decimal("22.00"))

        invoice_detail = reverse("invoice-detail", args=[invoice_id])
        self.client.force_authenticate(user=self.user)
        sent_res = self.client.patch(invoice_detail, {"status": "Sent"}, format="json")
        self.assertEqual(sent_res.status_code, status.HTTP_200_OK)

        db_item = Item.objects.get(pk=item["id"])
        self.assertEqual(db_item.stock_quantity, 3)

        self.client.force_authenticate(user=self.user)
        receipt_res = self.client.post(
            reverse("receipt-list"),
            {
                "invoice": invoice_id,
                "amount_paid": "22.00",
                "payment_method": "Cash",
                "reference_number": "R-1",
            },
            format="json",
        )
        self.assertEqual(receipt_res.status_code, status.HTTP_201_CREATED)
        invoice = Invoice.objects.get(pk=invoice_id)
        self.assertEqual(invoice.status, "Paid")

    def test_items_read_after_write_includes_new_item(self):
        create_res = self.client.post(
            reverse("item-list"),
            {"name": "Widget A", "sku": "W-A", "unit_price": "12.00", "tax_rate": "0", "stock_quantity": 7},
            format="json",
        )
        self.assertEqual(create_res.status_code, status.HTTP_201_CREATED)
        created_id = create_res.data["id"]

        list_res = self.client.get(reverse("item-list"))
        self.assertEqual(list_res.status_code, status.HTTP_200_OK)
        self.assertEqual(list_res.get("Cache-Control"), "no-store")

        payload = list_res.data
        results = payload.get("results") if isinstance(payload, dict) else payload
        ids = [row.get("id") for row in (results or []) if isinstance(row, dict)]
        self.assertIn(created_id, ids)

    def test_customer_detail_and_list_support_metrics_filtering_and_sorting(self):
        customer = self.client.post(
            reverse("customer-list"),
            {
                "name": "Alpha Buyer",
                "email": "alpha@example.com",
                "phone": "5551234",
                "billing_address": "123 Main St",
            },
            format="json",
        ).data
        item = self.client.post(
            reverse("item-list"),
            {
                "name": "Customer Metrics Item",
                "sku": "CM-001",
                "unit_price": "15.00",
                "tax_rate": "0",
                "stock_quantity": 10,
                "warehouse_location": "A-01",
                "last_restock_date": "2026-06-01",
            },
            format="json",
        ).data
        invoice = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 2}]},
            format="json",
        ).data

        list_res = self.client.get(f"{reverse('customer-list')}?q=Alpha&ordering=name")
        self.assertEqual(list_res.status_code, status.HTTP_200_OK)
        self.assertEqual(list_res.data["results"][0]["name"], "Alpha Buyer")
        self.assertEqual(list_res.data["results"][0]["invoice_count"], 1)
        self.assertEqual(Decimal(list_res.data["results"][0]["lifetime_value"]), Decimal("30.00"))

        detail_res = self.client.get(reverse("customer-detail", args=[customer["id"]]))
        self.assertEqual(detail_res.status_code, status.HTTP_200_OK)
        self.assertEqual(detail_res.data["email"], "alpha@example.com")
        self.assertEqual(detail_res.data["order_history"][0]["invoice_number"], invoice["invoice_number"])

    def test_item_detail_and_list_support_filters_sorting_and_usage_history(self):
        customer = self.client.post(reverse("customer-list"), {"name": "Inventory Buyer"}, format="json").data
        item = self.client.post(
            reverse("item-list"),
            {
                "name": "Warehouse Widget",
                "sku": "WH-001",
                "category": "Hardware",
                "description": "Detailed spec",
                "unit_price": "8.00",
                "tax_rate": "5",
                "tax_category": "standard",
                "unit_of_measure": "box",
                "stock_quantity": 9,
                "warehouse_location": "Rack-22",
                "last_restock_date": "2026-05-15",
            },
            format="json",
        ).data
        self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        )

        list_res = self.client.get(f"{reverse('item-list')}?warehouse_location=Rack&stock_min=5&ordering=-last_restock_date")
        self.assertEqual(list_res.status_code, status.HTTP_200_OK)
        self.assertEqual(list_res.data["results"][0]["warehouse_location"], "Rack-22")
        self.assertEqual(list_res.data["results"][0]["stock_status"], "in_stock")
        self.assertEqual(list_res.data["results"][0]["category"], "Hardware")

        detail_res = self.client.get(reverse("item-detail", args=[item["id"]]))
        self.assertEqual(detail_res.status_code, status.HTTP_200_OK)
        self.assertEqual(detail_res.data["specifications"]["unit_of_measure"], "box")
        self.assertEqual(detail_res.data["specifications"]["category"], "Hardware")
        self.assertEqual(detail_res.data["recent_invoice_usage"][0]["invoice_status"], "Draft")

        category_list_res = self.client.get(f"{reverse('item-list')}?category=Hard&q=Hardware")
        self.assertEqual(category_list_res.status_code, status.HTTP_200_OK)
        self.assertEqual(category_list_res.data["results"][0]["id"], item["id"])

    def test_invoice_detail_and_list_support_payment_filters_sorting_and_customer_context(self):
        customer = self.client.post(reverse("customer-list"), {"name": "Invoice Buyer", "email": "invoice@example.com"}, format="json").data
        item = self.client.post(
            reverse("item-list"),
            {"name": "Invoice Widget", "sku": "INV-001", "unit_price": "20.00", "tax_rate": "0", "stock_quantity": 4},
            format="json",
        ).data
        invoice = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Sent", "items": [{"item": item["id"], "quantity": 2}]},
            format="json",
        ).data
        self.client.post(
            reverse("receipt-list"),
            {
                "invoice": invoice["id"],
                "amount_paid": "10.00",
                "payment_method": "Cash",
                "reference_number": "",
            },
            format="json",
        )

        list_res = self.client.get(f"{reverse('invoice-list')}?payment_status=partial&ordering=-amount_paid")
        self.assertEqual(list_res.status_code, status.HTTP_200_OK)
        self.assertEqual(list_res.data["results"][0]["customer_name"], "Invoice Buyer")
        self.assertEqual(list_res.data["results"][0]["payment_status"], "partial")
        self.assertEqual(Decimal(list_res.data["results"][0]["balance_due"]), Decimal("30.00"))

        detail_res = self.client.get(reverse("invoice-detail", args=[invoice["id"]]))
        self.assertEqual(detail_res.status_code, status.HTTP_200_OK)
        self.assertEqual(detail_res.data["customer_email"], "invoice@example.com")
        self.assertEqual(detail_res.data["line_item_count"], 1)
        self.assertEqual(detail_res.data["invoice_items"][0]["item_name"], "Invoice Widget")

    def test_receipt_detail_and_list_support_filters_sorting_and_linked_invoice_context(self):
        customer = self.client.post(reverse("customer-list"), {"name": "Receipt Buyer"}, format="json").data
        item = self.client.post(
            reverse("item-list"),
            {"name": "Receipt Widget", "sku": "RCT-001", "unit_price": "12.00", "tax_rate": "0", "stock_quantity": 5},
            format="json",
        ).data
        invoice = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Sent", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data
        receipt = self.client.post(
            reverse("receipt-list"),
            {
                "invoice": invoice["id"],
                "amount_paid": "12.00",
                "payment_date": "2026-06-30",
                "payment_method": "Bank Transfer",
                "reference_number": "BANK-123",
            },
            format="json",
        ).data

        list_res = self.client.get(f"{reverse('receipt-list')}?invoice_number={invoice['invoice_number']}&ordering=-payment_date")
        self.assertEqual(list_res.status_code, status.HTTP_200_OK)
        self.assertEqual(list_res.data["results"][0]["customer_name"], "Receipt Buyer")
        self.assertEqual(list_res.data["results"][0]["invoice_number"], invoice["invoice_number"])

        detail_res = self.client.get(reverse("receipt-detail", args=[receipt["id"]]))
        self.assertEqual(detail_res.status_code, status.HTTP_200_OK)
        self.assertEqual(detail_res.data["linked_invoice"]["invoice_number"], invoice["invoice_number"])
        self.assertEqual(detail_res.data["transaction_timestamp"], detail_res.data["updated_at"])

    def test_entity_read_views_reject_invalid_ordering_and_handle_missing_or_unauthorized_requests(self):
        customer = self.client.post(reverse("customer-list"), {"name": "Missing Target"}, format="json").data
        item = self.client.post(
            reverse("item-list"),
            {"name": "Missing Widget", "unit_price": "1.00", "stock_quantity": 1},
            format="json",
        ).data
        invoice = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data
        receipt = self.client.post(
            reverse("receipt-list"),
            {"invoice": invoice["id"], "amount_paid": "1.00", "payment_method": "Cash"},
            format="json",
        ).data

        bad_customer_sort = self.client.get(f"{reverse('customer-list')}?ordering=unknown")
        self.assertEqual(bad_customer_sort.status_code, status.HTTP_400_BAD_REQUEST)

        bad_receipt_sort = self.client.get(f"{reverse('receipt-list')}?ordering=unknown")
        self.assertEqual(bad_receipt_sort.status_code, status.HTTP_400_BAD_REQUEST)

        self.assertEqual(self.client.get(reverse("customer-detail", args=[999999])).status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(self.client.get(reverse("item-detail", args=[999999])).status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(self.client.get(reverse("invoice-detail", args=[999999])).status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(self.client.get(reverse("receipt-detail", args=[999999])).status_code, status.HTTP_404_NOT_FOUND)

        self.client.force_authenticate(user=None)
        self.assertEqual(self.client.get(reverse("customer-list")).status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(self.client.get(reverse("item-list")).status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(self.client.get(reverse("invoice-list")).status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(self.client.get(reverse("receipt-list")).status_code, status.HTTP_403_FORBIDDEN)

    def test_prevent_negative_stock_deduction(self):
        customer = self.client.post(
            reverse("customer-list"),
            {"name": "Buyer", "email": "buyer@example.com"},
            format="json",
        ).data

        item = self.client.post(
            reverse("item-list"),
            {"name": "Widget", "sku": "W-002", "unit_price": "10.00", "tax_rate": "0", "stock_quantity": 1},
            format="json",
        ).data

        invoice_res = self.client.post(
            reverse("invoice-list"),
            {
                "customer": customer["id"],
                "status": "Draft",
                "items": [{"item": item["id"], "quantity": 2}],
            },
            format="json",
        )
        self.assertEqual(invoice_res.status_code, status.HTTP_201_CREATED)
        invoice_id = invoice_res.data["id"]

        invoice_detail = reverse("invoice-detail", args=[invoice_id])
        self.client.force_authenticate(user=self.user)
        res = self.client.patch(invoice_detail, {"status": "Sent"}, format="json")
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Item.objects.get(pk=item["id"]).stock_quantity, 1)

    def test_dashboard_metrics(self):
        customer = self.client.post(
            reverse("customer-list"),
            {"name": "Buyer", "email": "buyer@example.com"},
            format="json",
        ).data

        item = self.client.post(
            reverse("item-list"),
            {"name": "Widget", "sku": "W-003", "unit_price": "10.00", "tax_rate": "0", "stock_quantity": 5},
            format="json",
        ).data

        invoice_res = self.client.post(
            reverse("invoice-list"),
            {
                "customer": customer["id"],
                "status": "Sent",
                "items": [{"item": item["id"], "quantity": 2}],
            },
            format="json",
        )
        self.assertEqual(invoice_res.status_code, status.HTTP_201_CREATED)

        dash_res = self.client.get(reverse("dashboard-list"))
        self.assertEqual(dash_res.status_code, status.HTTP_200_OK)
        self.assertEqual(dash_res.data["outstanding_invoices_count"], 1)

        receipt_res = self.client.post(
            reverse("receipt-list"),
            {"invoice": invoice_res.data["id"], "amount_paid": "20.00", "payment_method": "Cash"},
            format="json",
        )
        self.assertEqual(receipt_res.status_code, status.HTTP_201_CREATED)

        dash_res = self.client.get(reverse("dashboard-list"))
        self.assertEqual(dash_res.status_code, status.HTTP_200_OK)
        self.assertEqual(dash_res.data["outstanding_invoices_count"], 0)

    def test_reports_summary(self):
        customer = self.client.post(
            reverse("customer-list"),
            {"name": "Buyer", "email": "buyer@example.com"},
            format="json",
        ).data

        item = self.client.post(
            reverse("item-list"),
            {"name": "Widget", "sku": "W-004", "unit_price": "10.00", "tax_rate": "0", "stock_quantity": 5},
            format="json",
        ).data

        invoice_res = self.client.post(
            reverse("invoice-list"),
            {
                "customer": customer["id"],
                "status": "Sent",
                "items": [{"item": item["id"], "quantity": 2}],
            },
            format="json",
        )
        self.assertEqual(invoice_res.status_code, status.HTTP_201_CREATED)

        receipt_res = self.client.post(
            reverse("receipt-list"),
            {"invoice": invoice_res.data["id"], "amount_paid": "20.00", "payment_method": "Cash"},
            format="json",
        )
        self.assertEqual(receipt_res.status_code, status.HTTP_201_CREATED)

        res = self.client.get(reverse("reports-list"))
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn("revenue_total", res.data)
        self.assertIn("invoice_status", res.data)
        self.assertIn("top_items", res.data)

    def test_soft_delete_customer_blocked_when_invoices_exist(self):
        customer = Customer.objects.create(name="C1")
        Invoice.objects.create(invoice_number="INV-2026-9999", customer=customer)
        self.client.force_authenticate(user=self.user)
        res = self.client.delete(reverse("customer-detail", args=[customer.id]))
        self.assertEqual(res.status_code, status.HTTP_409_CONFLICT)
        customer.refresh_from_db()
        self.assertFalse(customer.is_deleted)

    def test_customer_delete_allows_access_token_role_user_without_django_model_perms(self):
        user2 = User.objects.create_user(username="u2", password=_test_secret())
        role = Role.objects.get(name="user")
        token_row = AccessToken.objects.create(user=user2, role=role, key="tok_u2_user")
        customer = Customer.objects.create(name="C-TOKEN")

        self.client.force_authenticate(user=None)
        res = self.client.delete(
            reverse("customer-detail", args=[customer.id]),
            HTTP_AUTHORIZATION=f"Token {token_row.key}",
        )
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)
        customer.refresh_from_db()
        self.assertTrue(customer.is_deleted)

    def test_customer_delete_denies_session_auth_without_model_permission(self):
        limited_role = Role.objects.create(name="limited", description="Limited")
        user2 = User.objects.create_user(username="u3", password=_test_secret())
        UserRole.objects.filter(user=user2).delete()
        UserRole.objects.create(user=user2, role=limited_role)

        customer = Customer.objects.create(name="C-NO-PERM")
        self.client.force_authenticate(user=user2)
        res = self.client.delete(reverse("customer-detail", args=[customer.id]))
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_invoice_soft_delete_cascades_invoice_items_and_receipts_and_audits(self):
        customer = Customer.objects.create(name="C2")
        item = Item.objects.create(name="W", unit_price=Decimal("10.00"), stock_quantity=10)
        inv = Invoice.objects.create(invoice_number="INV-2026-8888", customer=customer, status="Draft")
        InvoiceItem.objects.create(invoice=inv, item=item, quantity=1, unit_price=item.unit_price, line_total=Decimal("10.00"))
        Receipt.objects.create(invoice=inv, amount_paid=Decimal("5.00"), payment_method="Cash")

        self.client.force_authenticate(user=self.user)
        res = self.client.delete(reverse("invoice-detail", args=[inv.id]))
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)
        inv.refresh_from_db()
        self.assertTrue(inv.is_deleted)
        self.assertEqual(InvoiceItem.objects.filter(invoice=inv, is_deleted=False).count(), 0)
        self.assertEqual(Receipt.objects.filter(invoice=inv, is_deleted=False).count(), 0)
        self.assertTrue(AuditLog.objects.filter(action="delete", object_id=str(inv.id)).exists())

    def test_receipt_update_recomputes_invoice_paid_status(self):
        customer = Customer.objects.create(name="C3")
        item = Item.objects.create(name="W", unit_price=Decimal("10.00"), stock_quantity=10)
        inv = Invoice.objects.create(invoice_number="INV-2026-7777", customer=customer, status="Sent", subtotal=Decimal("10.00"), total_amount=Decimal("10.00"))
        r = Receipt.objects.create(invoice=inv, amount_paid=Decimal("10.00"), payment_method="Cash")
        inv.refresh_from_db()
        self.assertEqual(inv.status, "Sent")

        self.client.force_authenticate(user=self.user)
        update_res = self.client.patch(
            reverse("receipt-detail", args=[r.id]),
            {"amount_paid": "10.00"},
            format="json",
        )
        self.assertEqual(update_res.status_code, status.HTTP_200_OK)
        inv.refresh_from_db()
        self.assertEqual(inv.status, "Paid")

        update_res = self.client.patch(
            reverse("receipt-detail", args=[r.id]),
            {"amount_paid": "1.00"},
            format="json",
        )
        self.assertEqual(update_res.status_code, status.HTTP_200_OK)
        inv.refresh_from_db()
        self.assertEqual(inv.status, "Sent")

    def test_evaluate_invoice_payment_status_paid_when_payments_cover_total(self):
        self.assertEqual(evaluate_invoice_payment_status("10.00", "10.00"), "paid")
        self.assertEqual(evaluate_invoice_payment_status(Decimal("10.00"), Decimal("10.01")), "paid")
        self.assertEqual(evaluate_invoice_payment_status("0.00", None), "paid")

    def test_evaluate_invoice_payment_status_bal_when_outstanding(self):
        self.assertEqual(evaluate_invoice_payment_status("10.00", "9.99"), "bal")
        self.assertEqual(evaluate_invoice_payment_status("10.00", None), "bal")

    def test_evaluate_invoice_payment_status_invalid_inputs(self):
        with self.assertRaises(ValueError):
            evaluate_invoice_payment_status(None, "0.00")
        with self.assertRaises(ValueError):
            evaluate_invoice_payment_status("-1.00", "0.00")
        with self.assertRaises(ValueError):
            evaluate_invoice_payment_status("10.00", "-0.01")
        with self.assertRaises(ValueError):
            evaluate_invoice_payment_status("abc", "0.00")
        with self.assertRaises(ValueError):
            evaluate_invoice_payment_status("10.00", "NaN")

    def test_resolve_invoice_discount_and_context_service(self):
        usd, _ = Currency.objects.get_or_create(
            code="USD",
            defaults={"name": "US Dollar", "symbol": "$", "decimal_places": 2},
        )
        discount_type, discount_value, discount_amount = resolve_invoice_discount(
            Decimal("200.00"),
            Invoice.DISCOUNT_TYPE_PERCENTAGE,
            "12.5",
        )
        self.assertEqual(discount_type, Invoice.DISCOUNT_TYPE_PERCENTAGE)
        self.assertEqual(discount_value, Decimal("12.50"))
        self.assertEqual(discount_amount, Decimal("25.00"))

        invoice = Invoice(
            invoice_number="INV-DISCOUNT-SVC",
            customer=Customer(name="Svc Customer"),
            subtotal=Decimal("200.00"),
            discount_type=discount_type,
            discount_value=discount_value,
            discount_amount=discount_amount,
            tax_total=Decimal("0.00"),
            total_amount=Decimal("175.00"),
        )
        context = invoice_discount_context(
            invoice,
            usd,
            "1,234.56",
            "prefix",
            format_money=format_money,
        )
        self.assertTrue(context["has_discount"])
        self.assertEqual(context["discount_value_label"], "12.5%")
        self.assertEqual(context["discount_amount_fmt"], "$25.00")

    def test_payment_sync_service_updates_invoice_status(self):
        customer = Customer.objects.create(name="Svc Status Customer")
        invoice = Invoice.objects.create(
            invoice_number="INV-SYNC-1",
            customer=customer,
            status="Paid",
            subtotal=Decimal("10.00"),
            total_amount=Decimal("10.00"),
        )
        self.assertEqual(outstanding_invoice_amount("10.00", "12.00"), Decimal("0.00"))
        self.assertEqual(sync_invoice_status_with_payments(invoice, Decimal("1.00")), "Sent")
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, "Sent")
        self.assertEqual(sync_invoice_status_with_payments(invoice, Decimal("10.00")), "Paid")
        invoice.refresh_from_db()
        self.assertEqual(invoice.status, "Paid")

    def test_rendering_services_resolve_identity_templates_and_region(self):
        profile, _ = UserProfile.objects.get_or_create(user=self.user)
        profile.company_legal_name = "Profile Company"
        profile.save(update_fields=["company_legal_name", "updated_at"])
        self.assertEqual(profile.company_legal_name, "Profile Company")
        identity = effective_company_identity_for_user(self.user)
        self.assertEqual(identity["company_name"], "Profile Company")

        templates = effective_templates_for_user(self.user)
        self.assertEqual(templates["global_appearance"]["company_name"], "Profile Company")
        self.assertEqual(templates["invoice_template"]["layout"], "classic")

        request_stub = type(
            "Req",
            (),
            {
                "user": self.user,
                "query_params": {},
                "headers": {"accept-language": "de-DE,de;q=0.9"},
            },
        )()
        self.assertEqual(detect_country(request_stub), "DE")

        region = effective_region_settings_for_user(request_stub, self.user, detect_country=detect_country)
        self.assertEqual(region["country"], "DE")
        self.assertEqual(region["currency_code"], "NGN")
        self.assertEqual(region["date_format"], "YYYY-MM-DD")
        self.assertEqual(region["number_format"], "1,234.56")

    def test_item_delete_blocked_when_referenced_by_active_invoice(self):
        customer = self.client.post(reverse("customer-list"), {"name": "B"}, format="json").data
        item = self.client.post(
            reverse("item-list"),
            {"name": "Widget", "sku": "WX", "unit_price": "10.00", "tax_rate": "0", "stock_quantity": 5},
            format="json",
        ).data
        inv = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data
        self.assertIsNotNone(inv["id"])

        self.client.force_authenticate(user=self.user)
        res = self.client.delete(reverse("item-detail", args=[item["id"]]))
        self.assertEqual(res.status_code, status.HTTP_409_CONFLICT)

    def test_bulk_delete_blocks_customers_with_invoices(self):
        c1 = self.client.post(reverse("customer-list"), {"name": "C1"}, format="json").data
        c2 = self.client.post(reverse("customer-list"), {"name": "C2"}, format="json").data
        Invoice.objects.create(invoice_number="INV-2026-1111", customer_id=c1["id"])

        self.client.force_authenticate(user=self.user)
        res = self.client.post(reverse("customer-bulk-delete"), {"ids": [c1["id"], c2["id"]]}, format="json")
        self.assertEqual(res.status_code, status.HTTP_409_CONFLICT)
        self.assertIn("blocked_ids", res.data)

    def test_bulk_delete_receipts_recomputes_invoice_status(self):
        customer = Customer.objects.create(name="C4")
        inv = Invoice.objects.create(
            invoice_number="INV-2026-2222",
            customer=customer,
            status="Paid",
            subtotal=Decimal("10.00"),
            total_amount=Decimal("10.00"),
        )
        r1 = Receipt.objects.create(invoice=inv, amount_paid=Decimal("6.00"), payment_method="Cash")
        r2 = Receipt.objects.create(invoice=inv, amount_paid=Decimal("4.00"), payment_method="Cash")

        self.client.force_authenticate(user=self.user)
        res = self.client.post(reverse("receipt-bulk-delete"), {"ids": [r1.id]}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        inv.refresh_from_db()
        self.assertEqual(inv.status, "Sent")

        res = self.client.post(reverse("receipt-bulk-delete"), {"ids": [r2.id]}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        inv.refresh_from_db()
        self.assertEqual(inv.status, "Sent")

    def test_concurrency_conflict_returns_409(self):
        cust = self.client.post(reverse("customer-list"), {"name": "C"}, format="json").data
        detail_url = reverse("customer-detail", args=[cust["id"]])
        self.client.force_authenticate(user=self.user)
        res = self.client.patch(detail_url, {"phone": "1", "updated_at": "2000-01-01T00:00:00Z"}, format="json")
        self.assertEqual(res.status_code, status.HTTP_409_CONFLICT)

    def test_rbac_denies_without_permission(self):
        limited_role = Role.objects.create(name="limited2", description="Limited2")
        user2 = User.objects.create_user(username="u2", password=_test_secret())
        UserRole.objects.filter(user=user2).delete()
        UserRole.objects.create(user=user2, role=limited_role)
        cust = self.client.post(reverse("customer-list"), {"name": "C"}, format="json").data
        detail_url = reverse("customer-detail", args=[cust["id"]])
        self.client.force_authenticate(user=user2)
        res = self.client.patch(detail_url, {"phone": "1"}, format="json")
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_item_tax_rate_validation(self):
        res = self.client.post(
            reverse("item-list"),
            {"name": "BadTax", "unit_price": "1.00", "tax_rate": "200", "stock_quantity": 1},
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_item_category_defaults_and_blank_category_is_rejected(self):
        created = self.client.post(
            reverse("item-list"),
            {"name": "Default Category Item", "unit_price": "1.00", "stock_quantity": 1},
            format="json",
        )
        self.assertEqual(created.status_code, status.HTTP_201_CREATED)
        self.assertEqual(created.data["category"], "General")

        denied = self.client.post(
            reverse("item-list"),
            {"name": "Blank Category Item", "category": "   ", "unit_price": "1.00", "stock_quantity": 1},
            format="json",
        )
        self.assertEqual(denied.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("category", denied.data)

    def test_service_cannot_have_stock(self):
        res = self.client.post(
            reverse("item-list"),
            {"type": "service", "name": "Consulting", "unit_price": "100.00", "stock_quantity": 1},
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_adjust_stock_success(self):
        item = self.client.post(
            reverse("item-list"),
            {"name": "Widget", "unit_price": "1.00", "stock_quantity": 1},
            format="json",
        ).data
        self.client.force_authenticate(user=self.user)
        res = self.client.post(reverse("item-adjust-stock", args=[item["id"]]), {"adjustment": 2}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data["new_stock_quantity"], 3)

    def test_bulk_delete_items_success(self):
        i1 = self.client.post(reverse("item-list"), {"name": "A", "unit_price": "1.00", "stock_quantity": 1}, format="json").data
        i2 = self.client.post(reverse("item-list"), {"name": "B", "unit_price": "1.00", "stock_quantity": 1}, format="json").data
        self.client.force_authenticate(user=self.user)
        res = self.client.post(reverse("item-bulk-delete"), {"ids": [i1["id"], i2["id"]]}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(Item.objects.filter(is_deleted=False).count(), 0)

    def test_bulk_delete_invoices_success(self):
        customer = self.client.post(reverse("customer-list"), {"name": "B"}, format="json").data
        item = self.client.post(reverse("item-list"), {"name": "W", "unit_price": "1.00", "stock_quantity": 10}, format="json").data
        inv1 = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data
        inv2 = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data
        self.client.force_authenticate(user=self.user)
        res = self.client.post(reverse("invoice-bulk-delete"), {"ids": [inv1["id"], inv2["id"]]}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(Invoice.objects.filter(is_deleted=False).count(), 0)

    def test_receipt_amount_paid_validation(self):
        customer = self.client.post(reverse("customer-list"), {"name": "B"}, format="json").data
        item = self.client.post(reverse("item-list"), {"name": "W", "unit_price": "1.00", "stock_quantity": 1}, format="json").data
        inv = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data
        res = self.client.post(
            reverse("receipt-list"),
            {"invoice": inv["id"], "amount_paid": "0", "payment_method": "Cash"},
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_invoice_pay_success_and_idempotency(self):
        customer = self.client.post(reverse("customer-list"), {"name": "PayC"}, format="json").data
        item = self.client.post(reverse("item-list"), {"name": "PayW", "unit_price": "1.00", "stock_quantity": 1}, format="json").data
        inv = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data

        self.client.force_authenticate(user=self.user)
        url = reverse("invoice-pay", args=[inv["id"]])
        res1 = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Cash"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-1",
        )
        self.assertEqual(res1.status_code, status.HTTP_201_CREATED)
        receipt_id = res1.data["id"]

        res2 = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Cash"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-1",
        )
        self.assertEqual(res2.status_code, status.HTTP_200_OK)
        self.assertEqual(res2.data["id"], receipt_id)

    def test_invoice_pay_declined_and_gateway_errors(self):
        customer = self.client.post(reverse("customer-list"), {"name": "PayC2"}, format="json").data
        item = self.client.post(reverse("item-list"), {"name": "PayW2", "unit_price": "1.00", "stock_quantity": 1}, format="json").data
        inv = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data

        self.client.force_authenticate(user=self.user)
        url = reverse("invoice-pay", args=[inv["id"]])
        declined = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Card", "reference_number": "DECLINE"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-2",
        )
        self.assertEqual(declined.status_code, 402)

        unavailable = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Card", "reference_number": "NETWORK"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-3",
        )
        self.assertEqual(unavailable.status_code, 503)

        timeout = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Card", "reference_number": "TIMEOUT"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-4",
        )
        self.assertEqual(timeout.status_code, 504)

    def test_invoice_pay_supports_transaction_date_and_bank_transfer_reference_validation(self):
        customer = self.client.post(reverse("customer-list"), {"name": "PayC4"}, format="json").data
        item = self.client.post(reverse("item-list"), {"name": "PayW4", "unit_price": "1.00", "stock_quantity": 1}, format="json").data
        inv = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data

        self.client.force_authenticate(user=self.user)
        url = reverse("invoice-pay", args=[inv["id"]])

        missing_ref = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Bank Transfer", "payment_date": "2026-05-01"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-bt-1",
        )
        self.assertEqual(missing_ref.status_code, status.HTTP_400_BAD_REQUEST)

        ok = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Bank Transfer", "reference_number": "TRX-123", "payment_date": "2026-05-01"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-bt-2",
        )
        self.assertEqual(ok.status_code, status.HTTP_201_CREATED)
        r = Receipt.objects.get(pk=ok.data["id"])
        self.assertEqual(str(r.payment_date), "2026-05-01")

    def test_invoice_pay_rejects_card_number_in_reference_and_handles_card_edge_cases(self):
        customer = self.client.post(reverse("customer-list"), {"name": "PayC5"}, format="json").data
        item = self.client.post(reverse("item-list"), {"name": "PayW5", "unit_price": "1.00", "stock_quantity": 1}, format="json").data
        inv = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data

        self.client.force_authenticate(user=self.user)
        url = reverse("invoice-pay", args=[inv["id"]])

        bad_ref = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Card", "reference_number": "4111 1111 1111 1111"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-card-1",
        )
        self.assertEqual(bad_ref.status_code, status.HTTP_400_BAD_REQUEST)

        invalid = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Card", "reference_number": "INVALID"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-card-2",
        )
        self.assertEqual(invalid.status_code, 402)

        expired = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Card", "reference_number": "EXPIRED"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-card-3",
        )
        self.assertEqual(expired.status_code, 402)

        insufficient = self.client.post(
            url,
            {"amount_paid": "1.00", "payment_method": "Card", "reference_number": "INSUFFICIENT"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-card-4",
        )
        self.assertEqual(insufficient.status_code, 402)

    def test_payments_report_api(self):
        customer = Customer.objects.create(name="RC")
        inv = Invoice.objects.create(invoice_number="INV-2026-RPT", customer=customer, status="Sent", subtotal=Decimal("10.00"), total_amount=Decimal("10.00"))
        Receipt.objects.create(invoice=inv, amount_paid=Decimal("3.00"), payment_method="Cash", payment_date=date(2026, 5, 1))
        Receipt.objects.create(invoice=inv, amount_paid=Decimal("7.00"), payment_method="Bank Transfer", reference_number="BT-1", payment_date=date(2026, 5, 2))
        PaymentTransaction.objects.create(invoice=inv, created_by=self.user, provider="paystack", status="pending", amount=Decimal("10.00"), currency_code="NGN", reference="REF-RPT-1")

        self.client.force_authenticate(user=self.user)
        res = self.client.get(reverse("payments-report") + "?start=2026-05-01&end=2026-05-31")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        methods = {r["payment_method"]: r for r in res.data["receipts_by_method"]}
        self.assertIn("Cash", methods)
        self.assertIn("Bank Transfer", methods)

    @override_settings(BANK_TRANSFER_RECONCILIATION_URL="http://bank.example/reconcile", BANK_TRANSFER_RECONCILIATION_SECRET="s")
    @patch("core.views.urllib.request.urlopen")
    def test_bank_transfer_reconciliation_network_failure_is_recorded(self, urlopen_mock):
        from urllib.error import URLError

        urlopen_mock.side_effect = URLError("down")
        customer = Customer.objects.create(name="BC")
        inv = Invoice.objects.create(invoice_number="INV-2026-BANK", customer=customer, status="Sent", subtotal=Decimal("10.00"), total_amount=Decimal("10.00"))

        self.client.force_authenticate(user=self.user)
        res = self.client.post(
            reverse("payment-transaction-list"),
            {"provider": "bank_transfer", "invoice": inv.id, "amount": "10.00", "currency_code": "NGN"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-bank-1",
        )
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        tx = PaymentTransaction.objects.get(pk=res.data["id"])
        self.assertEqual(tx.status, "pending")
        self.assertIn("reconciliation_error", tx.metadata or {})

    def test_invoice_pay_overpayment_rejected(self):
        customer = self.client.post(reverse("customer-list"), {"name": "PayC3"}, format="json").data
        item = self.client.post(reverse("item-list"), {"name": "PayW3", "unit_price": "1.00", "stock_quantity": 1}, format="json").data
        inv = self.client.post(
            reverse("invoice-list"),
            {"customer": customer["id"], "status": "Draft", "items": [{"item": item["id"], "quantity": 1}]},
            format="json",
        ).data

        self.client.force_authenticate(user=self.user)
        url = reverse("invoice-pay", args=[inv["id"]])
        res = self.client.post(
            url,
            {"amount_paid": "2.00", "payment_method": "Cash"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-pay-5",
        )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_finance_overview_and_activity_and_top_products(self):
        today = timezone.localdate()
        customer = Customer.objects.create(name="C5")
        item = Item.objects.create(name="Widget", unit_price=Decimal("10.00"), stock_quantity=10)
        inv = Invoice.objects.create(
            invoice_number="INV-2026-3333",
            customer=customer,
            status="Sent",
            issue_date=today,
            subtotal=Decimal("10.00"),
            tax_total=Decimal("0.00"),
            total_amount=Decimal("10.00"),
        )
        InvoiceItem.objects.create(
            invoice=inv,
            item=item,
            quantity=2,
            unit_price=Decimal("10.00"),
            line_subtotal=Decimal("20.00"),
            line_tax=Decimal("0.00"),
            line_total=Decimal("20.00"),
        )
        Receipt.objects.create(invoice=inv, amount_paid=Decimal("10.00"), payment_date=today, payment_method="Cash")
        Expense.objects.create(amount=Decimal("3.00"), expense_date=today, category="Office", description="Paper")

        res = self.client.get(reverse("finance-list") + "?period=1m")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn("income_total", res.data)
        self.assertIn("expense_total", res.data)
        self.assertIn("points", res.data)

        activity_all = self.client.get(reverse("finance-activity") + "?type=all&limit=15")
        self.assertEqual(activity_all.status_code, status.HTTP_200_OK)
        self.assertGreaterEqual(len(activity_all.data["events"]), 2)

        activity_income = self.client.get(reverse("finance-activity") + "?type=income&limit=15")
        self.assertEqual(activity_income.status_code, status.HTTP_200_OK)
        self.assertTrue(all(e["type"] == "income" for e in activity_income.data["events"]))

        top = self.client.get(reverse("finance-top-products") + "?period=1m")
        self.assertEqual(top.status_code, status.HTTP_200_OK)
        self.assertIn("products", top.data)
        self.assertGreaterEqual(len(top.data["products"]), 1)

    def test_finance_overview_returns_zero_expenses_and_reflects_expense_updates(self):
        today = timezone.localdate()
        initial = self.client.get(reverse("finance-list") + "?period=1m")
        self.assertEqual(initial.status_code, status.HTTP_200_OK)
        self.assertEqual(initial.data["expense_total"], "0.00")
        self.assertTrue(all(point["expense"] == "0.00" for point in initial.data["points"]))

        expense = Expense.objects.create(amount=Decimal("12.50"), expense_date=today, category="Office", description="Printer paper")
        after_create = self.client.get(reverse("finance-list") + "?period=1m")
        self.assertEqual(after_create.status_code, status.HTTP_200_OK)
        self.assertEqual(after_create.data["expense_total"], "12.50")
        self.assertIn("12.50", {point["expense"] for point in after_create.data["points"]})

        expense.amount = Decimal("18.75")
        expense.save(update_fields=["amount", "updated_at"])
        after_update = self.client.get(reverse("finance-list") + "?period=1m")
        self.assertEqual(after_update.status_code, status.HTTP_200_OK)
        self.assertEqual(after_update.data["expense_total"], "18.75")
        self.assertIn("18.75", {point["expense"] for point in after_update.data["points"]})


class SettingsTests(APITestCase):
    def setUp(self):
        super().setUp()
        self.admin = User.objects.create_user(username="admin", password=_test_secret(), is_staff=True, is_superuser=True)
        self.staff = User.objects.create_user(username="staff", password=_test_secret(), is_staff=True)
        self.user = User.objects.create_user(username="u1", password=_test_secret())

    def test_country_defaults(self):
        res = self.client.get("/api/settings/country-defaults/?country=DE")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data["country"], "DE")
        self.assertEqual(res.data["defaults"]["currency"], "EUR")

        bad = self.client.get("/api/settings/country-defaults/?country=ZZ")
        self.assertEqual(bad.status_code, status.HTTP_400_BAD_REQUEST)

    def test_currency_conversion_direct_and_inverse(self):
        ExchangeRate.objects.create(base_code="USD", quote_code="EUR", rate=Decimal("0.9"))
        res = self.client.get("/api/settings/convert/?amount=10&from=USD&to=EUR")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data["converted"], "9.00")

        ExchangeRate.objects.create(base_code="GBP", quote_code="JPY", rate=Decimal("200"))
        inv = self.client.get("/api/settings/convert/?amount=200&from=JPY&to=GBP")
        self.assertEqual(inv.status_code, status.HTTP_200_OK)
        self.assertEqual(inv.data["converted"], "1.00")

    def test_global_settings_admin_only_and_audited(self):
        self.client.force_authenticate(user=self.user)
        res = self.client.get("/api/settings/global/")
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(user=self.staff)
        staff_get = self.client.get("/api/settings/global/")
        self.assertEqual(staff_get.status_code, status.HTTP_200_OK)
        self.assertIsNone(staff_get.data.get("tax_identification_number"))

        cur = Currency.objects.get(code="USD")
        put_denied = self.client.put(
            "/api/settings/global/",
            {"default_currency": cur.id, "allow_user_overrides": True, "tax_configuration": {"default_rate": "5"}, "appearance": {"company_name": "X"}},
            format="json",
        )
        self.assertEqual(put_denied.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(user=self.admin)
        cur = Currency.objects.get(code="USD")
        put = self.client.put(
            "/api/settings/global/",
            {
                "default_currency": cur.id,
                "allow_user_overrides": True,
                "tax_configuration": {"default_rate": "5"},
                "appearance": {"company_name": "X"},
                "tax_identification_number": "TIN-ADMIN-123",
            },
            format="json",
        )
        self.assertEqual(put.status_code, status.HTTP_200_OK)
        gs = GlobalSettings.objects.get(singleton_key="global")
        self.assertEqual(gs.default_currency_id, cur.id)
        self.assertEqual(gs.tax_identification_number, "TIN-ADMIN-123")
        self.assertTrue(AuditLog.objects.filter(object_id=str(gs.id), action="update").exists())

    def test_superuser_without_userrole_still_has_admin_access(self):
        UserRole.objects.filter(user=self.admin).delete()

        self.client.force_authenticate(user=self.admin)
        me = self.client.get("/api/auth/me/")
        self.assertEqual(me.status_code, status.HTTP_200_OK)
        self.assertIn("admin", me.data.get("roles", []))
        self.assertIn("settings.global.write", me.data.get("permissions", []))

        global_get = self.client.get("/api/settings/global/")
        self.assertEqual(global_get.status_code, status.HTTP_200_OK)

        admin_users = self.client.get("/api/admin/users/?page=1")
        self.assertEqual(admin_users.status_code, status.HTTP_200_OK)

    def test_effective_settings_hides_tax_id_for_non_admin(self):
        gs = GlobalSettings.objects.get_or_create(singleton_key="global")[0]
        gs.tax_identification_number = "SECRET-TIN"
        gs.save(update_fields=["tax_identification_number"])

        res = self.client.get("/api/settings/effective/")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIsNone((res.data.get("global") or {}).get("tax_identification_number"))

        self.client.force_authenticate(user=self.admin)
        res2 = self.client.get("/api/settings/effective/")
        self.assertEqual(res2.status_code, status.HTTP_200_OK)
        self.assertEqual((res2.data.get("global") or {}).get("tax_identification_number"), "SECRET-TIN")

    def test_user_settings_update_and_rollback(self):
        self.client.force_authenticate(user=self.user)
        me = self.client.get("/api/settings/me/")
        self.assertEqual(me.status_code, status.HTTP_200_OK)
        settings_id = me.data["id"]

        patch = self.client.patch("/api/settings/me/", {"date_format": "DD/MM/YYYY"}, format="json")
        self.assertEqual(patch.status_code, status.HTTP_200_OK)
        us = UserSettings.objects.get(pk=settings_id)
        self.assertEqual(us.date_format, "DD/MM/YYYY")

        log = AuditLog.objects.filter(object_id=str(us.id), action="update").order_by("-id").first()
        self.assertIsNotNone(log)
        rollback = self.client.post("/api/settings/rollback/", {"audit_log_id": log.id}, format="json")
        self.assertEqual(rollback.status_code, status.HTTP_200_OK)
        us.refresh_from_db()
        self.assertNotEqual(us.date_format, "DD/MM/YYYY")

    def test_effective_settings_includes_user_when_authenticated(self):
        self.client.force_authenticate(user=self.user)
        res = self.client.get("/api/settings/effective/")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIsNotNone(res.data["user"])

    def test_effective_settings_load_many_times(self):
        self.client.force_authenticate(user=self.user)
        for _ in range(50):
            res = self.client.get("/api/settings/effective/")
            self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_effective_settings_uses_global_currency_by_default(self):
        res = self.client.get("/api/settings/effective/", HTTP_X_COUNTRY_CODE="US")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data["effective"]["currency_code"], "NGN")


class AdminLogoUploadTests(APITestCase):
    def setUp(self):
        super().setUp()
        self.admin = User.objects.create_user(username="admin_logo", password=_test_secret(), is_staff=True, is_superuser=True)
        self.user = User.objects.create_user(username="u_logo", password=_test_secret())

    def test_non_admin_cannot_upload_logo(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/"):
                file = _png_logo_upload_file()
                self.client.force_authenticate(user=self.user)
                res = self.client.post("/api/admin/logo/upload/", {"file": file}, format="multipart")
                self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_upload_png_logo_creates_thumbnail_and_updates_global_settings(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/"):
                file = _png_logo_upload_file()
                self.client.force_authenticate(user=self.admin)
                res = self.client.post("/api/admin/logo/upload/", {"file": file}, format="multipart")
                self.assertEqual(res.status_code, status.HTTP_201_CREATED)
                self.assertTrue(str(res.data.get("logo_url", "")).startswith("/media/uploads/logos/"))
                self.assertTrue(str(res.data.get("thumbnail_url", "")).startswith("/media/uploads/logos/"))
                self.assertEqual(res.data.get("scope"), "global_appearance")
                self.assertEqual(len(str(res.data.get("sha256") or "")), 64)

                logo_path = str(res.data["logo_url"]).replace("/media/", "", 1)
                thumb_path = str(res.data["thumbnail_url"]).replace("/media/", "", 1)
                self.assertTrue(os.path.exists(os.path.join(tmp, logo_path)))
                self.assertTrue(os.path.exists(os.path.join(tmp, thumb_path)))
                gs = GlobalSettings.objects.get(singleton_key="global")
                self.assertIsNotNone(gs.appearance_logo_id)
                self.assertTrue(LogoAsset.objects.filter(pk=gs.appearance_logo_id, scope="global_appearance").exists())

    def test_upload_rejects_unsupported_type(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/"):
                file = SimpleUploadedFile("logo.gif", b"GIF89a", content_type="image/gif")
                self.client.force_authenticate(user=self.admin)
                res = self.client.post("/api/admin/logo/upload/", {"file": file}, format="multipart")
                self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_upload_accepts_webp_logo(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/"):
                file = _webp_logo_upload_file()
                self.client.force_authenticate(user=self.admin)
                res = self.client.post("/api/admin/logo/upload/", {"file": file}, format="multipart")
                self.assertEqual(res.status_code, status.HTTP_201_CREATED)
                self.assertIn(".webp", str(res.data.get("logo_url", "")))
                self.assertTrue(str(res.data.get("thumbnail_url", "")).endswith(".png"))

    def test_upload_rejects_oversized_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/"):
                file = SimpleUploadedFile("huge.png", b"\x89PNG\r\n\x1a\n" + (b"x" * (5 * 1024 * 1024)), content_type="image/png")
                self.client.force_authenticate(user=self.admin)
                res = self.client.post("/api/admin/logo/upload/", {"file": file}, format="multipart")
                self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
                self.assertIn("5MB", str(res.data))

    def test_upload_rejects_excessive_pixels(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/", LOGO_UPLOAD_MAX_PIXELS=10_000):
                buf = io.BytesIO()
                Image.new("RGB", (200, 200), (10, 10, 10)).save(buf, format="PNG")
                file = SimpleUploadedFile("logo.png", buf.getvalue(), content_type="image/png")
                self.client.force_authenticate(user=self.admin)
                res = self.client.post("/api/admin/logo/upload/", {"file": file}, format="multipart")
                self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_upload_rejects_svg_with_script(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/"):
                svg = b'<svg xmlns="http://www.w3.org/2000/svg"><script>alert(1)</script></svg>'
                file = SimpleUploadedFile("logo.svg", svg, content_type="image/svg+xml")
                self.client.force_authenticate(user=self.admin)
                res = self.client.post("/api/admin/logo/upload/", {"file": file}, format="multipart")
                self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)


class SettingsLogoUploadTests(APITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(username="u_settings_logo", password=_test_secret())

    def test_user_can_upload_invoice_logo_and_effective_settings_use_it(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/"):
                self.client.force_authenticate(user=self.user)
                res = self.client.post(
                    "/api/settings/logo/upload/",
                    {"file": _png_logo_upload_file(), "scope": "invoice_template"},
                    format="multipart",
                )
                self.assertEqual(res.status_code, status.HTTP_201_CREATED)
                us = UserSettings.objects.get(user=self.user)
                self.assertIsNotNone(us.invoice_logo_id)
                effective = self.client.get("/api/settings/effective/")
                self.assertEqual(effective.status_code, status.HTTP_200_OK)
                self.assertEqual(effective.data["user"]["invoice_template"]["logo_url"], res.data["logo_url"])
                self.assertEqual(effective.data["effective"]["templates"]["invoice_template"]["logo_url"], res.data["logo_url"])

    def test_user_can_upload_receipt_logo(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/"):
                self.client.force_authenticate(user=self.user)
                res = self.client.post(
                    "/api/settings/logo/upload/",
                    {"file": _webp_logo_upload_file(), "scope": "receipt_template"},
                    format="multipart",
                )
                self.assertEqual(res.status_code, status.HTTP_201_CREATED)
                us = UserSettings.objects.get(user=self.user)
                self.assertIsNotNone(us.receipt_logo_id)
                self.assertTrue(str(res.data["logo_url"]).endswith(".webp"))

    def test_uploaded_logos_render_in_invoice_and_receipt_documents(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/", BACKEND_PUBLIC_BASE_URL="https://backend.example"):
                self.client.force_authenticate(user=self.user)
                invoice_logo = self.client.post(
                    "/api/settings/logo/upload/",
                    {"file": _png_logo_upload_file("invoice-logo.png"), "scope": "invoice_template"},
                    format="multipart",
                )
                receipt_logo = self.client.post(
                    "/api/settings/logo/upload/",
                    {"file": _webp_logo_upload_file("receipt-logo.webp"), "scope": "receipt_template"},
                    format="multipart",
                )
                self.assertEqual(invoice_logo.status_code, status.HTTP_201_CREATED)
                self.assertEqual(receipt_logo.status_code, status.HTTP_201_CREATED)
                patch_res = self.client.patch(
                    "/api/settings/me/",
                    {
                        "invoice_template": {
                            "layout": "compact",
                            "show_item_description": True,
                            "footer_text": "Compact invoice footer",
                        },
                        "receipt_template": {
                            "layout": "compact",
                            "show_items": False,
                            "show_item_description": True,
                            "header_text": "Compact Receipt",
                            "footer_text": "Compact receipt footer",
                        },
                    },
                    format="json",
                )
                self.assertEqual(patch_res.status_code, status.HTTP_200_OK)

                customer = Customer.objects.create(name="Logo Buyer", email="logo@example.com")
                item = Item.objects.create(name="Logo Widget", description="Visible description", unit_price=Decimal("10.00"), stock_quantity=10)
                invoice = Invoice.objects.create(
                    invoice_number="INV-2026-LOGO",
                    customer=customer,
                    status="Draft",
                    issue_date=timezone.localdate(),
                    subtotal=Decimal("10.00"),
                    tax_total=Decimal("0.00"),
                    total_amount=Decimal("10.00"),
                )
                InvoiceItem.objects.create(
                    invoice=invoice,
                    item=item,
                    quantity=1,
                    unit_price=Decimal("10.00"),
                    line_subtotal=Decimal("10.00"),
                    line_tax=Decimal("0.00"),
                    line_total=Decimal("10.00"),
                )
                receipt = Receipt.objects.create(
                    invoice=invoice,
                    amount_paid=Decimal("10.00"),
                    payment_method="Cash",
                    reference_number="LOGO-RCPT-1",
                )

                request_stub = type("Req", (), {"user": self.user, "query_params": {}, "headers": {}})()
                rendered_invoice = render_invoice(request_stub, invoice, "html").content.decode("utf-8")
                rendered_receipt = render_receipt(request_stub, receipt, "html").content.decode("utf-8")
                self.assertIn('class="invoice-layout-compact"', rendered_invoice)
                self.assertIn('class="receipt-layout-compact"', rendered_receipt)
                self.assertIn("data:image/png;base64,", rendered_invoice)
                self.assertIn("data:image/webp;base64,", rendered_receipt)
                self.assertIn("Visible description", rendered_invoice)
                self.assertIn("Compact invoice footer", rendered_invoice)
                self.assertIn("Compact Receipt", rendered_receipt)
                self.assertIn("Compact receipt footer", rendered_receipt)
                self.assertNotIn("<table>", rendered_receipt)

                receipt_html = self.client.get(f"/api/receipts/{receipt.id}/print_html/")
                self.assertEqual(receipt_html.status_code, status.HTTP_200_OK)
                self.assertContains(receipt_html, 'class="receipt-layout-compact"')
                self.assertContains(receipt_html, "data:image/webp;base64,")

    def test_user_settings_patch_ignores_manual_logo_url(self):
        self.client.force_authenticate(user=self.user)
        res = self.client.patch(
            "/api/settings/me/",
            {"invoice_template": {"logo_url": "https://evil.example/logo.png", "footer_text": "hello"}},
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        us = UserSettings.objects.get(user=self.user)
        self.assertEqual(us.invoice_template.get("footer_text"), "hello")
        self.assertNotIn("logo_url", us.invoice_template)

    def test_user_settings_patch_blocks_currency_and_template_overrides_when_disabled(self):
        self.client.force_authenticate(user=self.user)
        gs = GlobalSettings.objects.get_or_create(singleton_key="global")[0]
        gs.allow_user_overrides = False
        gs.save(update_fields=["allow_user_overrides", "updated_at"])

        res = self.client.patch(
            "/api/settings/me/",
            {"currency": None, "invoice_template": {"footer_text": "blocked"}},
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("disabled", str(res.data).lower())
        self.assertIn("invoice_template", str(res.data.get("fields", [])))

    def test_user_settings_patch_allows_personal_preferences_when_overrides_disabled(self):
        self.client.force_authenticate(user=self.user)
        gs = GlobalSettings.objects.get_or_create(singleton_key="global")[0]
        gs.allow_user_overrides = False
        gs.save(update_fields=["allow_user_overrides", "updated_at"])

        res = self.client.patch(
            "/api/settings/me/",
            {"language": "fr", "date_format": "DD/MM/YYYY", "notifications": {"email": True}},
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        us = UserSettings.objects.get(user=self.user)
        self.assertEqual(us.language, "fr")
        self.assertEqual(us.date_format, "DD/MM/YYYY")
        self.assertEqual(us.notifications.get("email"), True)

    def test_user_logo_upload_blocked_when_admin_disables_overrides(self):
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(MEDIA_ROOT=tmp, MEDIA_URL="/media/"):
                self.client.force_authenticate(user=self.user)
                gs = GlobalSettings.objects.get_or_create(singleton_key="global")[0]
                gs.allow_user_overrides = False
                gs.save(update_fields=["allow_user_overrides", "updated_at"])

                res = self.client.post(
                    "/api/settings/logo/upload/",
                    {"file": _png_logo_upload_file(), "scope": "invoice_template"},
                    format="multipart",
                )
                self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
                self.assertIn("disabled", str(res.data).lower())


class InvoiceDiscountTests(APITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(username="discount_user", password=_test_secret(), email="discount@example.com")
        self.client.force_authenticate(user=self.user)
        self.customer = Customer.objects.create(name="Discount Buyer", email="discount-buyer@example.com")
        self.item = Item.objects.create(
            type="product",
            name="Discount Widget",
            sku="DISC-001",
            unit_price=Decimal("10.00"),
            tax_rate=Decimal("10.00"),
            stock_quantity=20,
        )

    def test_percentage_discount_calculates_total_and_persists(self):
        res = self.client.post(
            "/api/invoices/",
            {
                "customer": self.customer.id,
                "status": "Draft",
                "discount_type": "percentage",
                "discount_value": "10",
                "items": [{"item": self.item.id, "quantity": 2}],
            },
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Decimal(res.data["subtotal"]), Decimal("20.00"))
        self.assertEqual(Decimal(res.data["discount_amount"]), Decimal("2.00"))
        self.assertEqual(Decimal(res.data["tax_total"]), Decimal("2.00"))
        self.assertEqual(Decimal(res.data["total_amount"]), Decimal("20.00"))

        invoice = Invoice.objects.get(pk=res.data["id"])
        self.assertEqual(invoice.discount_type, Invoice.DISCOUNT_TYPE_PERCENTAGE)
        self.assertEqual(invoice.discount_value, Decimal("10.00"))
        self.assertEqual(invoice.discount_amount, Decimal("2.00"))

    def test_fixed_discount_equal_to_subtotal_results_in_zero_total_when_tax_is_zero(self):
        zero_tax_item = Item.objects.create(
            type="service",
            name="Zero Tax Service",
            sku="DISC-002",
            unit_price=Decimal("25.00"),
            tax_rate=Decimal("0.00"),
            stock_quantity=0,
        )
        res = self.client.post(
            "/api/invoices/",
            {
                "customer": self.customer.id,
                "status": "Draft",
                "discount_type": "fixed",
                "discount_value": "25.00",
                "items": [{"item": zero_tax_item.id, "quantity": 1}],
            },
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Decimal(res.data["subtotal"]), Decimal("25.00"))
        self.assertEqual(Decimal(res.data["discount_amount"]), Decimal("25.00"))
        self.assertEqual(Decimal(res.data["total_amount"]), Decimal("0.00"))

    def test_discount_validation_rejects_out_of_range_values(self):
        percentage_res = self.client.post(
            "/api/invoices/",
            {
                "customer": self.customer.id,
                "status": "Draft",
                "discount_type": "percentage",
                "discount_value": "150",
                "items": [{"item": self.item.id, "quantity": 1}],
            },
            format="json",
        )
        self.assertEqual(percentage_res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("discount", str(percentage_res.data).lower())

        fixed_res = self.client.post(
            "/api/invoices/",
            {
                "customer": self.customer.id,
                "status": "Draft",
                "discount_type": "fixed",
                "discount_value": "25.00",
                "items": [{"item": self.item.id, "quantity": 2}],
            },
            format="json",
        )
        self.assertEqual(fixed_res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("subtotal", str(fixed_res.data).lower())

    def test_invoice_discount_patch_updates_total_amount(self):
        create_res = self.client.post(
            "/api/invoices/",
            {
                "customer": self.customer.id,
                "status": "Draft",
                "items": [{"item": self.item.id, "quantity": 2}],
            },
            format="json",
        )
        self.assertEqual(create_res.status_code, status.HTTP_201_CREATED)
        invoice_id = create_res.data["id"]

        patch_res = self.client.patch(
            f"/api/invoices/{invoice_id}/",
            {
                "discount_type": "fixed",
                "discount_value": "5.00",
                "updated_at": create_res.data["updated_at"],
            },
            format="json",
        )
        self.assertEqual(patch_res.status_code, status.HTTP_200_OK)
        self.assertEqual(Decimal(patch_res.data["discount_amount"]), Decimal("5.00"))
        self.assertEqual(Decimal(patch_res.data["total_amount"]), Decimal("17.00"))

    def test_invoice_create_accepts_unit_price_override_and_recomputes_totals(self):
        res = self.client.post(
            "/api/invoices/",
            {
                "customer": self.customer.id,
                "status": "Draft",
                "items": [{"item": self.item.id, "quantity": 2, "unit_price": "12.50"}],
            },
            format="json",
        )

        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Decimal(res.data["subtotal"]), Decimal("25.00"))
        self.assertEqual(Decimal(res.data["tax_total"]), Decimal("2.50"))
        self.assertEqual(Decimal(res.data["total_amount"]), Decimal("27.50"))
        self.assertEqual(Decimal(res.data["invoice_items"][0]["unit_price"]), Decimal("12.50"))
        self.assertEqual(Decimal(res.data["invoice_items"][0]["line_total"]), Decimal("27.50"))

    def test_invoice_create_rejects_negative_or_invalid_unit_price_override(self):
        invalid_res = self.client.post(
            "/api/invoices/",
            {
                "customer": self.customer.id,
                "status": "Draft",
                "items": [{"item": self.item.id, "quantity": 1, "unit_price": "abc"}],
            },
            format="json",
        )
        self.assertEqual(invalid_res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("unit_price", str(invalid_res.data).lower())

        negative_res = self.client.post(
            "/api/invoices/",
            {
                "customer": self.customer.id,
                "status": "Draft",
                "items": [{"item": self.item.id, "quantity": 1, "unit_price": "-1.00"}],
            },
            format="json",
        )
        self.assertEqual(negative_res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("unit_price", str(negative_res.data).lower())

    def test_invoice_create_persists_description_overrides_and_blank_values(self):
        product = Item.objects.create(
            type="product",
            name="Override Product",
            sku="DESC-001",
            description="Catalog product description",
            unit_price=Decimal("12.00"),
            tax_rate=Decimal("5.00"),
            stock_quantity=10,
        )
        service = Item.objects.create(
            type="service",
            name="Override Service",
            sku="DESC-002",
            description="Catalog service description",
            unit_price=Decimal("8.00"),
            tax_rate=Decimal("0.00"),
            stock_quantity=0,
        )

        res = self.client.post(
            "/api/invoices/",
            {
                "customer": self.customer.id,
                "status": "Draft",
                "items": [
                    {"item": product.id, "quantity": 1, "description": "Custom product description"},
                    {"item": service.id, "quantity": 1, "description": ""},
                ],
            },
            format="json",
        )

        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        invoice_items = res.data["invoice_items"]
        self.assertEqual(invoice_items[0]["description"], "Custom product description")
        self.assertIsNone(invoice_items[1]["description"])

        persisted_items = list(InvoiceItem.objects.filter(invoice_id=res.data["id"], is_deleted=False).order_by("id"))
        self.assertEqual(persisted_items[0].description, "Custom product description")
        self.assertIsNone(persisted_items[1].description)

    def test_invoice_create_rejects_description_longer_than_500_characters(self):
        res = self.client.post(
            "/api/invoices/",
            {
                "customer": self.customer.id,
                "status": "Draft",
                "items": [{"item": self.item.id, "quantity": 1, "description": "x" * 501}],
            },
            format="json",
        )

        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("description", str(res.data).lower())

    def test_invoice_render_includes_discount_details(self):
        invoice = Invoice.objects.create(
            invoice_number="INV-DISC-1",
            customer=self.customer,
            status="Draft",
            subtotal=Decimal("20.00"),
            discount_type=Invoice.DISCOUNT_TYPE_PERCENTAGE,
            discount_value=Decimal("10.00"),
            discount_amount=Decimal("2.00"),
            tax_total=Decimal("2.00"),
            total_amount=Decimal("20.00"),
        )
        InvoiceItem.objects.create(
            invoice=invoice,
            item=self.item,
            quantity=2,
            unit_price=Decimal("10.00"),
            tax_rate=Decimal("10.00"),
            line_subtotal=Decimal("20.00"),
            line_tax=Decimal("2.00"),
            line_total=Decimal("22.00"),
        )
        request_stub = type("Req", (), {"user": self.user, "query_params": {}, "headers": {}})()
        rendered = render_invoice(request_stub, invoice, "html").content.decode("utf-8")
        self.assertIn("Discount Percentage (10%):", rendered)
        self.assertIn("Total:", rendered)


class InvoiceFiltersAndSavedViewsTests(APITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(username="u_inv", password=_test_secret())
        self.client.force_authenticate(user=self.user)
        self.c1 = Customer.objects.create(name="Alice", email="a@example.com", created_at=timezone.now())
        self.c2 = Customer.objects.create(name="Bob", email="b@example.com", created_at=timezone.now())

        Invoice.objects.create(
            invoice_number="INV-1",
            customer=self.c1,
            issue_date=date(2026, 5, 1),
            due_date=date(2026, 5, 10),
            subtotal=Decimal("100.00"),
            tax_total=Decimal("0.00"),
            total_amount=Decimal("100.00"),
            status="Draft",
        )
        Invoice.objects.create(
            invoice_number="INV-2",
            customer=self.c1,
            issue_date=date(2026, 5, 2),
            due_date=date(2026, 5, 12),
            subtotal=Decimal("250.00"),
            tax_total=Decimal("0.00"),
            total_amount=Decimal("250.00"),
            status="Paid",
        )
        Invoice.objects.create(
            invoice_number="INV-3",
            customer=self.c2,
            issue_date=date(2026, 4, 15),
            due_date=date(2026, 5, 1),
            subtotal=Decimal("75.00"),
            tax_total=Decimal("0.00"),
            total_amount=Decimal("75.00"),
            status="Sent",
        )

    def test_invoice_filters_status_and_amount_range(self):
        res = self.client.get("/api/invoices/?page=1&status=Paid")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data["count"], 1)
        self.assertEqual(res.data["results"][0]["invoice_number"], "INV-2")

        res2 = self.client.get("/api/invoices/?page=1&total_min=80&total_max=200")
        self.assertEqual(res2.status_code, status.HTTP_200_OK)
        nums = [r["invoice_number"] for r in res2.data["results"]]
        self.assertIn("INV-1", nums)
        self.assertNotIn("INV-2", nums)
        self.assertNotIn("INV-3", nums)

    def test_invoice_filters_dates_and_customer(self):
        res = self.client.get(f"/api/invoices/?page=1&customer={self.c2.id}")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data["count"], 1)
        self.assertEqual(res.data["results"][0]["invoice_number"], "INV-3")

        res2 = self.client.get("/api/invoices/?page=1&issue_date_from=2026-05-01&issue_date_to=2026-05-31")
        self.assertEqual(res2.status_code, status.HTTP_200_OK)
        nums = [r["invoice_number"] for r in res2.data["results"]]
        self.assertIn("INV-1", nums)
        self.assertIn("INV-2", nums)
        self.assertNotIn("INV-3", nums)

    def test_invoice_export_csv_uses_filters(self):
        res = self.client.get("/api/invoices/export/?status=Paid")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn("text/csv", res.get("Content-Type", ""))
        body = b"".join(res.streaming_content).decode("utf-8")
        self.assertIn("invoice_number", body)
        self.assertIn("INV-2", body)
        self.assertNotIn("INV-1", body)

    def test_invoice_export_xlsx_and_fields(self):
        res = self.client.get("/api/invoices/export/?status=Paid&file_format=xlsx&fields=invoice_number,total_amount")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            res.get("Content-Type", ""),
        )
        wb = load_workbook(io.BytesIO(res.content), read_only=True, data_only=True)
        ws = wb.active
        header = list(next(ws.iter_rows(values_only=True)))
        self.assertEqual(header, ["invoice_number", "total_amount"])
        rows = list(ws.iter_rows(values_only=True))
        flat = [r[0] for r in rows]
        self.assertIn("INV-2", flat)
        self.assertNotIn("INV-1", flat)

    def test_invoice_export_pdf(self):
        UserProfile.objects.update_or_create(user=self.user, defaults={"company_legal_name": "AmbienteSoft LTD"})
        res = self.client.get("/api/invoices/export/?status=Paid&file_format=pdf&fields=invoice_number,status")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(
            ("application/pdf" in (res.get("Content-Type") or "")) or ("text/html" in (res.get("Content-Type") or "")),
            msg=f"unexpected content-type: {res.get('Content-Type')}",
        )
        if "text/html" in (res.get("Content-Type") or ""):
            self.assertContains(res, "AmbienteSoft LTD")
            self.assertContains(res, "Invoices Export")

    def test_saved_invoice_views_crud(self):
        create = self.client.post("/api/invoices/views/", {"name": "Paid", "filters": {"status": "Paid"}, "is_default": True}, format="json")
        self.assertEqual(create.status_code, status.HTTP_201_CREATED)
        view_id = create.data["id"]

        ls = self.client.get("/api/invoices/views/")
        self.assertEqual(ls.status_code, status.HTTP_200_OK)
        self.assertEqual(len(ls.data["results"]), 1)
        self.assertEqual(ls.data["results"][0]["name"], "Paid")
        self.assertTrue(ls.data["results"][0]["is_default"])

        patch = self.client.patch(f"/api/invoices/views/{view_id}/", {"name": "Paid invoices"}, format="json")
        self.assertEqual(patch.status_code, status.HTTP_200_OK)

        ls2 = self.client.get("/api/invoices/views/")
        self.assertEqual(ls2.status_code, status.HTTP_200_OK)
        self.assertEqual(ls2.data["results"][0]["name"], "Paid invoices")

        delete = self.client.delete(f"/api/invoices/views/{view_id}/")
        self.assertEqual(delete.status_code, status.HTTP_204_NO_CONTENT)
        ls3 = self.client.get("/api/invoices/views/")
        self.assertEqual(ls3.status_code, status.HTTP_200_OK)
        self.assertEqual(len(ls3.data["results"]), 0)


class ImportExportTests(APITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(username="u_impexp", password=_test_secret())
        perms = Permission.objects.filter(codename__in=["add_item", "change_item", "add_invoice", "change_invoice"])
        self.user.user_permissions.add(*list(perms))
        self.client.force_authenticate(user=self.user)

    def test_inventory_export_csv(self):
        Item.objects.create(type="product", name="Widget", sku="W-1", unit_price=Decimal("10.00"), stock_quantity=5)
        res = self.client.get("/api/items/export/?file_format=csv&fields=sku,name,unit_price,stock_quantity")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        body = b"".join(res.streaming_content).decode("utf-8")
        self.assertIn("sku", body)
        self.assertIn("W-1", body)

    def test_inventory_export_pdf(self):
        UserProfile.objects.update_or_create(user=self.user, defaults={"company_legal_name": "AmbienteSoft LTD"})
        Item.objects.create(type="product", name="Widget", sku="W-1", unit_price=Decimal("10.00"), stock_quantity=5)
        res = self.client.get("/api/items/export/?file_format=pdf&fields=sku,name,unit_price,stock_quantity")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(
            ("application/pdf" in (res.get("Content-Type") or "")) or ("text/html" in (res.get("Content-Type") or "")),
            msg=f"unexpected content-type: {res.get('Content-Type')}",
        )
        if "text/html" in (res.get("Content-Type") or ""):
            self.assertContains(res, "AmbienteSoft LTD")
            self.assertContains(res, "Inventory Export")

    def test_inventory_import_template_downloads(self):
        csv_res = self.client.get("/api/items/import_template/?file_format=csv")
        self.assertEqual(csv_res.status_code, status.HTTP_200_OK)
        self.assertIn("text/csv", csv_res.get("Content-Type", ""))
        self.assertIn("inventory_import_template.csv", csv_res.get("Content-Disposition", ""))
        self.assertIn("unit_price", csv_res.content.decode("utf-8", errors="ignore"))
        self.assertIn("category", csv_res.content.decode("utf-8", errors="ignore"))

        xlsx_res = self.client.get("/api/items/import_template/?file_format=xlsx")
        self.assertEqual(xlsx_res.status_code, status.HTTP_200_OK)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", xlsx_res.get("Content-Type", ""))
        self.assertIn("inventory_import_template.xlsx", xlsx_res.get("Content-Disposition", ""))

    def test_inventory_import_rollback_and_error_log(self):
        Item.objects.create(type="product", name="Existing", sku="DUP-1", unit_price=Decimal("1.00"), stock_quantity=1)
        csv_body = "type,sku,name,unit_price,tax_rate,stock_quantity\nproduct,DUP-1,New Item,10.00,0,5\nproduct,,Bad Price,xx,0,1\n"
        up = SimpleUploadedFile("items.csv", csv_body.encode("utf-8"), content_type="text/csv")
        res = self.client.post("/api/items/import/", {"file": up, "rollback_on_error": "true"}, format="multipart")
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("error_log_token", res.data)
        self.assertFalse(Item.objects.filter(name="New Item", is_deleted=False).exists())
        token = res.data["error_log_token"]
        dl = self.client.get(f"/api/imports/error-log/{token}/")
        self.assertEqual(dl.status_code, status.HTTP_200_OK)
        self.assertIn("text/csv", dl.get("Content-Type", ""))
        self.assertIn("sku already exists", dl.content.decode("utf-8", errors="ignore"))

    def test_inventory_import_success(self):
        csv_body = "type,sku,name,category,unit_price,tax_rate,stock_quantity\nproduct,SKU-100,Imported,Hardware,12.50,5,10\n"
        up = SimpleUploadedFile("items.csv", csv_body.encode("utf-8"), content_type="text/csv")
        res = self.client.post("/api/items/import/", {"file": up, "rollback_on_error": "true"}, format="multipart")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data["imported"], 1)
        self.assertTrue(Item.objects.filter(sku="SKU-100", is_deleted=False).exists())
        self.assertEqual(Item.objects.get(sku="SKU-100", is_deleted=False).category, "Hardware")

    def test_inventory_import_dry_run(self):
        csv_body = "type,sku,name,category,unit_price,tax_rate,stock_quantity\nproduct,SKU-DRY,DryRun,General,1.00,0,1\n"
        up = SimpleUploadedFile("items.csv", csv_body.encode("utf-8"), content_type="text/csv")
        res = self.client.post("/api/items/import/", {"file": up, "dry_run": "true", "rollback_on_error": "true"}, format="multipart")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(res.data.get("dry_run"))
        self.assertEqual(res.data.get("would_create"), 1)
        self.assertFalse(Item.objects.filter(sku="SKU-DRY", is_deleted=False).exists())

    def test_inventory_import_xlsx_success(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["type", "sku", "name", "category", "unit_price", "tax_rate", "stock_quantity"])
        ws.append(["product", "SKU-XLSX", "FromXLSX", "Supplies", "2.50", "0", 4])
        out = io.BytesIO()
        wb.save(out)
        up = SimpleUploadedFile(
            "items.xlsx",
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        res = self.client.post("/api/items/import/", {"file": up}, format="multipart")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(Item.objects.filter(sku="SKU-XLSX", is_deleted=False).exists())
        self.assertEqual(Item.objects.get(sku="SKU-XLSX", is_deleted=False).category, "Supplies")

    def test_invoice_import_groups_rows_and_deducts_inventory(self):
        customer = Customer.objects.create(name="Buyer", email="buyer@example.com")
        item = Item.objects.create(type="product", name="Widget", sku="W-001", unit_price=Decimal("10.00"), tax_rate=Decimal("0.00"), stock_quantity=10)
        csv_body = (
            "invoice_key,invoice_number,customer_email,customer_name,status,issue_date,due_date,item_sku,quantity,unit_price,tax_rate,description,unit_of_measure\n"
            "B1,,buyer@example.com,,Sent,2026-05-01,2026-05-10,W-001,2,,,Line 1,pcs\n"
            "B1,,buyer@example.com,,Sent,2026-05-01,2026-05-10,W-001,1,,,Line 2,pcs\n"
        )
        up = SimpleUploadedFile("invoices.csv", csv_body.encode("utf-8"), content_type="text/csv")
        res = self.client.post("/api/invoices/import/", {"file": up, "rollback_on_error": "true"}, format="multipart")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data["imported_invoices"], 1)
        self.assertEqual(res.data["imported_invoice_items"], 2)
        item.refresh_from_db()
        self.assertEqual(item.stock_quantity, 7)
        inv = Invoice.objects.order_by("-id").first()
        self.assertIsNotNone(inv.inventory_deducted_at)
        self.assertEqual(inv.customer_id, customer.id)

    def test_invoice_import_template_downloads(self):
        csv_res = self.client.get("/api/invoices/import_template/?file_format=csv")
        self.assertEqual(csv_res.status_code, status.HTTP_200_OK)
        self.assertIn("text/csv", csv_res.get("Content-Type", ""))
        self.assertIn("invoice_import_template.csv", csv_res.get("Content-Disposition", ""))
        self.assertIn("item_sku", csv_res.content.decode("utf-8", errors="ignore"))

        xlsx_res = self.client.get("/api/invoices/import_template/?file_format=xlsx")
        self.assertEqual(xlsx_res.status_code, status.HTTP_200_OK)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", xlsx_res.get("Content-Type", ""))
        self.assertIn("invoice_import_template.xlsx", xlsx_res.get("Content-Disposition", ""))

    def test_invoice_import_dry_run(self):
        customer = Customer.objects.create(name="Buyer2", email="buyer2@example.com")
        Item.objects.create(type="product", name="Widget2", sku="W-002", unit_price=Decimal("10.00"), tax_rate=Decimal("0.00"), stock_quantity=10)
        csv_body = (
            "invoice_key,customer_email,status,issue_date,item_sku,quantity\n"
            "B2,buyer2@example.com,Draft,2026-05-01,W-002,1\n"
        )
        up = SimpleUploadedFile("invoices.csv", csv_body.encode("utf-8"), content_type="text/csv")
        res = self.client.post("/api/invoices/import/", {"file": up, "dry_run": "true"}, format="multipart")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(res.data.get("dry_run"))
        self.assertEqual(res.data.get("would_create_invoices"), 1)
        self.assertEqual(res.data.get("would_create_invoice_items"), 1)
        self.assertFalse(Invoice.objects.filter(customer=customer, is_deleted=False).exists())

    def test_invoice_import_duplicate_invoice_number_rejected(self):
        customer = Customer.objects.create(name="Buyer3", email="buyer3@example.com")
        Item.objects.create(type="product", name="Widget3", sku="W-003", unit_price=Decimal("10.00"), tax_rate=Decimal("0.00"), stock_quantity=10)
        Invoice.objects.create(invoice_number="INV-DUP", customer=customer, status="Draft")
        csv_body = (
            "invoice_number,customer_email,status,issue_date,item_sku,quantity\n"
            "INV-DUP,buyer3@example.com,Draft,2026-05-01,W-003,1\n"
        )
        up = SimpleUploadedFile("invoices.csv", csv_body.encode("utf-8"), content_type="text/csv")
        res = self.client.post("/api/invoices/import/", {"file": up, "rollback_on_error": "true"}, format="multipart")
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("error_log_token", res.data)

    def test_invoice_import_missing_item_rejected(self):
        Customer.objects.create(name="Buyer4", email="buyer4@example.com")
        csv_body = (
            "invoice_key,customer_email,status,issue_date,item_sku,quantity\n"
            "B4,buyer4@example.com,Draft,2026-05-01,NO-SUCH-SKU,1\n"
        )
        up = SimpleUploadedFile("invoices.csv", csv_body.encode("utf-8"), content_type="text/csv")
        res = self.client.post("/api/invoices/import/", {"file": up, "rollback_on_error": "true"}, format="multipart")
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("error_log_token", res.data)


class CustomerExpenseImportExportTests(APITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_superuser(username="ops_admin", email="ops@example.com", password=_test_secret())
        self.client.force_authenticate(user=self.user)
        self.usd, _ = Currency.objects.get_or_create(
            code="USD",
            defaults={"name": "US Dollar", "symbol": "$", "decimal_places": 2},
        )
        self.petty1 = SourceAccount.objects.create(
            name="petty1",
            account_type=SourceAccount.TYPE_PETTY_CASH,
            initial_balance=Decimal("100.00"),
            currency=self.usd,
            status=SourceAccount.STATUS_ACTIVE,
        )
        self.petty2 = SourceAccount.objects.create(
            name="petty2",
            account_type=SourceAccount.TYPE_PETTY_CASH,
            initial_balance=Decimal("200.00"),
            currency=self.usd,
            status=SourceAccount.STATUS_ACTIVE,
        )

    def test_customer_export_and_import_support_transaction_fields(self):
        customer = Customer.objects.create(name="Export Buyer", email="buyer@example.com", phone="+2348012345678")
        item = Item.objects.create(type="product", name="Export Widget", sku="CUST-EXPORT-1", unit_price=Decimal("25.00"), stock_quantity=5)
        invoice = Invoice.objects.create(
            invoice_number="INV-CUST-EXPORT",
            customer=customer,
            status="Paid",
            subtotal=Decimal("25.00"),
            tax_total=Decimal("0.00"),
            total_amount=Decimal("25.00"),
        )
        InvoiceItem.objects.create(
            invoice=invoice,
            item=item,
            quantity=1,
            unit_price=Decimal("25.00"),
            line_subtotal=Decimal("25.00"),
            line_tax=Decimal("0.00"),
            line_total=Decimal("25.00"),
        )
        Receipt.objects.create(
            invoice=invoice,
            amount_paid=Decimal("25.00"),
            payment_date=timezone.localdate(),
            payment_method="Cash",
        )

        res = self.client.get(
            "/api/customers/export/?file_format=csv&fields=name,account_status,segment,invoice_count,total_paid_amount,order_history"
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        body = b"".join(res.streaming_content).decode("utf-8")
        self.assertIn("Export Buyer", body)
        self.assertIn("active", body)
        self.assertIn("INV-CUST-EXPORT:Paid:25.00", body)
        self.assertIn("25.00", body)

        template = self.client.get("/api/customers/import_template/?file_format=xlsx")
        self.assertEqual(template.status_code, status.HTTP_200_OK)
        wb = load_workbook(io.BytesIO(template.content))
        self.assertEqual(wb.active["A1"].value, "name")

    def test_customer_import_rollback_and_success(self):
        Customer.objects.create(name="Existing", email="existing@example.com")
        bad_csv = "name,email,phone,billing_address\nNew One,existing@example.com,+2348000000000,Addr\nBroken,not-an-email,,\n"
        bad_upload = SimpleUploadedFile("customers.csv", bad_csv.encode("utf-8"), content_type="text/csv")
        bad_res = self.client.post("/api/customers/import/", {"file": bad_upload, "rollback_on_error": "true"}, format="multipart")
        self.assertEqual(bad_res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("error_log_token", bad_res.data)
        log_res = self.client.get(f"/api/imports/error-log/{bad_res.data['error_log_token']}/")
        self.assertEqual(log_res.status_code, status.HTTP_200_OK)
        self.assertIn("email already exists", log_res.content.decode("utf-8", errors="ignore"))

        good_csv = "name,email,phone,billing_address\nImported Person,imported@example.com,+2348010000000,12 Palm Ave\n"
        good_upload = SimpleUploadedFile("customers.csv", good_csv.encode("utf-8"), content_type="text/csv")
        good_res = self.client.post("/api/customers/import/", {"file": good_upload}, format="multipart")
        self.assertEqual(good_res.status_code, status.HTTP_200_OK)
        self.assertEqual(good_res.data["imported"], 1)
        self.assertTrue(Customer.objects.filter(email="imported@example.com", is_deleted=False).exists())

    def test_customer_import_xlsx_rejects_overlong_contact_fields_without_500(self):
        upload = _xlsx_upload_file(
            "customers.xlsx",
            ["name", "email", "phone", "billing_address"],
            [["Valid Name", "valid@example.com", "1" * 21, "12 Palm Ave"]],
        )
        response = self.client.post(
            "/api/customers/import/",
            {"file": upload, "rollback_on_error": "true"},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data["imported"], 0)
        self.assertTrue(
            any(
                err["field"] == "phone" and "at most 20 characters" in err["message"]
                for err in response.data["errors"]
            )
        )
        self.assertFalse(Customer.objects.filter(email="valid@example.com", is_deleted=False).exists())

    def test_source_account_crud_delete_dependencies_and_permissions(self):
        list_url = "/api/source-accounts/"
        create_res = self.client.post(
            list_url,
            {
                "name": "Operations Wallet",
                "account_type": "mobile_money",
                "initial_balance": "500.00",
                "currency": self.usd.id,
                "status": "active",
            },
            format="json",
        )
        self.assertEqual(create_res.status_code, status.HTTP_201_CREATED)
        account_id = create_res.data["id"]
        self.assertEqual(create_res.data["currency_code"], "USD")

        patch_res = self.client.patch(
            f"{list_url}{account_id}/",
            {
                "name": "Operations Wallet Updated",
                "status": "inactive",
                "updated_at": create_res.data["updated_at"],
            },
            format="json",
        )
        self.assertEqual(patch_res.status_code, status.HTTP_200_OK)
        self.assertEqual(patch_res.data["name"], "Operations Wallet Updated")

        linked_expense = Expense.objects.create(
            amount=Decimal("25.00"),
            expense_date=timezone.localdate(),
            category="Travel",
            description="Ride",
            vendor="Vendor",
            merchant_reference="SRC-DEL-1",
            project_code="PRJ-SA",
            source_account=SourceAccount.objects.get(pk=account_id),
            assigned_to=self.user,
            created_by=self.user,
        )
        delete_res = self.client.delete(
            f"{list_url}{account_id}/",
            {"confirm_keyword": "DELETE", "updated_at": patch_res.data["updated_at"]},
            format="json",
        )
        self.assertEqual(delete_res.status_code, status.HTTP_200_OK)
        self.assertEqual(delete_res.data["active_expense_count"], 1)
        deleted = SourceAccount.objects.get(pk=account_id)
        self.assertTrue(deleted.is_deleted)
        self.assertEqual(deleted.status, SourceAccount.STATUS_CLOSED)
        linked_expense.refresh_from_db()
        self.assertEqual(linked_expense.source_account_id, account_id)

        create_audit = AuditLog.objects.filter(action="create", object_id=str(account_id)).exists()
        update_audit = AuditLog.objects.filter(action="update", object_id=str(account_id)).exists()
        delete_audit = AuditLog.objects.filter(action="delete", object_id=str(account_id)).first()
        self.assertTrue(create_audit)
        self.assertTrue(update_audit)
        self.assertIsNotNone(delete_audit)
        self.assertEqual(delete_audit.changes["active_expense_count"], 1)

        no_role_user = User.objects.create_user(username="readonly_source_account", password=_test_secret())
        UserRole.objects.filter(user=no_role_user).delete()
        self.client.force_authenticate(user=no_role_user)
        denied = self.client.post(
            list_url,
            {
                "name": "Denied Account",
                "account_type": "other",
                "initial_balance": "10.00",
                "currency": self.usd.id,
                "status": "active",
            },
            format="json",
        )
        self.assertEqual(denied.status_code, status.HTTP_403_FORBIDDEN)
        self.client.force_authenticate(user=self.user)

    def test_source_account_list_filters_and_validation_edges(self):
        Expense.objects.create(
            amount=Decimal("15.00"),
            expense_date=timezone.localdate(),
            category="Office",
            description="Pens",
            vendor="Stationery Hub",
            merchant_reference="SRC-LIST-1",
            project_code="PRJ-LIST",
            source_account=self.petty1,
            assigned_to=self.user,
            created_by=self.user,
        )

        list_res = self.client.get("/api/source-accounts/?q=petty&status=active")
        self.assertEqual(list_res.status_code, status.HTTP_200_OK)
        self.assertEqual(len(list_res.data), 2)
        petty1_row = next(row for row in list_res.data if row["id"] == self.petty1.id)
        self.assertEqual(petty1_row["active_expense_count"], 1)
        self.assertEqual(Decimal(petty1_row["current_balance"]), Decimal("85.00"))

        missing_confirm = self.client.delete(f"/api/source-accounts/{self.petty1.id}/", {}, format="json")
        self.assertEqual(missing_confirm.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("confirm_keyword", missing_confirm.data)

        blank_name = self.client.post(
            "/api/source-accounts/",
            {
                "name": "   ",
                "account_type": "other",
                "initial_balance": "5.00",
                "currency": self.usd.id,
                "status": "active",
            },
            format="json",
        )
        self.assertEqual(blank_name.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("name", blank_name.data)

        negative_balance = self.client.post(
            "/api/source-accounts/",
            {
                "name": "Negative Balance Account",
                "account_type": "bank",
                "initial_balance": "-1.00",
                "currency": self.usd.id,
                "status": "active",
            },
            format="json",
        )
        self.assertEqual(negative_balance.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("initial_balance", negative_balance.data)

    def test_expense_project_code_generation_duplicate_blocking_and_balance_export(self):
        next_code_res = self.client.get("/api/expenses/next_project_code/")
        self.assertEqual(next_code_res.status_code, status.HTTP_200_OK)
        self.assertRegex(next_code_res.data["project_code"], r"^PRJ-\d{4}-\d{4}$")

        generated_res = self.client.post(
            "/api/expenses/",
            {
                "amount": "25.00",
                "expense_date": str(timezone.localdate()),
                "category": "Travel",
                "source_account": self.petty1.id,
            },
            format="json",
        )
        self.assertEqual(generated_res.status_code, status.HTTP_201_CREATED)
        self.assertRegex(generated_res.data["project_code"], r"^PRJ-\d{4}-\d{4}$")
        self.assertEqual(generated_res.data["source_account_currency_code"], "USD")
        create_audit = AuditLog.objects.filter(action="create", object_id=str(generated_res.data["id"])).first()
        self.assertIsNotNone(create_audit)
        self.assertEqual(create_audit.changes["project_code"]["to"], generated_res.data["project_code"])
        self.assertTrue(create_audit.changes["project_code"]["generated"])

        duplicate_res = self.client.post(
            "/api/expenses/",
            {
                "amount": "10.00",
                "expense_date": str(timezone.localdate()),
                "category": "Travel",
                "project_code": generated_res.data["project_code"].lower(),
                "source_account": self.petty1.id,
            },
            format="json",
        )
        self.assertEqual(duplicate_res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("project_code", duplicate_res.data)

        export_res = self.client.get("/api/expenses/export/?file_format=csv&fields=source_account,source_account_balance,project_code")
        self.assertEqual(export_res.status_code, status.HTTP_200_OK)
        export_body = b"".join(export_res.streaming_content).decode("utf-8")
        self.assertIn("petty1", export_body)
        self.assertIn("75.00", export_body)
        self.assertIn(generated_res.data["project_code"], export_body)

    def test_expense_crud_source_account_filters_export_and_removed_endpoints(self):
        create_res = self.client.post(
            "/api/expenses/",
            {
                "amount": "250.00",
                "expense_date": str(timezone.localdate()),
                "category": "Travel",
                "description": "Airport taxi",
                "vendor": "Ride Co",
                "merchant_reference": "TRAVEL-1",
                "project_code": "PROJECT-ALPHA",
                "source_account": self.petty1.id,
            },
            format="json",
        )
        self.assertEqual(create_res.status_code, status.HTTP_201_CREATED)
        expense_id = create_res.data["id"]
        self.assertEqual(create_res.data["created_by"], self.user.id)
        self.assertEqual(create_res.data["assigned_to"], self.user.id)
        self.assertEqual(create_res.data["source_account"], self.petty1.id)
        self.assertEqual(create_res.data["source_account_name"], "petty1")

        filtered = self.client.get(f"/api/expenses/?source_account={self.petty1.id}&category=Travel")
        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.data["count"], 1)

        export_res = self.client.get("/api/expenses/export/?file_format=csv&fields=expense_date,amount,category,source_account,assigned_to")
        self.assertEqual(export_res.status_code, status.HTTP_200_OK)
        body = b"".join(export_res.streaming_content).decode("utf-8")
        self.assertIn("Travel", body)
        self.assertIn("petty1", body)
        self.assertIn(self.user.username, body)

        approve_res = self.client.post(f"/api/expenses/{expense_id}/approve/", {"approval_status": "approved"}, format="json")
        self.assertEqual(approve_res.status_code, status.HTTP_404_NOT_FOUND)
        receipt_res = self.client.get(f"/api/expenses/{expense_id}/receipt/")
        self.assertEqual(receipt_res.status_code, status.HTTP_404_NOT_FOUND)

    def test_expense_storage_is_encrypted_at_rest_and_decrypted_in_api_and_export(self):
        create_res = self.client.post(
            "/api/expenses/",
            {
                "amount": "1500.00",
                "expense_date": str(timezone.localdate()),
                "category": "Compliance",
                "description": "Background verification",
                "vendor": "Secure Vendor",
                "merchant_reference": "SEC-42",
                "project_code": "PROJECT-GAMMA",
                "source_account": self.petty2.id,
            },
            format="json",
        )
        self.assertEqual(create_res.status_code, status.HTTP_201_CREATED)
        expense = Expense.objects.get(pk=create_res.data["id"])
        self.assertTrue(is_encrypted_expense_value(expense.description))
        self.assertTrue(is_encrypted_expense_value(expense.merchant_reference))
        self.assertEqual(decrypt_expense_text(expense.description), "Background verification")
        self.assertEqual(decrypt_expense_text(expense.merchant_reference), "SEC-42")

        detail_res = self.client.get("/api/expenses/?q=verification")
        self.assertEqual(detail_res.status_code, status.HTTP_200_OK)
        self.assertEqual(detail_res.data["count"], 1)
        self.assertEqual(detail_res.data["results"][0]["description"], "Background verification")
        self.assertEqual(detail_res.data["results"][0]["merchant_reference"], "SEC-42")

        export_res = self.client.get("/api/expenses/export/?file_format=csv&fields=description,merchant_reference,source_account")
        self.assertEqual(export_res.status_code, status.HTTP_200_OK)
        export_body = b"".join(export_res.streaming_content).decode("utf-8")
        self.assertIn("Background verification", export_body)
        self.assertIn("SEC-42", export_body)
        self.assertIn("petty2", export_body)

    def test_expense_import_error_log_without_approval_or_policy(self):
        dry_run_csv = (
            "amount,expense_date,category,description,vendor,merchant_reference,project_code,cost_center,source_account,assigned_to\n"
            f"1250.00,{timezone.localdate()},Software,Annual license,Vendor Ltd,SOFT-1,PROJECT-BETA,,petty1,{self.user.username}\n"
        )
        dry_run_upload = SimpleUploadedFile("expenses.csv", dry_run_csv.encode("utf-8"), content_type="text/csv")
        dry_run_res = self.client.post("/api/expenses/import/", {"file": dry_run_upload, "dry_run": "true"}, format="multipart")
        self.assertEqual(dry_run_res.status_code, status.HTTP_200_OK)
        self.assertTrue(dry_run_res.data["dry_run"])
        self.assertNotIn("flags", dry_run_res.data)

        bad_csv = (
            "amount,expense_date,category,description,vendor,merchant_reference,project_code,cost_center,source_account,assigned_to\n"
            "12.00,2026-01-01,,Bad Row,Vendor Ltd,REF-1,,,petty2,unknown-user\n"
        )
        bad_upload = SimpleUploadedFile("expenses.csv", bad_csv.encode("utf-8"), content_type="text/csv")
        bad_res = self.client.post("/api/expenses/import/", {"file": bad_upload, "rollback_on_error": "true"}, format="multipart")
        self.assertEqual(bad_res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("error_log_token", bad_res.data)
        self.assertFalse(Expense.objects.filter(vendor="Vendor Ltd", is_deleted=False).exists())

    def test_import_items_function_covers_validation_edges(self):
        upload = _csv_upload_file(
            "items.csv",
            (
                "type,sku,name,unit_price,tax_rate,stock_quantity,description,unit_of_measure,tax_category\n"
                "invalid,SKU-INV,Bad Type,10.00,0,1,,,\n"
                "product,SKU-NONAME,,10.00,0,1,,,\n"
                "product,SKU-DUP,First Dup,10.00,0,1,,,\n"
                "product,SKU-DUP,Second Dup,10.00,0,1,,,\n"
                "product,,Repeat Name,10.00,0,1,,,\n"
                "product,,Repeat Name,10.00,0,1,,,\n"
                "product,SKU-NOPRICE,No Price,,0,1,,,\n"
                "product,SKU-NEGPRICE,Negative Price,-1.00,0,1,,,\n"
                "product,SKU-BADTAXTXT,Bad Tax Text,10.00,oops,1,,,\n"
                "product,SKU-BADTAXRANGE,Bad Tax Range,10.00,101,1,,,\n"
                "product,SKU-BADQTYTXT,Bad Qty Text,10.00,0,nope,,,\n"
                "product,SKU-BADQTYNEG,Bad Qty Negative,10.00,0,-5,,,\n"
                "service,SKU-SVC,Service Import,20.00,,9,Consulting,hour,reduced\n"
            ),
        )

        status_code, payload = import_items_from_upload(upload, dry_run=True, rollback_on_error=False)

        self.assertEqual(status_code, 200)
        self.assertTrue(payload["dry_run"])
        self.assertEqual(payload["would_create"], 3)
        messages = {(err["field"], err["message"]) for err in payload["errors"]}
        self.assertIn(("type", "Invalid type"), messages)
        self.assertIn(("name", "Name is required"), messages)
        self.assertIn(("sku", "Duplicate sku in file"), messages)
        self.assertIn(("name", "Duplicate name/type in file"), messages)
        self.assertIn(("unit_price", "unit_price is required"), messages)
        self.assertIn(("unit_price", "unit_price must be >= 0"), messages)
        self.assertIn(("tax_rate", "Invalid tax_rate"), messages)
        self.assertIn(("tax_rate", "tax_rate must be between 0 and 100"), messages)
        self.assertIn(("stock_quantity", "stock_quantity must be an integer"), messages)
        self.assertIn(("stock_quantity", "stock_quantity must be >= 0"), messages)
        self.assertEqual(len(payload["successful_items"]), 3)
        self.assertEqual(len(payload["failed_items"]), 10)
        self.assertFalse(Item.objects.filter(sku="SKU-SVC", is_deleted=False).exists())

    def test_process_batch_import_separates_successful_and_failed_items(self):
        raw_items = [
            {"_row": 2, "code": "GOOD-1", "amount": "12.00"},
            {"_row": 3, "code": "BAD-1", "amount": "-5.00"},
        ]

        def validate_item(row):
            try:
                amount = Decimal(str(row["amount"]))
            except Exception:
                return None, [{"row": row["_row"], "field": "amount", "message": "Invalid amount"}]
            if amount < 0:
                return None, [{"row": row["_row"], "field": "amount", "message": "amount must be >= 0"}]
            return {"code": row["code"], "amount": amount}, []

        def persist_valid_items(valid_items):
            return len(valid_items), [{"row": 2, "item": {"code": valid_items[0]["code"], "amount": str(valid_items[0]["amount"])}}]

        status_code, payload = process_batch_import(
            raw_items,
            validate_item=validate_item,
            persist_valid_items=persist_valid_items,
            dry_run=False,
            rollback_on_error=False,
        )

        self.assertEqual(status_code, 200)
        self.assertEqual(payload["imported"], 1)
        self.assertEqual(len(payload["successful_items"]), 1)
        self.assertEqual(payload["successful_items"][0]["item"]["code"], "GOOD-1")
        self.assertEqual(len(payload["failed_items"]), 1)
        self.assertEqual(payload["failed_items"][0]["item"]["code"], "BAD-1")
        self.assertEqual(payload["failed_items"][0]["errors"][0]["message"], "amount must be >= 0")

    def test_import_invoices_function_covers_name_lookup_and_validation_edges(self):
        Customer.objects.create(name="Lookup Buyer", email="lookup@example.com")
        Item.objects.create(type="product", name="Widget A", sku="W-A", unit_price=Decimal("10.00"), tax_rate=Decimal("7.50"), stock_quantity=5)
        success_upload = _csv_upload_file(
            "invoices-success.csv",
            (
                "invoice_key,invoice_number,customer_email,customer_name,status,issue_date,due_date,item_sku,quantity,unit_price,tax_rate,description,unit_of_measure\n"
                "OK-NAME,,lookup@example.com,,Draft,2026-05-01,,W-A,2,,,Email lookup line,pcs\n"
            ),
        )

        success_status, success_payload = import_invoices_from_upload(
            success_upload,
            dry_run=True,
            rollback_on_error=False,
            deduct_inventory_for_invoice=lambda invoice: None,
        )

        self.assertEqual(success_status, 200)
        self.assertTrue(success_payload["dry_run"])

        upload = _csv_upload_file(
            "invoices.csv",
            (
                "invoice_key,invoice_number,customer_email,customer_name,status,issue_date,due_date,item_sku,quantity,unit_price,tax_rate,description,unit_of_measure\n"
                "ERR-MISSING-CUSTOMER,,,,Draft,2026-05-01,,W-A,1,,,Missing customer,pcs\n"
                "ERR-MISSING-SKU,,lookup@example.com,,Draft,2026-05-01,,,1,,,Missing sku,pcs\n"
                "ERR-BAD-STATUS,,lookup@example.com,,Unknown,2026-05-01,,W-A,1,,,Bad status,pcs\n"
                "ERR-MIXED,,lookup@example.com,,Draft,2026-05-01,,W-A,1,,,First mixed,pcs\n"
                "ERR-MIXED,,lookup@example.com,,Sent,2026-05-01,,W-A,1,,,Second mixed,pcs\n"
                "ERR-CUSTOMER-NOT-FOUND,,ghost@example.com,,Draft,2026-05-01,,W-A,1,,,Missing customer record,pcs\n"
                "ERR-BAD-DATE,,lookup@example.com,,Draft,not-a-date,bad-date,W-A,1,,,Bad dates,pcs\n"
            ),
        )

        status_code, payload = import_invoices_from_upload(
            upload,
            dry_run=True,
            rollback_on_error=False,
            deduct_inventory_for_invoice=lambda invoice: None,
        )

        self.assertEqual(status_code, 200)
        self.assertTrue(payload["dry_run"])
        messages = {(err["field"], err["message"]) for err in payload["errors"]}
        self.assertIn(("customer_email", "customer_email or customer_name is required"), messages)
        self.assertIn(("item_sku", "item_sku is required"), messages)
        self.assertIn(("status", "Invalid status"), messages)
        self.assertIn(("status", "Mixed statuses within the same invoice group"), messages)
        self.assertIn(("customer", "Customer not found"), messages)
        self.assertIn(("issue_date", "Invalid date. Use YYYY-MM-DD"), messages)
        self.assertIn(("due_date", "Invalid date. Use YYYY-MM-DD"), messages)

        line_error_upload = _csv_upload_file(
            "invoices-line-errors.csv",
            (
                "invoice_key,invoice_number,customer_email,customer_name,status,issue_date,due_date,item_sku,quantity,unit_price,tax_rate,description,unit_of_measure\n"
                "ERR-MISSING-QTY,,lookup@example.com,,Draft,2026-05-01,,W-A,,,,Missing qty,pcs\n"
                "ERR-BAD-QTY,,lookup@example.com,,Draft,2026-05-01,,W-A,nope,,,Bad qty,pcs\n"
                "ERR-ZERO-QTY,,lookup@example.com,,Draft,2026-05-01,,W-A,0,,,Zero qty,pcs\n"
                "ERR-BAD-PRICE,,lookup@example.com,,Draft,2026-05-01,,W-A,1,nope,,Bad price,pcs\n"
                "ERR-NEG-PRICE,,lookup@example.com,,Draft,2026-05-01,,W-A,1,-2.00,,Negative price,pcs\n"
                "ERR-BAD-TAX,,lookup@example.com,,Draft,2026-05-01,,W-A,1,,nope,Bad tax,pcs\n"
                "ERR-RANGE-TAX,,lookup@example.com,,Draft,2026-05-01,,W-A,1,,120,Bad tax range,pcs\n"
            ),
        )

        line_status, line_payload = import_invoices_from_upload(
            line_error_upload,
            dry_run=True,
            rollback_on_error=False,
            deduct_inventory_for_invoice=lambda invoice: None,
        )

        self.assertEqual(line_status, 200)
        line_messages = {(err["field"], err["message"]) for err in line_payload["errors"]}
        self.assertIn(("quantity", "quantity is required"), line_messages)
        self.assertIn(("quantity", "quantity must be an integer"), line_messages)
        self.assertIn(("quantity", "quantity must be >= 1"), line_messages)
        self.assertIn(("unit_price", "Invalid unit_price"), line_messages)
        self.assertIn(("unit_price", "unit_price must be >= 0"), line_messages)
        self.assertIn(("tax_rate", "Invalid tax_rate"), line_messages)
        self.assertIn(("tax_rate", "tax_rate must be between 0 and 100"), line_messages)

    def test_import_customers_function_covers_duplicate_name_email_and_dry_run(self):
        upload = _csv_upload_file(
            "customers.csv",
            (
                "name,email,phone,billing_address\n"
                ",missing@example.com,+2348010000000,No name\n"
                "Same Name,first@example.com,+2348010000001,Addr 1\n"
                "Same Name,second@example.com,+2348010000002,Addr 2\n"
                "Unique Person,dup@example.com,+2348010000003,Addr 3\n"
                "Another Person,dup@example.com,+2348010000004,Addr 4\n"
            ),
        )

        status_code, payload = import_customers_from_upload(upload, dry_run=True, rollback_on_error=False)

        self.assertEqual(status_code, 200)
        self.assertTrue(payload["dry_run"])
        self.assertEqual(payload["would_create"], 2)
        messages = {(err["field"], err["message"]) for err in payload["errors"]}
        self.assertIn(("name", "name is required"), messages)
        self.assertIn(("name", "Duplicate name in file"), messages)
        self.assertIn(("email", "Duplicate email in file"), messages)

    def test_import_expenses_function_covers_validation_and_success_creation(self):
        missing_amount = _csv_upload_file(
            "expenses.csv",
            "amount,expense_date,category,description,vendor,merchant_reference,project_code,cost_center,source_account,assigned_to\n"
            ",,,Missing everything,Vendor,REF-0,,,,,\n",
        )
        status_code, payload = import_expenses_from_upload(
            missing_amount,
            dry_run=True,
            rollback_on_error=False,
            actor=self.user,
        )
        self.assertEqual(status_code, 200)
        self.assertIn(("amount", "amount is required"), {(err["field"], err["message"]) for err in payload["errors"]})

        upload = _csv_upload_file(
            "expenses.csv",
            (
                "amount,expense_date,category,description,vendor,merchant_reference,project_code,cost_center,source_account,assigned_to\n"
                "oops,2026-05-01,Office,Bad amount,Vendor,REF-1,PRJ-1,,,petty1,\n"
                "-5.00,2026-05-01,Office,Negative amount,Vendor,REF-2,PRJ-1,,,petty1,\n"
                "10.00,not-a-date,Office,Bad date,Vendor,REF-3,PRJ-1,,,petty1,\n"
                "15.00,2026-05-01,Office,Generated project code,Vendor,REF-4,,,petty1,\n"
                "25.00,2026-05-01,Office,Unknown assignee,Vendor,REF-6,PRJ-1,,petty1,ghost-user\n"
                f"1500.00,{timezone.localdate() + timedelta(days=2)},Travel,Future expense,Vendor,REF-7,PRJ-2,,petty2,{self.user.username}\n"
                f"35.00,{timezone.localdate()},Meals,Created expense,Vendor,REF-8,PRJ-3,,petty2,{self.user.username}\n"
                f"45.00,{timezone.localdate()},Meals,Duplicate project code,Vendor,REF-9,PRJ-3,,petty2,{self.user.username}\n"
            ),
        )

        status_code, payload = import_expenses_from_upload(
            upload,
            dry_run=False,
            rollback_on_error=False,
            actor=self.user,
        )

        self.assertEqual(status_code, 200)
        self.assertEqual(payload["imported"], 2)
        messages = {(err["field"], err["message"]) for err in payload["errors"]}
        self.assertIn(("amount", "Invalid amount"), messages)
        self.assertIn(("amount", "amount must be > 0"), messages)
        self.assertIn(("expense_date", "Invalid date. Use YYYY-MM-DD"), messages)
        self.assertIn(("project_code", "project_code is duplicated in the import file"), messages)
        self.assertIn(("expense_date", "expense_date cannot be more than one day in the future"), messages)
        self.assertIn(("assigned_to", "Assigned user not found"), messages)
        self.assertGreaterEqual(len(payload.get("generated_project_codes") or []), 1)

        created = Expense.objects.filter(merchant_reference__isnull=False, is_deleted=False).order_by("id")
        self.assertEqual(created.count(), 2)
        self.assertTrue(all(exp.created_by == self.user for exp in created))
        created_by_reference = {decrypt_expense_text(exp.merchant_reference): exp for exp in created}
        self.assertEqual(created_by_reference["REF-8"].source_account_id, self.petty2.id)
        generated_imported = created_by_reference["REF-4"]
        self.assertRegex(generated_imported.project_code or "", r"^PRJ-\d{4}-\d{4}$")


@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
class RegistrationTests(APITestCase):
    def setUp(self):
        super().setUp()
        cache.clear()

    def test_register_and_verify_email_flow(self):
        secret = _test_secret()

        with override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend", DEFAULT_FROM_EMAIL="no-reply@test.local"):
            res = self.client.post(
                "/api/auth/register/",
                {
                    "email": "new@example.com",
                    "password": secret,
                    "password_confirm": secret,
                    "company_name": "NewCo Ltd",
                    "country_code": "+234",
                    "phone_number": "8012345678",
                    "country": "Nigeria",
                    "accept_terms": True,
                    "website": "",
                },
                format="json",
            )
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertTrue(res.data.get("verification_sent", False))
        u = User.objects.get(username="new@example.com")
        self.assertFalse(u.is_active)
        self.assertTrue(u.password.startswith("bcrypt_sha256$"))
        profile = UserProfile.objects.get(user=u)
        self.assertIsNotNone(profile.terms_accepted_at)
        self.assertEqual(profile.company_legal_name, "NewCo Ltd")
        self.assertEqual(profile.phone, "+2348012345678")
        self.assertIsNone(profile.email_verified_at)
        self.assertGreaterEqual(len(mail.outbox), 1)
        self.assertIn("verify-email?token=", (mail.outbox[-1].body or "").lower())

        token_row = EmailVerificationToken.objects.get(user=u)
        verify = self.client.post("/api/auth/verify-email/", {"token": token_row.token}, format="json")
        self.assertEqual(verify.status_code, status.HTTP_200_OK)
        u.refresh_from_db()
        self.assertTrue(u.is_active)
        profile.refresh_from_db()
        self.assertIsNotNone(profile.email_verified_at)

    def test_verified_registration_grants_business_write_access_and_can_persist_records(self):
        secret = _test_secret()

        with override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend", DEFAULT_FROM_EMAIL="no-reply@test.local"):
            res = self.client.post(
                "/api/auth/register/",
                {
                    "email": "ops@example.com",
                    "password": secret,
                    "password_confirm": secret,
                    "company_name": "OpsCo Ltd",
                    "country_code": "+234",
                    "phone_number": "8099999999",
                    "country": "Nigeria",
                    "accept_terms": True,
                    "website": "",
                },
                format="json",
            )
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)

        user = User.objects.get(username="ops@example.com")
        verify_row = EmailVerificationToken.objects.get(user=user)
        verify = self.client.post("/api/auth/verify-email/", {"token": verify_row.token}, format="json")
        self.assertEqual(verify.status_code, status.HTTP_200_OK)

        login = self.client.post("/api/auth/token/", {"username": "ops@example.com", "password": secret}, format="json")
        self.assertEqual(login.status_code, status.HTTP_200_OK)
        token = login.data["token"]
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {token}")

        me = self.client.get("/api/auth/me/")
        self.assertEqual(me.status_code, status.HTTP_200_OK)
        self.assertIn("editor", me.data.get("roles", []))
        self.assertIn("data.customers.write", me.data.get("permissions", []))
        self.assertIn("data.invoices.write", me.data.get("permissions", []))
        self.assertIn("data.receipts.write", me.data.get("permissions", []))

        customer_res = self.client.post(
            "/api/customers/",
            {"name": "Ops Buyer", "email": "buyer@example.com", "phone": "08012345678"},
            format="json",
        )
        self.assertEqual(customer_res.status_code, status.HTTP_201_CREATED)

        item_res = self.client.post(
            "/api/items/",
            {"name": "Ops Widget", "unit_price": "10.00", "tax_rate": "0", "stock_quantity": 5},
            format="json",
        )
        self.assertEqual(item_res.status_code, status.HTTP_201_CREATED)

        invoice_res = self.client.post(
            "/api/invoices/",
            {
                "customer": customer_res.data["id"],
                "status": "Draft",
                "items": [{"item": item_res.data["id"], "quantity": 1}],
            },
            format="json",
        )
        self.assertEqual(invoice_res.status_code, status.HTTP_201_CREATED)

        receipt_res = self.client.post(
            f"/api/invoices/{invoice_res.data['id']}/pay/",
            {"amount_paid": "10.00", "payment_method": "Cash"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-reg-editor-1",
        )
        self.assertEqual(receipt_res.status_code, status.HTTP_201_CREATED)
        self.assertTrue(Customer.objects.filter(pk=customer_res.data["id"], is_deleted=False).exists())
        self.assertTrue(Invoice.objects.filter(pk=invoice_res.data["id"], is_deleted=False).exists())
        self.assertTrue(Receipt.objects.filter(pk=receipt_res.data["id"], is_deleted=False).exists())

    def test_resend_verification_sends_email(self):
        email = "resend@example.com"
        u = User.objects.create_user(username=email, email=email, password=_test_secret())
        u.is_active = False
        u.save(update_fields=["is_active"])

        mail.outbox.clear()
        res = self.client.post("/api/auth/resend-verification/", {"email": email}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(res.data.get("sent", False))
        self.assertGreaterEqual(len(mail.outbox), 1)

    def test_register_requires_terms(self):
        secret = _test_secret()
        res = self.client.post(
            "/api/auth/register/",
            {
                "email": "t@example.com",
                "password": secret,
                "password_confirm": secret,
                "company_name": "TermsCo Ltd",
                "country_code": "+234",
                "phone_number": "8011111111",
                "country": "Nigeria",
                "accept_terms": False,
            },
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_register_password_strength(self):
        weak_cred = "weak"
        res = self.client.post(
            "/api/auth/register/",
            {
                "email": "weak@example.com",
                "password": weak_cred,
                "password_confirm": weak_cred,
                "company_name": "WeakCo Ltd",
                "country_code": "+234",
                "phone_number": "8022222222",
                "country": "Nigeria",
                "accept_terms": True,
            },
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_register_rejects_honeypot_submission(self):
        secret = _test_secret()
        res = self.client.post(
            "/api/auth/register/",
            {
                "email": "c@example.com",
                "password": secret,
                "password_confirm": secret,
                "company_name": "BotCo Ltd",
                "country_code": "+234",
                "phone_number": "8033333333",
                "country": "Nigeria",
                "accept_terms": True,
                "website": "https://spam.example.com",
            },
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)


@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend", DEFAULT_FROM_EMAIL="no-reply@test.local")
class PasswordResetTests(APITestCase):
    def setUp(self):
        super().setUp()
        cache.clear()
        mail.outbox.clear()

    def test_password_reset_request_sends_email_for_existing_user(self):
        User.objects.create_user(username="u1", email="u1@example.com", password=_test_secret())
        res = self.client.post("/api/auth/password-reset/", {"email": "u1@example.com"}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(res.data.get("sent", False))
        self.assertGreaterEqual(len(mail.outbox), 1)
        body = mail.outbox[-1].body or ""
        self.assertIn("/reset-password?uid=", body)
        self.assertIn("&token=", body)

    def test_password_reset_confirm_sets_password_and_returns_token(self):
        old_secret = _test_secret()
        new_secret = _test_secret()
        u = User.objects.create_user(username="u2", email="u2@example.com", password=old_secret, is_active=False)
        role = Role.objects.get(name="user")
        existing = AccessToken.objects.create(user=u, role=role, key="existing_token_1")
        uid = urlsafe_base64_encode(force_bytes(u.pk))
        token = PasswordResetTokenGenerator().make_token(u)
        res = self.client.post(
            "/api/auth/password-reset-confirm/",
            {"uid": uid, "token": token, "new_password": new_secret, "new_password_confirm": new_secret},
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(res.data.get("reset", False))
        self.assertTrue(res.data.get("token"))
        u.refresh_from_db()
        self.assertTrue(u.check_password(new_secret))
        self.assertTrue(u.is_active)
        existing.refresh_from_db()
        self.assertIsNotNone(existing.revoked_at)
        self.assertTrue(AccessToken.objects.filter(user=u, revoked_at__isnull=True, role__name="user").exists())


@override_settings(
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    DEFAULT_FROM_EMAIL="no-reply@test.local",
    FRONTEND_BASE_URL="http://frontend.test",
)
class AdminInvitationOnboardingTests(APITestCase):
    def setUp(self):
        super().setUp()
        cache.clear()
        mail.outbox.clear()
        self.admin = User.objects.create_user(
            username="admin_inviter",
            email="admin_inviter@example.com",
            password=_test_secret(),
            is_active=True,
            is_staff=True,
            is_superuser=True,
        )
        self.client.force_authenticate(user=self.admin)

    def _create_invited_user(self, *, username="invited_user", email="invited_user@example.com"):
        return self.client.post(
            "/api/admin/users/",
            {
                "username": username,
                "password": _test_secret(),
                "email": email,
                "company_name": "Invited Co",
                "phone": "8034567890",
                "primary_role": "user",
                "custom_roles": [],
            },
            format="json",
        )

    def _invitation_token_from_email(self) -> str:
        self.assertGreaterEqual(len(mail.outbox), 1)
        body = mail.outbox[-1].body or ""
        token_fragment = body.split("token=", 1)[1].split("&", 1)[0].strip()
        return urllib.parse.unquote(token_fragment)

    def test_admin_create_user_sends_signed_invitation_and_keeps_user_inactive(self):
        res = self._create_invited_user()
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertTrue(res.data.get("invitation_sent"))

        user = User.objects.get(username="invited_user")
        self.assertFalse(user.is_active)
        profile = UserProfile.objects.get(user=user)
        self.assertIsNone(profile.email_verified_at)

        invitation = AdminUserInvitation.objects.get(user=user)
        self.assertIsNotNone(invitation.confirmation_email_sent_at)
        self.assertGreater(invitation.expires_at, timezone.now() + timedelta(hours=71))
        self.assertLess(invitation.expires_at, timezone.now() + timedelta(hours=73, minutes=5))
        self.assertIn("Accept account invitation", mail.outbox[-1].alternatives[0][0])
        self.assertIn("token_type=admin_invitation", mail.outbox[-1].body)

        users = self.client.get("/api/admin/users/?page=1")
        self.assertEqual(users.status_code, status.HTTP_200_OK)
        created_row = next(row for row in users.data["results"] if row["id"] == user.id)
        self.assertEqual(created_row["invitation_status"], "pending_acceptance")
        self.assertFalse(created_row["is_active"])

    def test_admin_invitation_flow_accepts_then_activates_after_password_setup(self):
        create_res = self._create_invited_user(username="workflow_user", email="workflow_user@example.com")
        self.assertEqual(create_res.status_code, status.HTTP_201_CREATED)
        self.client.force_authenticate(user=None)

        blocked_reset = self.client.post("/api/auth/password-reset/", {"email": "workflow_user@example.com"}, format="json")
        self.assertEqual(blocked_reset.status_code, status.HTTP_200_OK)
        self.assertEqual(len(mail.outbox), 1)

        invite_token = self._invitation_token_from_email()
        accept_res = self.client.post(
            "/api/auth/verify-email/",
            {"token": invite_token, "token_type": "admin_invitation"},
            format="json",
        )
        self.assertEqual(accept_res.status_code, status.HTTP_200_OK)
        self.assertTrue(accept_res.data.get("invitation_accepted"))

        user = User.objects.get(username="workflow_user")
        user.refresh_from_db()
        self.assertFalse(user.is_active)
        profile = UserProfile.objects.get(user=user)
        self.assertIsNotNone(profile.email_verified_at)

        new_secret = _test_secret()
        confirm_res = self.client.post(
            "/api/auth/password-reset-confirm/",
            {
                "uid": accept_res.data["reset_uid"],
                "token": accept_res.data["reset_token"],
                "new_password": new_secret,
                "new_password_confirm": new_secret,
            },
            format="json",
        )
        self.assertEqual(confirm_res.status_code, status.HTTP_200_OK)
        self.assertTrue(confirm_res.data.get("activated"))
        self.assertTrue(confirm_res.data.get("activation_email_sent"))

        user.refresh_from_db()
        self.assertTrue(user.is_active)
        self.assertTrue(user.check_password(new_secret))
        invitation = AdminUserInvitation.objects.get(user=user)
        self.assertIsNotNone(invitation.accepted_at)
        self.assertIsNotNone(invitation.password_reset_completed_at)
        self.assertIsNotNone(invitation.activated_at)
        self.assertIsNotNone(invitation.welcome_email_sent_at)
        self.assertEqual(len(mail.outbox), 2)
        self.assertIn("Your account is now active", mail.outbox[-1].subject)
        self.assertTrue(
            AuditLog.objects.filter(object_id=str(user.pk), changes__event="admin_invitation_accepted").exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(object_id=str(user.pk), changes__event="password_reset_completed").exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(object_id=str(user.pk), changes__event="account_activated").exists()
        )

        login = self.client.post(
            "/api/auth/token/",
            {"username": "workflow_user", "password": new_secret},
            format="json",
        )
        self.assertEqual(login.status_code, status.HTTP_200_OK)

    def test_admin_invitation_rejects_invalid_or_expired_verification(self):
        res = self._create_invited_user(username="expired_user", email="expired_user@example.com")
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        invitation = AdminUserInvitation.objects.get(user__username="expired_user")
        invitation.expires_at = timezone.now() - timedelta(minutes=1)
        invitation.save(update_fields=["expires_at", "updated_at"])

        invite_token = self._invitation_token_from_email()
        accept_res = self.client.post(
            "/api/auth/verify-email/",
            {"token": invite_token, "token_type": "admin_invitation"},
            format="json",
        )
        self.assertEqual(accept_res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("expired", str(accept_res.data).lower())

    def test_admin_invitation_email_failure_returns_warning_and_logs_event(self):
        with patch("core.views._send_html_email_message", side_effect=RuntimeError("smtp down")):
            res = self._create_invited_user(username="mail_fail_user", email="mail_fail_user@example.com")
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertFalse(res.data.get("invitation_sent"))
        self.assertIn("could not be delivered", str(res.data.get("detail", "")).lower())
        user = User.objects.get(username="mail_fail_user")
        self.assertTrue(
            AuditLog.objects.filter(action="security", object_id=str(user.pk), changes__event="admin_invitation_email_failed").exists()
        )

class LogoutTests(APITestCase):
    def test_logout_deletes_token(self):
        u = User.objects.create_user(username="u3", password=_test_secret(), email="u3@example.com")
        role = Role.objects.get(name="user")
        token = AccessToken.objects.create(user=u, role=role, key="logout_token_1")
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {token.key}")
        res = self.client.post("/api/auth/logout/", {}, format="json")
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)
        token.refresh_from_db()
        self.assertIsNotNone(token.revoked_at)


class OAuthRedirectTests(APITestCase):
    def setUp(self):
        super().setUp()
        cache.clear()

    def test_google_start_not_configured_redirects_to_frontend_callback(self):
        with override_settings(FRONTEND_BASE_URL="http://frontend.test", GOOGLE_OAUTH_CLIENT_ID="", GOOGLE_OAUTH_CLIENT_SECRET=""):
            res = self.client.get("/api/auth/google/start/")
        self.assertEqual(res.status_code, 302)
        self.assertTrue(res["Location"].startswith("http://frontend.test/auth/callback#"))
        self.assertIn("provider=google", res["Location"])
        self.assertIn("error=not_configured", res["Location"])

    def test_google_start_missing_secret_redirects_to_frontend_callback(self):
        with override_settings(FRONTEND_BASE_URL="http://frontend.test", GOOGLE_OAUTH_CLIENT_ID="cid123", GOOGLE_OAUTH_CLIENT_SECRET=""):
            res = self.client.get("/api/auth/google/start/")
        self.assertEqual(res.status_code, 302)
        self.assertTrue(res["Location"].startswith("http://frontend.test/auth/callback#"))
        self.assertIn("provider=google", res["Location"])
        self.assertIn("error=not_configured", res["Location"])

    def test_google_start_configured_redirects_to_google(self):
        with override_settings(FRONTEND_BASE_URL="http://frontend.test", GOOGLE_OAUTH_CLIENT_ID="cid123", GOOGLE_OAUTH_CLIENT_SECRET="sec123"):
            res = self.client.get("/api/auth/google/start/")
        self.assertEqual(res.status_code, 302)
        loc = res["Location"]
        self.assertTrue(loc.startswith("https://accounts.google.com/o/oauth2/v2/auth?"))
        self.assertIn("client_id=cid123", loc)
        self.assertIn("redirect_uri=http%3A%2F%2Ftestserver%2Fapi%2Fauth%2Fgoogle%2Fcallback%2F", loc)
        self.assertIn("response_type=code", loc)
        self.assertIn("scope=openid+email+profile", loc)

    def test_google_callback_cancelled_redirects_to_frontend(self):
        with override_settings(FRONTEND_BASE_URL="http://frontend.test", GOOGLE_OAUTH_CLIENT_ID="x", GOOGLE_OAUTH_CLIENT_SECRET="y"):
            state = "state123"
            cache.set(f"google_oauth_state:{state}", {"ip": "127.0.0.1"}, timeout=600)
            res = self.client.get(f"/api/auth/google/callback/?error=access_denied&state={state}")
        self.assertEqual(res.status_code, 302)
        self.assertIn("provider=google", res["Location"])
        self.assertIn("error=cancelled", res["Location"])

    def test_facebook_start_not_configured_redirects_to_frontend_callback(self):
        with override_settings(FRONTEND_BASE_URL="http://frontend.test", FACEBOOK_OAUTH_CLIENT_ID="", FACEBOOK_OAUTH_CLIENT_SECRET=""):
            res = self.client.get("/api/auth/facebook/start/")
        self.assertEqual(res.status_code, 302)
        self.assertTrue(res["Location"].startswith("http://frontend.test/auth/callback#"))
        self.assertIn("provider=facebook", res["Location"])
        self.assertIn("error=not_configured", res["Location"])

    def test_facebook_callback_cancelled_redirects_to_frontend(self):
        with override_settings(FRONTEND_BASE_URL="http://frontend.test", FACEBOOK_OAUTH_CLIENT_ID="x", FACEBOOK_OAUTH_CLIENT_SECRET="y"):
            state = "statefb"
            cache.set(f"facebook_oauth_state:{state}", {"ip": "127.0.0.1"}, timeout=600)
            res = self.client.get(f"/api/auth/facebook/callback/?error=access_denied&state={state}")
        self.assertEqual(res.status_code, 302)
        self.assertIn("provider=facebook", res["Location"])
        self.assertIn("error=cancelled", res["Location"])


class ApiCoverageTests(APITestCase):
    def setUp(self):
        super().setUp()
        cache.clear()
        self.admin = User.objects.create_user(username="admin2", password=_test_secret(), is_staff=True, is_superuser=True, email="admin2@example.com")
        self.staff = User.objects.create_user(username="staff2", password=_test_secret(), is_staff=True, email="staff2@example.com")
        self.user = User.objects.create_user(username="u_cov", password=_test_secret(), email="u_cov@example.com")

    def test_geo_and_audit_and_admin_users(self):
        geo = self.client.get("/api/settings/geo/", HTTP_ACCEPT_LANGUAGE="de")
        self.assertEqual(geo.status_code, status.HTTP_200_OK)

        self.client.force_authenticate(user=self.staff)
        audit = self.client.get("/api/settings/audit/?scope=all&limit=5")
        self.assertEqual(audit.status_code, status.HTTP_200_OK)

        users_denied = self.client.get("/api/admin/users/?page=1")
        self.assertEqual(users_denied.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(user=self.admin)
        users = self.client.get("/api/admin/users/?page=1")
        self.assertEqual(users.status_code, status.HTTP_200_OK)

        self.client.force_authenticate(user=self.user)
        denied = self.client.get("/api/admin/users/?page=1")
        self.assertEqual(denied.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(user=self.admin)
        created_secret = _test_secret()
        created = self.client.post(
            "/api/admin/users/",
            {
                "username": "created_user",
                "password": created_secret,
                "email": "created_user@example.com",
                "company_name": "Created User Co",
                "phone": "8044444444",
                "primary_role": "user",
                "custom_roles": [],
                "is_active": True,
            },
            format="json",
        )
        self.assertEqual(created.status_code, status.HTTP_201_CREATED)
        self.assertTrue(created.data.get("invitation_sent"))
        self.assertTrue(AuditLog.objects.filter(action="create", object_id=str(created.data["id"])).exists())
        patched = self.client.patch("/api/admin/users/", {"id": created.data["id"], "primary_role": "staff"}, format="json")
        self.assertEqual(patched.status_code, status.HTTP_200_OK)
        self.assertTrue(AuditLog.objects.filter(action="update", object_id=str(created.data["id"])).exists())

    def test_admin_user_delete_enforces_permissions_and_cleans_personal_artifacts(self):
        target = User.objects.create_user(
            username="delete_me",
            password=_test_secret(),
            email="delete_me@example.com",
            is_active=False,
        )
        UserProfile.objects.create(
            user=target,
            full_name="Delete Me",
            phone="+2348011111111",
            company_legal_name="Delete Me Co",
        )
        customer = Customer.objects.create(name="Delete User Customer", email="delete-user-customer@example.com")
        invoice = Invoice.objects.create(
            invoice_number="INV-DELETE-0001",
            customer=customer,
            status="Draft",
            issue_date=timezone.localdate(),
            subtotal=Decimal("10.00"),
            tax_total=Decimal("0.00"),
            total_amount=Decimal("10.00"),
        )
        invitation = AdminUserInvitation.objects.create(
            user=target,
            invited_by=self.admin,
            token_key=secrets.token_urlsafe(16),
            expires_at=timezone.now() + timedelta(days=1),
        )
        SavedInvoiceView.objects.create(user=target, name="Delete View", filters={"status": "Draft"})
        DocumentDelivery.objects.create(
            user=target,
            document_type="invoice",
            invoice=invoice,
            channel="share",
            format="pdf",
            status="queued",
        )
        SavedDocument.objects.create(
            user=target,
            document_type="invoice",
            invoice=invoice,
            original_filename="delete-me.pdf",
            file=SimpleUploadedFile("delete-me.pdf", b"%PDF-1.4 test", content_type="application/pdf"),
            sha256="a" * 64,
            size_bytes=12,
            metadata={"ok": True},
        )
        SocialAuthConnection.objects.create(user=target, provider="google", provider_user_id="sub-delete-me")
        EmailVerificationToken.objects.create(user=target, token="verify-delete-me", expires_at=timezone.now() + timedelta(hours=1))
        role = Role.objects.create(name="ops_delete_test", description="Delete flow role")
        UserRole.objects.create(user=target, role=role)
        AccessToken.objects.create(key="tok_delete_me", user=target, role=role_for_name("user"))

        denied_anon = self.client.delete(
            "/api/admin/users/",
            {"id": target.id, "confirm_keyword": "DELETE", "confirm_email": target.email},
            format="json",
        )
        self.assertIn(denied_anon.status_code, {status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN})

        self.client.force_authenticate(user=self.user)
        denied_user = self.client.delete(
            "/api/admin/users/",
            {"id": target.id, "confirm_keyword": "DELETE", "confirm_email": target.email},
            format="json",
        )
        self.assertEqual(denied_user.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(user=self.admin)
        missing_confirm = self.client.delete(
            "/api/admin/users/",
            {"id": target.id, "confirm_keyword": "DELETE", "confirm_email": "wrong@example.com"},
            format="json",
        )
        self.assertEqual(missing_confirm.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("confirm_email", missing_confirm.data)

        self_delete = self.client.delete(
            "/api/admin/users/",
            {"id": self.admin.id, "confirm_keyword": "DELETE", "confirm_email": self.admin.email},
            format="json",
        )
        self.assertEqual(self_delete.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("id", self_delete.data)

        response = self.client.delete(
            "/api/admin/users/",
            {"id": target.id, "confirm_keyword": "DELETE", "confirm_email": target.email},
            format="json",
            HTTP_X_FORWARDED_FOR="203.0.113.55",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data["deleted"])
        self.assertFalse(User.objects.filter(pk=target.id).exists())
        self.assertFalse(UserProfile.objects.filter(user_id=target.id).exists())
        self.assertFalse(AdminUserInvitation.objects.filter(pk=invitation.pk).exists())
        self.assertFalse(SavedInvoiceView.objects.filter(user_id=target.id).exists())
        self.assertFalse(DocumentDelivery.objects.filter(user_id=target.id).exists())
        self.assertFalse(SavedDocument.objects.filter(user_id=target.id).exists())
        self.assertFalse(SocialAuthConnection.objects.filter(user_id=target.id).exists())
        self.assertFalse(EmailVerificationToken.objects.filter(user_id=target.id).exists())
        self.assertFalse(UserRole.objects.filter(user_id=target.id).exists())
        self.assertFalse(AccessToken.objects.filter(user_id=target.id).exists())

        delete_log = AuditLog.objects.filter(action="delete", object_id=f"deleted-user:{target.id}").first()
        self.assertIsNotNone(delete_log)
        self.assertEqual(delete_log.changes["deleted_user"]["email"], "delete_me@example.com")
        self.assertEqual(delete_log.changes["actor_admin_id"], self.admin.id)
        self.assertEqual(delete_log.changes["ip"], "203.0.113.55")

    def test_admin_user_delete_missing_user_and_database_errors_are_structured(self):
        self.client.force_authenticate(user=self.admin)

        missing = self.client.delete(
            "/api/admin/users/",
            {"id": 999999, "confirm_keyword": "DELETE", "confirm_email": "missing@example.com"},
            format="json",
        )
        self.assertEqual(missing.status_code, status.HTTP_404_NOT_FOUND)

        target = User.objects.create_user(username="delete_db_error", password=_test_secret(), email="delete_db_error@example.com")
        with patch("core.views._anonymize_user_artifacts_for_deletion", side_effect=DatabaseError("db down")):
            errored = self.client.delete(
                "/api/admin/users/",
                {"id": target.id, "confirm_keyword": "DELETE", "confirm_email": target.email},
                format="json",
            )
        self.assertEqual(errored.status_code, status.HTTP_503_SERVICE_UNAVAILABLE)
        self.assertEqual(errored.data["code"], "user_delete_failed")
        self.assertTrue(User.objects.filter(pk=target.id).exists())

    def test_admin_runtime_diagnostics_is_admin_only_and_reports_db_and_cache(self):
        anon = self.client.get("/api/admin/runtime/diagnostics/")
        self.assertIn(anon.status_code, {status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN})

        self.client.force_authenticate(user=self.user)
        denied = self.client.get("/api/admin/runtime/diagnostics/")
        self.assertEqual(denied.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(user=self.admin)
        ok = self.client.get("/api/admin/runtime/diagnostics/")
        self.assertEqual(ok.status_code, status.HTTP_200_OK)
        self.assertIn("db", ok.data)
        self.assertIn("cache", ok.data)
        self.assertTrue(ok.data["db"].get("engine"))
        self.assertTrue(ok.data["cache"].get("class"))

    def test_admin_created_standard_user_gets_default_business_role_and_can_persist_records(self):
        self.client.force_authenticate(user=self.admin)
        secret = _test_secret()
        create_res = self.client.post(
            "/api/admin/users/",
            {
                "username": "ops_admin_created@example.com",
                "email": "ops_admin_created@example.com",
                "password": secret,
                "company_name": "Ops Admin Created Co",
                "phone": "8077777777",
                "primary_role": "user",
                "custom_roles": [],
                "is_active": True,
            },
            format="json",
        )
        self.assertEqual(create_res.status_code, status.HTTP_201_CREATED)
        self.assertTrue(create_res.data.get("invitation_sent"))

        self.client.force_authenticate(user=None)
        invitation_email = mail.outbox[-1].body or ""
        invite_token = urllib.parse.unquote(invitation_email.split("token=", 1)[1].split("&", 1)[0].strip())
        accept = self.client.post(
            "/api/auth/verify-email/",
            {"token": invite_token, "token_type": "admin_invitation"},
            format="json",
        )
        self.assertEqual(accept.status_code, status.HTTP_200_OK)
        reset_secret = _test_secret()
        confirm = self.client.post(
            "/api/auth/password-reset-confirm/",
            {
                "uid": accept.data["reset_uid"],
                "token": accept.data["reset_token"],
                "new_password": reset_secret,
                "new_password_confirm": reset_secret,
            },
            format="json",
        )
        self.assertEqual(confirm.status_code, status.HTTP_200_OK)
        login = self.client.post(
            "/api/auth/token/",
            {"username": "ops_admin_created@example.com", "password": reset_secret},
            format="json",
        )
        self.assertEqual(login.status_code, status.HTTP_200_OK)
        token = login.data["token"]
        self.client.credentials(HTTP_AUTHORIZATION=f"Token {token}")

        me = self.client.get("/api/auth/me/")
        self.assertEqual(me.status_code, status.HTTP_200_OK)
        self.assertIn("editor", me.data.get("roles", []))
        self.assertIn("data.customers.write", me.data.get("permissions", []))
        self.assertIn("data.invoices.write", me.data.get("permissions", []))
        self.assertIn("data.receipts.write", me.data.get("permissions", []))

        customer_res = self.client.post(
            "/api/customers/",
            {"name": "Admin Created Buyer", "email": "buyer-admin-created@example.com", "phone": "08012340000"},
            format="json",
        )
        self.assertEqual(customer_res.status_code, status.HTTP_201_CREATED)

        item_res = self.client.post(
            "/api/items/",
            {"name": "Admin Created Widget", "unit_price": "15.00", "tax_rate": "0", "stock_quantity": 5},
            format="json",
        )
        self.assertEqual(item_res.status_code, status.HTTP_201_CREATED)

        invoice_res = self.client.post(
            "/api/invoices/",
            {
                "customer": customer_res.data["id"],
                "status": "Draft",
                "items": [{"item": item_res.data["id"], "quantity": 1}],
            },
            format="json",
        )
        self.assertEqual(invoice_res.status_code, status.HTTP_201_CREATED)

        receipt_res = self.client.post(
            f"/api/invoices/{invoice_res.data['id']}/pay/",
            {"amount_paid": "15.00", "payment_method": "Cash"},
            format="json",
            HTTP_IDEMPOTENCY_KEY="k-admin-created-editor-1",
        )
        self.assertEqual(receipt_res.status_code, status.HTTP_201_CREATED)
        self.assertTrue(Customer.objects.filter(pk=customer_res.data["id"], is_deleted=False).exists())
        self.assertTrue(Invoice.objects.filter(pk=invoice_res.data["id"], is_deleted=False).exists())
        self.assertTrue(Receipt.objects.filter(pk=receipt_res.data["id"], is_deleted=False).exists())

    def test_currency_and_exchange_rate_crud(self):
        self.client.force_authenticate(user=self.staff)
        list_res = self.client.get("/api/currencies/")
        self.assertEqual(list_res.status_code, status.HTTP_200_OK)
        codes = [row.get("code") for row in list_res.data]
        self.assertIn("NGN", codes)
        ngn = next((row for row in list_res.data if row.get("code") == "NGN"), None)
        self.assertIsNotNone(ngn)
        self.assertEqual(ngn.get("name"), "Nigerian Naira")
        self.assertEqual(ngn.get("symbol"), "₦")
        self.assertEqual(int(ngn.get("decimal_places")), 2)

        self.client.force_authenticate(user=self.admin)
        created = self.client.post("/api/currencies/", {"code": "KES", "name": "Kenyan Shilling", "symbol": "KSh", "decimal_places": 2}, format="json")
        self.assertEqual(created.status_code, status.HTTP_201_CREATED)

        fx = self.client.post("/api/exchange-rates/", {"base_code": "USD", "quote_code": "KES", "rate": "129.50000000"}, format="json")
        self.assertEqual(fx.status_code, status.HTTP_201_CREATED)
        fx_id = fx.data["id"]

        fx_up = self.client.patch(f"/api/exchange-rates/{fx_id}/", {"rate": "130.00000000"}, format="json")
        self.assertEqual(fx_up.status_code, status.HTTP_200_OK)

        fx_del = self.client.delete(f"/api/exchange-rates/{fx_id}/")
        self.assertEqual(fx_del.status_code, status.HTTP_204_NO_CONTENT)

    def test_admin_roles_and_audit_logs_enforce_rbac(self):
        self.client.force_authenticate(user=self.staff)
        self.assertEqual(self.client.get("/api/admin/roles/").status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(self.client.get("/api/admin/audit-logs/?page=1").status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(user=self.user)
        self.assertEqual(self.client.get("/api/admin/roles/").status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(self.client.get("/api/admin/audit-logs/?page=1").status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(user=self.admin)
        roles = self.client.get("/api/admin/roles/")
        self.assertEqual(roles.status_code, status.HTTP_200_OK)
        self.assertIn("permissions", roles.data)

        created = self.client.post(
            "/api/admin/roles/",
            {
                "name": "ops_e2e",
                "description": "Operations role for regression coverage",
                "permission_codes": ["data.items.read", "data.invoices.read"],
            },
            format="json",
        )
        self.assertEqual(created.status_code, status.HTTP_201_CREATED)

        patched = self.client.patch(
            "/api/admin/roles/",
            {
                "id": created.data["id"],
                "description": "Updated operations role",
                "permission_codes": ["data.items.read"],
            },
            format="json",
        )
        self.assertEqual(patched.status_code, status.HTTP_200_OK)
        role = Role.objects.get(pk=created.data["id"])
        self.assertEqual(role.description, "Updated operations role")
        self.assertTrue(AuditLog.objects.filter(object_id=str(role.id), action="create").exists())
        self.assertTrue(AuditLog.objects.filter(object_id=str(role.id), action="update").exists())

        audit = self.client.get(f"/api/admin/audit-logs/?page=1&q={role.id}")
        self.assertEqual(audit.status_code, status.HTTP_200_OK)
        self.assertGreaterEqual(audit.data["count"], 1)
        self.assertTrue(any(row["object_id"] == str(role.id) for row in audit.data["results"]))

    def test_invoice_pdf_and_receipt_print(self):
        self.client.force_authenticate(user=self.user)
        UserProfile.objects.update_or_create(user=self.user, defaults={"company_legal_name": "AmbienteSoft LTD"})
        customer = Customer.objects.create(name="PDF Buyer", email="p@example.com")
        item = Item.objects.create(name="Widget", unit_price=Decimal("10.00"), stock_quantity=10)
        inv = Invoice.objects.create(
            invoice_number="INV-2026-4444",
            customer=customer,
            status="Draft",
            issue_date=timezone.localdate(),
            subtotal=Decimal("10.00"),
            tax_total=Decimal("0.00"),
            total_amount=Decimal("10.00"),
        )
        InvoiceItem.objects.create(
            invoice=inv,
            item=item,
            quantity=1,
            unit_price=Decimal("10.00"),
            line_subtotal=Decimal("10.00"),
            line_tax=Decimal("0.00"),
            line_total=Decimal("10.00"),
        )

        request_stub = type("Req", (), {"user": self.user, "query_params": {}, "headers": {}})()
        rendered_invoice = render_invoice(request_stub, inv, "html")
        self.assertIn("AmbienteSoft LTD", rendered_invoice.content.decode("utf-8"))

        pdf = self.client.get(f"/api/invoices/{inv.id}/download_pdf/")
        self.assertEqual(pdf.status_code, status.HTTP_200_OK)
        self.assertIn("X-PDF-Backend", pdf)
        self.assertTrue(
            ("application/pdf" in (pdf.get("Content-Type") or "")) or ("text/html" in (pdf.get("Content-Type") or "")),
            msg=f"unexpected content-type: {pdf.get('Content-Type')}",
        )
        invoice_html = self.client.get(f"/api/invoices/{inv.id}/print_html/")
        self.assertEqual(invoice_html.status_code, status.HTTP_200_OK)
        self.assertIn("text/html", invoice_html["Content-Type"])
        self.assertContains(invoice_html, "AmbienteSoft LTD")

        receipt = Receipt.objects.create(invoice=inv, amount_paid=Decimal("10.00"), payment_method="Cash", reference_number="R1")
        rendered_receipt = render_receipt(request_stub, receipt, "html")
        receipt_body = rendered_receipt.content.decode("utf-8")
        self.assertIn("AmbienteSoft LTD", receipt_body)
        self.assertIn("Receipt", receipt_body)
        html = self.client.get(f"/api/receipts/{receipt.id}/print_html/")
        self.assertEqual(html.status_code, status.HTTP_200_OK)
        self.assertIn("text/html", html["Content-Type"])
        self.assertContains(html, "AmbienteSoft LTD")

    def test_invoice_pdf_falls_back_to_html_when_weasyprint_render_fails(self):
        self.client.force_authenticate(user=self.user)
        customer = Customer.objects.create(name="PDF Buyer2", email="p2@example.com")
        item = Item.objects.create(name="Widget2", unit_price=Decimal("10.00"), stock_quantity=10)
        inv = Invoice.objects.create(
            invoice_number="INV-2026-4445",
            customer=customer,
            status="Draft",
            issue_date=timezone.localdate(),
            subtotal=Decimal("10.00"),
            tax_total=Decimal("0.00"),
            total_amount=Decimal("10.00"),
        )
        InvoiceItem.objects.create(
            invoice=inv,
            item=item,
            quantity=1,
            unit_price=Decimal("10.00"),
            line_subtotal=Decimal("10.00"),
            line_tax=Decimal("0.00"),
            line_total=Decimal("10.00"),
        )

        try:
            import weasyprint
        except Exception:
            res = self.client.get(f"/api/invoices/{inv.id}/download_pdf/")
            self.assertEqual(res.status_code, status.HTTP_200_OK)
            self.assertIn("text/html", res.get("Content-Type", ""))
            self.assertEqual(res.get("X-PDF-Backend"), "unavailable")
            return

        with patch("weasyprint.HTML.write_pdf", side_effect=OSError("boom")):
            res = self.client.get(f"/api/invoices/{inv.id}/download_pdf/")
            self.assertEqual(res.status_code, status.HTTP_200_OK)
            self.assertIn("text/html", res.get("Content-Type", ""))
            self.assertEqual(res.get("X-PDF-Backend"), "failed")


class AuthApiTests(APITestCase):
    def setUp(self):
        super().setUp()
        cache.clear()

    def test_rate_limit_counts_over_window(self):
        cache.clear()
        key = "test:ratelimit"
        self.assertFalse(_rate_limit(key, limit=2, window_seconds=60))
        self.assertFalse(_rate_limit(key, limit=2, window_seconds=60))
        self.assertTrue(_rate_limit(key, limit=2, window_seconds=60))

    def test_login_requires_username_and_password(self):
        res = self.client.post("/api/auth/token/", {}, format="json")
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("detail", res.data)

    def test_login_invalid_credentials_returns_detail(self):
        secret = _test_secret()
        wrong_cred = "wrong"
        User.objects.create_user(username="u1", password=secret, is_active=True)
        res = self.client.post("/api/auth/token/", {"username": "u1", "password": wrong_cred}, format="json")
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(str(res.data.get("detail")), "Unable to log in with provided credentials.")

    def test_login_inactive_user_returns_403(self):
        secret = _test_secret()
        User.objects.create_user(username="u2", password=secret, is_active=False)
        res = self.client.post("/api/auth/token/", {"username": "u2", "password": secret}, format="json")
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn("verify your email", str(res.data.get("detail", "")).lower())

    def test_login_success_returns_token(self):
        secret = _test_secret()
        user = User.objects.create_user(username="u3", password=secret, is_active=True)
        res = self.client.post("/api/auth/token/", {"username": "u3", "password": secret}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn("token", res.data)
        self.assertEqual(res.data.get("role"), "user")
        self.assertTrue(AccessToken.objects.filter(user=user, role__name="user").exists())

    def test_privileged_accounts_can_login_via_standard_token(self):
        staff_secret = _test_secret()
        admin_secret = _test_secret()
        staff = User.objects.create_user(username="staff_user", password=staff_secret, is_active=True, is_staff=True)
        admin = User.objects.create_user(username="admin_user", password=admin_secret, is_active=True, is_staff=True, is_superuser=True)

        staff_res = self.client.post("/api/auth/token/", {"username": "staff_user", "password": staff_secret}, format="json")
        self.assertEqual(staff_res.status_code, status.HTTP_200_OK)
        self.assertEqual(staff_res.data.get("role"), "staff")
        self.assertTrue(AccessToken.objects.filter(user=staff, role__name="staff").exists())

        admin_res = self.client.post("/api/auth/token/", {"username": "admin_user", "password": admin_secret}, format="json")
        self.assertEqual(admin_res.status_code, status.HTTP_200_OK)
        self.assertEqual(admin_res.data.get("role"), "admin")
        self.assertTrue(AccessToken.objects.filter(user=admin, role__name="admin").exists())

    def test_removed_standalone_admin_auth_endpoints_return_404(self):
        self.assertEqual(self.client.post("/api/auth/staff/token/", {}, format="json").status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(self.client.post("/api/auth/admin/token/", {}, format="json").status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(self.client.post("/api/auth/admin/mfa/setup/", {}, format="json").status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(self.client.post("/api/auth/admin/mfa/confirm/", {}, format="json").status_code, status.HTTP_404_NOT_FOUND)

    def test_me_includes_company_name(self):
        user = User.objects.create_user(username="u_me", password=_test_secret(), is_active=True)
        UserProfile.objects.create(user=user, company_legal_name="Acme Incorporated")
        self.client.force_authenticate(user=user)
        res = self.client.get("/api/auth/me/")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data.get("company_name"), "Acme Incorporated")
        self.assertIn("roles", res.data)

    def test_me_includes_linked_social_accounts(self):
        user = User.objects.create_user(username="u_social", password=_test_secret(), is_active=True, email="u_social@example.com")
        SocialAuthConnection.objects.create(
            user=user,
            provider="google",
            provider_user_id="google-sub-1",
            email="u_social@example.com",
            display_name="U Social",
        )
        self.client.force_authenticate(user=user)
        res = self.client.get("/api/auth/me/")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        accounts = res.data.get("social_accounts") or []
        self.assertEqual(len(accounts), 1)
        self.assertEqual(accounts[0]["provider"], "google")

    def test_social_connections_endpoint_returns_provider_status(self):
        user = User.objects.create_user(username="u_links", password=_test_secret(), is_active=True, email="u_links@example.com")
        SocialAuthConnection.objects.create(
            user=user,
            provider="facebook",
            provider_user_id="fb-1",
            email="u_links@example.com",
            display_name="FB User",
        )
        self.client.force_authenticate(user=user)
        res = self.client.get("/api/auth/social/connections/")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        rows = {row["provider"]: row for row in res.data["results"]}
        self.assertTrue(rows["facebook"]["connected"])
        self.assertFalse(rows["google"]["connected"])

    def test_google_callback_links_social_account_to_existing_profile(self):
        user = User.objects.create_user(
            username="u_link_google",
            password=_test_secret(),
            is_active=True,
            email="u_link_google@example.com",
        )
        profile = UserProfile.objects.create(user=user)

        with override_settings(
            FRONTEND_BASE_URL="http://frontend.test",
            GOOGLE_OAUTH_CLIENT_ID="cid123",
            GOOGLE_OAUTH_CLIENT_SECRET="sec123",
        ):
            cache.set(
                "google_oauth_state:state-google-link-1",
                {"ip": "127.0.0.1", "intent": "link", "user_id": user.id, "remember": True},
                timeout=600,
            )
            with patch(
                "urllib.request.urlopen",
                side_effect=[
                    _MockJsonResponse({"id_token": "id-token-link-1"}),
                    _MockJsonResponse(
                        {
                            "sub": "google-sub-link-1",
                            "email": user.email,
                            "email_verified": "true",
                            "aud": "cid123",
                            "name": "Linked Google User",
                            "picture": "https://example.com/avatar.png",
                        }
                    ),
                ],
            ):
                res = self.client.get("/api/auth/google/callback/?code=abc&state=state-google-link-1")

        self.assertEqual(res.status_code, 302)
        self.assertIn("provider=google", res["Location"])
        self.assertIn("linked=1", res["Location"])

        connection = SocialAuthConnection.objects.get(provider="google", provider_user_id="google-sub-link-1")
        self.assertEqual(connection.user_id, user.id)
        self.assertEqual(connection.email, user.email)
        self.assertEqual(connection.display_name, "Linked Google User")
        self.assertIsNotNone(UserProfile.objects.get(pk=profile.pk).email_verified_at)
        self.assertTrue(
            AuditLog.objects.filter(
                action="update",
                object_id=str(connection.id),
                changes__event="social_account_linked",
            ).exists()
        )

    def test_register_and_duplicate_email(self):
        secret = _test_secret()
        ok = self.client.post(
            "/api/auth/register/",
            {
                "email": "new@example.com",
                "password": secret,
                "password_confirm": secret,
                "company_name": "NewCo Ltd",
                "country_code": "+234",
                "phone_number": "8055555555",
                "country": "Nigeria",
                "accept_terms": True,
                "website": "",
            },
            format="json",
        )
        self.assertEqual(ok.status_code, status.HTTP_201_CREATED)
        self.assertEqual(ok.data.get("registered"), True)

        dup = self.client.post(
            "/api/auth/register/",
            {
                "email": "new@example.com",
                "password": secret,
                "password_confirm": secret,
                "company_name": "NewCo Ltd",
                "country_code": "+234",
                "phone_number": "8066666666",
                "country": "Nigeria",
                "accept_terms": True,
                "website": "",
            },
            format="json",
        )
        self.assertEqual(dup.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("email", dup.data)

    def test_google_callback_blocks_privileged_account_from_social_sign_in(self):
        privileged = User.objects.create_user(
            username="admin_social",
            email="admin-social@example.com",
            password=_test_secret(),
            is_active=True,
            is_staff=True,
            is_superuser=True,
        )
        with override_settings(FRONTEND_BASE_URL="http://frontend.test", GOOGLE_OAUTH_CLIENT_ID="cid123", GOOGLE_OAUTH_CLIENT_SECRET="sec123"):
            cache.set(
                "google_oauth_state:state-google-1",
                {"ip": "127.0.0.1", "intent": "login", "remember": True},
                timeout=600,
            )
            with patch(
                "urllib.request.urlopen",
                side_effect=[
                    _MockJsonResponse({"id_token": "id-token-1"}),
                    _MockJsonResponse(
                        {
                            "sub": "google-sub-1",
                            "email": privileged.email,
                            "email_verified": "true",
                            "aud": "cid123",
                            "name": "Admin Social",
                        }
                    ),
                ],
            ):
                res = self.client.get("/api/auth/google/callback/?code=abc&state=state-google-1")
        self.assertEqual(res.status_code, 302)
        self.assertIn("error=privileged_account", res["Location"])
        self.assertFalse(SocialAuthConnection.objects.filter(provider="google", provider_user_id="google-sub-1").exists())


class DocumentDeliveryAndPaymentsTests(APITestCase):
    def setUp(self):
        super().setUp()
        self.user = User.objects.create_user(username="u_docs", password=_test_secret(), email="u_docs@example.com", is_active=True)
        self.client.force_authenticate(user=self.user)
        self.customer = Customer.objects.create(name="Buyer", email="buyer@example.com", phone="+2348012345678")
        self.item = Item.objects.create(name="Widget", unit_price=Decimal("10.00"), stock_quantity=10)
        self.invoice = Invoice.objects.create(
            invoice_number="INV-2026-DOCPAY",
            customer=self.customer,
            status="Sent",
            subtotal=Decimal("10.00"),
            tax_total=Decimal("0.00"),
            total_amount=Decimal("10.00"),
        )
        InvoiceItem.objects.create(
            invoice=self.invoice,
            item=self.item,
            quantity=1,
            unit_price=Decimal("10.00"),
            line_subtotal=Decimal("10.00"),
            line_tax=Decimal("0.00"),
            line_total=Decimal("10.00"),
        )

    def test_delivery_download_with_token_allows_anonymous_access(self):
        delivery, token = create_delivery(
            user=self.user,
            document_type="invoice",
            document_id=self.invoice.id,
            channel="share",
            fmt="text",
            ttl_minutes=60,
        )
        self.client.force_authenticate(user=None)
        res = self.client.get(f"/api/documents/deliveries/{delivery.id}/download/?token={token}")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(res["Content-Type"].startswith("text/plain"))
        self.assertIn("INVOICE", (res.content or b"").decode("utf-8", errors="replace"))

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_delivery_send_returns_success_report_for_email(self):
        res = self.client.post(
            "/api/documents/deliveries/",
            {
                "document_type": "invoice",
                "document_id": self.invoice.id,
                "channel": "email",
                "format": "pdf",
                "to_email": "buyer@example.com",
                "send_now": True,
            },
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertIn("delivery", res.data)
        self.assertIn("report", res.data)
        self.assertTrue(res.data["report"]["ok"])
        self.assertEqual(res.data["report"]["status"], "sent")
        self.assertEqual(res.data["report"]["channel"], "email")
        self.assertEqual(res.data["report"]["recipient"]["email"], "buyer@example.com")
        self.assertIsNotNone(res.data["report"]["last_attempt_at"])
        self.assertEqual(len(mail.outbox), 1)
        delivery_id = res.data["report"]["delivery_id"]
        self.assertTrue(AuditLog.objects.filter(object_id=str(delivery_id), action="create").exists())

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_delivery_send_supports_custom_email_templates(self):
        res = self.client.post(
            "/api/documents/deliveries/",
            {
                "document_type": "invoice",
                "document_id": self.invoice.id,
                "channel": "email",
                "format": "pdf",
                "to_email": "buyer@example.com",
                "email_subject_template": "Billing packet for {document_number}",
                "email_message_template": "Hello {customer_name}, download here: {download_url}",
                "send_now": True,
            },
            format="json",
        )
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].subject, "Billing packet for INV-2026-DOCPAY")
        self.assertIn("Hello Buyer, download here:", mail.outbox[0].body)
        self.assertIn("/api/documents/deliveries/", mail.outbox[0].body)

    def test_invoice_share_link_returns_download_url_and_token_works(self):
        res = self.client.post(f"/api/invoices/{self.invoice.id}/share_link/", {"ttl_minutes": 60}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn("download_url", res.data)
        download_url = str(res.data["download_url"])
        self.assertIn("/api/documents/deliveries/", download_url)
        self.assertIn("/download/?token=", download_url)

        from urllib.parse import urlsplit, parse_qs

        parts = urlsplit(download_url)
        token = (parse_qs(parts.query).get("token") or [""])[0]
        self.assertTrue(token)
        delivery_id = int(parts.path.strip("/").split("/")[-2])

        self.client.force_authenticate(user=None)
        dl = self.client.get(f"/api/documents/deliveries/{delivery_id}/download/?token={token}")
        self.assertEqual(dl.status_code, status.HTTP_200_OK)

    def test_receipt_share_link_returns_download_url(self):
        r = self.client.post(
            "/api/receipts/",
            {"invoice": self.invoice.id, "amount_paid": "5.00", "payment_method": "Cash", "payment_date": str(timezone.localdate())},
            format="json",
        )
        self.assertEqual(r.status_code, status.HTTP_201_CREATED)
        receipt_id = int(r.data["id"])
        res = self.client.post(f"/api/receipts/{receipt_id}/share_link/", {"ttl_minutes": 60}, format="json")
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn("download_url", res.data)

    def test_save_document_backup_creates_record_and_downloads_pdf(self):
        with patch(
            "core.documents.render_invoice",
            return_value=RenderedDocument(
                filename="invoice_INV-2026-DOCPAY.pdf",
                content_type="application/pdf",
                content=b"%PDF-1.4\nmock\n",
                backend="mock",
            ),
        ):
            res = self.client.post(
                "/api/documents/saved/",
                {
                    "document_type": "invoice",
                    "document_id": self.invoice.id,
                    "label": self.invoice.invoice_number,
                },
                format="json",
            )
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertIn("download_url", res.data)
        saved_id = int(res.data["saved_document"]["id"])
        saved = SavedDocument.objects.get(pk=saved_id)
        self.assertEqual(saved.document_type, "invoice")
        self.assertTrue(saved.file.name)
        dl = self.client.get(res.data["download_url"])
        self.assertEqual(dl.status_code, status.HTTP_200_OK)
        self.assertTrue(dl["Content-Type"].startswith("application/pdf"))

    def test_save_document_backup_reports_missing_weasyprint_runtime(self):
        with patch(
            "core.documents.render_invoice",
            return_value=RenderedDocument(
                filename="invoice_INV-2026-DOCPAY.html",
                content_type="text/html; charset=utf-8",
                content=b"<html></html>",
                backend="unavailable",
            ),
        ):
            res = self.client.post(
                "/api/documents/saved/",
                {
                    "document_type": "invoice",
                    "document_id": self.invoice.id,
                    "label": self.invoice.invoice_number,
                },
                format="json",
            )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            res.data["detail"],
            "PDF generation is unavailable on this server because required WeasyPrint system libraries are missing.",
        )

    @override_settings(PAYSTACK_SECRET_KEY=GW_SHARED)
    def test_paystack_webhook_marks_transaction_succeeded_and_creates_receipt(self):
        tx = PaymentTransaction.objects.create(
            invoice=self.invoice,
            created_by=self.user,
            provider="paystack",
            status="pending",
            amount=Decimal("10.00"),
            currency_code="NGN",
            reference="REFPAYSTACK123",
        )
        payload = {"event": "charge.success", "data": {"id": 991, "reference": tx.reference, "status": "success", "paid_at": timezone.now().isoformat()}}
        raw = json.dumps(payload).encode("utf-8")
        sig = hmac.new(GW_SHARED.encode("utf-8"), raw, hashlib.sha512).hexdigest()
        res = self.client.post("/api/payments/webhooks/paystack/", data=raw, content_type="application/json", HTTP_X_PAYSTACK_SIGNATURE=sig)
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        tx.refresh_from_db()
        self.assertEqual(tx.status, "succeeded")
        self.assertTrue(Receipt.objects.filter(invoice=self.invoice, is_deleted=False).exists())
