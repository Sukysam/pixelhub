from __future__ import annotations

import csv
import io
import logging
import os
import re
import uuid
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Callable

from django.core.cache import cache
from django.core.validators import validate_email
from django.core.exceptions import ValidationError as DjangoValidationError
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import Customer, Expense, Invoice, InvoiceItem, Item, SourceAccount
from .project_code_service import generate_next_project_code, normalize_project_code

CENTS = Decimal("0.01")
logger = logging.getLogger(__name__)


def _q2(value: Decimal) -> Decimal:
    return value.quantize(CENTS, rounding=ROUND_HALF_UP)


def _parse_upload_rows(upload) -> list[dict]:
    ext = os.path.splitext(getattr(upload, "name", "") or "")[1].lower()
    rows: list[dict] = []
    if ext == ".csv":
        text = io.TextIOWrapper(upload.file, encoding="utf-8-sig")
        reader = csv.DictReader(text)
        for idx, raw in enumerate(reader, start=2):
            rows.append({"_row": idx, **{(k or "").strip(): (v.strip() if isinstance(v, str) else v) for k, v in (raw or {}).items()}})
        return rows
    if ext == ".xlsx":
        wb = load_workbook(upload.file, read_only=True, data_only=True)
        ws = wb.active
        header = []
        for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                header = [str(c or "").strip() for c in r]
                continue
            data = {}
            for k, v in zip(header, r):
                if not k:
                    continue
                data[k] = v
            rows.append({"_row": i, **data})
        return rows
    raise ValueError("Unsupported file type")


def _cache_error_log(errors: list[dict]) -> str:
    token = uuid.uuid4()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["row", "field", "message"])
    for e in errors:
        w.writerow([e.get("row"), e.get("field"), e.get("message")])
    cache.set(f"import_error_log:{token}", out.getvalue(), timeout=60 * 60)
    return str(token)


def process_batch_import(
    raw_items: list[dict],
    *,
    validate_item: Callable[[dict], tuple[Any | None, list[dict]]],
    persist_valid_items: Callable[[list[Any]], tuple[int, list[dict] | None]],
    dry_run: bool,
    rollback_on_error: bool,
    imported_key: str = "imported",
    planned_key: str = "would_create",
    empty_success_value: int = 0,
) -> tuple[int, dict]:
    valid_items: list[Any] = []
    errors: list[dict] = []
    successful_items: list[dict] = []
    failed_items: list[dict] = []

    for raw_item in raw_items:
        row_num = int(raw_item.get("_row") or 0)
        try:
            normalized, item_errors = validate_item(raw_item)
        except Exception as exc:
            logger.exception("batch_import.item_exception row=%s error=%s", row_num, exc)
            normalized = None
            item_errors = [{"row": row_num, "field": "non_field_errors", "message": "Unexpected import error"}]

        if item_errors:
            errors.extend(item_errors)
            failed_items.append(
                {
                    "row": row_num,
                    "item": {k: v for k, v in raw_item.items() if k != "_row"},
                    "errors": item_errors,
                }
            )
            logger.warning("batch_import.item_failed row=%s errors=%s", row_num, item_errors)
            continue

        valid_items.append(normalized)
        successful_items.append(
            {
                "row": row_num,
                "item": {k: v for k, v in raw_item.items() if k != "_row"},
            }
        )

    if dry_run:
        logger.info(
            "batch_import.dry_run rows=%s planned=%s failed_items=%s",
            len(raw_items),
            len(valid_items),
            len(failed_items),
        )
        return 200, {
            "dry_run": True,
            "rows": len(raw_items),
            planned_key: len(valid_items),
            "errors": errors,
            "successful_items": successful_items,
            "failed_items": failed_items,
        }

    if errors and rollback_on_error:
        token = _cache_error_log(errors)
        logger.info(
            "batch_import.rollback rows=%s valid_items=%s failed_items=%s",
            len(raw_items),
            len(valid_items),
            len(failed_items),
        )
        return 400, {
            imported_key: empty_success_value,
            "rows": len(raw_items),
            "errors": errors[:200],
            "successful_items": [],
            "failed_items": failed_items[:200],
            "error_log_token": token,
            "rolled_back": True,
        }

    created_count, persisted_items = persist_valid_items(valid_items)
    logger.info(
        "batch_import.complete rows=%s created=%s failed_items=%s",
        len(raw_items),
        created_count,
        len(failed_items),
    )
    return 200, {
        imported_key: created_count,
        "rows": len(raw_items),
        "errors": errors,
        "successful_items": persisted_items if persisted_items is not None else successful_items,
        "failed_items": failed_items,
    }


def import_items_from_upload(upload, dry_run: bool, rollback_on_error: bool) -> tuple[int, dict]:
    rows = _parse_upload_rows(upload)

    def _to_decimal(value, field: str):
        if value in (None, ""):
            return None, None
        try:
            return Decimal(str(value)), None
        except (InvalidOperation, TypeError):
            return None, f"Invalid {field}"

    allowed_types = {t for t, _ in Item.TYPE_CHOICES}
    seen_skus: set[str] = set()
    seen_name_keys: set[str] = set()

    sku_values: list[str] = []
    for r in rows:
        sku = str(r.get("sku") or "").strip() or None
        if sku:
            sku_values.append(sku)
    existing_skus = set(Item.objects.filter(sku__in=sku_values, is_deleted=False).values_list("sku", flat=True))

    def _validate_row(r: dict) -> tuple[Item | None, list[dict]]:
        row_num = int(r.get("_row") or 0)
        raw_type = str(r.get("type") or "product").strip() or "product"
        raw_name = str(r.get("name") or "").strip()
        raw_category = re.sub(r"\s+", " ", str(r.get("category") or "General").strip()) or "General"
        raw_sku = str(r.get("sku") or "").strip() or None
        raw_desc = r.get("description")
        desc = str(raw_desc).strip() if raw_desc not in (None, "") else None
        uom = str(r.get("unit_of_measure") or "pcs").strip() or "pcs"
        tax_category = str(r.get("tax_category") or "standard").strip() or "standard"
        row_errors: list[dict] = []

        if raw_type not in allowed_types:
            row_errors.append({"row": row_num, "field": "type", "message": "Invalid type"})
        if not raw_name:
            row_errors.append({"row": row_num, "field": "name", "message": "Name is required"})
        if not raw_category:
            row_errors.append({"row": row_num, "field": "category", "message": "category is required"})

        if raw_sku:
            if raw_sku in seen_skus:
                row_errors.append({"row": row_num, "field": "sku", "message": "Duplicate sku in file"})
            else:
                seen_skus.add(raw_sku)
                if raw_sku in existing_skus:
                    row_errors.append({"row": row_num, "field": "sku", "message": "sku already exists"})
        else:
            key = f"{raw_type}:{raw_name.lower()}"
            if key in seen_name_keys:
                row_errors.append({"row": row_num, "field": "name", "message": "Duplicate name/type in file"})
            else:
                seen_name_keys.add(key)

        unit_price, unit_price_err = _to_decimal(r.get("unit_price"), "unit_price")
        if unit_price_err:
            row_errors.append({"row": row_num, "field": "unit_price", "message": unit_price_err})
        if unit_price is None:
            row_errors.append({"row": row_num, "field": "unit_price", "message": "unit_price is required"})
        elif unit_price < 0:
            row_errors.append({"row": row_num, "field": "unit_price", "message": "unit_price must be >= 0"})

        tax_rate, tax_rate_err = _to_decimal(r.get("tax_rate", 0), "tax_rate")
        if tax_rate_err:
            row_errors.append({"row": row_num, "field": "tax_rate", "message": tax_rate_err})
        if tax_rate is None:
            tax_rate = Decimal("0")
        if tax_rate < 0 or tax_rate > 100:
            row_errors.append({"row": row_num, "field": "tax_rate", "message": "tax_rate must be between 0 and 100"})

        stock_qty_raw = r.get("stock_quantity", 0)
        try:
            stock_qty = int(stock_qty_raw) if stock_qty_raw not in (None, "") else 0
        except (TypeError, ValueError):
            stock_qty = None
            row_errors.append({"row": row_num, "field": "stock_quantity", "message": "stock_quantity must be an integer"})
        if stock_qty is not None and stock_qty < 0:
            row_errors.append({"row": row_num, "field": "stock_quantity", "message": "stock_quantity must be >= 0"})
        if raw_type == "service":
            stock_qty = 0

        if row_errors:
            return None, row_errors

        return (
            Item(
                type=raw_type,
                sku=raw_sku,
                name=raw_name,
                category=raw_category,
                description=desc,
                unit_price=unit_price,
                tax_rate=tax_rate,
                tax_category=tax_category,
                unit_of_measure=uom,
                stock_quantity=stock_qty or 0,
            ),
            [],
        )

    def _persist(valid_items: list[Item]) -> tuple[int, list[dict] | None]:
        created = 0
        with transaction.atomic():
            batch_size = 1000
            for i in range(0, len(valid_items), batch_size):
                Item.objects.bulk_create(valid_items[i : i + batch_size], batch_size=batch_size)
                created += len(valid_items[i : i + batch_size])
        return created, None

    return process_batch_import(
        rows,
        validate_item=_validate_row,
        persist_valid_items=_persist,
        dry_run=dry_run,
        rollback_on_error=rollback_on_error,
    )


def import_invoices_from_upload(
    upload,
    dry_run: bool,
    rollback_on_error: bool,
    deduct_inventory_for_invoice: Callable[[Invoice], None],
) -> tuple[int, dict]:
    rows = _parse_upload_rows(upload)
    errors: list[dict] = []

    allowed_status = {s for s, _ in Invoice.STATUS_CHOICES}

    customer_emails: set[str] = set()
    customer_names: set[str] = set()
    item_skus: set[str] = set()
    invoice_numbers: set[str] = set()

    def _clean(value) -> str:
        return str(value or "").strip()

    def _parse_date(value, field: str, row_num: int):
        raw = _clean(value)
        if not raw:
            return None
        try:
            return date.fromisoformat(raw)
        except ValueError:
            errors.append({"row": row_num, "field": field, "message": "Invalid date. Use YYYY-MM-DD"})
            return None

    def _parse_int(value, field: str, row_num: int):
        if value in (None, ""):
            errors.append({"row": row_num, "field": field, "message": f"{field} is required"})
            return None
        try:
            v = int(value)
        except (TypeError, ValueError):
            errors.append({"row": row_num, "field": field, "message": f"{field} must be an integer"})
            return None
        return v

    def _parse_decimal(value, field: str, row_num: int):
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, TypeError):
            errors.append({"row": row_num, "field": field, "message": f"Invalid {field}"})
            return None

    groups: dict[str, list[dict]] = {}
    for r in rows:
        row_num = int(r.get("_row") or 0)
        inv_no = _clean(r.get("invoice_number"))
        inv_key = _clean(r.get("invoice_key"))
        key = inv_no or inv_key or f"row:{row_num}"
        groups.setdefault(key, []).append(r)

        if inv_no:
            invoice_numbers.add(inv_no)

        email = _clean(r.get("customer_email"))
        name = _clean(r.get("customer_name"))
        if email:
            customer_emails.add(email.lower())
        elif name:
            customer_names.add(name.lower())
        else:
            errors.append({"row": row_num, "field": "customer_email", "message": "customer_email or customer_name is required"})

        sku = _clean(r.get("item_sku"))
        if not sku:
            errors.append({"row": row_num, "field": "item_sku", "message": "item_sku is required"})
        else:
            item_skus.add(sku)

    if invoice_numbers:
        existing = set(Invoice.objects.filter(invoice_number__in=list(invoice_numbers), is_deleted=False).values_list("invoice_number", flat=True))
    else:
        existing = set()
    for inv_no in existing:
        errors.append({"row": 0, "field": "invoice_number", "message": f"invoice_number already exists: {inv_no}"})

    customers_by_email = {c.email.lower(): c for c in Customer.objects.filter(email__in=list(customer_emails), is_deleted=False) if c.email}
    customers_by_name = {c.name.lower(): c for c in Customer.objects.filter(name__in=list(customer_names), is_deleted=False)}
    items_by_sku = {i.sku: i for i in Item.objects.filter(sku__in=list(item_skus), is_deleted=False) if i.sku}

    planned_invoices = []
    planned_items_total = 0

    for _, group_rows in groups.items():
        first = group_rows[0]
        row_num = int(first.get("_row") or 0)
        inv_no = _clean(first.get("invoice_number")) or None

        status_val = _clean(first.get("status") or "Draft") or "Draft"
        if status_val not in allowed_status:
            errors.append({"row": row_num, "field": "status", "message": "Invalid status"})
            continue
        for r in group_rows[1:]:
            other = _clean(r.get("status") or status_val) or status_val
            if other != status_val:
                errors.append({"row": int(r.get("_row") or 0), "field": "status", "message": "Mixed statuses within the same invoice group"})

        issue_date = _parse_date(first.get("issue_date"), "issue_date", row_num) or timezone.localdate()
        due_date = _parse_date(first.get("due_date"), "due_date", row_num)

        email = _clean(first.get("customer_email"))
        name = _clean(first.get("customer_name"))
        customer = None
        if email:
            customer = customers_by_email.get(email.lower())
        elif name:
            customer = customers_by_name.get(name.lower())
        if customer is None:
            errors.append({"row": row_num, "field": "customer", "message": "Customer not found"})
            continue

        normalized_lines = []
        for r in group_rows:
            rnum = int(r.get("_row") or 0)
            sku = _clean(r.get("item_sku"))
            item = items_by_sku.get(sku)
            if item is None:
                errors.append({"row": rnum, "field": "item_sku", "message": "Item not found"})
                continue
            qty = _parse_int(r.get("quantity"), "quantity", rnum)
            if qty is None:
                continue
            if qty < 1:
                errors.append({"row": rnum, "field": "quantity", "message": "quantity must be >= 1"})
                continue
            unit_price_override = _parse_decimal(r.get("unit_price"), "unit_price", rnum)
            if unit_price_override is not None and unit_price_override < 0:
                errors.append({"row": rnum, "field": "unit_price", "message": "unit_price must be >= 0"})
                continue
            tax_rate_override = _parse_decimal(r.get("tax_rate"), "tax_rate", rnum)
            if tax_rate_override is not None and (tax_rate_override < 0 or tax_rate_override > 100):
                errors.append({"row": rnum, "field": "tax_rate", "message": "tax_rate must be between 0 and 100"})
                continue
            desc = _clean(r.get("description")) or None
            uom = _clean(r.get("unit_of_measure")) or None
            normalized_lines.append(
                {
                    "row": rnum,
                    "item": item,
                    "quantity": qty,
                    "unit_price": unit_price_override,
                    "tax_rate": tax_rate_override,
                    "description": desc,
                    "unit_of_measure": uom,
                }
            )

        if normalized_lines:
            planned_invoices.append(
                {
                    "invoice_number": inv_no,
                    "customer": customer,
                    "status": status_val,
                    "issue_date": issue_date,
                    "due_date": due_date,
                    "lines": normalized_lines,
                }
            )
            planned_items_total += len(normalized_lines)

    if errors and rollback_on_error:
        token = _cache_error_log(errors)
        return 400, {
            "imported_invoices": 0,
            "imported_invoice_items": 0,
            "rows": len(rows),
            "errors": errors[:200],
            "error_log_token": token,
            "rolled_back": True,
        }

    if dry_run:
        return 200, {
            "dry_run": True,
            "rows": len(rows),
            "would_create_invoices": len(planned_invoices),
            "would_create_invoice_items": planned_items_total,
            "errors": errors,
        }

    created_invoices = 0
    created_items = 0
    with transaction.atomic():
        for plan in planned_invoices:
            inv = Invoice.objects.create(
                invoice_number=plan["invoice_number"] or "",
                customer=plan["customer"],
                status=plan["status"],
                issue_date=plan["issue_date"],
                due_date=plan["due_date"],
                subtotal=Decimal("0.00"),
                tax_rate=Decimal("0.00"),
                tax_total=Decimal("0.00"),
                total_amount=Decimal("0.00"),
            )

            subtotal = Decimal("0.00")
            tax_total = Decimal("0.00")
            created_lines: list[InvoiceItem] = []
            for li in plan["lines"]:
                item: Item = li["item"]
                qty: int = li["quantity"]
                unit_price = li["unit_price"] if li["unit_price"] is not None else item.unit_price
                line_subtotal = _q2(Decimal(unit_price) * qty)
                line_tax_rate = li["tax_rate"] if li["tax_rate"] is not None else item.tax_rate
                line_tax = _q2((line_subtotal * Decimal(line_tax_rate)) / Decimal("100"))
                line_total = _q2(line_subtotal + line_tax)
                subtotal += line_subtotal
                tax_total += line_tax
                created_lines.append(
                    InvoiceItem(
                        invoice=inv,
                        item=item,
                        description=li["description"] if li["description"] is not None else item.description,
                        unit_of_measure=li["unit_of_measure"] if li["unit_of_measure"] is not None else item.unit_of_measure,
                        quantity=qty,
                        unit_price=unit_price,
                        tax_rate=line_tax_rate,
                        line_subtotal=line_subtotal,
                        line_tax=line_tax,
                        line_total=line_total,
                    )
                )
            InvoiceItem.objects.bulk_create(created_lines, batch_size=1000)
            created_items += len(created_lines)

            subtotal = _q2(subtotal)
            tax_total = _q2(tax_total)
            total_amount = _q2(subtotal + tax_total)
            computed_tax_rate = _q2((tax_total / subtotal) * Decimal("100")) if subtotal > 0 else Decimal("0.00")
            Invoice.objects.filter(pk=inv.pk).update(subtotal=subtotal, tax_rate=computed_tax_rate, tax_total=tax_total, total_amount=total_amount)
            inv.refresh_from_db()
            if inv.status in ["Sent", "Paid"]:
                deduct_inventory_for_invoice(inv)
            created_invoices += 1

    return 200, {"imported_invoices": created_invoices, "imported_invoice_items": created_items, "rows": len(rows), "errors": errors}


def import_customers_from_upload(upload, dry_run: bool, rollback_on_error: bool) -> tuple[int, dict]:
    rows = _parse_upload_rows(upload)
    seen_emails: set[str] = set()
    seen_names: set[str] = set()
    name_max_length = int(Customer._meta.get_field("name").max_length or 0)
    email_max_length = int(Customer._meta.get_field("email").max_length or 0)
    phone_max_length = int(Customer._meta.get_field("phone").max_length or 0)

    incoming_emails = []
    for r in rows:
        email = str(r.get("email") or "").strip().lower()
        if email:
            incoming_emails.append(email)
    existing_emails = set(
        Customer.objects.filter(email__in=incoming_emails, is_deleted=False).values_list("email", flat=True)
    )
    existing_emails = {str(v or "").lower() for v in existing_emails if v}

    def _validate_row(r: dict) -> tuple[Customer | None, list[dict]]:
        row_num = int(r.get("_row") or 0)
        name = str(r.get("name") or "").strip()
        email = str(r.get("email") or "").strip() or None
        phone = str(r.get("phone") or "").strip() or None
        billing_address = str(r.get("billing_address") or "").strip() or None
        row_errors: list[dict] = []

        if not name:
            row_errors.append({"row": row_num, "field": "name", "message": "name is required"})
        if name_max_length and len(name) > name_max_length:
            row_errors.append({"row": row_num, "field": "name", "message": f"name must be at most {name_max_length} characters"})
        name_key = name.lower()
        if name_key in seen_names:
            row_errors.append({"row": row_num, "field": "name", "message": "Duplicate name in file"})
        else:
            seen_names.add(name_key)

        if email:
            if email_max_length and len(email) > email_max_length:
                row_errors.append({"row": row_num, "field": "email", "message": f"email must be at most {email_max_length} characters"})
            email_key = email.lower()
            if email_key in seen_emails:
                row_errors.append({"row": row_num, "field": "email", "message": "Duplicate email in file"})
            else:
                seen_emails.add(email_key)
            if email_key in existing_emails:
                row_errors.append({"row": row_num, "field": "email", "message": "email already exists"})
            try:
                validate_email(email)
            except DjangoValidationError:
                row_errors.append({"row": row_num, "field": "email", "message": "Invalid email"})
        if phone and phone_max_length and len(phone) > phone_max_length:
            row_errors.append({"row": row_num, "field": "phone", "message": f"phone must be at most {phone_max_length} characters"})

        if row_errors:
            return None, row_errors

        return (
            Customer(
                name=name,
                email=email,
                phone=phone,
                billing_address=billing_address,
            ),
            [],
        )

    def _persist(valid_items: list[Customer]) -> tuple[int, list[dict] | None]:
        created = 0
        with transaction.atomic():
            for i in range(0, len(valid_items), 1000):
                batch = valid_items[i : i + 1000]
                Customer.objects.bulk_create(batch, batch_size=1000)
                created += len(batch)
        return created, None

    return process_batch_import(
        rows,
        validate_item=_validate_row,
        persist_valid_items=_persist,
        dry_run=dry_run,
        rollback_on_error=rollback_on_error,
    )


def import_expenses_from_upload(upload, *, dry_run: bool, rollback_on_error: bool, actor=None) -> tuple[int, dict]:
    rows = _parse_upload_rows(upload)
    User = get_user_model()

    usernames = {str(r.get("assigned_to") or "").strip() for r in rows if str(r.get("assigned_to") or "").strip()}
    users_by_name = {u.username: u for u in User.objects.filter(username__in=list(usernames))}
    source_account_names = {str(r.get("source_account") or "").strip().lower() for r in rows if str(r.get("source_account") or "").strip()}
    source_accounts_by_name = {}
    if source_account_names:
        for account in SourceAccount.objects.filter(is_deleted=False):
            lowered = str(account.name or "").strip().lower()
            if lowered in source_account_names:
                source_accounts_by_name[lowered] = account
    reserved_project_codes: set[str] = set()
    existing_project_codes = {
        normalize_project_code(code)
        for code in Expense.objects.filter(is_deleted=False, project_code__isnull=False)
        .exclude(project_code="")
        .values_list("project_code", flat=True)
    }
    generated_project_codes: list[str] = []

    def _clean(value) -> str:
        return str(value or "").strip()

    def _parse_date(value, field: str, row_num: int, row_errors: list[dict]):
        raw = _clean(value)
        if not raw:
            return None
        try:
            return date.fromisoformat(raw)
        except ValueError:
            row_errors.append({"row": row_num, "field": field, "message": "Invalid date. Use YYYY-MM-DD"})
            return None

    def _parse_decimal(value, field: str, row_num: int, row_errors: list[dict]):
        if value in (None, ""):
            row_errors.append({"row": row_num, "field": field, "message": f"{field} is required"})
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, TypeError):
            row_errors.append({"row": row_num, "field": field, "message": f"Invalid {field}"})
            return None

    def _validate_row(r: dict) -> tuple[Expense | None, list[dict]]:
        row_num = int(r.get("_row") or 0)
        row_errors: list[dict] = []
        amount = _parse_decimal(r.get("amount"), "amount", row_num, row_errors)
        if amount is not None and amount <= 0:
            row_errors.append({"row": row_num, "field": "amount", "message": "amount must be > 0"})

        expense_date = _parse_date(r.get("expense_date"), "expense_date", row_num, row_errors) or timezone.localdate()
        if expense_date and expense_date > timezone.localdate() + timedelta(days=1):
            row_errors.append({"row": row_num, "field": "expense_date", "message": "expense_date cannot be more than one day in the future"})
        category = _clean(r.get("category"))
        if not category:
            row_errors.append({"row": row_num, "field": "category", "message": "category is required"})
        project_code = normalize_project_code(r.get("project_code")) or None
        cost_center = _clean(r.get("cost_center")) or None
        if project_code:
            if project_code in existing_project_codes:
                row_errors.append({"row": row_num, "field": "project_code", "message": "project_code already exists"})

        assigned_to_name = _clean(r.get("assigned_to")) or None
        assigned_to = users_by_name.get(assigned_to_name) if assigned_to_name else actor
        if assigned_to_name and assigned_to is None:
            row_errors.append({"row": row_num, "field": "assigned_to", "message": "Assigned user not found"})

        source_account_name = _clean(r.get("source_account")) or None
        source_account = None
        if source_account_name:
            source_account = source_accounts_by_name.get(source_account_name.lower())
            if source_account is None:
                row_errors.append({"row": row_num, "field": "source_account", "message": "Source account not found"})
            elif source_account.status != SourceAccount.STATUS_ACTIVE:
                row_errors.append({"row": row_num, "field": "source_account", "message": "Source account must be active"})

        if project_code and project_code in reserved_project_codes:
            row_errors.append({"row": row_num, "field": "project_code", "message": "project_code is duplicated in the import file"})

        if row_errors:
            return None, row_errors

        if project_code:
            reserved_project_codes.add(project_code)

        return (
            Expense(
                amount=amount,
                expense_date=expense_date,
                category=category,
                description=_clean(r.get("description")) or None,
                vendor=_clean(r.get("vendor")) or None,
                merchant_reference=_clean(r.get("merchant_reference")) or None,
                project_code=project_code,
                cost_center=cost_center,
                source_account=source_account,
                assigned_to=assigned_to,
                created_by=actor if getattr(actor, "is_authenticated", False) else None,
            ),
            [],
        )

    def _persist(valid_items: list[Expense]) -> tuple[int, list[dict] | None]:
        created = 0
        with transaction.atomic():
            for expense in valid_items:
                if not expense.project_code:
                    expense.project_code = generate_next_project_code(year=(expense.expense_date or timezone.localdate()).year)
                    generated_project_codes.append(expense.project_code)
                expense.save()
                created += 1
        return created, None

    status_code, payload = process_batch_import(
        rows,
        validate_item=_validate_row,
        persist_valid_items=_persist,
        dry_run=dry_run,
        rollback_on_error=rollback_on_error,
    )
    if generated_project_codes:
        payload["generated_project_codes"] = generated_project_codes
    return status_code, payload
