from __future__ import annotations

import logging
import io
import os
import uuid
import csv
from datetime import date, timedelta
import calendar
import re
import secrets
import hashlib
import hmac
import time
import json
import urllib.parse
import urllib.error
import urllib.request
from typing import Any, Optional, Tuple
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from django.core import signing

from PIL import Image, ImageDraw, UnidentifiedImageError
from rest_framework import viewsets, status, permissions
from rest_framework import parsers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView
from django.core.cache import cache
from django.core.mail import EmailMultiAlternatives, send_mail
from django.conf import settings
from django.db import transaction
from django.db.utils import DatabaseError
from django.db import models
from django.db.models import F, Sum, Q, Max
from django.db.models.functions import Coalesce, TruncDay, TruncMonth
from django.db.models import Count
from django.contrib.contenttypes.models import ContentType
from django.utils import timezone
from django.template.loader import get_template
from django.http import HttpResponse, HttpResponseRedirect, StreamingHttpResponse
from rest_framework.exceptions import ValidationError, NotFound, APIException, PermissionDenied
from django.contrib.auth import get_user_model, authenticate
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from rest_framework import authentication
from openpyxl import Workbook, load_workbook
from .models import (
    Customer,
    Item,
    Invoice,
    InvoiceItem,
    Receipt,
    Expense,
    SourceAccount,
    AuditLog,
    Currency,
    ExchangeRate,
    GlobalSettings,
    UserSettings,
    UserProfile,
    SocialAuthConnection,
    SavedInvoiceView,
    EmailVerificationToken,
    AdminUserInvitation,
    AccessToken,
    AdminMfaDevice,
    Role,
    Permission,
    RolePermission,
    UserRole,
    DocumentDelivery,
    SavedDocument,
    PaymentTransaction,
    PaymentWebhookEvent,
    BusinessAccount,
    BusinessMembership,
    LogoAsset,
)
from .rbac import user_role_names, user_has_permission
from .auth_service import ensure_user_role, issue_access_token, revoke_token, role_for_name
from .logo_uploads import (
    LOGO_SCOPE_GLOBAL,
    LOGO_SCOPE_INVOICE,
    LOGO_SCOPE_RECEIPT,
    cleanup_logo_asset_if_unreferenced,
    create_logo_asset,
    normalize_logo_scope,
)
from .expense_security import decrypt_expense_text
from .finance_services import (
    CENTS,
    outstanding_invoice_amount,
    q2 as _q2,
    resolve_invoice_discount as _resolve_invoice_discount,
    sync_invoice_status_with_payments,
)
from .rendering_service import (
    country_config as _country_config,
    currency_for_code as _currency_for_code,
    effective_company_identity_for_user as _effective_company_identity_for_user,
    effective_region_settings_for_user,
    effective_templates_for_user as _effective_templates_for_user,
    get_global_settings as _get_global_settings,
    get_user_settings as _get_user_settings,
)
from .serializers import (
    CustomerListSerializer,
    CustomerSerializer,
    CustomerDetailSerializer,
    ItemSerializer,
    ItemDetailSerializer,
    InvoiceSerializer,
    InvoiceItemSerializer,
    ReceiptSerializer,
    ReceiptDetailSerializer,
    ExpenseSerializer,
    SourceAccountSerializer,
    CurrencySerializer,
    ExchangeRateSerializer,
    GlobalSettingsSerializer,
    UserSettingsSerializer,
    BusinessAccountSerializer,
    BusinessMembershipSerializer,
    SettingsRollbackSerializer,
    RegisterSerializer,
    VerifyEmailSerializer,
    PasswordResetRequestSerializer,
    PasswordResetConfirmSerializer,
    DocumentDeliverySerializer,
    SavedDocumentSerializer,
    PaymentTransactionSerializer,
    normalize_signup_phone,
    validate_application_password,
    NIGERIA_COUNTRY_NAME,
    DEFAULT_COUNTRY_CODE,
)

logger = logging.getLogger(__name__)

class PaymentDeclined(APIException):
    status_code = 402
    default_detail = "Payment was declined"
    default_code = "payment_declined"


class PaymentGatewayUnavailable(APIException):
    status_code = 503
    default_detail = "Payment gateway unavailable"
    default_code = "payment_gateway_unavailable"


class PaymentGatewayTimeout(APIException):
    status_code = 504
    default_detail = "Payment gateway timeout"
    default_code = "payment_gateway_timeout"


def _parse_iso_datetime(value: str):
    try:
        dt = timezone.datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone=timezone.utc)
    return dt


def _check_concurrency(request, instance):
    supplied = request.data.get("updated_at") if hasattr(request, "data") else None
    if supplied in (None, ""):
        supplied = request.query_params.get("updated_at")
    if supplied in (None, ""):
        return
    supplied_dt = _parse_iso_datetime(str(supplied))
    if supplied_dt is None:
        raise ValidationError({"updated_at": "Invalid updated_at; expected ISO datetime"})
    current = instance.updated_at
    if timezone.is_naive(current):
        current = timezone.make_aware(current, timezone=timezone.utc)
    if supplied_dt != current:
        raise ConflictError("Conflict: record was modified by another user")


class ConflictError(APIException):
    status_code = status.HTTP_409_CONFLICT
    default_detail = "Conflict"
    default_code = "conflict"


def _log_audit(user, action: str, instance, changes: dict):
    AuditLog.objects.create(
        user=user if getattr(user, "is_authenticated", False) else None,
        action=action,
        content_type=ContentType.objects.get_for_model(instance.__class__),
        object_id=str(instance.pk),
        changes=changes or {},
    )


def _log_security_event(user, content_type_model, object_id: str, changes: dict):
    AuditLog.objects.create(
        user=user if getattr(user, "is_authenticated", False) else None,
        action="security",
        content_type=ContentType.objects.get_for_model(content_type_model),
        object_id=str(object_id),
        changes=changes or {},
    )


def _log_operation(user, action: str, content_type_model, object_id: str, changes: dict):
    AuditLog.objects.create(
        user=user if getattr(user, "is_authenticated", False) else None,
        action=action,
        content_type=ContentType.objects.get_for_model(content_type_model),
        object_id=str(object_id),
        changes=changes or {},
    )


SYSTEM_ROLE_NAMES = {"user", "staff", "admin"}


def _anonymize_user_artifacts_for_deletion(user) -> None:
    user_id = getattr(user, "id", None)
    if not user_id:
        return

    settings_row = UserSettings.objects.filter(user=user).first()
    if settings_row is not None:
        invoice_logo = settings_row.invoice_logo
        receipt_logo = settings_row.receipt_logo
        settings_row.invoice_logo = None
        settings_row.receipt_logo = None
        settings_row.save(update_fields=["invoice_logo", "receipt_logo", "updated_at"])
        cleanup_logo_asset_if_unreferenced(invoice_logo)
        cleanup_logo_asset_if_unreferenced(receipt_logo)

    owned_logos = list(LogoAsset.objects.filter(owner=user))
    for asset in owned_logos:
        if asset.file:
            asset.file.delete(save=False)
        if asset.thumbnail:
            asset.thumbnail.delete(save=False)
        asset.delete()

    for saved_doc in SavedDocument.objects.filter(user=user):
        if saved_doc.file:
            saved_doc.file.delete(save=False)
        saved_doc.delete()

    DocumentDelivery.objects.filter(user=user).delete()
    SavedInvoiceView.objects.filter(user=user).delete()
    SocialAuthConnection.objects.filter(user=user).delete()
    EmailVerificationToken.objects.filter(user=user).delete()
    AccessToken.objects.filter(user=user).delete()
    AdminMfaDevice.objects.filter(user=user).delete()
    UserRole.objects.filter(user=user).delete()
    AdminUserInvitation.objects.filter(user=user).delete()
    BusinessMembership.objects.filter(user=user).delete()

    Expense.objects.filter(created_by=user).update(created_by=None)
    Expense.objects.filter(assigned_to=user).update(assigned_to=None)
    PaymentTransaction.objects.filter(created_by=user).update(created_by=None)
    BusinessAccount.objects.filter(owner=user).delete()

    login_content_type = ContentType.objects.get_for_model(user.__class__)
    AuditLog.objects.filter(content_type=login_content_type, object_id=str(user_id), action="security").update(
        object_id=f"deleted-user:{user_id}",
        changes={
            "event": "deleted_user_activity_redacted",
            "user_id": user_id,
        },
    )


def _permission_codes_for_user(user) -> list[str]:
    if not getattr(user, "is_authenticated", False):
        return []
    if bool(getattr(user, "is_superuser", False)):
        return list(Permission.objects.values_list("code", flat=True).distinct().order_by("code"))
    return list(
        Permission.objects.filter(permission_roles__role__role_users__user=user)
        .values_list("code", flat=True)
        .distinct()
        .order_by("code")
    )


def _role_permission_codes(role: Role) -> list[str]:
    return list(
        Permission.objects.filter(permission_roles__role=role)
        .values_list("code", flat=True)
        .distinct()
        .order_by("code")
    )


def _latest_login_timestamps_by_user(user_ids: list[int]) -> dict[int, Optional[str]]:
    if not user_ids:
        return {}
    ct = ContentType.objects.get_for_model(get_user_model())
    rows = (
        AuditLog.objects.filter(
            action="security",
            content_type=ct,
            object_id__in=[str(user_id) for user_id in user_ids],
            changes__event__in=["login_success", "staff_login_success", "admin_login_success"],
        )
        .order_by("object_id", "-created_at")
        .values("object_id", "created_at")
    )
    latest: dict[int, Optional[str]] = {}
    for row in rows:
        try:
            key = int(str(row["object_id"]))
        except (TypeError, ValueError):
            continue
        if key not in latest:
            created_at = row.get("created_at")
            latest[key] = created_at.isoformat() if created_at else None
    return latest


def _serialize_admin_user(user, *, latest_login_at: Optional[str] = None) -> dict[str, Any]:
    profile = UserProfile.objects.filter(user=user).only("full_name", "phone", "company_legal_name").first()
    invitation = _admin_invitation_for_user(user)
    roles = user_role_names(user)
    system_role = "user"
    if "admin" in roles or bool(getattr(user, "is_superuser", False)):
        system_role = "admin"
    elif "staff" in roles or bool(getattr(user, "is_staff", False)):
        system_role = "staff"
    custom_roles = [role for role in roles if role not in SYSTEM_ROLE_NAMES]
    return {
        "id": user.id,
        "username": getattr(user, "username", ""),
        "email": getattr(user, "email", ""),
        "is_active": bool(getattr(user, "is_active", True)),
        "is_staff": bool(getattr(user, "is_staff", False)),
        "is_superuser": bool(getattr(user, "is_superuser", False)),
        "primary_role": system_role,
        "roles": roles,
        "custom_roles": custom_roles,
        "full_name": getattr(profile, "full_name", None) if profile else None,
        "phone": getattr(profile, "phone", None) if profile else None,
        "company_name": getattr(profile, "company_legal_name", None) if profile else None,
        "invitation_status": _admin_invitation_status_for_user(user, invitation),
        "invitation_expires_at": invitation.expires_at.isoformat() if invitation and invitation.expires_at else None,
        "invitation_accepted_at": invitation.accepted_at.isoformat() if invitation and invitation.accepted_at else None,
        "last_login_at": latest_login_at,
    }


def _login_role_config_for_user(user) -> tuple[Any, str, int]:
    if user_has_permission(user, "settings.global.write"):
        return role_for_name("admin"), "admin_login_success", 12 * 3600
    if user_has_permission(user, "settings.global.read") or user_has_permission(user, "admin.users.read"):
        return role_for_name("staff"), "staff_login_success", 12 * 3600
    return role_for_name("user"), "login_success", 7 * 24 * 3600


class EditDeletePermission(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated)


class CurrencyPermission(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return bool(request.user and request.user.is_authenticated)
        return bool(request.user and request.user.is_authenticated and user_has_permission(request.user, "currency.write"))


class ExchangeRatePermission(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return bool(request.user and request.user.is_authenticated and user_has_permission(request.user, "fx.read"))
        return bool(request.user and request.user.is_authenticated and user_has_permission(request.user, "fx.write"))


def _detect_country(request) -> Optional[str]:
    explicit = request.query_params.get("country")
    if explicit:
        cc = str(explicit).upper()
        if re.fullmatch(r"[A-Z]{2}", cc):
            return cc

    for header in ("CF-IPCountry", "X-Country-Code", "X-Country"):
        value = request.headers.get(header)
        if value:
            cc = str(value).upper()
            if re.fullmatch(r"[A-Z]{2}", cc):
                return cc

    accept = (request.headers.get("accept-language") or "").lower()
    if accept.startswith("en-gb"):
        return "GB"
    if accept.startswith("de"):
        return "DE"
    if accept.startswith("fr"):
        return "FR"
    if accept.startswith("ja"):
        return "JP"
    if accept.startswith("en-au"):
        return "AU"
    if accept.startswith("en-ca"):
        return "CA"
    if accept.startswith("en-us"):
        return "US"
    return None


def _client_ip(request) -> str:
    xff = request.headers.get("x-forwarded-for") or request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR") or ""


def _rate_limit(key: str, limit: int, window_seconds: int) -> bool:
    now_bucket = int(timezone.now().timestamp() // window_seconds)
    cache_key = f"rl:{key}:{now_bucket}"
    try:
        count = cache.incr(cache_key)
    except ValueError:
        cache.set(cache_key, 1, timeout=window_seconds + 5)
        count = 1
    return count > limit


def _username_for_email(email: str) -> str:
    return str(email).strip().lower()


def _hash_email(email: str) -> str:
    return hashlib.sha256(str(email).strip().lower().encode("utf-8")).hexdigest()


ADMIN_INVITATION_TOKEN_TYPE = "admin_invitation"
ADMIN_INVITATION_SIGNING_SALT = "core.admin_user_invitation"
ADMIN_INVITATION_MAX_AGE_SECONDS = 72 * 3600


def _frontend_base_url() -> str:
    return str(getattr(settings, "FRONTEND_BASE_URL", "http://127.0.0.1:3000")).rstrip("/")


def _password_reset_link_for_user(user) -> tuple[str, str, str]:
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = PasswordResetTokenGenerator().make_token(user)
    link = f"{_frontend_base_url()}/reset-password?uid={urllib.parse.quote(uid)}&token={urllib.parse.quote(token)}"
    return uid, token, link


def _sign_admin_invitation(invitation: AdminUserInvitation) -> str:
    payload = {
        "invitation_id": int(invitation.pk),
        "user_id": int(invitation.user_id),
        "token_key": str(invitation.token_key),
        "token_type": ADMIN_INVITATION_TOKEN_TYPE,
    }
    return signing.dumps(payload, salt=ADMIN_INVITATION_SIGNING_SALT)


def _get_admin_invitation_by_signed_token(token: str) -> AdminUserInvitation:
    try:
        payload = signing.loads(token, salt=ADMIN_INVITATION_SIGNING_SALT, max_age=ADMIN_INVITATION_MAX_AGE_SECONDS)
    except signing.SignatureExpired as exc:
        raise ValidationError({"token": "Token expired"}) from exc
    except signing.BadSignature as exc:
        raise ValidationError({"token": "Invalid token"}) from exc

    try:
        invitation_id = int(payload.get("invitation_id"))
        user_id = int(payload.get("user_id"))
    except (TypeError, ValueError):
        raise ValidationError({"token": "Invalid token"})

    token_key = str(payload.get("token_key") or "").strip()
    if not token_key:
        raise ValidationError({"token": "Invalid token"})

    try:
        invitation = AdminUserInvitation.objects.select_related("user", "invited_by").get(pk=invitation_id, user_id=user_id)
    except AdminUserInvitation.DoesNotExist as exc:
        raise ValidationError({"token": "Invalid token"}) from exc

    if invitation.token_key != token_key:
        raise ValidationError({"token": "Invalid token"})
    if invitation.expires_at <= timezone.now():
        raise ValidationError({"token": "Token expired"})
    return invitation


def _admin_invitation_for_user(user) -> Optional[AdminUserInvitation]:
    if getattr(user, "pk", None) is None:
        return None
    return AdminUserInvitation.objects.filter(user=user).first()


def _admin_invitation_status_for_user(user, invitation: Optional[AdminUserInvitation] = None) -> Optional[str]:
    invitation = invitation or _admin_invitation_for_user(user)
    if invitation is None:
        return None
    if invitation.activated_at is not None or bool(getattr(user, "is_active", False)):
        return "active"
    if invitation.password_reset_completed_at is not None:
        return "password_set"
    if invitation.accepted_at is not None:
        return "accepted_pending_password"
    if invitation.expires_at <= timezone.now():
        return "expired"
    return "pending_acceptance"


def _send_html_email_message(subject: str, recipient: str, text_body: str, html_body: str) -> bool:
    message = EmailMultiAlternatives(
        subject,
        text_body,
        getattr(settings, "DEFAULT_FROM_EMAIL", "no-reply@example.com"),
        [recipient],
    )
    message.attach_alternative(html_body, "text/html")
    sent = message.send(fail_silently=False)
    return int(sent) > 0


def _send_verification_email(email: str, token: str, *, user=None, ip: Optional[str] = None, source: str = "register") -> bool:
    link = f"{_frontend_base_url()}/verify-email?token={token}"
    subject = "Verify your email"
    message = f"Please verify your email by opening this link:\n\n{link}\n"
    email_backend = getattr(settings, "EMAIL_BACKEND", "")
    User = get_user_model()
    object_id = str(getattr(user, "pk", "")) if getattr(user, "pk", None) else _hash_email(email)

    try:
        sent = send_mail(
            subject,
            message,
            getattr(settings, "DEFAULT_FROM_EMAIL", "no-reply@example.com"),
            [email],
            fail_silently=False,
        )
        ok = int(sent) > 0
        logger.info("verification_email_sent ok=%s backend=%s source=%s object_id=%s ip=%s", ok, email_backend, source, object_id, ip or "unknown")
        _log_security_event(
            user if getattr(user, "is_authenticated", False) else None,
            User,
            object_id=object_id,
            changes={"event": "verification_email_sent", "ok": ok, "backend": email_backend, "source": source, "ip": ip},
        )
        return ok
    except Exception as e:
        logger.exception("verification_email_failed backend=%s source=%s object_id=%s ip=%s", email_backend, source, object_id, ip or "unknown")
        _log_security_event(
            user if getattr(user, "is_authenticated", False) else None,
            User,
            object_id=object_id,
            changes={"event": "verification_email_failed", "ok": False, "backend": email_backend, "source": source, "ip": ip, "error": e.__class__.__name__},
        )
        return False


def _send_admin_invitation_email(invitation: AdminUserInvitation, *, ip: Optional[str] = None) -> bool:
    email = str(getattr(invitation.user, "email", "") or "").strip().lower()
    if not email:
        return False
    link = f"{_frontend_base_url()}/verify-email?token={urllib.parse.quote(_sign_admin_invitation(invitation))}&token_type={ADMIN_INVITATION_TOKEN_TYPE}"
    subject = "Accept your account invitation"
    text_body = (
        "An administrator created an account for you on PXL INVOICE.\n\n"
        f"Accept your invitation: {link}\n\n"
        "This invitation expires in 72 hours."
    )
    html_body = (
        "<p>An administrator created an account for you on <strong>PXL INVOICE</strong>.</p>"
        "<p>Accept the invitation to confirm your account and continue to your initial password setup.</p>"
        f'<p><a href="{link}" style="display:inline-block;padding:12px 18px;background:#2563eb;color:#ffffff;text-decoration:none;border-radius:6px;font-weight:600;">Accept account invitation</a></p>'
        "<p>This invitation expires in 72 hours.</p>"
    )
    email_backend = getattr(settings, "EMAIL_BACKEND", "")
    object_id = str(invitation.user_id)
    try:
        ok = _send_html_email_message(subject, email, text_body, html_body)
        logger.info("admin_invitation_email_sent ok=%s backend=%s user_id=%s ip=%s", ok, email_backend, object_id, ip or "unknown")
        if ok:
            invitation.confirmation_email_sent_at = timezone.now()
            invitation.save(update_fields=["confirmation_email_sent_at", "updated_at"])
        _log_security_event(
            invitation.invited_by,
            get_user_model(),
            object_id=object_id,
            changes={"event": "admin_invitation_email_sent", "ok": ok, "backend": email_backend, "ip": ip},
        )
        return ok
    except Exception as e:
        logger.exception("admin_invitation_email_failed backend=%s user_id=%s ip=%s", email_backend, object_id, ip or "unknown")
        _log_security_event(
            invitation.invited_by,
            get_user_model(),
            object_id=object_id,
            changes={"event": "admin_invitation_email_failed", "ok": False, "backend": email_backend, "ip": ip, "error": e.__class__.__name__},
        )
        return False


def _send_account_activation_email(user, *, ip: Optional[str] = None, invitation: Optional[AdminUserInvitation] = None) -> bool:
    email = str(getattr(user, "email", "") or "").strip().lower()
    if not email:
        return False
    subject = "Your account is now active"
    text_body = (
        "Welcome to PXL INVOICE.\n\n"
        "Your account has been activated successfully and your assigned access is now available."
    )
    html_body = (
        "<p>Welcome to <strong>PXL INVOICE</strong>.</p>"
        "<p>Your account has been activated successfully and your assigned access is now available.</p>"
    )
    email_backend = getattr(settings, "EMAIL_BACKEND", "")
    object_id = str(getattr(user, "pk", ""))
    try:
        ok = _send_html_email_message(subject, email, text_body, html_body)
        logger.info("account_activation_email_sent ok=%s backend=%s user_id=%s ip=%s", ok, email_backend, object_id, ip or "unknown")
        if ok and invitation is not None:
            invitation.welcome_email_sent_at = timezone.now()
            invitation.save(update_fields=["welcome_email_sent_at", "updated_at"])
        _log_security_event(
            user,
            get_user_model(),
            object_id=object_id,
            changes={"event": "account_activation_email_sent", "ok": ok, "backend": email_backend, "ip": ip},
        )
        return ok
    except Exception as e:
        logger.exception("account_activation_email_failed backend=%s user_id=%s ip=%s", email_backend, object_id, ip or "unknown")
        _log_security_event(
            user,
            get_user_model(),
            object_id=object_id,
            changes={"event": "account_activation_email_failed", "ok": False, "backend": email_backend, "ip": ip, "error": e.__class__.__name__},
        )
        return False


def _oauth_frontend_redirect(*, provider: str, token: Optional[str] = None, error: Optional[str] = None, extras: Optional[dict[str, Any]] = None):
    frontend_base = getattr(settings, "FRONTEND_BASE_URL", "http://127.0.0.1:3000")
    base = frontend_base.rstrip("/") + "/auth/callback"
    fragment = {"provider": str(provider)}
    if token:
        fragment["token"] = str(token)
    if error:
        fragment["error"] = str(error)
    if extras:
        for key, value in extras.items():
            if value is None:
                continue
            fragment[str(key)] = str(value)
    return HttpResponseRedirect(base + "#" + urllib.parse.urlencode(fragment))


def _oauth_state_cache_key(provider: str, state: str) -> str:
    return f"{provider}_oauth_state:{state}"


def _bool_query(value: Any, *, default: bool = False) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return default


def _oauth_state_payload(request, *, intent_default: str = "login") -> dict[str, Any]:
    intent = str(request.query_params.get("intent") or intent_default).strip().lower() or intent_default
    if intent not in {"login", "link"}:
        intent = intent_default
    payload: dict[str, Any] = {
        "ip": _client_ip(request),
        "intent": intent,
        "remember": _bool_query(request.query_params.get("remember"), default=True),
    }
    if intent == "link":
        if not getattr(request.user, "is_authenticated", False):
            raise PermissionDenied("Please sign in before linking a social account.")
        payload["user_id"] = int(request.user.pk)
    return payload


def _pop_oauth_state(provider: str, state: str) -> dict[str, Any] | None:
    key = _oauth_state_cache_key(provider, state)
    row = cache.get(key)
    if row is not None:
        cache.delete(key)
    return row


def _social_provider_label(provider: str) -> str:
    return "Google" if provider == "google" else "Facebook" if provider == "facebook" else str(provider).title()


def _social_identity_error(provider: str, code: str):
    return _oauth_frontend_redirect(provider=provider, error=code)


def _social_connections_for_user(user) -> list[dict[str, Any]]:
    rows = []
    for conn in SocialAuthConnection.objects.filter(user=user).order_by("provider"):
        rows.append(
            {
                "provider": conn.provider,
                "label": _social_provider_label(conn.provider),
                "email": conn.email,
                "display_name": conn.display_name,
                "avatar_url": conn.avatar_url,
                "created_at": conn.created_at.isoformat(),
                "last_login_at": conn.last_login_at.isoformat(),
            }
        )
    return rows


def _is_privileged_account(user) -> bool:
    return bool(
        user_has_permission(user, "settings.global.read")
        or user_has_permission(user, "settings.global.write")
        or user_has_permission(user, "admin.users.read")
        or user_has_permission(user, "admin.users.write")
    )


DEFAULT_BUSINESS_ROLE_NAME = "editor"


def _ensure_default_business_role(user) -> None:
    if user is None or not getattr(user, "pk", None):
        return
    if bool(getattr(user, "is_superuser", False)) or bool(getattr(user, "is_staff", False)):
        return
    existing_roles = set(user_role_names(user))
    if "admin" in existing_roles or "staff" in existing_roles:
        return
    if any(role_name not in SYSTEM_ROLE_NAMES for role_name in existing_roles):
        return
    role = Role.objects.filter(name=DEFAULT_BUSINESS_ROLE_NAME).first()
    if role is None:
        return
    UserRole.objects.get_or_create(user=user, role=role)


def _log_record_save_failure(*, request, entity: str, operation: str, exc: Exception) -> None:
    user_id = getattr(getattr(request, "user", None), "id", None)
    if isinstance(exc, (ValidationError, NotFound, PermissionDenied, ConflictError, APIException)):
        logger.warning(
            "save.%s_failed entity=%s user_id=%s detail=%s",
            operation,
            entity,
            user_id,
            getattr(exc, "detail", str(exc)),
        )
        return
    logger.exception("save.%s_failed entity=%s user_id=%s", operation, entity, user_id)


def _raise_record_save_failure(*, request, entity: str, operation: str, exc: Exception) -> None:
    _log_record_save_failure(request=request, entity=entity, operation=operation, exc=exc)
    if isinstance(exc, (ValidationError, NotFound, PermissionDenied, ConflictError, APIException)):
        raise exc
    raise APIException(f"Unable to {operation} {entity}. Please try again.")


def _touch_social_connection(
    *,
    user,
    provider: str,
    provider_user_id: str,
    email: str,
    display_name: str,
    avatar_url: str,
) -> SocialAuthConnection:
    connection, _ = SocialAuthConnection.objects.update_or_create(
        provider=provider,
        provider_user_id=str(provider_user_id),
        defaults={
            "user": user,
            "email": email or None,
            "display_name": display_name or None,
            "avatar_url": avatar_url or None,
            "last_login_at": timezone.now(),
        },
    )
    ensure_user_role(user, "user")
    _ensure_default_business_role(user)
    return connection


def _complete_social_login(*, user, provider: str, ip: Optional[str], remember: bool, created: bool) -> HttpResponseRedirect:
    if _is_privileged_account(user):
        _log_security_event(
            user,
            get_user_model(),
            object_id=str(user.pk),
            changes={"event": f"{provider}_social_login_blocked_for_privileged_account", "ip": ip},
        )
        return _social_identity_error(provider, "privileged_account")

    role = role_for_name("user")
    token_row = issue_access_token(user=user, role=role, expires_seconds=7 * 24 * 3600)
    _log_security_event(user, get_user_model(), object_id=str(user.pk), changes={"event": f"{provider}_login", "created": created, "ip": ip})
    resp = _oauth_frontend_redirect(provider=provider)
    _set_auth_cookie(resp, token_row, remember=remember)
    return resp


def _complete_social_link(
    *,
    user_id: int,
    provider: str,
    provider_user_id: str,
    email: str,
    display_name: str,
    avatar_url: str,
    ip: Optional[str],
) -> HttpResponseRedirect:
    User = get_user_model()
    target = User.objects.filter(pk=int(user_id)).first()
    if target is None:
        return _social_identity_error(provider, "link_target_missing")

    existing = SocialAuthConnection.objects.filter(provider=provider, provider_user_id=str(provider_user_id)).first()
    if existing is not None and existing.user_id != target.id:
        return _social_identity_error(provider, "already_linked")

    if email and getattr(target, "email", "") and str(target.email).strip().lower() != email:
        return _social_identity_error(provider, "email_mismatch")

    if email and not getattr(target, "email", ""):
        target.email = email
        target.save(update_fields=["email"])

    profile, _ = UserProfile.objects.get_or_create(user=target)
    fields_to_update = []
    if profile.email_verified_at is None and email:
        profile.email_verified_at = timezone.now()
        fields_to_update.extend(["email_verified_at", "updated_at"])
    if fields_to_update:
        profile.save(update_fields=fields_to_update)

    connection = _touch_social_connection(
        user=target,
        provider=provider,
        provider_user_id=str(provider_user_id),
        email=email,
        display_name=display_name,
        avatar_url=avatar_url,
    )
    _log_audit(target, "update", connection, {"event": "social_account_linked", "provider": provider, "ip": ip})
    _log_security_event(target, get_user_model(), object_id=str(target.pk), changes={"event": f"{provider}_linked", "ip": ip})
    return _oauth_frontend_redirect(provider=provider, extras={"linked": "1"})


def _effective_region_settings_for_user(request, user) -> dict:
    return effective_region_settings_for_user(request, user, detect_country=_detect_country)


def _apply_ordering_query(queryset, *, params, allowed_fields: dict[str, str], default: tuple[str, ...]):
    raw = params.get("ordering") or params.get("sort") or ""
    parts = [part.strip() for part in str(raw).split(",") if part.strip()]
    if not parts:
        return queryset.order_by(*default)

    order_by: list[str] = []
    invalid: list[str] = []
    for part in parts:
        descending = part.startswith("-")
        key = part[1:] if descending else part
        target = allowed_fields.get(key)
        if not target:
            invalid.append(key)
            continue
        order_by.append(f"-{target}" if descending else target)
    if invalid:
        raise ValidationError({"ordering": f"Invalid ordering field(s): {', '.join(invalid)}"})
    return queryset.order_by(*order_by)



def _serialize_settings_for_audit(instance) -> dict:
    data = {}
    for field in instance._meta.fields:
        name = field.name
        if name in ("id",):
            continue
        value = getattr(instance, name)
        if hasattr(value, "pk"):
            data[name] = value.pk
        elif isinstance(value, Decimal):
            data[name] = str(value)
        elif isinstance(value, (timezone.datetime, date)):
            data[name] = value.isoformat() if value is not None else None
        else:
            data[name] = value
    return data


def _convert_currency(amount: Decimal, base_code: str, quote_code: str) -> Decimal:
    base = (base_code or "").upper()
    quote = (quote_code or "").upper()
    if base == quote:
        return amount
    try:
        rate = ExchangeRate.objects.get(base_code=base, quote_code=quote).rate
        return (amount * rate).quantize(CENTS, rounding=ROUND_HALF_UP)
    except ExchangeRate.DoesNotExist:
        pass
    try:
        inverse = ExchangeRate.objects.get(base_code=quote, quote_code=base).rate
        return (amount / inverse).quantize(CENTS, rounding=ROUND_HALF_UP)
    except ExchangeRate.DoesNotExist:
        raise NotFound("Exchange rate not found")


class SoftDeleteModelViewSet(viewsets.ModelViewSet):
    permission_classes = [EditDeletePermission]
    pagination_class = None

    def _disable_response_cache(self, resp: Response) -> Response:
        resp["Cache-Control"] = "no-store"
        resp["Pragma"] = "no-cache"
        resp["Expires"] = "0"
        return resp

    def get_queryset(self):
        return super().get_queryset().filter(is_deleted=False)

    def _require_model_perm(self, action: str) -> None:
        user = self.request.user
        if not getattr(user, "is_authenticated", False):
            raise PermissionDenied()
        model = self.get_queryset().model
        prefix = None
        mn = model._meta.model_name
        if mn == "customer":
            prefix = "data.customers"
        elif mn == "item":
            prefix = "data.items"
        elif mn == "invoice":
            prefix = "data.invoices"
        elif mn == "invoiceitem":
            prefix = "data.invoices"
        elif mn == "receipt":
            prefix = "data.receipts"
        elif mn == "expense":
            prefix = "data.expenses"
        elif mn == "sourceaccount":
            prefix = "data.source_accounts"

        if prefix:
            code = f"{prefix}.read" if action == "view" else f"{prefix}.write"
            if user_has_permission(user, code):
                return
            raise PermissionDenied()

        app_label = model._meta.app_label
        model_name = model._meta.model_name
        if action in ("view", "add", "change", "delete"):
            perm = f"{app_label}.{action}_{model_name}"
        else:
            raise ValidationError({"detail": "Invalid permission check"})
        if not user.has_perm(perm):
            raise PermissionDenied()

    def _require_any_model_perm(self, actions: list[str]) -> None:
        user = self.request.user
        if not getattr(user, "is_authenticated", False):
            raise PermissionDenied()
        model = self.get_queryset().model
        prefix = None
        mn = model._meta.model_name
        if mn == "customer":
            prefix = "data.customers"
        elif mn == "item":
            prefix = "data.items"
        elif mn == "invoice":
            prefix = "data.invoices"
        elif mn == "invoiceitem":
            prefix = "data.invoices"
        elif mn == "receipt":
            prefix = "data.receipts"
        elif mn == "expense":
            prefix = "data.expenses"
        elif mn == "sourceaccount":
            prefix = "data.source_accounts"

        if prefix:
            for action in actions:
                code = f"{prefix}.read" if action == "view" else f"{prefix}.write"
                if user_has_permission(user, code):
                    return
            raise PermissionDenied()

        app_label = model._meta.app_label
        model_name = model._meta.model_name
        for action in actions:
            if action not in ("view", "add", "change", "delete"):
                continue
            if user.has_perm(f"{app_label}.{action}_{model_name}"):
                return
        raise PermissionDenied()

    def list(self, request, *args, **kwargs):
        self._require_model_perm("view")
        return self._disable_response_cache(super().list(request, *args, **kwargs))

    def retrieve(self, request, *args, **kwargs):
        self._require_model_perm("view")
        return self._disable_response_cache(super().retrieve(request, *args, **kwargs))

    def create(self, request, *args, **kwargs):
        model = self.get_queryset().model
        try:
            self._require_model_perm("add")
            with transaction.atomic():
                serializer = self.get_serializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                instance = serializer.save()
                _log_audit(request.user, "create", instance, {})
            headers = self.get_success_headers(serializer.data)
            return self._disable_response_cache(Response(self.get_serializer(instance).data, status=status.HTTP_201_CREATED, headers=headers))
        except Exception as exc:
            _raise_record_save_failure(request=request, entity=model._meta.model_name, operation="create", exc=exc)

    def update(self, request, *args, **kwargs):
        model = self.get_queryset().model
        try:
            self._require_model_perm("change")
            partial = kwargs.pop("partial", False)
            with transaction.atomic():
                instance = model.objects.select_for_update().get(pk=self.get_object().pk, is_deleted=False)
                _check_concurrency(request, instance)
                serializer = self.get_serializer(instance, data=request.data, partial=partial)
                serializer.is_valid(raise_exception=True)
                before = {k: getattr(instance, k, None) for k in serializer.validated_data.keys()}
                updated = serializer.save()
                after = {k: getattr(updated, k, None) for k in serializer.validated_data.keys()}
                changes = {k: {"from": str(before[k]) if before[k] is not None else None, "to": str(after[k]) if after[k] is not None else None} for k in before.keys() if before[k] != after[k]}
                if changes:
                    _log_audit(request.user, "update", updated, changes)
            return self._disable_response_cache(Response(self.get_serializer(updated).data, status=status.HTTP_200_OK))
        except Exception as exc:
            _raise_record_save_failure(request=request, entity=model._meta.model_name, operation="update", exc=exc)

    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        self._require_model_perm("delete")
        model = self.get_queryset().model
        with transaction.atomic():
            instance = model.objects.select_for_update().get(pk=self.get_object().pk, is_deleted=False)
            _check_concurrency(request, instance)
            instance.is_deleted = True
            instance.deleted_at = timezone.now()
            instance.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
            _log_audit(request.user, "delete", instance, {})
        return self._disable_response_cache(Response(status=status.HTTP_204_NO_CONTENT))

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        self._require_model_perm("delete")
        ids = request.data.get("ids", [])
        if not isinstance(ids, list) or not ids:
            raise ValidationError({"ids": "ids must be a non-empty list"})
        model = self.get_queryset().model
        now = timezone.now()
        with transaction.atomic():
            qs = model.objects.select_for_update().filter(id__in=ids, is_deleted=False)
            found = list(qs)
            if len(found) != len(set(ids)):
                raise NotFound("One or more records not found")
            for obj in found:
                obj.is_deleted = True
                obj.deleted_at = now
                obj.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
                _log_audit(request.user, "bulk_delete", obj, {"bulk": True})
        return self._disable_response_cache(Response({"deleted": len(ids)}, status=status.HTTP_200_OK))


class OptionalPageNumberPagination(PageNumberPagination):
    def paginate_queryset(self, queryset, request, view=None):
        self.page_size = getattr(settings, "API_PAGE_SIZE", 25)
        return super().paginate_queryset(queryset, request, view=view)


def _subtract_months(d: date, months: int) -> date:
    year = d.year
    month = d.month - months
    while month <= 0:
        month += 12
        year -= 1
    last_day = calendar.monthrange(year, month)[1]
    day = min(d.day, last_day)
    return date(year, month, day)


def _period_range(period: str) -> Tuple[date, date, date, date, str]:
    end = timezone.localdate()
    if period == "1m":
        start = end - timedelta(days=30)
        prev_end = start - timedelta(days=1)
        prev_start = prev_end - timedelta(days=30)
        granularity = "day"
    elif period == "6m":
        start = _subtract_months(end, 6)
        prev_end = start - timedelta(days=1)
        prev_start = _subtract_months(prev_end, 6)
        granularity = "month"
    elif period == "12m":
        start = _subtract_months(end, 12)
        prev_end = start - timedelta(days=1)
        prev_start = _subtract_months(prev_end, 12)
        granularity = "month"
    else:
        raise ValidationError({"period": "Invalid period. Use 1m, 6m, or 12m"})
    return start, end, prev_start, prev_end, granularity


def _pct_change(current: Decimal, previous: Decimal) -> Optional[str]:
    if previous == 0:
        return None
    return str(_q2(((current - previous) / previous) * Decimal("100")))


class FinanceViewSet(viewsets.ViewSet):
    def list(self, request):
        period = request.query_params.get("period", "6m")
        start, end, prev_start, prev_end, granularity = _period_range(period)

        receipts = Receipt.objects.filter(is_deleted=False, payment_date__gte=start, payment_date__lte=end)
        expenses = Expense.objects.filter(is_deleted=False, expense_date__gte=start, expense_date__lte=end)
        receipts_prev = Receipt.objects.filter(is_deleted=False, payment_date__gte=prev_start, payment_date__lte=prev_end)
        expenses_prev = Expense.objects.filter(is_deleted=False, expense_date__gte=prev_start, expense_date__lte=prev_end)

        income_total = receipts.aggregate(total=Coalesce(Sum("amount_paid"), Decimal("0.00")))["total"]
        expense_total = expenses.aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))["total"]
        income_prev_total = receipts_prev.aggregate(total=Coalesce(Sum("amount_paid"), Decimal("0.00")))["total"]
        expense_prev_total = expenses_prev.aggregate(total=Coalesce(Sum("amount"), Decimal("0.00")))["total"]

        if granularity == "day":
            income_series = list(
                receipts.annotate(bucket=TruncDay("payment_date"))
                .values("bucket")
                .annotate(total=Coalesce(Sum("amount_paid"), Decimal("0.00")))
                .order_by("bucket")
            )
            expense_series = list(
                expenses.annotate(bucket=TruncDay("expense_date"))
                .values("bucket")
                .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
                .order_by("bucket")
            )
            label_fmt = "%Y-%m-%d"
        else:
            income_series = list(
                receipts.annotate(bucket=TruncMonth("payment_date"))
                .values("bucket")
                .annotate(total=Coalesce(Sum("amount_paid"), Decimal("0.00")))
                .order_by("bucket")
            )
            expense_series = list(
                expenses.annotate(bucket=TruncMonth("expense_date"))
                .values("bucket")
                .annotate(total=Coalesce(Sum("amount"), Decimal("0.00")))
                .order_by("bucket")
            )
            label_fmt = "%Y-%m"

        income_by_bucket = {r["bucket"].date() if hasattr(r["bucket"], "date") else r["bucket"]: r["total"] for r in income_series}
        expense_by_bucket = {r["bucket"].date() if hasattr(r["bucket"], "date") else r["bucket"]: r["total"] for r in expense_series}

        if granularity == "day":
            buckets = []
            cur = start
            while cur <= end:
                buckets.append(cur)
                cur += timedelta(days=1)
        else:
            buckets = []
            cur = date(start.year, start.month, 1)
            last = date(end.year, end.month, 1)
            while cur <= last:
                buckets.append(cur)
                y = cur.year + (cur.month // 12)
                m = (cur.month % 12) + 1
                cur = date(y, m, 1)

        points = []
        for b in buckets:
            points.append(
                {
                    "label": b.strftime(label_fmt),
                    "income": str(_q2(Decimal(income_by_bucket.get(b, Decimal("0.00"))))),
                    "expense": str(_q2(Decimal(expense_by_bucket.get(b, Decimal("0.00"))))),
                }
            )

        return Response(
            {
                "period": period,
                "range": {"start": str(start), "end": str(end)},
                "income_total": str(_q2(Decimal(income_total))),
                "expense_total": str(_q2(Decimal(expense_total))),
                "income_change_pct": _pct_change(Decimal(income_total), Decimal(income_prev_total)),
                "expense_change_pct": _pct_change(Decimal(expense_total), Decimal(expense_prev_total)),
                "points": points,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"])
    def activity(self, request):
        activity_type = request.query_params.get("type", "all")
        try:
            limit = int(request.query_params.get("limit", "15"))
        except ValueError:
            raise ValidationError({"limit": "Invalid limit"})
        limit = max(1, min(limit, 50))

        events = []
        if activity_type in ("all", "income"):
            for r in Receipt.objects.filter(is_deleted=False).order_by("-payment_date", "-id")[:limit]:
                events.append(
                    {
                        "type": "income",
                        "amount": str(r.amount_paid),
                        "date": str(r.payment_date),
                        "description": f"Receipt {r.reference_number or ''}".strip() or "Receipt",
                    }
                )
        if activity_type in ("all", "expense"):
            for e in Expense.objects.filter(is_deleted=False).order_by("-expense_date", "-id")[:limit]:
                events.append(
                    {
                        "type": "expense",
                        "amount": str(e.amount),
                        "date": str(e.expense_date),
                        "description": e.description or e.category or e.vendor or "Expense",
                    }
                )

        events.sort(key=lambda x: (x["date"], x["type"]), reverse=True)
        return Response({"events": events[:limit]}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"])
    def top_products(self, request):
        period = request.query_params.get("period", "6m")
        start, end, _, _, _ = _period_range(period)

        items_qs = (
            InvoiceItem.objects.filter(
                is_deleted=False,
                invoice__is_deleted=False,
                invoice__status__in=["Sent", "Paid"],
                invoice__issue_date__gte=start,
                invoice__issue_date__lte=end,
            )
            .select_related("item")
        )

        totals = items_qs.aggregate(total=Coalesce(Sum("line_total"), Decimal("0.00")))["total"]
        rows = list(
            items_qs.values("item_id", "item__name")
            .annotate(units_sold=Coalesce(Sum("quantity"), 0), revenue=Coalesce(Sum("line_total"), Decimal("0.00")))
            .order_by("-revenue", "item__name")[:10]
        )

        total_sales = Decimal(totals)
        products = []
        for r in rows:
            revenue = Decimal(r["revenue"])
            pct = None if total_sales == 0 else str(_q2((revenue / total_sales) * Decimal("100")))
            products.append(
                {
                    "item_id": r["item_id"],
                    "name": r["item__name"],
                    "units_sold": int(r["units_sold"]),
                    "revenue": str(_q2(revenue)),
                    "pct_of_total_sales": pct,
                }
            )

        return Response(
            {
                "period": period,
                "range": {"start": str(start), "end": str(end)},
                "total_sales": str(_q2(total_sales)),
                "products": products,
            },
            status=status.HTTP_200_OK,
        )


class DashboardViewSet(viewsets.ViewSet):
    def list(self, request):
        total_revenue = Receipt.objects.filter(is_deleted=False).aggregate(total=Coalesce(Sum("amount_paid"), Decimal("0.00")))[
            "total"
        ]

        outstanding_qs = (
            Invoice.objects.filter(is_deleted=False, status__in=["Sent", "Overdue"])
            .annotate(paid=Coalesce(Sum("receipts__amount_paid", filter=Q(receipts__is_deleted=False)), Decimal("0.00")))
            .annotate(outstanding=F("total_amount") - F("paid"))
            .filter(outstanding__gt=Decimal("0.00"))
        )
        outstanding_invoices_count = outstanding_qs.count()
        outstanding_amount = outstanding_qs.aggregate(
            total=Coalesce(Sum("outstanding"), Decimal("0.00"))
        )["total"]

        low_stock_qs = Item.objects.filter(is_deleted=False, type="product", stock_quantity__lt=5).order_by("stock_quantity", "id")
        low_stock_count = low_stock_qs.count()
        low_stock_items = list(
            low_stock_qs.values("id", "name", "sku", "stock_quantity")[:10]
        )

        payload = {
            "total_revenue": str(total_revenue),
            "outstanding_invoices_count": outstanding_invoices_count,
            "outstanding_amount": str(outstanding_amount),
            "low_stock_count": low_stock_count,
            "low_stock_items": low_stock_items,
        }
        logger.info(
            "dashboard.metrics revenue=%s outstanding_count=%s outstanding_amount=%s low_stock=%s",
            payload["total_revenue"],
            outstanding_invoices_count,
            payload["outstanding_amount"],
            low_stock_count,
        )
        return Response(payload, status=status.HTTP_200_OK)


class ReportsViewSet(viewsets.ViewSet):
    def list(self, request):
        start = request.query_params.get("start")
        end = request.query_params.get("end")

        start_date = None
        end_date = None
        if start:
            try:
                start_date = date.fromisoformat(start)
            except ValueError:
                raise ValidationError({"start": "Invalid date. Use YYYY-MM-DD"})
        if end:
            try:
                end_date = date.fromisoformat(end)
            except ValueError:
                raise ValidationError({"end": "Invalid date. Use YYYY-MM-DD"})

        receipts_qs = Receipt.objects.filter(is_deleted=False)
        invoices_qs = Invoice.objects.filter(is_deleted=False)
        invoice_items_qs = InvoiceItem.objects.filter(is_deleted=False, invoice__is_deleted=False).select_related("invoice", "item")

        if start_date:
            receipts_qs = receipts_qs.filter(payment_date__gte=start_date)
            invoices_qs = invoices_qs.filter(issue_date__gte=start_date)
            invoice_items_qs = invoice_items_qs.filter(invoice__issue_date__gte=start_date)
        if end_date:
            receipts_qs = receipts_qs.filter(payment_date__lte=end_date)
            invoices_qs = invoices_qs.filter(issue_date__lte=end_date)
            invoice_items_qs = invoice_items_qs.filter(invoice__issue_date__lte=end_date)

        revenue_total = receipts_qs.aggregate(total=Coalesce(Sum("amount_paid"), Decimal("0.00")))[
            "total"
        ]

        revenue_by_day = list(
            receipts_qs.values("payment_date")
            .annotate(total=Coalesce(Sum("amount_paid"), Decimal("0.00")))
            .order_by("payment_date")
        )
        revenue_by_day = [{"day": str(r["payment_date"]), "total": str(r["total"])} for r in revenue_by_day]

        invoice_status = list(
            invoices_qs.values("status").annotate(count=Count("id")).order_by("status")
        )

        top_items = list(
            invoice_items_qs.filter(invoice__status__in=["Sent", "Paid"])
            .values("item_id", "item__name", "item__sku", "item__type")
            .annotate(quantity=Coalesce(Sum("quantity"), 0))
            .order_by("-quantity", "item__name")[:10]
        )
        top_items = [
            {
                "item_id": r["item_id"],
                "name": r["item__name"],
                "sku": r["item__sku"],
                "type": r["item__type"],
                "quantity": int(r["quantity"]),
            }
            for r in top_items
        ]

        payload = {
            "range": {"start": str(start_date) if start_date else None, "end": str(end_date) if end_date else None},
            "revenue_total": str(revenue_total),
            "revenue_by_day": revenue_by_day,
            "invoice_status": invoice_status,
            "top_items": top_items,
        }
        logger.info("reports.summary start=%s end=%s", payload["range"]["start"], payload["range"]["end"])
        return Response(payload, status=status.HTTP_200_OK)


class PaymentsReportApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        start = request.query_params.get("start")
        end = request.query_params.get("end")

        start_date = None
        end_date = None
        if start:
            try:
                start_date = date.fromisoformat(start)
            except ValueError:
                raise ValidationError({"start": "Invalid date. Use YYYY-MM-DD"})
        if end:
            try:
                end_date = date.fromisoformat(end)
            except ValueError:
                raise ValidationError({"end": "Invalid date. Use YYYY-MM-DD"})

        receipts_qs = Receipt.objects.filter(is_deleted=False)
        if start_date:
            receipts_qs = receipts_qs.filter(payment_date__gte=start_date)
        if end_date:
            receipts_qs = receipts_qs.filter(payment_date__lte=end_date)

        by_method = list(
            receipts_qs.values("payment_method")
            .annotate(count=Count("id"), total=Coalesce(Sum("amount_paid"), Decimal("0.00")))
            .order_by("payment_method")
        )
        by_method = [{"payment_method": r["payment_method"], "count": int(r["count"]), "total": str(r["total"])} for r in by_method]

        tx_qs = PaymentTransaction.objects.all()
        if start_date:
            tx_qs = tx_qs.filter(created_at__date__gte=start_date)
        if end_date:
            tx_qs = tx_qs.filter(created_at__date__lte=end_date)

        by_provider_status = list(tx_qs.values("provider", "status").annotate(count=Count("id")).order_by("provider", "status"))
        payload = {
            "range": {"start": str(start_date) if start_date else None, "end": str(end_date) if end_date else None},
            "receipts_by_method": by_method,
            "transactions_by_provider_status": [{"provider": r["provider"], "status": r["status"], "count": int(r["count"])} for r in by_provider_status],
        }
        logger.info("payments.report start=%s end=%s", payload["range"]["start"], payload["range"]["end"])
        return Response(payload, status=status.HTTP_200_OK)


class CustomerViewSet(SoftDeleteModelViewSet):
    queryset = Customer.objects.all().order_by("-id")
    serializer_class = CustomerSerializer
    pagination_class = OptionalPageNumberPagination

    def get_serializer_class(self):
        if getattr(self, "action", None) == "retrieve":
            return CustomerDetailSerializer
        if getattr(self, "action", None) == "list":
            return CustomerListSerializer
        return CustomerSerializer

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .annotate(
                invoice_count=Count("invoices", filter=Q(invoices__is_deleted=False), distinct=True),
                lifetime_value=Coalesce(Sum("invoices__total_amount", filter=Q(invoices__is_deleted=False)), Decimal("0.00")),
                last_invoice_date=Max("invoices__issue_date", filter=Q(invoices__is_deleted=False)),
                total_paid_amount=Coalesce(
                    Sum("invoices__receipts__amount_paid", filter=Q(invoices__is_deleted=False, invoices__receipts__is_deleted=False)),
                    Decimal("0.00"),
                ),
            )
        )
        p = self.request.query_params

        q = str(p.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(name__icontains=q)
                | Q(email__icontains=q)
                | Q(phone__icontains=q)
                | Q(billing_address__icontains=q)
            )

        email = str(p.get("email") or "").strip()
        if email:
            qs = qs.filter(email__icontains=email)

        phone = str(p.get("phone") or "").strip()
        if phone:
            qs = qs.filter(phone__icontains=phone)

        def _date_param(key: str):
            raw = p.get(key)
            if raw in (None, ""):
                return None
            try:
                return date.fromisoformat(str(raw))
            except ValueError:
                raise ValidationError({key: "Invalid date. Use YYYY-MM-DD"})

        created_from = _date_param("created_from")
        created_to = _date_param("created_to")
        if created_from:
            qs = qs.filter(created_at__date__gte=created_from)
        if created_to:
            qs = qs.filter(created_at__date__lte=created_to)

        account_status = str(p.get("account_status") or "").strip().lower()
        if account_status:
            if account_status == "active":
                qs = qs.filter(invoice_count__gt=0)
            elif account_status in ("prospect", "inactive"):
                qs = qs.filter(invoice_count=0)
            else:
                raise ValidationError({"account_status": "Invalid account_status. Use active, prospect, or inactive"})

        segment = str(p.get("segment") or "").strip().lower()
        if segment:
            if segment == "vip":
                qs = qs.filter(Q(invoice_count__gte=5) | Q(lifetime_value__gte=Decimal("10000.00")))
            elif segment == "standard":
                qs = qs.filter(invoice_count__gt=0).exclude(Q(invoice_count__gte=5) | Q(lifetime_value__gte=Decimal("10000.00")))
            elif segment == "prospect":
                qs = qs.filter(invoice_count=0)
            else:
                raise ValidationError({"segment": "Invalid segment. Use prospect, standard, or vip"})

        return _apply_ordering_query(
            qs,
            params=p,
            allowed_fields={
                "id": "id",
                "name": "name",
                "email": "email",
                "phone": "phone",
                "created_at": "created_at",
                "updated_at": "updated_at",
                "invoice_count": "invoice_count",
                "lifetime_value": "lifetime_value",
                "last_invoice_date": "last_invoice_date",
            },
            default=("-id",),
        )

    def destroy(self, request, *args, **kwargs):
        self._require_model_perm("delete")
        with transaction.atomic():
            customer = Customer.objects.select_for_update().get(pk=self.get_object().pk, is_deleted=False)
            _check_concurrency(request, customer)
            if Invoice.objects.filter(customer=customer, is_deleted=False).exists():
                return Response(
                    {"detail": "Cannot delete customer with existing invoices."},
                    status=status.HTTP_409_CONFLICT,
                )
            customer.is_deleted = True
            customer.deleted_at = timezone.now()
            customer.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
            _log_audit(request.user, "delete", customer, {})
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        self._require_model_perm("delete")
        ids = request.data.get("ids", [])
        if not isinstance(ids, list) or not ids:
            raise ValidationError({"ids": "ids must be a non-empty list"})
        with transaction.atomic():
            customers = list(Customer.objects.select_for_update().filter(id__in=ids, is_deleted=False))
            if len(customers) != len(set(ids)):
                raise NotFound("One or more records not found")
            blocked = list(
                Invoice.objects.filter(customer_id__in=ids, is_deleted=False)
                .values_list("customer_id", flat=True)
                .distinct()
            )
            if blocked:
                return Response(
                    {"detail": "Cannot delete customers with existing invoices.", "blocked_ids": blocked},
                    status=status.HTTP_409_CONFLICT,
                )
            now = timezone.now()
            for customer in customers:
                customer.is_deleted = True
                customer.deleted_at = now
                customer.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
                _log_audit(request.user, "bulk_delete", customer, {"bulk": True})
        return Response({"deleted": len(ids)}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        self._require_model_perm("view")

        fmt = str(request.query_params.get("file_format") or request.query_params.get("export_format") or "csv").strip().lower()
        if fmt not in ("csv", "xlsx"):
            raise ValidationError({"file_format": "Invalid file_format. Use csv or xlsx"})

        allowed_fields = [
            "id",
            "name",
            "email",
            "phone",
            "billing_address",
            "account_status",
            "segment",
            "invoice_count",
            "lifetime_value",
            "total_paid_amount",
            "last_invoice_date",
            "order_history",
            "created_at",
            "updated_at",
        ]
        fields_raw = request.query_params.get("fields")
        if fields_raw in (None, ""):
            fields = [
                "name",
                "email",
                "phone",
                "account_status",
                "segment",
                "invoice_count",
                "lifetime_value",
                "total_paid_amount",
                "last_invoice_date",
            ]
        else:
            parts = [p.strip() for p in str(fields_raw).split(",") if p.strip()]
            invalid = [p for p in parts if p not in allowed_fields]
            if invalid:
                raise ValidationError({"fields": f"Invalid fields: {', '.join(invalid)}"})
            fields = parts or allowed_fields

        rows_limit = 50000
        try:
            limit_raw = request.query_params.get("limit")
            if limit_raw not in (None, ""):
                rows_limit = int(limit_raw)
        except ValueError:
            raise ValidationError({"limit": "Invalid limit"})
        rows_limit = max(1, min(rows_limit, 50000))

        qs = self.filter_queryset(self.get_queryset()).order_by("-id")[:rows_limit]

        def _csv_cell(value: str) -> str:
            v = str(value or "")
            if v and v[0] in ("=", "+", "-", "@"):
                return "'" + v
            return v

        def _account_status(customer: Customer) -> str:
            return "active" if int(getattr(customer, "invoice_count", 0) or 0) > 0 else "prospect"

        def _segment(customer: Customer) -> str:
            invoice_count = int(getattr(customer, "invoice_count", 0) or 0)
            lifetime_value = Decimal(str(getattr(customer, "lifetime_value", Decimal("0.00")) or Decimal("0.00")))
            if invoice_count == 0:
                return "prospect"
            if invoice_count >= 5 or lifetime_value >= Decimal("10000.00"):
                return "vip"
            return "standard"

        def _order_history(customer: Customer) -> str:
            invoices = customer.invoices.filter(is_deleted=False).order_by("-issue_date", "-id")[:20]
            parts = [f"{inv.invoice_number}:{inv.status}:{inv.total_amount}" for inv in invoices]
            return " | ".join(parts)

        def _value_for(customer: Customer, field: str) -> str:
            if field == "account_status":
                return _account_status(customer)
            if field == "segment":
                return _segment(customer)
            if field == "invoice_count":
                return str(int(getattr(customer, "invoice_count", 0) or 0))
            if field in ("lifetime_value", "total_paid_amount"):
                return str(getattr(customer, field, "") or "")
            if field == "last_invoice_date":
                value = getattr(customer, field, None)
                return str(value or "")
            if field == "order_history":
                return _order_history(customer)
            if field in ("created_at", "updated_at"):
                dt = getattr(customer, field, None)
                return dt.isoformat() if dt else ""
            return str(getattr(customer, field, "") or "")

        filename_base = "customers"
        _log_operation(request.user, "export", Customer, "customers_export", {"format": fmt, "fields": fields, "limit": rows_limit})

        if fmt == "csv":
            class _Echo:
                def write(self, value):
                    return value

            def _iter_rows():
                yield writer.writerow(fields)
                for customer in qs.iterator(chunk_size=2000):
                    yield writer.writerow([_csv_cell(_value_for(customer, f)) for f in fields])

            pseudo_buffer = _Echo()
            writer = csv.writer(pseudo_buffer)
            resp = StreamingHttpResponse(_iter_rows(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
            return resp

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Customers")
        ws.append(fields)
        for customer in qs.iterator(chunk_size=2000):
            ws.append([_value_for(customer, f) for f in fields])
        out = io.BytesIO()
        wb.save(out)
        resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    @action(detail=False, methods=["get"], url_path="import_template")
    def import_template(self, request):
        self._require_any_model_perm(["add", "change"])

        fmt = str(request.query_params.get("file_format") or "xlsx").strip().lower()
        if fmt not in ("csv", "xlsx"):
            raise ValidationError({"file_format": "Invalid file_format. Use csv or xlsx"})

        header = ["name", "email", "phone", "billing_address"]
        example = ["Example Customer", "buyer@example.com", "+2348012345678", "12 Marina Road, Lagos"]
        filename_base = "customer_import_template"

        if fmt == "csv":
            out = io.StringIO()
            w = csv.writer(out)
            w.writerow(header)
            w.writerow(example)
            resp = HttpResponse(out.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
            return resp

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Template")
        ws.append(header)
        ws.append(example)
        out = io.BytesIO()
        wb.save(out)
        resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    @action(detail=False, methods=["post"], url_path="import")
    def import_customers(self, request):
        self._require_any_model_perm(["add", "change"])

        upload = request.FILES.get("file")
        if upload is None:
            raise ValidationError({"file": "file is required"})

        dry_run = str(request.data.get("dry_run") or "").strip().lower() in ("1", "true", "yes", "on")
        rollback_on_error = str(request.data.get("rollback_on_error") or "").strip().lower() not in ("0", "false", "no", "off")
        from .import_export import import_customers_from_upload

        try:
            status_code, payload = import_customers_from_upload(upload, dry_run=dry_run, rollback_on_error=rollback_on_error)
        except ValueError:
            raise ValidationError({"file": "Unsupported file type. Use .csv or .xlsx"})

        if status_code >= 400:
            _log_operation(
                request.user,
                "import",
                Customer,
                "customers_import_failed",
                {"dry_run": dry_run, "rows": int(payload.get("rows") or 0), "errors": len(payload.get("errors") or [])},
            )
            return Response(payload, status=status_code)

        if payload.get("dry_run"):
            _log_operation(
                request.user,
                "import",
                Customer,
                "customers_import_dry_run",
                {"rows": int(payload.get("rows") or 0), "would_create": int(payload.get("would_create") or 0), "errors": len(payload.get("errors") or [])},
            )
            return Response(payload, status=status_code)

        _log_operation(
            request.user,
            "import",
            Customer,
            "customers_import",
            {"rows": int(payload.get("rows") or 0), "created": int(payload.get("imported") or 0), "errors": len(payload.get("errors") or [])},
        )
        return Response(payload, status=status_code)


class ItemViewSet(SoftDeleteModelViewSet):
    queryset = Item.objects.all().order_by("-id")
    serializer_class = ItemSerializer
    pagination_class = OptionalPageNumberPagination
    parser_classes = [parsers.JSONParser, parsers.MultiPartParser, parsers.FormParser]

    def get_serializer_class(self):
        if getattr(self, "action", None) == "retrieve":
            return ItemDetailSerializer
        return ItemSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        p = self.request.query_params

        q = str(p.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(sku__icontains=q))

        type_val = str(p.get("type") or "").strip()
        if type_val:
            allowed = {t for t, _ in Item.TYPE_CHOICES}
            if type_val not in allowed:
                raise ValidationError({"type": "Invalid type"})
            qs = qs.filter(type=type_val)

        warehouse_location = str(p.get("warehouse_location") or "").strip()
        if warehouse_location:
            qs = qs.filter(warehouse_location__icontains=warehouse_location)

        def _date_param(key: str):
            raw = p.get(key)
            if raw in (None, ""):
                return None
            try:
                return date.fromisoformat(str(raw))
            except ValueError:
                raise ValidationError({key: "Invalid date. Use YYYY-MM-DD"})

        created_from = _date_param("created_from")
        created_to = _date_param("created_to")
        if created_from:
            qs = qs.filter(created_at__date__gte=created_from)
        if created_to:
            qs = qs.filter(created_at__date__lte=created_to)

        restock_from = _date_param("last_restock_from")
        restock_to = _date_param("last_restock_to")
        if restock_from:
            qs = qs.filter(last_restock_date__gte=restock_from)
        if restock_to:
            qs = qs.filter(last_restock_date__lte=restock_to)

        def _int_param(key: str):
            raw = p.get(key)
            if raw in (None, ""):
                return None
            try:
                return int(str(raw))
            except (TypeError, ValueError):
                raise ValidationError({key: "Invalid integer"})

        stock_min = _int_param("stock_min")
        stock_max = _int_param("stock_max")
        if stock_min is not None:
            qs = qs.filter(stock_quantity__gte=stock_min)
        if stock_max is not None:
            qs = qs.filter(stock_quantity__lte=stock_max)

        return _apply_ordering_query(
            qs,
            params=p,
            allowed_fields={
                "id": "id",
                "type": "type",
                "sku": "sku",
                "name": "name",
                "unit_price": "unit_price",
                "tax_rate": "tax_rate",
                "stock_quantity": "stock_quantity",
                "warehouse_location": "warehouse_location",
                "last_restock_date": "last_restock_date",
                "created_at": "created_at",
                "updated_at": "updated_at",
            },
            default=("-id",),
        )

    def destroy(self, request, *args, **kwargs):
        self._require_model_perm("delete")
        with transaction.atomic():
            item = Item.objects.select_for_update().get(pk=self.get_object().pk, is_deleted=False)
            _check_concurrency(request, item)
            if InvoiceItem.objects.filter(item=item, is_deleted=False, invoice__is_deleted=False).exists():
                return Response(
                    {"detail": "Cannot delete item that is referenced by active invoices."},
                    status=status.HTTP_409_CONFLICT,
                )
            item.is_deleted = True
            item.deleted_at = timezone.now()
            item.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
            _log_audit(request.user, "delete", item, {})
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        self._require_model_perm("delete")
        ids = request.data.get("ids", [])
        if not isinstance(ids, list) or not ids:
            raise ValidationError({"ids": "ids must be a non-empty list"})
        with transaction.atomic():
            items = list(Item.objects.select_for_update().filter(id__in=ids, is_deleted=False))
            if len(items) != len(set(ids)):
                raise NotFound("One or more records not found")
            blocked = list(
                InvoiceItem.objects.filter(item_id__in=ids, is_deleted=False, invoice__is_deleted=False)
                .values_list("item_id", flat=True)
                .distinct()
            )
            if blocked:
                return Response(
                    {"detail": "Cannot delete items referenced by active invoices.", "blocked_ids": blocked},
                    status=status.HTTP_409_CONFLICT,
                )
            now = timezone.now()
            for item in items:
                item.is_deleted = True
                item.deleted_at = now
                item.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
                _log_audit(request.user, "bulk_delete", item, {"bulk": True})
        return Response({"deleted": len(ids)}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def adjust_stock(self, request, pk=None):
        self._require_model_perm("change")
        adjustment = request.data.get("adjustment", 0)
        try:
            adjustment = int(adjustment)
        except (TypeError, ValueError):
            raise ValidationError({"adjustment": "Adjustment must be an integer"})

        with transaction.atomic():
            try:
                item = Item.objects.select_for_update().get(pk=pk, is_deleted=False)
            except Item.DoesNotExist:
                raise NotFound("Item not found")
            if item.type != "product":
                raise ValidationError({"detail": "Stock can only be adjusted for products"})
            new_qty = item.stock_quantity + adjustment
            if new_qty < 0:
                raise ValidationError({"stock_quantity": "stock_quantity cannot go below 0"})
            item.stock_quantity = new_qty
            item.save(update_fields=["stock_quantity", "updated_at"])

        logger.info("inventory.adjust_stock item_id=%s adjustment=%s new_stock=%s", item.id, adjustment, item.stock_quantity)
        return Response(
            {
                'item': ItemSerializer(item).data,
                'adjustment': adjustment,
                'new_stock_quantity': item.stock_quantity
            },
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        if not getattr(request.user, "is_authenticated", False):
            raise PermissionDenied()

        fmt = str(request.query_params.get("file_format") or request.query_params.get("export_format") or "csv").strip().lower()
        if fmt not in ("csv", "xlsx", "pdf"):
            raise ValidationError({"file_format": "Invalid file_format. Use csv, xlsx, or pdf"})

        allowed_fields = [
            "id",
            "type",
            "sku",
            "name",
            "description",
            "unit_price",
            "tax_rate",
            "tax_category",
            "unit_of_measure",
            "stock_quantity",
            "created_at",
            "updated_at",
        ]
        fields_raw = request.query_params.get("fields")
        if fields_raw in (None, ""):
            fields = ["type", "sku", "name", "unit_price", "tax_rate", "stock_quantity", "updated_at"]
        else:
            parts = [p.strip() for p in str(fields_raw).split(",") if p.strip()]
            invalid = [p for p in parts if p not in allowed_fields]
            if invalid:
                raise ValidationError({"fields": f"Invalid fields: {', '.join(invalid)}"})
            fields = parts or ["type", "sku", "name", "unit_price", "tax_rate", "stock_quantity", "updated_at"]

        rows_limit = 50000 if fmt in ("csv", "xlsx") else 2000
        try:
            limit_raw = request.query_params.get("limit")
            if limit_raw not in (None, ""):
                rows_limit = int(limit_raw)
        except ValueError:
            raise ValidationError({"limit": "Invalid limit"})
        rows_limit = max(1, min(rows_limit, 50000 if fmt in ("csv", "xlsx") else 5000))

        qs = self.filter_queryset(self.get_queryset()).order_by("-id")[:rows_limit]

        def _csv_cell(value: str) -> str:
            v = str(value or "")
            if v and v[0] in ("=", "+", "-", "@"):
                return "'" + v
            return v

        def _value_for(it: Item, field: str) -> str:
            if field in ("created_at", "updated_at"):
                dt = getattr(it, field, None)
                return dt.isoformat() if dt else ""
            if field == "unit_price":
                return str(it.unit_price)
            if field == "tax_rate":
                return str(it.tax_rate)
            if field == "stock_quantity":
                return str(it.stock_quantity)
            return str(getattr(it, field, "") or "")

        filename_base = "inventory_items"
        _log_operation(request.user, "export", Item, "items_export", {"format": fmt, "fields": fields, "limit": rows_limit})

        if fmt == "csv":
            class _Echo:
                def write(self, value):
                    return value

            def _iter_rows():
                yield writer.writerow(fields)
                for it in qs.iterator(chunk_size=2000):
                    yield writer.writerow([_csv_cell(_value_for(it, f)) for f in fields])

            pseudo_buffer = _Echo()
            writer = csv.writer(pseudo_buffer)
            resp = StreamingHttpResponse(_iter_rows(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
            return resp

        if fmt == "xlsx":
            wb = Workbook(write_only=True)
            ws = wb.create_sheet("Items")
            ws.append(fields)
            for it in qs.iterator(chunk_size=2000):
                ws.append([_value_for(it, f) for f in fields])
            out = io.BytesIO()
            wb.save(out)
            resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
            return resp

        rows = []
        for it in qs.iterator(chunk_size=2000):
            row = {f: _value_for(it, f) for f in fields}
            rows.append(row)
        company_name = str(_effective_company_identity_for_user(request.user)["company_name"] or "PIXELHUB").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        html = "<html><head><meta charset='utf-8' /><style>body{font-family:Arial, sans-serif;font-size:10pt;}table{border-collapse:collapse;width:100%;}th,td{border:1px solid #ddd;padding:4px;}th{background:#f3f3f3;text-align:left;}.doc-company{font-size:18pt;font-weight:700;margin:0 0 4px;color:#1a4d8e;}.doc-title{font-size:12pt;font-weight:700;margin:0 0 12px;}</style></head><body>"
        html += f"<div class='doc-company'>{company_name}</div><div class='doc-title'>Inventory Export</div>"
        html += "<table><thead><tr>" + "".join([f"<th>{f}</th>" for f in fields]) + "</tr></thead><tbody>"
        for r in rows:
            html += "<tr>" + "".join([f"<td>{str(r.get(f, '')).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')}</td>" for f in fields]) + "</tr>"
        html += "</tbody></table></body></html>"
        try:
            from weasyprint import HTML

            pdf = HTML(string=html).write_pdf()
            resp = HttpResponse(pdf, content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.pdf"'
            resp["X-PDF-Backend"] = "weasyprint"
            return resp
        except (OSError, ValueError) as e:
            logger.warning("pdf.weasyprint_render_failed error=%s items_export=1", e)
            resp = HttpResponse(html, content_type="text/html; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.html"'
            resp["X-PDF-Backend"] = "failed"
            return resp

    @action(detail=False, methods=["get"], url_path="import_template")
    def import_template(self, request):
        self._require_any_model_perm(["add", "change"])

        fmt = str(request.query_params.get("file_format") or "xlsx").strip().lower()
        if fmt not in ("csv", "xlsx"):
            raise ValidationError({"file_format": "Invalid file_format. Use csv or xlsx"})

        header = [
            "type",
            "sku",
            "name",
            "description",
            "unit_price",
            "tax_rate",
            "tax_category",
            "unit_of_measure",
            "stock_quantity",
        ]
        example = ["product", "SKU-001", "Example Item", "Optional description", "10.00", "0", "standard", "pcs", "5"]
        filename_base = "inventory_import_template"

        if fmt == "csv":
            out = io.StringIO()
            w = csv.writer(out)
            w.writerow(header)
            w.writerow(example)
            resp = HttpResponse(out.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
            return resp

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Template")
        ws.append(header)
        ws.append(example)
        out = io.BytesIO()
        wb.save(out)
        resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    @action(detail=False, methods=["post"], url_path="import")
    def import_items(self, request):
        self._require_any_model_perm(["add", "change"])

        upload = request.FILES.get("file")
        if upload is None:
            raise ValidationError({"file": "file is required"})

        dry_run = str(request.data.get("dry_run") or "").strip().lower() in ("1", "true", "yes", "on")
        rollback_on_error = str(request.data.get("rollback_on_error") or "").strip().lower() not in ("0", "false", "no", "off")
        from .import_export import import_items_from_upload

        try:
            status_code, payload = import_items_from_upload(upload, dry_run=dry_run, rollback_on_error=rollback_on_error)
        except ValueError:
            raise ValidationError({"file": "Unsupported file type. Use .csv or .xlsx"})

        if status_code >= 400:
            _log_operation(
                request.user,
                "import",
                Item,
                "items_import_failed",
                {"dry_run": dry_run, "rows": int(payload.get("rows") or 0), "errors": len(payload.get("errors") or [])},
            )
            return Response(payload, status=status_code)

        if payload.get("dry_run"):
            _log_operation(
                request.user,
                "import",
                Item,
                "items_import_dry_run",
                {"rows": int(payload.get("rows") or 0), "would_create": int(payload.get("would_create") or 0), "errors": len(payload.get("errors") or [])},
            )
            return Response(payload, status=status_code)

        _log_operation(
            request.user,
            "import",
            Item,
            "items_import",
            {"rows": int(payload.get("rows") or 0), "created": int(payload.get("imported") or 0), "errors": len(payload.get("errors") or [])},
        )
        return Response(payload, status=status_code)


class InvoiceViewSet(SoftDeleteModelViewSet):
    queryset = Invoice.objects.all().select_related("customer").prefetch_related("invoice_items", "invoice_items__item").order_by("-id")
    serializer_class = InvoiceSerializer
    pagination_class = OptionalPageNumberPagination
    parser_classes = [parsers.JSONParser, parsers.MultiPartParser, parsers.FormParser]

    def get_queryset(self):
        qs = super().get_queryset().annotate(
            amount_paid=Coalesce(Sum("receipts__amount_paid", filter=Q(receipts__is_deleted=False)), Decimal("0.00")),
            line_item_count=Count("invoice_items", filter=Q(invoice_items__is_deleted=False), distinct=True),
            last_payment_date=Max("receipts__payment_date", filter=Q(receipts__is_deleted=False)),
        )
        p = self.request.query_params

        q = str(p.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(invoice_number__icontains=q) | Q(customer__name__icontains=q))

        invoice_number = str(p.get("invoice_number") or "").strip()
        if invoice_number:
            qs = qs.filter(invoice_number__icontains=invoice_number)

        customer_name = str(p.get("customer_name") or "").strip()
        if customer_name:
            qs = qs.filter(customer__name__icontains=customer_name)

        customer_id = p.get("customer")
        if customer_id not in (None, ""):
            try:
                customer_id_int = int(customer_id)
            except (TypeError, ValueError):
                raise ValidationError({"customer": "Invalid customer"})
            qs = qs.filter(customer_id=customer_id_int)

        status_val = str(p.get("status") or "").strip()
        if status_val:
            allowed = {s for s, _ in Invoice.STATUS_CHOICES}
            if status_val not in allowed:
                raise ValidationError({"status": "Invalid status"})
            qs = qs.filter(status=status_val)

        payment_status = str(p.get("payment_status") or "").strip().lower()
        if payment_status:
            if payment_status == "paid":
                qs = qs.filter(amount_paid__gte=F("total_amount"))
            elif payment_status == "partial":
                qs = qs.filter(amount_paid__gt=Decimal("0.00"), amount_paid__lt=F("total_amount"))
            elif payment_status == "unpaid":
                qs = qs.filter(amount_paid__lte=Decimal("0.00"))
            else:
                raise ValidationError({"payment_status": "Invalid payment_status. Use paid, partial, or unpaid"})

        def _date_param(key: str):
            raw = p.get(key)
            if raw in (None, ""):
                return None
            try:
                return date.fromisoformat(str(raw))
            except ValueError:
                raise ValidationError({key: "Invalid date. Use YYYY-MM-DD"})

        issue_from = _date_param("issue_date_from")
        issue_to = _date_param("issue_date_to")
        if issue_from:
            qs = qs.filter(issue_date__gte=issue_from)
        if issue_to:
            qs = qs.filter(issue_date__lte=issue_to)

        due_from = _date_param("due_date_from")
        due_to = _date_param("due_date_to")
        if due_from:
            qs = qs.filter(due_date__gte=due_from)
        if due_to:
            qs = qs.filter(due_date__lte=due_to)

        def _dec_param(key: str):
            raw = p.get(key)
            if raw in (None, ""):
                return None
            try:
                v = Decimal(str(raw))
            except (InvalidOperation, TypeError):
                raise ValidationError({key: "Invalid amount"})
            if v < 0:
                raise ValidationError({key: "Amount must be >= 0"})
            return v

        total_min = _dec_param("total_min")
        total_max = _dec_param("total_max")
        if total_min is not None:
            qs = qs.filter(total_amount__gte=total_min)
        if total_max is not None:
            qs = qs.filter(total_amount__lte=total_max)

        return _apply_ordering_query(
            qs,
            params=p,
            allowed_fields={
                "id": "id",
                "invoice_number": "invoice_number",
                "customer_name": "customer__name",
                "issue_date": "issue_date",
                "due_date": "due_date",
                "status": "status",
                "subtotal": "subtotal",
                "tax_total": "tax_total",
                "total_amount": "total_amount",
                "amount_paid": "amount_paid",
                "balance_due": "total_amount",
                "payment_date": "last_payment_date",
                "updated_at": "updated_at",
            },
            default=("-id",),
        )

    def create(self, request, *args, **kwargs):
        try:
            self._require_model_perm("add")
            payload = request.data.copy()
            items_payload = payload.pop('items', payload.pop('invoice_items', []))

            if not isinstance(items_payload, list) or len(items_payload) == 0:
                return Response({'error': 'items must be a non-empty list'}, status=status.HTTP_400_BAD_REQUEST)

            requested_status = payload.get("status") or "Draft"

            with transaction.atomic():
                customer_id = payload.get("customer")
                if customer_id:
                    if not Customer.objects.filter(pk=customer_id, is_deleted=False).exists():
                        raise NotFound("Customer not found")
                invoice_serializer = self.get_serializer(data={**payload, "status": requested_status})
                invoice_serializer.is_valid(raise_exception=True)
                invoice = invoice_serializer.save(
                    subtotal=Decimal("0.00"),
                    tax_rate=Decimal("0.00"),
                    tax_total=Decimal("0.00"),
                    total_amount=Decimal("0.00"),
                )

                item_ids = []
                normalized_items = []
                for raw in items_payload:
                    if not isinstance(raw, dict):
                        raise ValidationError("Each item must be an object")
                    item_id = raw.get("item") or raw.get("item_id") or raw.get("itemId")
                    qty = raw.get("quantity")
                    override_tax_rate = raw.get("tax_rate")
                    override_unit_price = raw.get("unit_price")
                    try:
                        item_id = int(item_id)
                    except (TypeError, ValueError):
                        raise ValidationError("Each item must include a valid item id")
                    try:
                        qty = int(qty)
                    except (TypeError, ValueError):
                        raise ValidationError("Each item must include a valid quantity")
                    if qty < 1:
                        raise ValidationError("quantity must be >= 1")
                    if override_unit_price in (None, ""):
                        normalized_unit_price = None
                    else:
                        try:
                            normalized_unit_price = Decimal(str(override_unit_price))
                        except (InvalidOperation, TypeError):
                            raise ValidationError({"unit_price": "Each item must include a valid unit_price"})
                        if not normalized_unit_price.is_finite():
                            raise ValidationError({"unit_price": "Each item must include a valid unit_price"})
                        if normalized_unit_price < 0:
                            raise ValidationError({"unit_price": "unit_price must be >= 0"})
                    item_ids.append(item_id)
                    normalized_items.append((item_id, qty, override_tax_rate, normalized_unit_price))

                items_by_id = {i.id: i for i in Item.objects.select_for_update().filter(id__in=item_ids, is_deleted=False)}
                if len(items_by_id) != len(set(item_ids)):
                    raise ValidationError("One or more items do not exist")

                subtotal = Decimal("0.00")
                tax_total = Decimal("0.00")
                created_lines = []
                for item_id, qty, override_tax_rate, override_unit_price in normalized_items:
                    db_item = items_by_id[item_id]
                    unit_price = _q2(override_unit_price if override_unit_price is not None else db_item.unit_price)
                    line_subtotal = _q2(unit_price * qty)

                    if override_tax_rate is None or override_tax_rate == "":
                        line_tax_rate = db_item.tax_rate
                    else:
                        try:
                            line_tax_rate = Decimal(str(override_tax_rate))
                        except (InvalidOperation, TypeError):
                            raise ValidationError({"tax_rate": "Invalid tax_rate"})
                    if line_tax_rate < 0 or line_tax_rate > 100:
                        raise ValidationError({"tax_rate": "tax_rate must be between 0 and 100"})

                    line_tax = _q2((line_subtotal * line_tax_rate) / Decimal("100"))
                    line_total = _q2(line_subtotal + line_tax)

                    subtotal += line_subtotal
                    tax_total += line_tax
                    created_lines.append(
                        InvoiceItem(
                            invoice=invoice,
                            item=db_item,
                            description=db_item.description,
                            unit_of_measure=db_item.unit_of_measure,
                            quantity=qty,
                            unit_price=unit_price,
                            tax_rate=line_tax_rate,
                            line_subtotal=line_subtotal,
                            line_tax=line_tax,
                            line_total=line_total,
                        )
                    )

                InvoiceItem.objects.bulk_create(created_lines)

                subtotal = _q2(subtotal)
                tax_total = _q2(tax_total)
                discount_type, discount_value, discount_amount = _resolve_invoice_discount(
                    subtotal, invoice.discount_type, invoice.discount_value
                )
                total_amount = _q2(subtotal - discount_amount + tax_total)
                computed_tax_rate = _q2((tax_total / subtotal) * Decimal("100")) if subtotal > 0 else Decimal("0.00")
                Invoice.objects.filter(pk=invoice.pk).update(
                    subtotal=subtotal,
                    discount_type=discount_type,
                    discount_value=discount_value,
                    discount_amount=discount_amount,
                    tax_rate=computed_tax_rate,
                    tax_total=tax_total,
                    total_amount=total_amount,
                )
                invoice.refresh_from_db()

                if invoice.status in ["Sent", "Paid"]:
                    self._deduct_inventory_for_invoice(invoice)

            logger.info("invoice.create invoice_id=%s invoice_number=%s status=%s", invoice.id, invoice.invoice_number, invoice.status)
            return Response(InvoiceSerializer(invoice).data, status=status.HTTP_201_CREATED)
        except Exception as exc:
            _raise_record_save_failure(request=request, entity="invoice", operation="create", exc=exc)

    def update(self, request, *args, **kwargs):
        try:
            self._require_model_perm("change")
            partial = kwargs.pop("partial", False)
            with transaction.atomic():
                pk = kwargs.get("pk")
                try:
                    invoice = Invoice.objects.select_for_update().get(pk=pk, is_deleted=False)
                except Invoice.DoesNotExist:
                    raise NotFound("Invoice not found")
                _check_concurrency(request, invoice)
                old_status = invoice.status
                serializer = self.get_serializer(invoice, data=request.data, partial=partial)
                serializer.is_valid(raise_exception=True)
                before = {k: getattr(invoice, k, None) for k in serializer.validated_data.keys()}
                updated = serializer.save()
                discount_type, discount_value, discount_amount = _resolve_invoice_discount(
                    Decimal(updated.subtotal), updated.discount_type, updated.discount_value
                )
                recomputed_total = _q2(Decimal(updated.subtotal) - discount_amount + Decimal(updated.tax_total))
                updated_fields: list[str] = []
                if updated.discount_type != discount_type:
                    updated.discount_type = discount_type
                    updated_fields.append("discount_type")
                if updated.discount_value != discount_value:
                    updated.discount_value = discount_value
                    updated_fields.append("discount_value")
                if updated.discount_amount != discount_amount:
                    updated.discount_amount = discount_amount
                    updated_fields.append("discount_amount")
                if updated.total_amount != recomputed_total:
                    updated.total_amount = recomputed_total
                    updated_fields.append("total_amount")
                if updated_fields:
                    updated.save(update_fields=updated_fields + ["updated_at"])
                updated.refresh_from_db()
                after_keys = set(serializer.validated_data.keys())
                if {"discount_type", "discount_value"} & set(serializer.validated_data.keys()):
                    after_keys.update({"discount_amount", "total_amount"})
                    before.setdefault("discount_amount", getattr(invoice, "discount_amount", None))
                    before.setdefault("total_amount", getattr(invoice, "total_amount", None))
                after = {k: getattr(updated, k, None) for k in after_keys}
                changes = {k: {"from": str(before[k]) if before[k] is not None else None, "to": str(after[k]) if after[k] is not None else None} for k in before.keys() if before[k] != after[k]}
                if changes:
                    _log_audit(request.user, "update", updated, changes)

                if old_status not in ["Sent", "Paid"] and updated.status in ["Sent", "Paid"]:
                    self._deduct_inventory_for_invoice(updated)

            logger.info("invoice.update invoice_id=%s status_from=%s status_to=%s", invoice.id, old_status, updated.status)
            return Response(InvoiceSerializer(updated).data, status=status.HTTP_200_OK)
        except Exception as exc:
            _raise_record_save_failure(request=request, entity="invoice", operation="update", exc=exc)

    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    def _deduct_inventory_for_invoice(self, invoice: Invoice) -> None:
        if invoice.inventory_deducted_at is not None:
            return

        invoice_items = list(invoice.invoice_items.select_related("item").filter(is_deleted=False))
        item_ids = [li.item_id for li in invoice_items]
        items = {i.id: i for i in Item.objects.select_for_update().filter(id__in=item_ids, is_deleted=False)}

        for li in invoice_items:
            db_item = items[li.item_id]
            if db_item.type == "service":
                continue
            if db_item.stock_quantity < li.quantity:
                raise ValidationError(f"Insufficient stock for item {db_item.id}")

        for li in invoice_items:
            db_item = items[li.item_id]
            if db_item.type == "service":
                continue
            Item.objects.filter(pk=li.item_id).update(stock_quantity=F("stock_quantity") - li.quantity)

        Invoice.objects.filter(pk=invoice.pk, inventory_deducted_at__isnull=True).update(
            inventory_deducted_at=timezone.now()
        )
        invoice.refresh_from_db()
        logger.info("inventory.deduct invoice_id=%s deducted_at=%s", invoice.id, invoice.inventory_deducted_at)

    def destroy(self, request, *args, **kwargs):
        self._require_model_perm("delete")
        with transaction.atomic():
            invoice = Invoice.objects.select_for_update().get(pk=self.get_object().pk, is_deleted=False)
            _check_concurrency(request, invoice)
            now = timezone.now()
            InvoiceItem.objects.filter(invoice=invoice, is_deleted=False).update(is_deleted=True, deleted_at=now)
            Receipt.objects.filter(invoice=invoice, is_deleted=False).update(is_deleted=True, deleted_at=now)
            invoice.is_deleted = True
            invoice.deleted_at = now
            invoice.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
            _log_audit(request.user, "delete", invoice, {"cascade": ["invoice_items", "receipts"]})
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        self._require_model_perm("delete")
        ids = request.data.get("ids", [])
        if not isinstance(ids, list) or not ids:
            raise ValidationError({"ids": "ids must be a non-empty list"})
        now = timezone.now()
        with transaction.atomic():
            invoices = list(Invoice.objects.select_for_update().filter(id__in=ids, is_deleted=False))
            if len(invoices) != len(set(ids)):
                raise NotFound("One or more invoices not found")
            for inv in invoices:
                InvoiceItem.objects.filter(invoice=inv, is_deleted=False).update(is_deleted=True, deleted_at=now)
                Receipt.objects.filter(invoice=inv, is_deleted=False).update(is_deleted=True, deleted_at=now)
                inv.is_deleted = True
                inv.deleted_at = now
                inv.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
                _log_audit(request.user, "bulk_delete", inv, {"bulk": True, "cascade": ["invoice_items", "receipts"]})
        return Response({"deleted": len(ids)}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        invoice = self.get_object()
        from .documents import render_invoice

        rendered = render_invoice(request, invoice, "pdf")
        response = HttpResponse(rendered.content, content_type=rendered.content_type)
        response["Content-Disposition"] = f'attachment; filename="{rendered.filename}"'
        response["X-PDF-Backend"] = rendered.backend or ("weasyprint" if "application/pdf" in rendered.content_type else "failed")
        return response

    @action(detail=True, methods=["get"])
    def print_html(self, request, pk=None):
        invoice = self.get_object()
        from .documents import render_invoice

        rendered = render_invoice(request, invoice, "html")
        return HttpResponse(rendered.content, content_type=rendered.content_type)

    @action(detail=True, methods=["post"])
    def pay(self, request, pk=None):
        try:
            user = request.user
            if not getattr(user, "is_authenticated", False):
                raise PermissionDenied()

            invoice_id = pk
            idempotency_key = request.META.get("HTTP_IDEMPOTENCY_KEY") or request.headers.get("Idempotency-Key")
            if not idempotency_key or not isinstance(idempotency_key, str):
                raise ValidationError({"detail": "Idempotency-Key header is required"})

            cache_key = f"invoice.pay:{user.id}:{invoice_id}:{hashlib.sha256(idempotency_key.encode('utf-8')).hexdigest()}"
            cached_receipt_id = cache.get(cache_key)
            if cached_receipt_id:
                receipt = Receipt.objects.filter(pk=cached_receipt_id, is_deleted=False).select_related("invoice").first()
                if receipt:
                    logger.info("payment.idempotent_replay user_id=%s invoice_id=%s receipt_id=%s", user.id, invoice_id, receipt.id)
                    return Response(ReceiptSerializer(receipt).data, status=status.HTTP_200_OK)

            raw_amount = request.data.get("amount_paid")
            raw_method = request.data.get("payment_method")
            raw_reference = request.data.get("reference_number")
            raw_payment_date = request.data.get("payment_date")

            try:
                amount_paid = Decimal(str(raw_amount)).quantize(CENTS, rounding=ROUND_HALF_UP)
            except (InvalidOperation, TypeError):
                raise ValidationError({"amount_paid": "amount_paid must be a valid number"})
            if amount_paid <= 0:
                raise ValidationError({"amount_paid": "amount_paid must be > 0"})

            if raw_method not in dict(Receipt.PAYMENT_METHOD_CHOICES):
                raise ValidationError({"payment_method": "Invalid payment_method"})
            payment_method = str(raw_method)

            reference_number = None
            if raw_reference is not None and str(raw_reference).strip():
                reference_number = str(raw_reference).strip()

            if payment_method in ("Card", "Bank Transfer") and not reference_number:
                raise ValidationError({"reference_number": "reference_number is required for Card and Bank Transfer payments"})

            if reference_number:
                ref_upper = reference_number.upper()
                digits_only = "".join(ch for ch in reference_number if ch.isdigit())
                if payment_method == "Card" and len(digits_only) >= 12:
                    raise ValidationError({"reference_number": "Do not store card numbers. Use an authorization/reference code instead."})
                if "INVALID" in ref_upper:
                    raise PaymentDeclined("Invalid card number")
                if "EXPIRED" in ref_upper:
                    raise PaymentDeclined("Card has expired")
                if "INSUFFICIENT" in ref_upper:
                    raise PaymentDeclined("Insufficient funds")
                if "DECLINE" in ref_upper:
                    logger.warning("payment.declined user_id=%s invoice_id=%s", user.id, invoice_id)
                    raise PaymentDeclined("Card was declined")
                if "NETWORK" in ref_upper:
                    logger.warning("payment.gateway_unavailable user_id=%s invoice_id=%s", user.id, invoice_id)
                    raise PaymentGatewayUnavailable("Payment gateway unavailable")
                if "TIMEOUT" in ref_upper:
                    logger.warning("payment.gateway_timeout user_id=%s invoice_id=%s", user.id, invoice_id)
                    time.sleep(2)
                    raise PaymentGatewayTimeout("Payment gateway timeout")

            payment_date = None
            if raw_payment_date not in (None, ""):
                try:
                    payment_date = date.fromisoformat(str(raw_payment_date))
                except ValueError:
                    raise ValidationError({"payment_date": "Invalid date. Use YYYY-MM-DD"})

            with transaction.atomic():
                try:
                    invoice = Invoice.objects.select_for_update().get(pk=invoice_id, is_deleted=False)
                except Invoice.DoesNotExist:
                    raise NotFound("Invoice not found")

                total_paid = invoice.receipts.filter(is_deleted=False).aggregate(total=Sum("amount_paid"))["total"] or Decimal("0.00")
                outstanding = outstanding_invoice_amount(invoice.total_amount, total_paid)
                if outstanding <= 0:
                    raise ValidationError({"detail": "Invoice is already fully paid"})
                if amount_paid > outstanding:
                    raise ValidationError({"amount_paid": f"amount_paid cannot exceed outstanding amount ({outstanding})"})

                receipt = Receipt.objects.create(
                    invoice=invoice,
                    amount_paid=amount_paid,
                    payment_date=payment_date if payment_date else timezone.localdate(),
                    payment_method=payment_method,
                    reference_number=reference_number,
                )

                total_paid_after = (total_paid + amount_paid).quantize(CENTS, rounding=ROUND_HALF_UP)
                if sync_invoice_status_with_payments(invoice, total_paid_after) == "Paid":
                    logger.info("invoice.paid invoice_id=%s total_paid=%s", invoice.id, total_paid_after)

            cache.set(cache_key, receipt.id, timeout=15 * 60)
            logger.info(
                "payment.success user_id=%s invoice_id=%s receipt_id=%s amount_paid=%s method=%s",
                user.id,
                invoice_id,
                receipt.id,
                str(amount_paid),
                payment_method,
            )
            return Response(ReceiptSerializer(receipt).data, status=status.HTTP_201_CREATED)
        except Exception as exc:
            _raise_record_save_failure(request=request, entity="receipt", operation="create", exc=exc)

    @action(detail=False, methods=["post"])
    def bulk_update_status(self, request):
        self._require_model_perm("change")
        ids = request.data.get("ids", [])
        new_status = str(request.data.get("status") or "").strip()
        if not isinstance(ids, list) or not ids:
            raise ValidationError({"ids": "ids must be a non-empty list"})
        allowed = {s for s, _ in Invoice.STATUS_CHOICES}
        if new_status not in allowed:
            raise ValidationError({"status": "Invalid status"})

        with transaction.atomic():
            invoices = list(Invoice.objects.select_for_update().filter(id__in=ids, is_deleted=False).prefetch_related("invoice_items", "invoice_items__item"))
            if len(invoices) != len(set(ids)):
                raise NotFound("One or more invoices not found")
            for inv in invoices:
                old_status = inv.status
                inv.status = new_status
                inv.save(update_fields=["status", "updated_at"])
                _log_audit(request.user, "update", inv, {"status": {"from": old_status, "to": new_status}, "bulk": True})
                if old_status not in ["Sent", "Paid"] and new_status in ["Sent", "Paid"]:
                    self._deduct_inventory_for_invoice(inv)
        return Response({"updated": len(ids)}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        if not getattr(request.user, "is_authenticated", False):
            raise PermissionDenied()

        fmt = str(request.query_params.get("file_format") or request.query_params.get("export_format") or "csv").strip().lower()
        if fmt not in ("csv", "xlsx", "pdf"):
            raise ValidationError({"file_format": "Invalid file_format. Use csv, xlsx, or pdf"})

        allowed_fields = [
            "id",
            "invoice_number",
            "customer_name",
            "customer_email",
            "status",
            "issue_date",
            "due_date",
            "subtotal",
            "tax_total",
            "total_amount",
            "updated_at",
        ]
        fields_raw = request.query_params.get("fields")
        if fields_raw in (None, ""):
            fields = ["invoice_number", "customer_name", "status", "issue_date", "due_date", "subtotal", "tax_total", "total_amount"]
        else:
            parts = [p.strip() for p in str(fields_raw).split(",") if p.strip()]
            invalid = [p for p in parts if p not in allowed_fields]
            if invalid:
                raise ValidationError({"fields": f"Invalid fields: {', '.join(invalid)}"})
            fields = parts or ["invoice_number", "customer_name", "status", "issue_date", "due_date", "subtotal", "tax_total", "total_amount"]

        rows_limit = 50000 if fmt in ("csv", "xlsx") else 2000
        try:
            limit_raw = request.query_params.get("limit")
            if limit_raw not in (None, ""):
                rows_limit = int(limit_raw)
        except ValueError:
            raise ValidationError({"limit": "Invalid limit"})
        rows_limit = max(1, min(rows_limit, 50000 if fmt in ("csv", "xlsx") else 5000))

        qs = self.filter_queryset(self.get_queryset()).select_related("customer").order_by("-id")[:rows_limit]

        def _csv_cell(value: str) -> str:
            v = str(value or "")
            if v and v[0] in ("=", "+", "-", "@"):
                return "'" + v
            return v

        def _value_for(inv: Invoice, field: str) -> str:
            if field == "customer_name":
                return str(getattr(inv.customer, "name", "") or "")
            if field == "customer_email":
                return str(getattr(inv.customer, "email", "") or "")
            if field in ("issue_date", "due_date"):
                d = getattr(inv, field, None)
                return str(d) if d else ""
            if field in ("subtotal", "tax_total", "total_amount"):
                return str(getattr(inv, field, "") or "")
            if field == "updated_at":
                dt = getattr(inv, "updated_at", None)
                return dt.isoformat() if dt else ""
            return str(getattr(inv, field, "") or "")

        filename_base = "invoices"
        _log_operation(request.user, "export", Invoice, "invoices_export", {"format": fmt, "fields": fields, "limit": rows_limit})

        if fmt == "csv":
            class _Echo:
                def write(self, value):
                    return value

            def _iter_rows():
                yield writer.writerow(fields)
                for inv in qs.iterator(chunk_size=2000):
                    yield writer.writerow([_csv_cell(_value_for(inv, f)) for f in fields])

            pseudo_buffer = _Echo()
            writer = csv.writer(pseudo_buffer)
            resp = StreamingHttpResponse(_iter_rows(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
            return resp

        if fmt == "xlsx":
            wb = Workbook(write_only=True)
            ws = wb.create_sheet("Invoices")
            ws.append(fields)
            for inv in qs.iterator(chunk_size=2000):
                ws.append([_value_for(inv, f) for f in fields])
            out = io.BytesIO()
            wb.save(out)
            resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
            return resp

        rows = []
        for inv in qs.iterator(chunk_size=2000):
            row = {f: _value_for(inv, f) for f in fields}
            rows.append(row)
        company_name = str(_effective_company_identity_for_user(request.user)["company_name"] or "PIXELHUB").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        html = "<html><head><meta charset='utf-8' /><style>body{font-family:Arial, sans-serif;font-size:10pt;}table{border-collapse:collapse;width:100%;}th,td{border:1px solid #ddd;padding:4px;}th{background:#f3f3f3;text-align:left;}.doc-company{font-size:18pt;font-weight:700;margin:0 0 4px;color:#1a4d8e;}.doc-title{font-size:12pt;font-weight:700;margin:0 0 12px;}</style></head><body>"
        html += f"<div class='doc-company'>{company_name}</div><div class='doc-title'>Invoices Export</div>"
        html += "<table><thead><tr>" + "".join([f"<th>{f}</th>" for f in fields]) + "</tr></thead><tbody>"
        for r in rows:
            html += "<tr>" + "".join([f"<td>{str(r.get(f, '')).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')}</td>" for f in fields]) + "</tr>"
        html += "</tbody></table></body></html>"
        try:
            from weasyprint import HTML

            pdf = HTML(string=html).write_pdf()
            resp = HttpResponse(pdf, content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.pdf"'
            resp["X-PDF-Backend"] = "weasyprint"
            return resp
        except (OSError, ValueError) as e:
            logger.warning("pdf.weasyprint_render_failed error=%s invoices_export=1", e)
            resp = HttpResponse(html, content_type="text/html; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.html"'
            resp["X-PDF-Backend"] = "failed"
            return resp

    @action(detail=False, methods=["get"], url_path="import_template")
    def import_template(self, request):
        self._require_any_model_perm(["add", "change"])

        fmt = str(request.query_params.get("file_format") or "xlsx").strip().lower()
        if fmt not in ("csv", "xlsx"):
            raise ValidationError({"file_format": "Invalid file_format. Use csv or xlsx"})

        header = [
            "invoice_key",
            "invoice_number",
            "customer_email",
            "customer_name",
            "status",
            "issue_date",
            "due_date",
            "item_sku",
            "quantity",
            "unit_price",
            "tax_rate",
            "description",
            "unit_of_measure",
        ]
        example = [
            "BATCH-1",
            "",
            "buyer@example.com",
            "",
            "Draft",
            "2026-05-01",
            "2026-05-10",
            "SKU-001",
            "2",
            "",
            "",
            "Optional line description",
            "pcs",
        ]
        filename_base = "invoice_import_template"

        if fmt == "csv":
            out = io.StringIO()
            w = csv.writer(out)
            w.writerow(header)
            w.writerow(example)
            resp = HttpResponse(out.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
            return resp

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Template")
        ws.append(header)
        ws.append(example)
        out = io.BytesIO()
        wb.save(out)
        resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    @action(detail=False, methods=["post"], url_path="import")
    def import_invoices(self, request):
        self._require_any_model_perm(["add", "change"])

        upload = request.FILES.get("file")
        if upload is None:
            raise ValidationError({"file": "file is required"})

        dry_run = str(request.data.get("dry_run") or "").strip().lower() in ("1", "true", "yes", "on")
        rollback_on_error = str(request.data.get("rollback_on_error") or "").strip().lower() not in ("0", "false", "no", "off")
        from .import_export import import_invoices_from_upload

        try:
            status_code, payload = import_invoices_from_upload(
                upload,
                dry_run=dry_run,
                rollback_on_error=rollback_on_error,
                deduct_inventory_for_invoice=self._deduct_inventory_for_invoice,
            )
        except ValueError:
            raise ValidationError({"file": "Unsupported file type. Use .csv or .xlsx"})

        if status_code >= 400:
            _log_operation(
                request.user,
                "import",
                Invoice,
                "invoices_import_failed",
                {"dry_run": dry_run, "rows": int(payload.get("rows") or 0), "errors": len(payload.get("errors") or [])},
            )
            return Response(payload, status=status_code)

        if payload.get("dry_run"):
            _log_operation(
                request.user,
                "import",
                Invoice,
                "invoices_import_dry_run",
                {
                    "rows": int(payload.get("rows") or 0),
                    "would_create_invoices": int(payload.get("would_create_invoices") or 0),
                    "would_create_invoice_items": int(payload.get("would_create_invoice_items") or 0),
                    "errors": len(payload.get("errors") or []),
                },
            )
            return Response(payload, status=status_code)

        _log_operation(
            request.user,
            "import",
            Invoice,
            "invoices_import",
            {
                "rows": int(payload.get("rows") or 0),
                "created_invoices": int(payload.get("imported_invoices") or 0),
                "created_invoice_items": int(payload.get("imported_invoice_items") or 0),
                "errors": len(payload.get("errors") or []),
            },
        )
        return Response(payload, status=status_code)

    @action(detail=True, methods=["post"], url_path="share_link")
    def share_link(self, request, pk=None):
        if not getattr(request.user, "is_authenticated", False):
            raise PermissionDenied()
        invoice = self.get_object()
        ttl_minutes = request.data.get("ttl_minutes")
        try:
            ttl_minutes_int = int(ttl_minutes) if ttl_minutes not in (None, "") else (60 * 24 * 7)
        except (TypeError, ValueError):
            raise ValidationError({"ttl_minutes": "ttl_minutes must be an integer"})
        ttl_minutes_int = max(5, min(ttl_minutes_int, 60 * 24 * 30))

        from .documents import backend_public_base_url, create_delivery

        delivery, token = create_delivery(
            user=request.user,
            document_type="invoice",
            document_id=invoice.id,
            channel="share",
            fmt="pdf",
            ttl_minutes=ttl_minutes_int,
        )
        download_url = f"{backend_public_base_url()}/api/documents/deliveries/{delivery.id}/download/?token={token}"
        _log_audit(request.user, "create", delivery, {"channel": "share", "document_type": "invoice", "invoice_id": invoice.id})
        return Response(
            {
                "delivery_id": delivery.id,
                "download_url": download_url,
                "expires_at": delivery.download_expires_at.isoformat() if delivery.download_expires_at else None,
            },
            status=status.HTTP_200_OK,
        )


class InvoiceItemViewSet(SoftDeleteModelViewSet):
    queryset = InvoiceItem.objects.all().select_related("invoice", "item").order_by("-id")
    serializer_class = InvoiceItemSerializer
    pagination_class = OptionalPageNumberPagination


class ImportErrorLogDownloadApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, token: uuid.UUID):
        payload = cache.get(f"import_error_log:{token}")
        if not payload:
            raise NotFound("Error log not found or expired")
        resp = HttpResponse(payload, content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="import_errors_{token}.csv"'
        return resp


class InvoiceSavedViewsApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        qs = SavedInvoiceView.objects.filter(user=request.user).order_by("-is_default", "-created_at", "-id")
        rows = []
        for r in qs:
            rows.append({"id": r.id, "name": r.name, "filters": r.filters or {}, "is_default": bool(r.is_default)})
        return Response({"results": rows}, status=status.HTTP_200_OK)

    def post(self, request):
        name = str(request.data.get("name") or "").strip()
        filters = request.data.get("filters")
        is_default = bool(request.data.get("is_default"))
        if not name:
            raise ValidationError({"name": "Name is required"})
        if len(name) > 80:
            raise ValidationError({"name": "Name must be 80 characters or less"})
        if filters is None:
            filters = {}
        if not isinstance(filters, dict):
            raise ValidationError({"filters": "filters must be an object"})
        with transaction.atomic():
            if is_default:
                SavedInvoiceView.objects.filter(user=request.user, is_default=True).update(is_default=False)
            row = SavedInvoiceView.objects.create(user=request.user, name=name, filters=filters, is_default=is_default)
        return Response({"id": row.id, "created": True}, status=status.HTTP_201_CREATED)


class InvoiceSavedViewDetailApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, view_id: int):
        name = request.data.get("name")
        filters = request.data.get("filters")
        is_default = request.data.get("is_default")

        with transaction.atomic():
            row = SavedInvoiceView.objects.select_for_update().filter(id=view_id, user=request.user).first()
            if row is None:
                raise NotFound("Saved view not found")

            if name is not None:
                name_clean = str(name).strip()
                if not name_clean:
                    raise ValidationError({"name": "Name is required"})
                if len(name_clean) > 80:
                    raise ValidationError({"name": "Name must be 80 characters or less"})
                row.name = name_clean

            if filters is not None:
                if not isinstance(filters, dict):
                    raise ValidationError({"filters": "filters must be an object"})
                row.filters = filters

            if is_default is not None:
                is_default_bool = bool(is_default)
                if is_default_bool:
                    SavedInvoiceView.objects.filter(user=request.user, is_default=True).exclude(id=row.id).update(is_default=False)
                row.is_default = is_default_bool

            row.save()

        return Response({"updated": True}, status=status.HTTP_200_OK)

    def delete(self, request, view_id: int):
        deleted, _ = SavedInvoiceView.objects.filter(id=view_id, user=request.user).delete()
        if not deleted:
            raise NotFound("Saved view not found")
        return Response(status=status.HTTP_204_NO_CONTENT)


class ReceiptViewSet(SoftDeleteModelViewSet):
    queryset = Receipt.objects.all().select_related("invoice").order_by("-id")
    serializer_class = ReceiptSerializer
    pagination_class = OptionalPageNumberPagination

    def get_serializer_class(self):
        if getattr(self, "action", None) == "retrieve":
            return ReceiptDetailSerializer
        return ReceiptSerializer

    def get_queryset(self):
        qs = super().get_queryset().select_related("invoice", "invoice__customer")
        p = self.request.query_params

        q = str(p.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(reference_number__icontains=q)
                | Q(invoice__invoice_number__icontains=q)
                | Q(invoice__customer__name__icontains=q)
            )

        invoice_number = str(p.get("invoice_number") or "").strip()
        if invoice_number:
            qs = qs.filter(invoice__invoice_number__icontains=invoice_number)

        customer_name = str(p.get("customer_name") or "").strip()
        if customer_name:
            qs = qs.filter(invoice__customer__name__icontains=customer_name)

        payment_method = str(p.get("payment_method") or "").strip()
        if payment_method:
            allowed = {method for method, _ in Receipt.PAYMENT_METHOD_CHOICES}
            if payment_method not in allowed:
                raise ValidationError({"payment_method": "Invalid payment_method"})
            qs = qs.filter(payment_method=payment_method)

        def _date_param(key: str):
            raw = p.get(key)
            if raw in (None, ""):
                return None
            try:
                return date.fromisoformat(str(raw))
            except ValueError:
                raise ValidationError({key: "Invalid date. Use YYYY-MM-DD"})

        payment_from = _date_param("payment_date_from")
        payment_to = _date_param("payment_date_to")
        if payment_from:
            qs = qs.filter(payment_date__gte=payment_from)
        if payment_to:
            qs = qs.filter(payment_date__lte=payment_to)

        updated_from = _date_param("updated_from")
        updated_to = _date_param("updated_to")
        if updated_from:
            qs = qs.filter(updated_at__date__gte=updated_from)
        if updated_to:
            qs = qs.filter(updated_at__date__lte=updated_to)

        def _dec_param(key: str):
            raw = p.get(key)
            if raw in (None, ""):
                return None
            try:
                value = Decimal(str(raw))
            except (InvalidOperation, TypeError):
                raise ValidationError({key: "Invalid amount"})
            if value < 0:
                raise ValidationError({key: "Amount must be >= 0"})
            return value

        amount_min = _dec_param("amount_min")
        amount_max = _dec_param("amount_max")
        if amount_min is not None:
            qs = qs.filter(amount_paid__gte=amount_min)
        if amount_max is not None:
            qs = qs.filter(amount_paid__lte=amount_max)

        return _apply_ordering_query(
            qs,
            params=p,
            allowed_fields={
                "id": "id",
                "invoice_number": "invoice__invoice_number",
                "customer_name": "invoice__customer__name",
                "amount_paid": "amount_paid",
                "payment_date": "payment_date",
                "payment_method": "payment_method",
                "reference_number": "reference_number",
                "transaction_timestamp": "updated_at",
                "updated_at": "updated_at",
            },
            default=("-id",),
        )

    def create(self, request, *args, **kwargs):
        try:
            self._require_model_perm("add")
            user = request.user
            user_key = str(user.id) if getattr(user, "is_authenticated", False) else f"anon:{_client_ip(request)}"
            idempotency_key = request.META.get("HTTP_IDEMPOTENCY_KEY") or request.headers.get("Idempotency-Key")
            cache_key = None
            with transaction.atomic():
                invoice_id = request.data.get("invoice")
                try:
                    invoice_id = int(invoice_id)
                except (TypeError, ValueError):
                    return Response({'error': 'invoice is required'}, status=status.HTTP_400_BAD_REQUEST)

                if idempotency_key and isinstance(idempotency_key, str):
                    cache_key = f"receipt.create:{user_key}:{invoice_id}:{hashlib.sha256(idempotency_key.encode('utf-8')).hexdigest()}"
                    cached_receipt_id = cache.get(cache_key)
                    if cached_receipt_id:
                        receipt = Receipt.objects.filter(pk=cached_receipt_id, is_deleted=False).select_related("invoice").first()
                        if receipt:
                            logger.info("receipt.idempotent_replay user_id=%s invoice_id=%s receipt_id=%s", user_key, invoice_id, receipt.id)
                            return Response(ReceiptSerializer(receipt).data, status=status.HTTP_200_OK)

                try:
                    invoice = Invoice.objects.select_for_update().get(pk=invoice_id, is_deleted=False)
                except Invoice.DoesNotExist:
                    raise NotFound("Invoice not found")

                serializer = self.get_serializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                receipt = serializer.save()

                total_paid = invoice.receipts.filter(is_deleted=False).aggregate(total=Sum("amount_paid"))["total"] or Decimal("0.00")
                if sync_invoice_status_with_payments(invoice, total_paid) == "Paid":
                    logger.info("invoice.paid invoice_id=%s total_paid=%s", invoice.id, total_paid)

            if cache_key:
                cache.set(cache_key, receipt.id, timeout=15 * 60)
            if getattr(request.user, "is_authenticated", False):
                _log_audit(request.user, "create", receipt, {"invoice": str(invoice.id), "amount_paid": str(receipt.amount_paid), "payment_method": receipt.payment_method})
            logger.info("receipt.create receipt_id=%s invoice_id=%s amount_paid=%s", receipt.id, invoice.id, receipt.amount_paid)
            return Response(ReceiptSerializer(receipt).data, status=status.HTTP_201_CREATED)
        except Exception as exc:
            _raise_record_save_failure(request=request, entity="receipt", operation="create", exc=exc)

    @action(detail=True, methods=["post"], url_path="share_link")
    def share_link(self, request, pk=None):
        if not getattr(request.user, "is_authenticated", False):
            raise PermissionDenied()
        receipt = self.get_object()
        ttl_minutes = request.data.get("ttl_minutes")
        try:
            ttl_minutes_int = int(ttl_minutes) if ttl_minutes not in (None, "") else (60 * 24 * 7)
        except (TypeError, ValueError):
            raise ValidationError({"ttl_minutes": "ttl_minutes must be an integer"})
        ttl_minutes_int = max(5, min(ttl_minutes_int, 60 * 24 * 30))

        from .documents import backend_public_base_url, create_delivery

        delivery, token = create_delivery(
            user=request.user,
            document_type="receipt",
            document_id=receipt.id,
            channel="share",
            fmt="pdf",
            ttl_minutes=ttl_minutes_int,
        )
        download_url = f"{backend_public_base_url()}/api/documents/deliveries/{delivery.id}/download/?token={token}"
        _log_audit(request.user, "create", delivery, {"channel": "share", "document_type": "receipt", "receipt_id": receipt.id})
        return Response(
            {
                "delivery_id": delivery.id,
                "download_url": download_url,
                "expires_at": delivery.download_expires_at.isoformat() if delivery.download_expires_at else None,
            },
            status=status.HTTP_200_OK,
        )

    def update(self, request, *args, **kwargs):
        try:
            self._require_model_perm("change")
            partial = kwargs.pop("partial", False)
            with transaction.atomic():
                receipt = Receipt.objects.select_for_update().get(pk=self.get_object().pk, is_deleted=False)
                _check_concurrency(request, receipt)
                serializer = self.get_serializer(receipt, data=request.data, partial=partial)
                serializer.is_valid(raise_exception=True)
                before = {k: getattr(receipt, k, None) for k in serializer.validated_data.keys()}
                updated = serializer.save()
                after = {k: getattr(updated, k, None) for k in serializer.validated_data.keys()}
                changes = {k: {"from": str(before[k]) if before[k] is not None else None, "to": str(after[k]) if after[k] is not None else None} for k in before.keys() if before[k] != after[k]}
                if changes:
                    _log_audit(request.user, "update", updated, changes)

                try:
                    invoice = Invoice.objects.select_for_update().get(pk=updated.invoice_id, is_deleted=False)
                except Invoice.DoesNotExist:
                    invoice = None
                if invoice is not None:
                    total_paid = invoice.receipts.filter(is_deleted=False).aggregate(total=Sum("amount_paid"))["total"] or Decimal("0.00")
                    sync_invoice_status_with_payments(invoice, total_paid)

            return Response(self.get_serializer(updated).data, status=status.HTTP_200_OK)
        except Exception as exc:
            _raise_record_save_failure(request=request, entity="receipt", operation="update", exc=exc)

    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        self._require_model_perm("delete")
        with transaction.atomic():
            receipt = Receipt.objects.select_for_update().select_related("invoice").get(pk=self.get_object().pk, is_deleted=False)
            _check_concurrency(request, receipt)
            receipt.is_deleted = True
            receipt.deleted_at = timezone.now()
            receipt.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
            _log_audit(request.user, "delete", receipt, {})

            invoice = Invoice.objects.select_for_update().get(pk=receipt.invoice_id, is_deleted=False)
            total_paid = invoice.receipts.filter(is_deleted=False).aggregate(total=Sum("amount_paid"))["total"] or Decimal("0.00")
            sync_invoice_status_with_payments(invoice, total_paid)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        self._require_model_perm("delete")
        ids = request.data.get("ids", [])
        if not isinstance(ids, list) or not ids:
            raise ValidationError({"ids": "ids must be a non-empty list"})
        now = timezone.now()
        with transaction.atomic():
            receipts = list(Receipt.objects.select_for_update().filter(id__in=ids, is_deleted=False))
            if len(receipts) != len(set(ids)):
                raise NotFound("One or more receipts not found")
            invoice_ids = {r.invoice_id for r in receipts}
            for r in receipts:
                r.is_deleted = True
                r.deleted_at = now
                r.save(update_fields=["is_deleted", "deleted_at", "updated_at"])
                _log_audit(request.user, "bulk_delete", r, {"bulk": True})
            for invoice_id in invoice_ids:
                try:
                    inv = Invoice.objects.select_for_update().get(pk=invoice_id, is_deleted=False)
                except Invoice.DoesNotExist:
                    continue
                total_paid = inv.receipts.filter(is_deleted=False).aggregate(total=Sum("amount_paid"))["total"] or Decimal("0.00")
                sync_invoice_status_with_payments(inv, total_paid)
        return Response({"deleted": len(ids)}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"])
    def print_html(self, request, pk=None):
        receipt = self.get_object()
        from .documents import render_receipt

        rendered = render_receipt(request, receipt, "html")
        return HttpResponse(rendered.content, content_type=rendered.content_type)


class SourceAccountViewSet(SoftDeleteModelViewSet):
    queryset = SourceAccount.objects.all().select_related("currency").order_by("name", "id")
    serializer_class = SourceAccountSerializer
    pagination_class = None
    parser_classes = [parsers.JSONParser, parsers.FormParser]

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .select_related("currency")
            .annotate(active_expense_count=Count("expenses", filter=Q(expenses__is_deleted=False)))
        )
        params = self.request.query_params
        q = str(params.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(account_type__icontains=q) | Q(currency__code__icontains=q))
        status_value = str(params.get("status") or "").strip().lower()
        if status_value:
            qs = qs.filter(status=status_value)
        return _apply_ordering_query(
            qs,
            params=params,
            allowed_fields={
                "id": "id",
                "name": "name",
                "account_type": "account_type",
                "initial_balance": "initial_balance",
                "currency": "currency__code",
                "status": "status",
                "created_at": "created_at",
                "updated_at": "updated_at",
            },
            default=("name", "id"),
        )

    def destroy(self, request, *args, **kwargs):
        self._require_model_perm("delete")
        confirm_keyword = str(request.data.get("confirm_keyword") or request.query_params.get("confirm_keyword") or "").strip()
        if confirm_keyword != "DELETE":
            raise ValidationError({"confirm_keyword": "Type DELETE to confirm removal."})

        with transaction.atomic():
            instance = SourceAccount.objects.select_for_update().get(pk=self.get_object().pk, is_deleted=False)
            _check_concurrency(request, instance)
            dependency_count = Expense.objects.filter(source_account=instance, is_deleted=False).count()
            previous_status = instance.status
            instance.status = SourceAccount.STATUS_CLOSED
            instance.is_deleted = True
            instance.deleted_at = timezone.now()
            instance.save(update_fields=["status", "is_deleted", "deleted_at", "updated_at"])
            _log_audit(
                request.user,
                "delete",
                instance,
                {
                    "name": instance.name,
                    "status": {"from": previous_status, "to": instance.status},
                    "active_expense_count": dependency_count,
                    "dependency_handling": "historical expenses retained with deleted source account reference",
                },
            )
        return self._disable_response_cache(Response({"deleted": True, "active_expense_count": dependency_count}, status=status.HTTP_200_OK))


class ExpenseViewSet(SoftDeleteModelViewSet):
    queryset = Expense.objects.all().select_related("assigned_to", "created_by", "source_account", "source_account__currency").order_by("-expense_date", "-id")
    serializer_class = ExpenseSerializer
    pagination_class = OptionalPageNumberPagination
    parser_classes = [parsers.JSONParser, parsers.MultiPartParser, parsers.FormParser]

    def get_queryset(self):
        qs = super().get_queryset().select_related("assigned_to", "created_by", "source_account", "source_account__currency")
        p = self.request.query_params

        q = str(p.get("q") or "").strip()
        if q:
            db_match_ids = set(
                qs.filter(
                    Q(category__icontains=q)
                    | Q(vendor__icontains=q)
                    | Q(project_code__icontains=q)
                    | Q(cost_center__icontains=q)
                    | Q(assigned_to__username__icontains=q)
                ).values_list("id", flat=True)
            )
            lowered_q = q.lower()
            for expense_id, description_value, merchant_reference_value in qs.values_list(
                "id", "description", "merchant_reference"
            ).iterator(chunk_size=2000):
                description = decrypt_expense_text(description_value) or ""
                merchant_reference = decrypt_expense_text(merchant_reference_value) or ""
                if lowered_q in description.lower() or lowered_q in merchant_reference.lower():
                    db_match_ids.add(expense_id)
            qs = qs.filter(pk__in=db_match_ids) if db_match_ids else qs.none()

        category = str(p.get("category") or "").strip()
        if category:
            qs = qs.filter(category__icontains=category)

        source_account = str(p.get("source_account") or "").strip()
        if source_account:
            try:
                qs = qs.filter(source_account_id=int(source_account))
            except ValueError:
                qs = qs.filter(source_account__name__icontains=source_account)

        project_code = str(p.get("project_code") or "").strip()
        if project_code:
            qs = qs.filter(project_code__icontains=project_code)

        cost_center = str(p.get("cost_center") or "").strip()
        if cost_center:
            qs = qs.filter(cost_center__icontains=cost_center)

        assigned_to = str(p.get("assigned_to") or "").strip()
        if assigned_to:
            if assigned_to == "me" and getattr(self.request.user, "is_authenticated", False):
                qs = qs.filter(assigned_to=self.request.user)
            else:
                try:
                    qs = qs.filter(assigned_to_id=int(assigned_to))
                except ValueError:
                    qs = qs.filter(assigned_to__username__icontains=assigned_to)

        def _date_param(key: str):
            raw = p.get(key)
            if raw in (None, ""):
                return None
            try:
                return date.fromisoformat(str(raw))
            except ValueError:
                raise ValidationError({key: "Invalid date. Use YYYY-MM-DD"})

        expense_from = _date_param("expense_date_from")
        expense_to = _date_param("expense_date_to")
        if expense_from:
            qs = qs.filter(expense_date__gte=expense_from)
        if expense_to:
            qs = qs.filter(expense_date__lte=expense_to)

        def _decimal_param(key: str):
            raw = p.get(key)
            if raw in (None, ""):
                return None
            try:
                return Decimal(str(raw))
            except (TypeError, ValueError, InvalidOperation):
                raise ValidationError({key: "Invalid decimal"})

        amount_min = _decimal_param("amount_min")
        amount_max = _decimal_param("amount_max")
        if amount_min is not None:
            qs = qs.filter(amount__gte=amount_min)
        if amount_max is not None:
            qs = qs.filter(amount__lte=amount_max)

        return _apply_ordering_query(
            qs,
            params=p,
            allowed_fields={
                "id": "id",
                "amount": "amount",
                "expense_date": "expense_date",
                "category": "category",
                "vendor": "vendor",
                "source_account": "source_account__name",
                "assigned_to": "assigned_to__username",
                "project_code": "project_code",
                "cost_center": "cost_center",
                "created_at": "created_at",
                "updated_at": "updated_at",
            },
            default=("-expense_date", "-id"),
        )

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        self._require_model_perm("view")

        fmt = str(request.query_params.get("file_format") or request.query_params.get("export_format") or "csv").strip().lower()
        if fmt not in ("csv", "xlsx", "pdf"):
            raise ValidationError({"file_format": "Invalid file_format. Use csv, xlsx, or pdf"})

        allowed_fields = [
            "id",
            "amount",
            "expense_date",
            "category",
            "description",
            "vendor",
            "merchant_reference",
            "project_code",
            "cost_center",
            "source_account",
            "assigned_to",
            "created_by",
            "created_at",
            "updated_at",
        ]
        fields_raw = request.query_params.get("fields")
        if fields_raw in (None, ""):
            fields = [
                "expense_date",
                "amount",
                "category",
                "vendor",
                "project_code",
                "cost_center",
                "source_account",
                "assigned_to",
            ]
        else:
            parts = [p.strip() for p in str(fields_raw).split(",") if p.strip()]
            invalid = [p for p in parts if p not in allowed_fields]
            if invalid:
                raise ValidationError({"fields": f"Invalid fields: {', '.join(invalid)}"})
            fields = parts or allowed_fields

        rows_limit = 50000 if fmt in ("csv", "xlsx") else 2000
        try:
            limit_raw = request.query_params.get("limit")
            if limit_raw not in (None, ""):
                rows_limit = int(limit_raw)
        except ValueError:
            raise ValidationError({"limit": "Invalid limit"})
        rows_limit = max(1, min(rows_limit, 50000 if fmt in ("csv", "xlsx") else 5000))

        qs = self.filter_queryset(self.get_queryset())[:rows_limit]

        def _csv_cell(value: str) -> str:
            v = str(value or "")
            if v and v[0] in ("=", "+", "-", "@"):
                return "'" + v
            return v

        def _value_for(expense: Expense, field: str) -> str:
            if field == "assigned_to":
                return str(getattr(getattr(expense, "assigned_to", None), "username", "") or "")
            if field == "created_by":
                return str(getattr(getattr(expense, "created_by", None), "username", "") or "")
            if field == "source_account":
                return str(getattr(getattr(expense, "source_account", None), "name", "") or "")
            if field in ("description", "merchant_reference"):
                return str(decrypt_expense_text(getattr(expense, field, None)) or "")
            if field in ("expense_date",):
                return str(getattr(expense, field, "") or "")
            if field in ("amount",):
                return str(getattr(expense, field, "") or "")
            if field in ("created_at", "updated_at"):
                dt = getattr(expense, field, None)
                return dt.isoformat() if dt else ""
            return str(getattr(expense, field, "") or "")

        filename_base = "expenses"
        _log_operation(request.user, "export", Expense, "expenses_export", {"format": fmt, "fields": fields, "limit": rows_limit})

        if fmt == "csv":
            class _Echo:
                def write(self, value):
                    return value

            def _iter_rows():
                yield writer.writerow(fields)
                for expense in qs.iterator(chunk_size=2000):
                    yield writer.writerow([_csv_cell(_value_for(expense, f)) for f in fields])

            pseudo_buffer = _Echo()
            writer = csv.writer(pseudo_buffer)
            resp = StreamingHttpResponse(_iter_rows(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
            return resp

        if fmt == "xlsx":
            wb = Workbook(write_only=True)
            ws = wb.create_sheet("Expenses")
            ws.append(fields)
            for expense in qs.iterator(chunk_size=2000):
                ws.append([_value_for(expense, f) for f in fields])
            out = io.BytesIO()
            wb.save(out)
            resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
            return resp

        rows = []
        for expense in qs.iterator(chunk_size=2000):
            rows.append({f: _value_for(expense, f) for f in fields})
        company_name = str(_effective_company_identity_for_user(request.user)["company_name"] or "PIXELHUB").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        html = "<html><head><meta charset='utf-8' /><style>body{font-family:Arial, sans-serif;font-size:10pt;}table{border-collapse:collapse;width:100%;}th,td{border:1px solid #ddd;padding:4px;}th{background:#f3f3f3;text-align:left;}.doc-company{font-size:18pt;font-weight:700;margin:0 0 4px;color:#1a4d8e;}.doc-title{font-size:12pt;font-weight:700;margin:0 0 12px;}</style></head><body>"
        html += f"<div class='doc-company'>{company_name}</div><div class='doc-title'>Expenses Export</div>"
        html += "<table><thead><tr>" + "".join([f"<th>{f}</th>" for f in fields]) + "</tr></thead><tbody>"
        for row in rows:
            html += "<tr>" + "".join([f"<td>{str(row.get(f, '')).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')}</td>" for f in fields]) + "</tr>"
        html += "</tbody></table></body></html>"
        try:
            from weasyprint import HTML

            pdf = HTML(string=html).write_pdf()
            resp = HttpResponse(pdf, content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.pdf"'
            resp["X-PDF-Backend"] = "weasyprint"
            return resp
        except (OSError, ValueError) as e:
            logger.warning("pdf.weasyprint_render_failed error=%s expenses_export=1", e)
            resp = HttpResponse(html, content_type="text/html; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.html"'
            resp["X-PDF-Backend"] = "failed"
            return resp

    @action(detail=False, methods=["get"], url_path="import_template")
    def import_template(self, request):
        self._require_any_model_perm(["add", "change"])

        fmt = str(request.query_params.get("file_format") or "xlsx").strip().lower()
        if fmt not in ("csv", "xlsx"):
            raise ValidationError({"file_format": "Invalid file_format. Use csv or xlsx"})

        header = [
            "amount",
            "expense_date",
            "category",
            "description",
            "vendor",
            "merchant_reference",
            "project_code",
            "cost_center",
            "source_account",
            "assigned_to",
        ]
        example = [
            "250.00",
            str(timezone.localdate()),
            "Travel",
            "Client visit taxi fare",
            "Ride Service",
            "TRIP-1001",
            "PROJECT-ALPHA",
            "",
            "petty1",
            getattr(request.user, "username", ""),
        ]
        filename_base = "expense_import_template"

        if fmt == "csv":
            out = io.StringIO()
            w = csv.writer(out)
            w.writerow(header)
            w.writerow(example)
            resp = HttpResponse(out.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
            return resp

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Template")
        ws.append(header)
        ws.append(example)
        out = io.BytesIO()
        wb.save(out)
        resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    @action(detail=False, methods=["post"], url_path="import")
    def import_expenses(self, request):
        self._require_any_model_perm(["add", "change"])

        upload = request.FILES.get("file")
        if upload is None:
            raise ValidationError({"file": "file is required"})

        dry_run = str(request.data.get("dry_run") or "").strip().lower() in ("1", "true", "yes", "on")
        rollback_on_error = str(request.data.get("rollback_on_error") or "").strip().lower() not in ("0", "false", "no", "off")
        from .import_export import import_expenses_from_upload

        try:
            status_code, payload = import_expenses_from_upload(
                upload,
                dry_run=dry_run,
                rollback_on_error=rollback_on_error,
                actor=request.user,
            )
        except ValueError:
            raise ValidationError({"file": "Unsupported file type. Use .csv or .xlsx"})

        if status_code >= 400:
            _log_operation(
                request.user,
                "import",
                Expense,
                "expenses_import_failed",
                {
                    "dry_run": dry_run,
                    "rows": int(payload.get("rows") or 0),
                    "errors": len(payload.get("errors") or []),
                    "flags": len(payload.get("flags") or []),
                },
            )
            return Response(payload, status=status_code)

        if payload.get("dry_run"):
            _log_operation(
                request.user,
                "import",
                Expense,
                "expenses_import_dry_run",
                {
                    "rows": int(payload.get("rows") or 0),
                    "would_create": int(payload.get("would_create") or 0),
                    "errors": len(payload.get("errors") or []),
                    "flags": len(payload.get("flags") or []),
                },
            )
            return Response(payload, status=status_code)

        _log_operation(
            request.user,
            "import",
            Expense,
            "expenses_import",
            {
                "rows": int(payload.get("rows") or 0),
                "created": int(payload.get("imported") or 0),
                "errors": len(payload.get("errors") or []),
                "flags": len(payload.get("flags") or []),
            },
        )
        return Response(payload, status=status_code)


class CurrencyViewSet(viewsets.ModelViewSet):
    queryset = Currency.objects.all().order_by("code")
    serializer_class = CurrencySerializer
    permission_classes = [CurrencyPermission]

    def perform_create(self, serializer):
        obj = serializer.save()
        _log_audit(self.request.user, "update", obj, {"created": True})

    def perform_update(self, serializer):
        before = _serialize_settings_for_audit(serializer.instance)
        obj = serializer.save()
        after = _serialize_settings_for_audit(obj)
        _log_audit(self.request.user, "update", obj, {"before": before, "after": after})

    def perform_destroy(self, instance):
        _log_audit(self.request.user, "delete", instance, {})
        instance.delete()


class ExchangeRateViewSet(viewsets.ModelViewSet):
    queryset = ExchangeRate.objects.all().order_by("base_code", "quote_code")
    serializer_class = ExchangeRateSerializer
    permission_classes = [ExchangeRatePermission]

    def perform_create(self, serializer):
        obj = serializer.save()
        _log_audit(self.request.user, "update", obj, {"created": True})

    def perform_update(self, serializer):
        before = _serialize_settings_for_audit(serializer.instance)
        obj = serializer.save()
        after = _serialize_settings_for_audit(obj)
        _log_audit(self.request.user, "update", obj, {"before": before, "after": after})

    def perform_destroy(self, instance):
        _log_audit(self.request.user, "delete", instance, {})
        instance.delete()


class GlobalSettingsApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not user_has_permission(request.user, "settings.global.read"):
            raise PermissionDenied("You do not have permission to view admin settings.")
        gs = _get_global_settings()
        data = GlobalSettingsSerializer(gs).data
        if not user_has_permission(request.user, "settings.global.write"):
            data["tax_identification_number"] = None
        return Response(data, status=status.HTTP_200_OK)

    def put(self, request):
        if not user_has_permission(request.user, "settings.global.write"):
            raise PermissionDenied("You do not have permission to modify admin settings.")
        with transaction.atomic():
            gs = GlobalSettings.objects.select_for_update().get(singleton_key=_get_global_settings().singleton_key)
            before = _serialize_settings_for_audit(gs)
            serializer = GlobalSettingsSerializer(gs, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            updated = serializer.save(updated_by=request.user)
            after = _serialize_settings_for_audit(updated)
            _log_audit(request.user, "update", updated, {"before": before, "after": after, "scope": "global"})
        return Response(GlobalSettingsSerializer(updated).data, status=status.HTTP_200_OK)


class MySettingsApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        us = _get_user_settings(request.user)
        return Response(UserSettingsSerializer(us).data, status=status.HTTP_200_OK)

    def patch(self, request):
        gs = _get_global_settings()
        if not gs.allow_user_overrides:
            blocked_fields = [field for field in ("currency", "invoice_template", "receipt_template") if field in request.data]
            if blocked_fields:
                raise ValidationError(
                    {
                        "detail": (
                            "User template and currency overrides are disabled by an administrator. "
                            "Only personal preferences can be updated."
                        ),
                        "fields": blocked_fields,
                    }
                )
        with transaction.atomic():
            us = UserSettings.objects.select_for_update().get(pk=_get_user_settings(request.user).pk)
            before = _serialize_settings_for_audit(us)
            serializer = UserSettingsSerializer(us, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            updated = serializer.save()
            after = _serialize_settings_for_audit(updated)
            _log_audit(request.user, "update", updated, {"before": before, "after": after, "scope": "user"})
        return Response(UserSettingsSerializer(updated).data, status=status.HTTP_200_OK)


class SettingsRollbackApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        serializer = SettingsRollbackSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        audit_id = serializer.validated_data["audit_log_id"]
        log = AuditLog.objects.select_related("content_type", "user").get(pk=audit_id)
        model_class = log.content_type.model_class()
        if model_class not in (GlobalSettings, UserSettings):
            raise ValidationError({"detail": "Rollback not supported for this entry"})

        if model_class is GlobalSettings and not user_has_permission(request.user, "settings.global.write"):
            raise PermissionDenied()

        obj = model_class.objects.select_for_update().get(pk=log.object_id)
        if model_class is UserSettings and not (user_has_permission(request.user, "settings.global.write") or obj.user_id == request.user.id):
            raise PermissionDenied()

        before = (log.changes or {}).get("before")
        if not isinstance(before, dict):
            raise ValidationError({"detail": "Rollback data not available"})

        for field in obj._meta.fields:
            name = field.name
            if name in ("id", "updated_at"):
                continue
            if name not in before:
                continue
            value = before[name]
            if isinstance(field, models.ForeignKey):
                setattr(obj, f"{name}_id", value)
            else:
                setattr(obj, name, value)
        obj.save()
        _log_audit(request.user, "update", obj, {"rollback_from_audit_id": audit_id})
        return Response({"rolled_back": True}, status=status.HTTP_200_OK)


class CountryDefaultsApi(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        country = (request.query_params.get("country") or "").upper()
        cfg = _country_config(country)
        if not country or cfg is None:
            raise ValidationError({"country": "Unsupported country"})
        return Response(
            {
                "country": country,
                "defaults": {
                    "currency": cfg.get("currency"),
                    "date_format": (cfg.get("formats") or {}).get("date_format"),
                    "number_format": (cfg.get("formats") or {}).get("number_format"),
                    "tax": cfg.get("tax"),
                    "compliance": cfg.get("compliance"),
                },
            },
            status=status.HTTP_200_OK,
        )


class CurrencySuggestionApi(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        country = _detect_country(request) or "US"
        cfg = _country_config(country) or {}
        return Response({"country": country, "suggested_currency": cfg.get("currency") or "USD"}, status=status.HTTP_200_OK)


class GeoDetectApi(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        country = _detect_country(request)
        cfg = _country_config(country) if country else None
        return Response(
            {
                "country": country,
                "suggested_currency": (cfg or {}).get("currency"),
                "formats": (cfg or {}).get("formats"),
            },
            status=status.HTTP_200_OK,
        )


class EffectiveSettingsApi(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        region = _effective_region_settings_for_user(request, request.user)
        currency = _currency_for_code(region["currency_code"])
        currency_payload = None
        if currency is not None:
            currency_payload = {
                "code": currency.code,
                "name": currency.name,
                "symbol": currency.symbol,
                "decimal_places": currency.decimal_places,
            }
        templates = _effective_templates_for_user(request.user)
        gs = _get_global_settings()
        global_data = GlobalSettingsSerializer(gs).data
        if not getattr(request.user, "is_authenticated", False) or not user_has_permission(request.user, "settings.global.write"):
            global_data["tax_identification_number"] = None
        payload = {
            "global": global_data,
            "user": UserSettingsSerializer(_get_user_settings(request.user)).data if getattr(request.user, "is_authenticated", False) else None,
            "effective": {
                **region,
                "currency": currency_payload,
                "templates": templates,
            },
        }
        return Response(payload, status=status.HTTP_200_OK)


class SettingsAuditLogApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        scope = (request.query_params.get("scope") or "all").lower()
        try:
            limit = int(request.query_params.get("limit", "50"))
        except ValueError:
            raise ValidationError({"limit": "Invalid limit"})
        limit = max(1, min(limit, 200))

        qs = AuditLog.objects.select_related("content_type", "user").order_by("-created_at", "-id")
        settings_ct = ContentType.objects.get_for_models(GlobalSettings, UserSettings)
        allowed_ct_ids = [ct.id for ct in settings_ct.values()]
        qs = qs.filter(content_type_id__in=allowed_ct_ids)

        if not user_has_permission(request.user, "settings.global.write"):
            my_settings = _get_user_settings(request.user)
            qs = qs.filter(content_type=settings_ct[UserSettings], object_id=str(my_settings.pk))

        if scope == "global":
            qs = qs.filter(content_type=settings_ct[GlobalSettings])
        elif scope == "user":
            qs = qs.filter(content_type=settings_ct[UserSettings])
        elif scope != "all":
            raise ValidationError({"scope": "Invalid scope"})

        rows = []
        for log in qs[:limit]:
            rows.append(
                {
                    "id": log.id,
                    "action": log.action,
                    "model": log.content_type.model,
                    "object_id": log.object_id,
                    "user_id": log.user_id,
                    "created_at": log.created_at.isoformat(),
                    "changes": log.changes,
                }
            )
        return Response({"results": rows}, status=status.HTTP_200_OK)


class MeApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        u = request.user
        profile = UserProfile.objects.filter(user=u).only("company_legal_name").first()
        roles = user_role_names(u)
        permission_codes = _permission_codes_for_user(u)
        session_role = getattr(getattr(request, "auth", None), "role", None)
        return Response(
            {
                "id": u.id,
                "username": getattr(u, "username", ""),
                "email": getattr(u, "email", ""),
                "is_staff": bool(getattr(u, "is_staff", False)),
                "is_superuser": bool(getattr(u, "is_superuser", False)),
                "roles": roles,
                "permissions": permission_codes,
                "session_role": getattr(session_role, "name", None),
                "company_name": (getattr(profile, "company_legal_name", None) or None),
                "social_accounts": _social_connections_for_user(u),
            },
            status=status.HTTP_200_OK,
        )


class SocialConnectionsApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        connected = {row["provider"]: row for row in _social_connections_for_user(request.user)}
        payload = []
        for provider in ("google", "facebook"):
            row = connected.get(provider)
            payload.append(
                {
                    "provider": provider,
                    "label": _social_provider_label(provider),
                    "connected": row is not None,
                    "display_name": row.get("display_name") if row else None,
                    "email": row.get("email") if row else None,
                    "avatar_url": row.get("avatar_url") if row else None,
                    "linked_at": row.get("created_at") if row else None,
                    "last_login_at": row.get("last_login_at") if row else None,
                }
            )
        return Response({"results": payload}, status=status.HTTP_200_OK)


def _set_auth_cookie(resp: Response, token_row: AccessToken, *, remember: bool) -> None:
    secure = not bool(getattr(settings, "DEBUG", False))
    max_age = None
    if remember and token_row.expires_at is not None:
        delta = token_row.expires_at - timezone.now()
        max_age = max(1, int(delta.total_seconds()))
    resp.set_cookie(
        "auth_token",
        token_row.key,
        httponly=True,
        secure=secure,
        samesite="Lax",
        max_age=max_age,
        path="/",
    )


def _clear_auth_cookie(resp: Response) -> None:
    resp.delete_cookie("auth_token", path="/", samesite="Lax")


class TokenApi(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        ip = _client_ip(request)
        if _rate_limit(f"login:ip:{ip}", limit=30, window_seconds=60):
            _log_security_event(None, get_user_model(), object_id=ip or "unknown", changes={"event": "rate_limit", "scope": "login_ip"})
            return Response({"detail": "Too many requests. Please try again later."}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        username = str(request.data.get("username") or "").strip()
        password = str(request.data.get("password") or "")
        if not username or not password:
            logger.info("login_validation_failed ip=%s has_username=%s has_password=%s", ip or "unknown", bool(username), bool(password))
            raise ValidationError({"detail": "username and password are required"})

        if _rate_limit(f"login:user:{_hash_email(username)}", limit=10, window_seconds=3600):
            _log_security_event(None, get_user_model(), object_id=_hash_email(username), changes={"event": "rate_limit", "scope": "login_user", "ip": ip})
            return Response({"detail": "Too many requests. Please try again later."}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        User = get_user_model()
        existing = User.objects.filter(username=username).first()
        if existing is not None and not bool(getattr(existing, "is_active", True)) and existing.check_password(password):
            invitation = _admin_invitation_for_user(existing)
            detail = "Account is not active. Please verify your email before signing in."
            if invitation is not None:
                status_name = _admin_invitation_status_for_user(existing, invitation)
                if status_name == "pending_acceptance":
                    detail = "Account invitation pending. Please accept the invitation sent to your email before signing in."
                elif status_name == "accepted_pending_password":
                    detail = "Account invitation accepted. Please finish setting your password before signing in."
            logger.info("login_inactive ip=%s user_id=%s", ip or "unknown", str(getattr(existing, "pk", "")))
            _log_security_event(existing, User, object_id=str(getattr(existing, "pk", "")), changes={"event": "login_inactive", "ip": ip, "invitation_status": _admin_invitation_status_for_user(existing, invitation)})
            raise PermissionDenied(detail)

        user = authenticate(request=request, username=username, password=password)
        if user is None:
            logger.info("login_failed ip=%s username=%s", ip or "unknown", username)
            _log_security_event(None, User, object_id=username or "unknown", changes={"event": "login_failed", "ip": ip})
            raise ValidationError({"detail": "Unable to log in with provided credentials."})

        role, event_name, expires_seconds = _login_role_config_for_user(user)
        token = issue_access_token(user=user, role=role, expires_seconds=expires_seconds)
        logger.info("%s ip=%s user_id=%s", event_name, ip or "unknown", str(getattr(user, "pk", "")))
        _log_security_event(user, User, object_id=str(getattr(user, "pk", "")), changes={"event": event_name, "ip": ip})
        remember = bool(request.data.get("remember"))
        resp = Response({"token": token.key, "role": role.name, "expires_at": token.expires_at.isoformat() if token.expires_at else None}, status=status.HTTP_200_OK)
        _set_auth_cookie(resp, token, remember=remember)
        return resp


class RegisterApi(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        ip = _client_ip(request)
        if _rate_limit(f"register:ip:{ip}", limit=10, window_seconds=60):
            _log_security_event(None, get_user_model(), object_id=ip or "unknown", changes={"event": "rate_limit", "scope": "register_ip"})
            return Response({"detail": "Too many requests. Please try again later."}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        serializer = RegisterSerializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
        except ValidationError as e:
            logger.info(
                "register_validation_failed ip=%s keys=%s errors=%s",
                ip or "unknown",
                sorted(list((request.data or {}).keys())),
                getattr(e, "detail", None),
            )
            raise
        data = serializer.validated_data
        email = str(data["email"]).strip().lower()
        phone_e164 = str(data["phone_e164"]).strip()
        if _rate_limit(f"register:email:{_hash_email(email)}", limit=3, window_seconds=3600):
            _log_security_event(None, get_user_model(), object_id=email, changes={"event": "rate_limit", "scope": "register_email"})
            return Response({"detail": "Too many requests. Please try again later."}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        User = get_user_model()
        username = _username_for_email(email)
        if User.objects.filter(username=username).exists() or User.objects.filter(email=email).exists():
            _log_security_event(None, User, object_id=email, changes={"event": "register_duplicate", "ip": ip})
            raise ValidationError({"email": "An account with this email already exists"})
        if UserProfile.objects.filter(phone=phone_e164).exists():
            _log_security_event(None, User, object_id=phone_e164, changes={"event": "register_duplicate_phone", "ip": ip})
            raise ValidationError({"phone_number": "An account with this phone number already exists"})

        with transaction.atomic():
            user = User(username=username, email=email, is_active=False)
            user.set_password(data["password"])
            user.save()
            _ensure_default_business_role(user)
            profile = UserProfile.objects.create(
                user=user,
                phone=phone_e164,
                company_legal_name=str(data.get("company_name") or "").strip() or None,
                terms_accepted_at=timezone.now(),
            )
            user_settings = _get_user_settings(user)
            user_settings.country = "NG"
            ngn = Currency.objects.filter(code="NGN").first()
            if ngn is not None:
                user_settings.currency = ngn
            user_settings.save(update_fields=["country", "currency", "updated_at"])
            token = secrets.token_urlsafe(48)[:120]
            EmailVerificationToken.objects.create(
                user=user,
                token=token,
                expires_at=timezone.now() + timedelta(days=2),
            )
            _log_audit(user, "create", user, {"event": "registered", "ip": ip, "company_name": profile.company_legal_name, "phone": phone_e164})

        verification_sent = _send_verification_email(email, token, user=user, ip=ip, source="register")
        payload = {"registered": True, "verification_sent": bool(verification_sent)}
        if not verification_sent:
            payload["detail"] = "Account created, but verification email could not be delivered. Please try resend verification."
        return Response(payload, status=status.HTTP_201_CREATED)


class VerifyEmailApi(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = VerifyEmailSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        token = str(serializer.validated_data["token"]).strip()
        token_type = str(serializer.validated_data.get("token_type") or "").strip().lower()

        if token_type == ADMIN_INVITATION_TOKEN_TYPE:
            try:
                invitation = _get_admin_invitation_by_signed_token(token)
            except ValidationError:
                _log_security_event(None, get_user_model(), object_id=_hash_email(token), changes={"event": "admin_invitation_invalid", "ip": _client_ip(request)})
                raise
            if invitation.accepted_at is not None:
                _log_security_event(invitation.user, get_user_model(), object_id=str(invitation.user_id), changes={"event": "admin_invitation_reused", "ip": _client_ip(request)})
                raise ValidationError({"token": "Invitation already accepted"})

            with transaction.atomic():
                invitation.accepted_at = timezone.now()
                invitation.save(update_fields=["accepted_at", "updated_at"])
                profile, _ = UserProfile.objects.get_or_create(user=invitation.user)
                if profile.email_verified_at is None:
                    profile.email_verified_at = timezone.now()
                    profile.save(update_fields=["email_verified_at", "updated_at"])
                _log_audit(None, "update", invitation.user, {"event": "admin_invitation_accepted"})
                _log_security_event(invitation.user, get_user_model(), object_id=str(invitation.user_id), changes={"event": "admin_invitation_accepted", "ip": _client_ip(request)})

            reset_uid, reset_token, reset_url = _password_reset_link_for_user(invitation.user)
            return Response(
                {
                    "verified": True,
                    "invitation_accepted": True,
                    "password_reset_unlocked": True,
                    "reset_uid": reset_uid,
                    "reset_token": reset_token,
                    "reset_url": reset_url,
                },
                status=status.HTTP_200_OK,
            )

        try:
            row = EmailVerificationToken.objects.select_related("user").get(token=token)
        except EmailVerificationToken.DoesNotExist:
            raise ValidationError({"token": "Invalid token"})
        if row.used_at is not None:
            raise ValidationError({"token": "Token already used"})
        if row.expires_at <= timezone.now():
            raise ValidationError({"token": "Token expired"})

        with transaction.atomic():
            row.used_at = timezone.now()
            row.save(update_fields=["used_at"])
            user = row.user
            user.is_active = True
            user.save(update_fields=["is_active"])
            profile, _ = UserProfile.objects.get_or_create(user=user)
            if profile.email_verified_at is None:
                profile.email_verified_at = timezone.now()
                profile.save(update_fields=["email_verified_at", "updated_at"])
            _log_audit(user, "update", user, {"event": "email_verified"})
            _log_security_event(user, get_user_model(), object_id=str(user.pk), changes={"event": "email_verified"})

        return Response({"verified": True}, status=status.HTTP_200_OK)


class ResendVerificationApi(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = str(request.data.get("email") or "").strip().lower()
        ip = _client_ip(request)
        if not email:
            raise ValidationError({"email": "Email is required"})
        if _rate_limit(f"resend_verify:ip:{ip}", limit=10, window_seconds=60):
            return Response({"detail": "Too many requests. Please try again later."}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        User = get_user_model()
        user = User.objects.filter(email=email).first()
        if user is None:
            return Response({"sent": True}, status=status.HTTP_200_OK)

        profile = UserProfile.objects.filter(user=user).first()
        if profile and profile.email_verified_at is not None:
            return Response({"sent": True}, status=status.HTTP_200_OK)

        token = secrets.token_urlsafe(48)[:120]
        EmailVerificationToken.objects.create(user=user, token=token, expires_at=timezone.now() + timedelta(days=2))
        sent = _send_verification_email(email, token, user=user, ip=ip, source="resend")
        _log_security_event(user, User, object_id=str(user.pk), changes={"event": "verification_resent", "ip": ip, "ok": bool(sent)})
        if not sent:
            return Response({"sent": False, "detail": "Unable to send verification email. Please try again later."}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        return Response({"sent": True}, status=status.HTTP_200_OK)


class LogoutApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        token = getattr(request, "auth", None)
        if isinstance(token, AccessToken):
            revoke_token(token)
        resp = Response(status=status.HTTP_204_NO_CONTENT)
        _clear_auth_cookie(resp)
        return resp


class PasswordResetApi(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        ip = _client_ip(request)
        if _rate_limit(f"pwreset:ip:{ip}", limit=10, window_seconds=60):
            return Response({"detail": "Too many requests. Please try again later."}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        serializer = PasswordResetRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = str(serializer.validated_data["email"]).strip().lower()
        if _rate_limit(f"pwreset:email:{_hash_email(email)}", limit=3, window_seconds=3600):
            return Response({"detail": "Too many requests. Please try again later."}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        User = get_user_model()
        user = User.objects.filter(email=email).first()
        if user is not None:
            invitation = _admin_invitation_for_user(user)
            if invitation is not None and invitation.accepted_at is None:
                logger.info("password_reset_blocked_pending_acceptance user_id=%s ip=%s", str(user.pk), ip or "unknown")
                _log_security_event(user, User, object_id=str(user.pk), changes={"event": "password_reset_blocked_pending_acceptance", "ip": ip})
                return Response({"sent": True}, status=status.HTTP_200_OK)
            uid, token, link = _password_reset_link_for_user(user)
            subject = "Reset your password"
            message = f"Open this link to reset your password:\n\n{link}\n"
            try:
                sent = send_mail(subject, message, getattr(settings, "DEFAULT_FROM_EMAIL", "no-reply@example.com"), [email], fail_silently=False)
                ok = int(sent) > 0
                logger.info("password_reset_email_sent ok=%s user_id=%s ip=%s", ok, str(user.pk), ip or "unknown")
                _log_security_event(user, User, object_id=str(user.pk), changes={"event": "password_reset_email_sent", "ok": ok, "ip": ip})
            except Exception as e:
                logger.exception("password_reset_email_failed user_id=%s ip=%s", str(getattr(user, "pk", "")), ip or "unknown")
                _log_security_event(user, User, object_id=str(user.pk), changes={"event": "password_reset_email_failed", "ok": False, "ip": ip, "error": e.__class__.__name__})

        return Response({"sent": True}, status=status.HTTP_200_OK)


class PasswordResetConfirmApi(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        ip = _client_ip(request)
        if _rate_limit(f"pwreset_confirm:ip:{ip}", limit=10, window_seconds=60):
            return Response({"detail": "Too many requests. Please try again later."}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        serializer = PasswordResetConfirmSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        uid = str(serializer.validated_data["uid"]).strip()
        token = str(serializer.validated_data["token"]).strip()
        new_password = str(serializer.validated_data["new_password"])

        User = get_user_model()
        try:
            user_id = force_str(urlsafe_base64_decode(uid))
            user = User.objects.get(pk=int(user_id))
        except Exception:
            raise ValidationError({"uid": "Invalid uid"})

        if not PasswordResetTokenGenerator().check_token(user, token):
            _log_security_event(user, User, object_id=str(user.pk), changes={"event": "password_reset_invalid_token", "ip": ip})
            raise ValidationError({"token": "Invalid or expired token"})

        invitation = _admin_invitation_for_user(user)
        if invitation is not None and invitation.accepted_at is None:
            _log_security_event(user, User, object_id=str(user.pk), changes={"event": "password_reset_blocked_pending_acceptance", "ip": ip})
            raise PermissionDenied("Account invitation must be accepted before setting a password.")

        activated_by_onboarding = False
        with transaction.atomic():
            user.set_password(new_password)
            if not bool(getattr(user, "is_active", True)):
                user.is_active = True
                activated_by_onboarding = invitation is not None
            user.save()
            profile, _ = UserProfile.objects.get_or_create(user=user)
            if profile.email_verified_at is None:
                profile.email_verified_at = timezone.now()
                profile.save(update_fields=["email_verified_at", "updated_at"])
            if invitation is not None:
                now = timezone.now()
                update_fields = ["updated_at"]
                if invitation.password_reset_completed_at is None:
                    invitation.password_reset_completed_at = now
                    update_fields.append("password_reset_completed_at")
                if invitation.activated_at is None:
                    invitation.activated_at = now
                    update_fields.append("activated_at")
                invitation.save(update_fields=update_fields)
                _ensure_default_business_role(user)
                _log_audit(None, "update", user, {"event": "password_reset_completed", "onboarding": "admin_invite"})
                _log_security_event(user, User, object_id=str(user.pk), changes={"event": "password_reset_completed", "ip": ip})
                _log_audit(None, "update", user, {"event": "account_activated", "roles": user_role_names(user)})
                _log_security_event(user, User, object_id=str(user.pk), changes={"event": "account_activated", "ip": ip, "roles": user_role_names(user)})
            else:
                _log_audit(user, "update", user, {"event": "password_reset"})
                _log_security_event(user, User, object_id=str(user.pk), changes={"event": "password_reset", "ip": ip})

        AccessToken.objects.filter(user=user, revoked_at__isnull=True).update(revoked_at=timezone.now())
        role, _, expires_seconds = _login_role_config_for_user(user)
        token_row = issue_access_token(user=user, role=role, expires_seconds=expires_seconds)
        remember = bool(request.data.get("remember"))
        activation_email_sent = False
        if activated_by_onboarding and invitation is not None:
            activation_email_sent = _send_account_activation_email(user, ip=ip, invitation=invitation)
        resp = Response(
            {
                "reset": True,
                "token": token_row.key,
                "role": role.name,
                "expires_at": token_row.expires_at.isoformat() if token_row.expires_at else None,
                "activated": bool(getattr(user, "is_active", False)),
                "activation_email_sent": activation_email_sent,
            },
            status=status.HTTP_200_OK,
        )
        _set_auth_cookie(resp, token_row, remember=remember)
        return resp


class GoogleOAuthStartApi(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        ip = _client_ip(request)
        if _rate_limit(f"google_oauth_start:ip:{ip}", limit=20, window_seconds=60):
            logger.info("google_oauth_start_rate_limited ip=%s", ip or "unknown")
            return _oauth_frontend_redirect(provider="google", error="rate_limited")

        client_id = getattr(settings, "GOOGLE_OAUTH_CLIENT_ID", "") or ""
        client_secret = getattr(settings, "GOOGLE_OAUTH_CLIENT_SECRET", "") or ""
        if not client_id or not client_secret:
            logger.info("google_oauth_not_configured ip=%s has_client_id=%s has_client_secret=%s", ip or "unknown", bool(client_id), bool(client_secret))
            return _oauth_frontend_redirect(provider="google", error="not_configured")

        state = secrets.token_urlsafe(24)
        try:
            state_payload = _oauth_state_payload(request)
        except PermissionDenied:
            return _oauth_frontend_redirect(provider="google", error="link_requires_login")
        cache.set(_oauth_state_cache_key("google", state), state_payload, timeout=600)
        redirect_uri = request.build_absolute_uri("/api/auth/google/callback/")
        params = {
            "client_id": client_id,
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "scope": "openid email profile",
            "state": state,
            "prompt": "select_account",
            "access_type": "online",
        }
        url = "https://accounts.google.com/o/oauth2/v2/auth?" + urllib.parse.urlencode(params)
        logger.info("google_oauth_start ip=%s", ip or "unknown")
        return HttpResponseRedirect(url)


class GoogleOAuthCallbackApi(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        ip = _client_ip(request)
        err = str(request.query_params.get("error") or "").strip()
        code = str(request.query_params.get("code") or "").strip()
        state = str(request.query_params.get("state") or "").strip()
        if err:
            mapped = "cancelled" if err in ("access_denied", "user_cancelled") else "failed"
            logger.info("google_oauth_cancelled ip=%s error=%s", ip or "unknown", err)
            return _oauth_frontend_redirect(provider="google", error=mapped)
        if not code:
            return _oauth_frontend_redirect(provider="google", error="missing_code")
        if not state:
            return _oauth_frontend_redirect(provider="google", error="missing_state")

        state_row = _pop_oauth_state("google", state)
        if state_row is None:
            logger.info("google_oauth_invalid_state ip=%s", ip or "unknown")
            return _oauth_frontend_redirect(provider="google", error="invalid_state")
        remember = bool(state_row.get("remember", True))
        intent = str(state_row.get("intent") or "login")

        client_id = getattr(settings, "GOOGLE_OAUTH_CLIENT_ID", "") or ""
        client_secret = getattr(settings, "GOOGLE_OAUTH_CLIENT_SECRET", "") or ""
        if not client_id or not client_secret:
            logger.info(
                "google_oauth_not_configured_callback ip=%s has_client_id=%s has_client_secret=%s",
                ip or "unknown",
                bool(client_id),
                bool(client_secret),
            )
            return _oauth_frontend_redirect(provider="google", error="not_configured")

        redirect_uri = request.build_absolute_uri("/api/auth/google/callback/")
        data = urllib.parse.urlencode(
            {
                "code": code,
                "client_id": client_id,
                "client_secret": client_secret,
                "redirect_uri": redirect_uri,
                "grant_type": "authorization_code",
            }
        ).encode("utf-8")

        try:
            req = urllib.request.Request("https://oauth2.googleapis.com/token", data=data, method="POST")
            req.add_header("Content-Type", "application/x-www-form-urlencoded")
            with urllib.request.urlopen(req, timeout=10) as resp:
                token_payload = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            logger.info("google_oauth_token_exchange_http_error ip=%s status=%s", ip or "unknown", getattr(e, "code", "unknown"))
            _log_security_event(None, get_user_model(), object_id=ip or "unknown", changes={"event": "google_oauth_exchange_failed", "ip": ip, "error": f"http_{getattr(e, 'code', 'unknown')}"})
            return _oauth_frontend_redirect(provider="google", error="exchange_failed")
        except Exception as e:
            logger.exception("google_oauth_token_exchange_failed ip=%s", ip or "unknown")
            _log_security_event(None, get_user_model(), object_id=ip or "unknown", changes={"event": "google_oauth_exchange_failed", "ip": ip, "error": e.__class__.__name__})
            return _oauth_frontend_redirect(provider="google", error="exchange_failed")

        id_token = str(token_payload.get("id_token") or "").strip()
        if not id_token:
            return _oauth_frontend_redirect(provider="google", error="missing_id_token")

        try:
            info_url = "https://oauth2.googleapis.com/tokeninfo?" + urllib.parse.urlencode({"id_token": id_token})
            with urllib.request.urlopen(info_url, timeout=10) as resp:
                info = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            logger.info("google_oauth_tokeninfo_http_error ip=%s status=%s", ip or "unknown", getattr(e, "code", "unknown"))
            _log_security_event(None, get_user_model(), object_id=ip or "unknown", changes={"event": "google_oauth_tokeninfo_failed", "ip": ip, "error": f"http_{getattr(e, 'code', 'unknown')}"})
            return _oauth_frontend_redirect(provider="google", error="identity_failed")
        except Exception as e:
            logger.exception("google_oauth_tokeninfo_failed ip=%s", ip or "unknown")
            _log_security_event(None, get_user_model(), object_id=ip or "unknown", changes={"event": "google_oauth_tokeninfo_failed", "ip": ip, "error": e.__class__.__name__})
            return _oauth_frontend_redirect(provider="google", error="identity_failed")

        email = str(info.get("email") or "").strip().lower()
        email_verified = str(info.get("email_verified") or "").lower() == "true"
        aud = str(info.get("aud") or "").strip()
        provider_user_id = str(info.get("sub") or "").strip()
        display_name = str(info.get("name") or "").strip()
        avatar_url = str(info.get("picture") or "").strip()
        if not email or not email_verified:
            return _oauth_frontend_redirect(provider="google", error="email_not_verified")
        if aud and aud != client_id:
            return _oauth_frontend_redirect(provider="google", error="invalid_audience")
        if not provider_user_id:
            return _oauth_frontend_redirect(provider="google", error="identity_failed")

        if intent == "link":
            return _complete_social_link(
                user_id=int(state_row.get("user_id")),
                provider="google",
                provider_user_id=provider_user_id,
                email=email,
                display_name=display_name,
                avatar_url=avatar_url,
                ip=ip,
            )

        User = get_user_model()
        connection = SocialAuthConnection.objects.filter(provider="google", provider_user_id=provider_user_id).select_related("user").first()
        user = connection.user if connection is not None else User.objects.filter(email=email).first()
        created = False
        if user is None:
            user = User(username=_username_for_email(email), email=email, is_active=True)
            user.set_unusable_password()
            user.save()
            created = True
        else:
            if not bool(getattr(user, "is_active", True)):
                user.is_active = True
                user.save(update_fields=["is_active"])

        profile, _ = UserProfile.objects.get_or_create(user=user)
        if profile.email_verified_at is None:
            profile.email_verified_at = timezone.now()
            profile.save(update_fields=["email_verified_at", "updated_at"])

        if _is_privileged_account(user):
            return _complete_social_login(user=user, provider="google", ip=ip, remember=remember, created=created)
        _touch_social_connection(
            user=user,
            provider="google",
            provider_user_id=provider_user_id,
            email=email,
            display_name=display_name,
            avatar_url=avatar_url,
        )
        logger.info("google_oauth_success ip=%s user_id=%s created=%s", ip or "unknown", str(user.pk), created)
        return _complete_social_login(user=user, provider="google", ip=ip, remember=remember, created=created)


class FacebookOAuthStartApi(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        ip = _client_ip(request)
        if _rate_limit(f"facebook_oauth_start:ip:{ip}", limit=20, window_seconds=60):
            logger.info("facebook_oauth_start_rate_limited ip=%s", ip or "unknown")
            return _oauth_frontend_redirect(provider="facebook", error="rate_limited")

        client_id = getattr(settings, "FACEBOOK_OAUTH_CLIENT_ID", "") or ""
        client_secret = getattr(settings, "FACEBOOK_OAUTH_CLIENT_SECRET", "") or ""
        if not client_id or not client_secret:
            logger.info("facebook_oauth_not_configured ip=%s has_client_id=%s has_client_secret=%s", ip or "unknown", bool(client_id), bool(client_secret))
            return _oauth_frontend_redirect(provider="facebook", error="not_configured")

        state = secrets.token_urlsafe(24)
        try:
            state_payload = _oauth_state_payload(request)
        except PermissionDenied:
            return _oauth_frontend_redirect(provider="facebook", error="link_requires_login")
        cache.set(_oauth_state_cache_key("facebook", state), state_payload, timeout=600)
        redirect_uri = request.build_absolute_uri("/api/auth/facebook/callback/")
        params = {
            "client_id": client_id,
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "scope": "email,public_profile",
            "state": state,
        }
        url = "https://www.facebook.com/v20.0/dialog/oauth?" + urllib.parse.urlencode(params)
        logger.info("facebook_oauth_start ip=%s", ip or "unknown")
        return HttpResponseRedirect(url)


class FacebookOAuthCallbackApi(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        ip = _client_ip(request)
        err = str(request.query_params.get("error") or "").strip()
        code = str(request.query_params.get("code") or "").strip()
        state = str(request.query_params.get("state") or "").strip()
        if err:
            mapped = "cancelled" if err in ("access_denied", "user_denied") else "failed"
            logger.info("facebook_oauth_cancelled ip=%s error=%s", ip or "unknown", err)
            return _oauth_frontend_redirect(provider="facebook", error=mapped)
        if not code:
            return _oauth_frontend_redirect(provider="facebook", error="missing_code")
        if not state:
            return _oauth_frontend_redirect(provider="facebook", error="missing_state")

        state_row = _pop_oauth_state("facebook", state)
        if state_row is None:
            logger.info("facebook_oauth_invalid_state ip=%s", ip or "unknown")
            return _oauth_frontend_redirect(provider="facebook", error="invalid_state")
        remember = bool(state_row.get("remember", True))
        intent = str(state_row.get("intent") or "login")

        client_id = getattr(settings, "FACEBOOK_OAUTH_CLIENT_ID", "") or ""
        client_secret = getattr(settings, "FACEBOOK_OAUTH_CLIENT_SECRET", "") or ""
        if not client_id or not client_secret:
            logger.info(
                "facebook_oauth_not_configured_callback ip=%s has_client_id=%s has_client_secret=%s",
                ip or "unknown",
                bool(client_id),
                bool(client_secret),
            )
            return _oauth_frontend_redirect(provider="facebook", error="not_configured")

        redirect_uri = request.build_absolute_uri("/api/auth/facebook/callback/")
        token_url = "https://graph.facebook.com/v20.0/oauth/access_token?" + urllib.parse.urlencode(
            {
                "client_id": client_id,
                "client_secret": client_secret,
                "redirect_uri": redirect_uri,
                "code": code,
            }
        )
        try:
            with urllib.request.urlopen(token_url, timeout=10) as resp:
                token_payload = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            logger.info("facebook_oauth_token_exchange_http_error ip=%s status=%s", ip or "unknown", getattr(e, "code", "unknown"))
            _log_security_event(None, get_user_model(), object_id=ip or "unknown", changes={"event": "facebook_oauth_exchange_failed", "ip": ip, "error": f"http_{getattr(e, 'code', 'unknown')}"})
            return _oauth_frontend_redirect(provider="facebook", error="exchange_failed")
        except Exception as e:
            logger.exception("facebook_oauth_token_exchange_failed ip=%s", ip or "unknown")
            _log_security_event(None, get_user_model(), object_id=ip or "unknown", changes={"event": "facebook_oauth_exchange_failed", "ip": ip, "error": e.__class__.__name__})
            return _oauth_frontend_redirect(provider="facebook", error="exchange_failed")

        access_token = str(token_payload.get("access_token") or "").strip()
        if not access_token:
            return _oauth_frontend_redirect(provider="facebook", error="missing_access_token")

        me_url = "https://graph.facebook.com/v20.0/me?" + urllib.parse.urlencode(
            {"fields": "id,name,email,picture.type(large)", "access_token": access_token}
        )
        try:
            with urllib.request.urlopen(me_url, timeout=10) as resp:
                info = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            logger.info("facebook_oauth_identity_http_error ip=%s status=%s", ip or "unknown", getattr(e, "code", "unknown"))
            _log_security_event(None, get_user_model(), object_id=ip or "unknown", changes={"event": "facebook_oauth_identity_failed", "ip": ip, "error": f"http_{getattr(e, 'code', 'unknown')}"})
            return _oauth_frontend_redirect(provider="facebook", error="identity_failed")
        except Exception as e:
            logger.exception("facebook_oauth_identity_failed ip=%s", ip or "unknown")
            _log_security_event(None, get_user_model(), object_id=ip or "unknown", changes={"event": "facebook_oauth_identity_failed", "ip": ip, "error": e.__class__.__name__})
            return _oauth_frontend_redirect(provider="facebook", error="identity_failed")

        provider_user_id = str(info.get("id") or "").strip()
        display_name = str(info.get("name") or "").strip()
        email = str(info.get("email") or "").strip().lower()
        picture = info.get("picture") if isinstance(info.get("picture"), dict) else {}
        picture_data = picture.get("data") if isinstance(picture, dict) else {}
        avatar_url = str(picture_data.get("url") or "").strip() if isinstance(picture_data, dict) else ""

        if not provider_user_id:
            return _oauth_frontend_redirect(provider="facebook", error="identity_failed")
        if not email:
            return _oauth_frontend_redirect(provider="facebook", error="email_unavailable")

        if intent == "link":
            if not state_row.get("user_id"):
                return _oauth_frontend_redirect(provider="facebook", error="link_target_missing")
            return _complete_social_link(
                user_id=int(state_row.get("user_id")),
                provider="facebook",
                provider_user_id=provider_user_id,
                email=email,
                display_name=display_name,
                avatar_url=avatar_url,
                ip=ip,
            )

        User = get_user_model()
        connection = SocialAuthConnection.objects.filter(provider="facebook", provider_user_id=provider_user_id).select_related("user").first()
        user = connection.user if connection is not None else User.objects.filter(email=email).first()
        created = False
        if user is None:
            user = User(username=_username_for_email(email), email=email, is_active=True)
            user.set_unusable_password()
            user.save()
            created = True
        else:
            updates = []
            if not bool(getattr(user, "is_active", True)):
                user.is_active = True
                updates.append("is_active")
            if email and str(getattr(user, "email", "") or "").strip().lower() != email:
                return _oauth_frontend_redirect(provider="facebook", error="email_mismatch")
            if updates:
                user.save(update_fields=updates)

        profile, _ = UserProfile.objects.get_or_create(user=user)
        if profile.email_verified_at is None:
            profile.email_verified_at = timezone.now()
            profile.save(update_fields=["email_verified_at", "updated_at"])

        if _is_privileged_account(user):
            return _complete_social_login(user=user, provider="facebook", ip=ip, remember=remember, created=created)
        _touch_social_connection(
            user=user,
            provider="facebook",
            provider_user_id=provider_user_id,
            email=email,
            display_name=display_name,
            avatar_url=avatar_url,
        )
        logger.info("facebook_oauth_success ip=%s user_id=%s created=%s", ip or "unknown", str(user.pk), created)
        return _complete_social_login(user=user, provider="facebook", ip=ip, remember=remember, created=created)


class AdminEmailVerificationMetricsApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not user_has_permission(request.user, "admin.email.test"):
            raise PermissionDenied("You do not have permission to view admin email metrics.")
        try:
            hours = int(request.query_params.get("hours", "24"))
        except ValueError:
            raise ValidationError({"hours": "Invalid hours"})
        hours = max(1, min(hours, 720))
        since = timezone.now() - timedelta(hours=hours)

        sent_ok = 0
        sent_fail = 0
        resent = 0
        verified = 0
        test_ok = 0
        test_fail = 0

        security_qs = AuditLog.objects.filter(action="security", created_at__gte=since).only("changes")
        for log in security_qs.iterator():
            changes = log.changes or {}
            event = changes.get("event")
            if event == "verification_email_sent":
                if bool(changes.get("ok")):
                    sent_ok += 1
                else:
                    sent_fail += 1
            elif event == "verification_email_failed":
                sent_fail += 1
            elif event == "verification_resent":
                resent += 1
            elif event == "email_verified":
                verified += 1
            elif event == "test_email_sent":
                if bool(changes.get("ok")):
                    test_ok += 1
                else:
                    test_fail += 1

        return Response(
            {
                "window_hours": hours,
                "since": since.isoformat(),
                "verification_emails_sent_ok": sent_ok,
                "verification_emails_sent_failed": sent_fail,
                "verification_emails_resent": resent,
                "emails_verified": verified,
                "test_emails_sent_ok": test_ok,
                "test_emails_sent_failed": test_fail,
            },
            status=status.HTTP_200_OK,
        )


class AdminSendTestEmailApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not user_has_permission(request.user, "admin.email.test"):
            raise PermissionDenied("You do not have permission to send test emails.")
        ip = _client_ip(request)
        if _rate_limit(f"admin_test_email:ip:{ip}", limit=5, window_seconds=60):
            return Response({"detail": "Too many requests. Please try again later."}, status=status.HTTP_429_TOO_MANY_REQUESTS)

        recipient = str(request.data.get("email") or "").strip()
        if not recipient:
            raise ValidationError({"email": "Email is required"})

        subject = str(request.data.get("subject") or "").strip() or "PIXELHUB test email"
        message = str(request.data.get("message") or "").strip() or f"Test email sent at {timezone.now().isoformat()}."

        User = get_user_model()
        object_id = _hash_email(recipient)
        try:
            sent = send_mail(
                subject,
                message,
                getattr(settings, "DEFAULT_FROM_EMAIL", "no-reply@example.com"),
                [recipient],
                fail_silently=False,
            )
            ok = int(sent) > 0
            logger.info("test_email_sent ok=%s recipient_hash=%s ip=%s", ok, object_id, ip or "unknown")
            _log_security_event(
                request.user,
                User,
                object_id=object_id,
                changes={"event": "test_email_sent", "ok": ok, "ip": ip},
            )
            if not ok:
                return Response({"sent": False, "detail": "Email backend did not report a successful send."}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
            return Response({"sent": True}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.exception("test_email_failed recipient_hash=%s ip=%s", object_id, ip or "unknown")
            _log_security_event(
                request.user,
                User,
                object_id=object_id,
                changes={"event": "test_email_sent", "ok": False, "ip": ip, "error": e.__class__.__name__},
            )
            return Response({"sent": False, "detail": "Unable to send test email."}, status=status.HTTP_503_SERVICE_UNAVAILABLE)


class CurrencyConvertApi(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        amount = request.query_params.get("amount")
        base = request.query_params.get("from")
        quote = request.query_params.get("to")
        if amount is None or base is None or quote is None:
            raise ValidationError({"detail": "amount, from, and to are required"})
        try:
            amount_d = Decimal(str(amount))
        except (InvalidOperation, TypeError):
            raise ValidationError({"amount": "Invalid amount"})
        converted = _convert_currency(amount_d, str(base), str(quote))
        return Response(
            {"amount": str(_q2(amount_d)), "from": str(base).upper(), "to": str(quote).upper(), "converted": str(converted)},
            status=status.HTTP_200_OK,
        )


def _parse_custom_role_names(raw_value) -> list[str]:
    if raw_value in (None, ""):
        return []
    if not isinstance(raw_value, list):
        raise ValidationError({"custom_roles": "custom_roles must be a list"})
    cleaned = sorted({str(value or "").strip() for value in raw_value if str(value or "").strip()})
    invalid = [name for name in cleaned if name in SYSTEM_ROLE_NAMES]
    if invalid:
        raise ValidationError({"custom_roles": f"Use primary_role for system roles: {', '.join(sorted(invalid))}"})
    existing = set(Role.objects.filter(name__in=cleaned).values_list("name", flat=True))
    missing = [name for name in cleaned if name not in existing]
    if missing:
        raise ValidationError({"custom_roles": f"Unknown roles: {', '.join(missing)}"})
    return cleaned


def _apply_admin_user_roles(user, *, primary_role: str, custom_role_names: list[str]) -> None:
    role_name = str(primary_role or "user").strip().lower()
    if role_name not in SYSTEM_ROLE_NAMES:
        raise ValidationError({"primary_role": "primary_role must be user, staff, or admin"})
    user.is_superuser = role_name == "admin"
    user.is_staff = role_name in {"staff", "admin"}
    user.save(update_fields=["is_superuser", "is_staff"])
    custom_roles = list(Role.objects.filter(name__in=custom_role_names))
    UserRole.objects.filter(user=user).exclude(role__name__in=SYSTEM_ROLE_NAMES).exclude(role__in=custom_roles).delete()
    for role in custom_roles:
        UserRole.objects.get_or_create(user=user, role=role)
    _ensure_default_business_role(user)


class AdminUsersApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not user_has_permission(request.user, "admin.users.read"):
            raise PermissionDenied("You do not have permission to view users.")
        User = get_user_model()
        qs = User.objects.all().order_by("-id")
        try:
            page = int(request.query_params.get("page", "1"))
        except ValueError:
            page = 1
        page = max(1, page)
        page_size = 25
        start = (page - 1) * page_size
        end = start + page_size
        page_users = list(qs[start:end])
        latest_logins = _latest_login_timestamps_by_user([u.id for u in page_users])
        results = [_serialize_admin_user(u, latest_login_at=latest_logins.get(u.id)) for u in page_users]
        return Response({"page": page, "results": results, "count": qs.count()}, status=status.HTTP_200_OK)

    def post(self, request):
        if not user_has_permission(request.user, "admin.users.write"):
            raise PermissionDenied("You do not have permission to manage users.")
        username = str(request.data.get("username") or "").strip()
        password = str(request.data.get("password") or "")
        email = str(request.data.get("email") or "").strip().lower()
        full_name = str(request.data.get("full_name") or "").strip()
        company_name = str(request.data.get("company_name") or "").strip()
        primary_role = str(request.data.get("primary_role") or "user").strip().lower()
        custom_role_names = _parse_custom_role_names(request.data.get("custom_roles"))
        if not username or not password:
            raise ValidationError({"detail": "username and password are required"})
        if not email:
            raise ValidationError({"email": "Email is required"})
        validate_application_password(password, min_length=6)
        phone_value = str(request.data.get("phone") or "").strip()
        country_code = str(request.data.get("country_code") or DEFAULT_COUNTRY_CODE).strip() or DEFAULT_COUNTRY_CODE
        phone_e164 = None
        if phone_value:
            try:
                phone_e164 = normalize_signup_phone(country_code, phone_value)
            except ValidationError as exc:
                raise ValidationError({"phone": exc.detail})
        User = get_user_model()
        if User.objects.filter(username=username).exists():
            raise ValidationError({"username": "Username already exists"})
        if email and User.objects.filter(email=email).exists():
            raise ValidationError({"email": "Email already exists"})
        if phone_e164 and UserProfile.objects.filter(phone=phone_e164).exists():
            raise ValidationError({"phone": "Phone number already exists"})
        invitation = None
        with transaction.atomic():
            u = User(username=username, email=email, is_active=False)
            u.set_password(password)
            u.save()
            _apply_admin_user_roles(u, primary_role=primary_role, custom_role_names=custom_role_names)
            UserProfile.objects.update_or_create(
                user=u,
                defaults={
                    "full_name": full_name or None,
                    "phone": phone_e164,
                    "company_legal_name": company_name or None,
                    "email_verified_at": None,
                },
            )
            invitation = AdminUserInvitation.objects.create(
                user=u,
                invited_by=request.user,
                token_key=secrets.token_urlsafe(24),
                expires_at=timezone.now() + timedelta(hours=72),
            )
            _log_audit(
                request.user,
                "create",
                u,
                {
                    "username": username,
                    "email": email or None,
                    "is_active": False,
                    "primary_role": primary_role,
                    "custom_roles": custom_role_names,
                    "onboarding": "admin_invitation_pending",
                    "invitation_expires_at": invitation.expires_at.isoformat(),
                },
            )
        invite_sent = _send_admin_invitation_email(invitation, ip=_client_ip(request)) if invitation is not None else False
        payload = {"id": u.id, "created": True, "invitation_sent": bool(invite_sent)}
        if not invite_sent:
            payload["detail"] = "User created, but the invitation email could not be delivered."
        return Response(payload, status=status.HTTP_201_CREATED)

    def patch(self, request):
        if not user_has_permission(request.user, "admin.users.write"):
            raise PermissionDenied("You do not have permission to manage users.")
        user_id = request.data.get("id")
        if user_id is None:
            raise ValidationError({"id": "id is required"})
        User = get_user_model()
        try:
            u = User.objects.get(pk=int(user_id))
        except (User.DoesNotExist, ValueError, TypeError):
            raise NotFound("User not found")

        before = _serialize_admin_user(u, latest_login_at=_latest_login_timestamps_by_user([u.id]).get(u.id))
        update_fields = []
        if "username" in request.data:
            username = str(request.data.get("username") or "").strip()
            if not username:
                raise ValidationError({"username": "Username is required"})
            if User.objects.exclude(pk=u.pk).filter(username=username).exists():
                raise ValidationError({"username": "Username already exists"})
            u.username = username
            update_fields.append("username")
        if "email" in request.data:
            email = str(request.data.get("email") or "").strip().lower()
            if email and User.objects.exclude(pk=u.pk).filter(email=email).exists():
                raise ValidationError({"email": "Email already exists"})
            u.email = email
            update_fields.append("email")
        if "is_active" in request.data:
            invitation = _admin_invitation_for_user(u)
            requested_active = bool(request.data.get("is_active"))
            if invitation is not None and invitation.activated_at is None and requested_active:
                raise ValidationError({"is_active": "Admin-invited users activate only after invitation acceptance and password setup."})
            u.is_active = requested_active
            update_fields.append("is_active")

        primary_role = before["primary_role"]
        if "primary_role" in request.data:
            primary_role = str(request.data.get("primary_role") or "").strip().lower()
        custom_role_names = before["custom_roles"]
        if "custom_roles" in request.data:
            custom_role_names = _parse_custom_role_names(request.data.get("custom_roles"))
        if int(request.user.id) == int(u.id) and primary_role != "admin":
            raise ValidationError({"primary_role": "You cannot remove your own admin access."})

        new_password = request.data.get("new_password")
        if new_password not in (None, ""):
            new_password = str(new_password)
            validate_application_password(new_password, min_length=6)

        full_name = str(request.data.get("full_name") or "").strip() if "full_name" in request.data else None
        company_name = str(request.data.get("company_name") or "").strip() if "company_name" in request.data else None
        phone_value = str(request.data.get("phone") or "").strip() if "phone" in request.data else None
        country_code = str(request.data.get("country_code") or DEFAULT_COUNTRY_CODE).strip() or DEFAULT_COUNTRY_CODE
        phone_e164 = None
        if phone_value is not None:
            if phone_value:
                try:
                    phone_e164 = normalize_signup_phone(country_code, phone_value)
                except ValidationError as exc:
                    raise ValidationError({"phone": exc.detail})
                if UserProfile.objects.exclude(user=u).filter(phone=phone_e164).exists():
                    raise ValidationError({"phone": "Phone number already exists"})

        with transaction.atomic():
            if update_fields:
                u.save(update_fields=sorted(set(update_fields)))
            _apply_admin_user_roles(u, primary_role=primary_role, custom_role_names=custom_role_names)
            if new_password:
                u.set_password(new_password)
                u.save(update_fields=["password"])
            profile, _ = UserProfile.objects.get_or_create(user=u)
            profile_updates = []
            if full_name is not None:
                profile.full_name = full_name or None
                profile_updates.append("full_name")
            if company_name is not None:
                profile.company_legal_name = company_name or None
                profile_updates.append("company_legal_name")
            if phone_value is not None:
                profile.phone = phone_e164
                profile_updates.append("phone")
            if profile_updates:
                profile.save(update_fields=sorted(set(profile_updates + ["updated_at"])))

        after = _serialize_admin_user(u, latest_login_at=_latest_login_timestamps_by_user([u.id]).get(u.id))
        changes = {"before": before, "after": after}
        if new_password:
            changes["password_reset"] = True
        if before != after or new_password:
            _log_audit(request.user, "update", u, changes)
        return Response({"updated": True}, status=status.HTTP_200_OK)

    def delete(self, request):
        if not user_has_permission(request.user, "admin.users.write"):
            raise PermissionDenied("You do not have permission to manage users.")

        user_id = request.data.get("id") or request.query_params.get("id")
        confirm_keyword = str(request.data.get("confirm_keyword") or request.query_params.get("confirm_keyword") or "").strip()
        confirm_email = str(request.data.get("confirm_email") or request.query_params.get("confirm_email") or "").strip().lower()

        if user_id is None:
            raise ValidationError({"id": "id is required"})
        if confirm_keyword != "DELETE":
            raise ValidationError({"confirm_keyword": "Type DELETE to confirm permanent removal."})

        User = get_user_model()
        try:
            target = User.objects.get(pk=int(user_id))
        except (User.DoesNotExist, ValueError, TypeError):
            raise NotFound("User not found")

        if int(request.user.id) == int(target.id):
            raise ValidationError({"id": "You cannot delete your own account."})
        if confirm_email != str(target.email or "").strip().lower():
            raise ValidationError({"confirm_email": "Confirmation email does not match the selected user."})

        target_snapshot = _serialize_admin_user(target, latest_login_at=_latest_login_timestamps_by_user([target.id]).get(target.id))
        ip = _client_ip(request)

        try:
            with transaction.atomic():
                _anonymize_user_artifacts_for_deletion(target)
                user_content_type = ContentType.objects.get_for_model(User)
                AuditLog.objects.filter(content_type=user_content_type, object_id=str(target.id)).exclude(action="security").delete()
                _log_operation(
                    request.user,
                    "delete",
                    User,
                    f"deleted-user:{target.id}",
                    {
                        "deleted_user": {
                            "id": target.id,
                            "email": target.email,
                            "username": target.username,
                            "primary_role": target_snapshot.get("primary_role"),
                            "custom_roles": target_snapshot.get("custom_roles", []),
                        },
                        "actor_admin_id": request.user.id,
                        "actor_email": getattr(request.user, "email", None) or getattr(request.user, "username", None),
                        "ip": ip,
                        "deleted_at": timezone.now().isoformat(),
                        "compliance": {
                            "personal_data": "removed_or_anonymized",
                            "retained_business_records": ["expenses", "payment_transactions"],
                        },
                    },
                )
                target.delete()
        except DatabaseError:
            return Response(
                {
                    "detail": "Unable to delete user due to a database error.",
                    "code": "user_delete_failed",
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        return Response({"deleted": True}, status=status.HTTP_200_OK)


class AdminRolesApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not user_has_permission(request.user, "admin.users.read"):
            raise PermissionDenied("You do not have permission to view roles.")
        roles = Role.objects.all().order_by("name")
        permissions_qs = Permission.objects.all().order_by("code")
        return Response(
            {
                "permissions": [{"code": perm.code, "description": perm.description} for perm in permissions_qs],
                "results": [
                    {
                        "id": role.id,
                        "name": role.name,
                        "description": role.description,
                        "is_system": role.name in SYSTEM_ROLE_NAMES,
                        "permission_codes": _role_permission_codes(role),
                    }
                    for role in roles
                ],
            },
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        if not user_has_permission(request.user, "admin.users.write"):
            raise PermissionDenied("You do not have permission to manage roles.")
        name = str(request.data.get("name") or "").strip().lower()
        description = str(request.data.get("description") or "").strip()
        permission_codes_raw = request.data.get("permission_codes") or []
        if not name:
            raise ValidationError({"name": "name is required"})
        if not re.fullmatch(r"[a-z][a-z0-9_\\-]{1,48}", name):
            raise ValidationError({"name": "Use lowercase letters, numbers, hyphen, or underscore"})
        if name in SYSTEM_ROLE_NAMES:
            raise ValidationError({"name": "System roles cannot be created here"})
        if not isinstance(permission_codes_raw, list):
            raise ValidationError({"permission_codes": "permission_codes must be a list"})
        permission_codes = sorted({str(code or "").strip() for code in permission_codes_raw if str(code or "").strip()})
        permissions_map = {perm.code: perm for perm in Permission.objects.filter(code__in=permission_codes)}
        missing = [code for code in permission_codes if code not in permissions_map]
        if missing:
            raise ValidationError({"permission_codes": f"Unknown permissions: {', '.join(missing)}"})
        role, created = Role.objects.get_or_create(name=name, defaults={"description": description or None})
        if not created:
            raise ValidationError({"name": "Role already exists"})
        role.description = description or None
        role.save(update_fields=["description"])
        for code in permission_codes:
            RolePermission.objects.get_or_create(role=role, permission=permissions_map[code])
        _log_audit(request.user, "create", role, {"permission_codes": permission_codes})
        return Response({"id": role.id, "created": True}, status=status.HTTP_201_CREATED)

    def patch(self, request):
        if not user_has_permission(request.user, "admin.users.write"):
            raise PermissionDenied("You do not have permission to manage roles.")
        role_id = request.data.get("id")
        if role_id is None:
            raise ValidationError({"id": "id is required"})
        try:
            role = Role.objects.get(pk=int(role_id))
        except (Role.DoesNotExist, ValueError, TypeError):
            raise NotFound("Role not found")
        if role.name in SYSTEM_ROLE_NAMES:
            raise ValidationError({"id": "System roles cannot be modified here"})
        description = str(request.data.get("description") or "").strip() if "description" in request.data else role.description
        permission_codes = _role_permission_codes(role)
        if "permission_codes" in request.data:
            raw_codes = request.data.get("permission_codes") or []
            if not isinstance(raw_codes, list):
                raise ValidationError({"permission_codes": "permission_codes must be a list"})
            permission_codes = sorted({str(code or "").strip() for code in raw_codes if str(code or "").strip()})
        permissions_map = {perm.code: perm for perm in Permission.objects.filter(code__in=permission_codes)}
        missing = [code for code in permission_codes if code not in permissions_map]
        if missing:
            raise ValidationError({"permission_codes": f"Unknown permissions: {', '.join(missing)}"})
        before = {"description": role.description, "permission_codes": _role_permission_codes(role)}
        with transaction.atomic():
            role.description = description or None
            role.save(update_fields=["description"])
            RolePermission.objects.filter(role=role).exclude(permission__code__in=permission_codes).delete()
            for code in permission_codes:
                RolePermission.objects.get_or_create(role=role, permission=permissions_map[code])
        after = {"description": role.description, "permission_codes": _role_permission_codes(role)}
        if before != after:
            _log_audit(request.user, "update", role, {"before": before, "after": after})
        return Response({"updated": True}, status=status.HTTP_200_OK)


class AdminAuditLogsApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not user_has_permission(request.user, "admin.users.read"):
            raise PermissionDenied("You do not have permission to view audit logs.")
        try:
            page = int(request.query_params.get("page", "1"))
        except ValueError:
            page = 1
        page = max(1, page)
        page_size = 25
        q = str(request.query_params.get("q") or "").strip().lower()
        qs = AuditLog.objects.select_related("user", "content_type").order_by("-created_at")
        if q:
            qs = qs.filter(Q(action__icontains=q) | Q(changes__icontains=q) | Q(object_id__icontains=q))
        total = qs.count()
        start = (page - 1) * page_size
        rows = []
        for row in qs[start : start + page_size]:
            rows.append(
                {
                    "id": row.id,
                    "action": row.action,
                    "object_id": row.object_id,
                    "content_type": getattr(row.content_type, "model", None),
                    "changes": row.changes,
                    "created_at": row.created_at.isoformat() if row.created_at else None,
                    "actor": getattr(getattr(row, "user", None), "email", None) or getattr(getattr(row, "user", None), "username", None),
                }
            )
        return Response({"page": page, "count": total, "results": rows}, status=status.HTTP_200_OK)


class AdminOAuthStatusApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not user_has_permission(request.user, "admin.oauth.status.read"):
            raise PermissionDenied("You do not have permission to view OAuth status.")
        ip = _client_ip(request)
        google_id = getattr(settings, "GOOGLE_OAUTH_CLIENT_ID", "") or ""
        google_secret = getattr(settings, "GOOGLE_OAUTH_CLIENT_SECRET", "") or ""
        facebook_id = getattr(settings, "FACEBOOK_OAUTH_CLIENT_ID", "") or ""
        facebook_secret = getattr(settings, "FACEBOOK_OAUTH_CLIENT_SECRET", "") or ""

        payload = {
            "google": {
                "configured": bool(google_id and google_secret),
                "has_client_id": bool(google_id),
                "has_client_secret": bool(google_secret),
                "callback_url": request.build_absolute_uri("/api/auth/google/callback/"),
                "expected_env": [
                    "DJANGO_GOOGLE_OAUTH_CLIENT_ID",
                    "DJANGO_GOOGLE_OAUTH_CLIENT_SECRET",
                    "GOOGLE_OAUTH_CLIENT_ID",
                    "GOOGLE_OAUTH_CLIENT_SECRET",
                ],
            },
            "facebook": {
                "configured": bool(facebook_id and facebook_secret),
                "has_client_id": bool(facebook_id),
                "has_client_secret": bool(facebook_secret),
                "callback_url": request.build_absolute_uri("/api/auth/facebook/callback/"),
                "expected_env": [
                    "DJANGO_FACEBOOK_OAUTH_CLIENT_ID",
                    "DJANGO_FACEBOOK_OAUTH_CLIENT_SECRET",
                    "FACEBOOK_OAUTH_CLIENT_ID",
                    "FACEBOOK_OAUTH_CLIENT_SECRET",
                ],
            },
        }
        logger.info("admin_oauth_status ip=%s google=%s facebook=%s", ip or "unknown", payload["google"]["configured"], payload["facebook"]["configured"])
        return Response(payload, status=status.HTTP_200_OK)


class AdminRuntimeDiagnosticsApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        allowed = bool(getattr(user, "is_superuser", False) or getattr(user, "is_staff", False))
        if not allowed:
            allowed = bool(
                user_has_permission(user, "settings.global.read")
                or user_has_permission(user, "admin.users.read")
                or user_has_permission(user, "admin.audit.read")
            )
        if not allowed:
            raise PermissionDenied("You do not have permission to view runtime diagnostics.")

        def _sanitize_url(value: Any) -> Any:
            if not isinstance(value, str):
                return value
            raw = value.strip()
            if "://" not in raw:
                return value
            parsed = urllib.parse.urlparse(raw)
            if not parsed.scheme or not parsed.netloc:
                return value
            host = parsed.hostname or ""
            port = f":{parsed.port}" if parsed.port else ""
            path = parsed.path or ""
            query = f"?{parsed.query}" if parsed.query else ""
            frag = f"#{parsed.fragment}" if parsed.fragment else ""
            return f"{parsed.scheme}://{host}{port}{path}{query}{frag}"

        from django.db import connection

        db_settings = connection.settings_dict or {}
        db_payload = {
            "engine": db_settings.get("ENGINE"),
            "name": str(db_settings.get("NAME") or ""),
            "host": db_settings.get("HOST"),
            "port": db_settings.get("PORT"),
            "conn_max_age": db_settings.get("CONN_MAX_AGE"),
            "conn_health_checks": db_settings.get("CONN_HEALTH_CHECKS"),
            "atomic_requests": db_settings.get("ATOMIC_REQUESTS"),
            "vendor": getattr(connection, "vendor", None),
            "in_atomic_block": bool(getattr(connection, "in_atomic_block", False)),
            "autocommit": bool(connection.get_autocommit()),
        }
        try:
            db_payload["is_usable"] = bool(connection.is_usable())
        except Exception as exc:
            db_payload["is_usable"] = False
            db_payload["usable_error"] = str(exc)

        cache_settings = (getattr(settings, "CACHES", {}) or {}).get("default", {}) or {}
        cache_payload = {
            "backend": cache_settings.get("BACKEND"),
            "timeout": cache_settings.get("TIMEOUT"),
            "key_prefix": cache_settings.get("KEY_PREFIX"),
            "location": _sanitize_url(cache_settings.get("LOCATION")),
            "class": f"{cache.__class__.__module__}.{cache.__class__.__name__}",
        }

        payload = {
            "service": "pixelhub",
            "server_time": timezone.now().isoformat(),
            "instance": {"hostname": os.environ.get("HOSTNAME") or os.uname().nodename, "pid": os.getpid()},
            "debug": bool(getattr(settings, "DEBUG", False)),
            "db": db_payload,
            "cache": cache_payload,
        }
        return Response(payload, status=status.HTTP_200_OK)


def _logo_upload_payload(asset: LogoAsset) -> dict[str, Any]:
    return {
        "asset_id": asset.id,
        "scope": asset.scope,
        "logo_url": asset.file_url,
        "thumbnail_url": asset.thumbnail_url,
        "logo_thumbnail_url": asset.thumbnail_url,
        "content_type": asset.content_type,
        "size_bytes": asset.size_bytes,
        "sha256": asset.sha256,
    }


def _assign_logo_asset_for_scope(*, user, scope: str, asset: LogoAsset) -> None:
    if scope == LOGO_SCOPE_GLOBAL:
        gs = GlobalSettings.objects.select_for_update().get(singleton_key=_get_global_settings().singleton_key)
        before = _serialize_settings_for_audit(gs)
        previous_asset = gs.appearance_logo
        gs.appearance_logo = asset
        gs.updated_by = user
        gs.save(update_fields=["appearance_logo", "updated_by", "updated_at"])
        after = _serialize_settings_for_audit(gs)
        _log_audit(user, "update", gs, {"before": before, "after": after, "scope": "global_logo_upload"})
        cleanup_logo_asset_if_unreferenced(previous_asset)
        return

    us = UserSettings.objects.select_for_update().get(pk=_get_user_settings(user).pk)
    before = _serialize_settings_for_audit(us)
    if scope == LOGO_SCOPE_INVOICE:
        previous_asset = us.invoice_logo
        us.invoice_logo = asset
        update_fields = ["invoice_logo", "updated_at"]
    elif scope == LOGO_SCOPE_RECEIPT:
        previous_asset = us.receipt_logo
        us.receipt_logo = asset
        update_fields = ["receipt_logo", "updated_at"]
    else:
        raise ValidationError({"scope": "Invalid logo scope"})
    us.save(update_fields=update_fields)
    after = _serialize_settings_for_audit(us)
    _log_audit(user, "update", us, {"before": before, "after": after, "scope": f"{scope}_logo_upload"})
    cleanup_logo_asset_if_unreferenced(previous_asset)


class AdminLogoUploadApi(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [parsers.MultiPartParser, parsers.FormParser]

    def post(self, request):
        if not user_has_permission(request.user, "admin.logo.upload"):
            raise PermissionDenied("You do not have permission to upload logos.")
        upload = request.FILES.get("file")
        if upload is None:
            raise ValidationError({"file": "file is required"})
        scope = normalize_logo_scope(request.data.get("scope") or LOGO_SCOPE_GLOBAL)
        if scope != LOGO_SCOPE_GLOBAL:
            raise ValidationError({"scope": "Admin logo uploads only support the global appearance scope."})
        asset = create_logo_asset(upload, scope=scope, owner=request.user)
        try:
            with transaction.atomic():
                _assign_logo_asset_for_scope(user=request.user, scope=scope, asset=asset)
        except Exception:
            cleanup_logo_asset_if_unreferenced(asset)
            raise
        return Response(_logo_upload_payload(asset), status=status.HTTP_201_CREATED)


class SettingsLogoUploadApi(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [parsers.MultiPartParser, parsers.FormParser]

    def post(self, request):
        if not _get_global_settings().allow_user_overrides:
            raise ValidationError(
                {
                    "detail": (
                        "User template overrides are currently disabled by an administrator. "
                        "Logo uploads for invoice and receipt templates are unavailable."
                    )
                }
            )
        upload = request.FILES.get("file")
        if upload is None:
            raise ValidationError({"file": "file is required"})
        scope = normalize_logo_scope(request.data.get("scope"))
        if scope == LOGO_SCOPE_GLOBAL:
            raise PermissionDenied("Global logo uploads must use the admin settings endpoint.")
        asset = create_logo_asset(upload, scope=scope, owner=request.user)
        try:
            with transaction.atomic():
                _assign_logo_asset_for_scope(user=request.user, scope=scope, asset=asset)
        except Exception:
            cleanup_logo_asset_if_unreferenced(asset)
            raise
        return Response(_logo_upload_payload(asset), status=status.HTTP_201_CREATED)


def _delivery_backoff_seconds(attempt_count: int) -> int:
    attempt = max(0, int(attempt_count))
    base = 60
    cap = 6 * 60 * 60
    delay = base * (2**attempt)
    return int(min(cap, max(base, delay)))


def _tx_idempotency_hash(key: str) -> str:
    return hashlib.sha256((key or "").encode("utf-8")).hexdigest()


def _decimal_places_for_currency(code: str) -> int:
    c = str(code or "").strip().upper()
    row = Currency.objects.filter(code=c).first()
    if row is None:
        return 2
    try:
        return int(row.decimal_places)
    except Exception:
        return 2


def _to_minor_units(amount: Decimal, currency_code: str) -> int:
    dp = max(0, min(6, _decimal_places_for_currency(currency_code)))
    scale = Decimal(10) ** dp
    v = (amount * scale).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return int(v)


def _http_json(method: str, url: str, headers: dict[str, str], body: dict | None, timeout: int = 20) -> tuple[int, dict, str]:
    data_bytes = None
    if body is not None:
        data_bytes = json.dumps(body).encode("utf-8")
        headers = {**headers, "Content-Type": "application/json"}
    req = urllib.request.Request(url=url, method=method.upper(), data=data_bytes, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            status_code = int(getattr(resp, "status", 200))
            raw = resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as e:
        status_code = int(getattr(e, "code", 500))
        raw = e.read().decode("utf-8", errors="replace")
    try:
        parsed = json.loads(raw or "{}")
        if not isinstance(parsed, dict):
            parsed = {}
    except json.JSONDecodeError:
        parsed = {}
    return status_code, parsed, raw


def _safe_webhook_headers(request) -> dict:
    out = {}
    for k, v in dict(getattr(request, "headers", {}) or {}).items():
        lk = str(k or "").lower()
        if lk in ("authorization", "cookie"):
            continue
        if lk.startswith("x-") or lk in ("content-type", "user-agent"):
            out[lk] = str(v)
    return out


def _settle_payment_transaction(*, tx: PaymentTransaction, provider_paid_at: str | None, provider_status: str | None, provider_tx_id: str | None, provider_ref: str | None) -> None:
    with transaction.atomic():
        tx = PaymentTransaction.objects.select_for_update().select_related("invoice").get(pk=tx.pk)
        if tx.status == "succeeded":
            return
        paid_at = _parse_iso_datetime(provider_paid_at) if provider_paid_at else None
        if paid_at is None:
            paid_at = timezone.now()
        tx.status = "succeeded"
        tx.paid_at = paid_at
        if provider_status:
            tx.metadata = {**(tx.metadata or {}), "provider_status": str(provider_status)}
        if provider_tx_id:
            tx.provider_transaction_id = str(provider_tx_id)
        if provider_ref:
            tx.provider_reference = str(provider_ref)
        tx.save(update_fields=["status", "paid_at", "provider_transaction_id", "provider_reference", "metadata", "updated_at"])

        invoice = tx.invoice
        total_paid = invoice.receipts.filter(is_deleted=False).aggregate(total=Sum("amount_paid"))["total"] or Decimal("0.00")
        outstanding = outstanding_invoice_amount(invoice.total_amount, total_paid)
        amount_paid = min(outstanding, tx.amount).quantize(CENTS, rounding=ROUND_HALF_UP)
        if amount_paid <= 0:
            return
        Receipt.objects.create(
            invoice=invoice,
            amount_paid=amount_paid,
            payment_method="Bank Transfer" if tx.provider == "bank_transfer" else "Card",
            reference_number=f"{tx.provider}:{tx.provider_transaction_id or tx.provider_reference or tx.reference}",
        )
        total_paid_after = (total_paid + amount_paid).quantize(CENTS, rounding=ROUND_HALF_UP)
        sync_invoice_status_with_payments(invoice, total_paid_after)
        _log_audit(tx.created_by, "update", tx, {"status": {"to": "succeeded"}, "receipt_amount": str(amount_paid)})


def _fail_payment_transaction(*, tx: PaymentTransaction, failure_code: str | None, failure_message: str | None, provider_status: str | None) -> None:
    with transaction.atomic():
        tx = PaymentTransaction.objects.select_for_update().get(pk=tx.pk)
        if tx.status in ("succeeded", "failed", "cancelled"):
            return
        tx.status = "failed"
        tx.failure_code = str(failure_code)[:80] if failure_code else None
        tx.failure_message = str(failure_message)[:255] if failure_message else None
        if provider_status:
            tx.metadata = {**(tx.metadata or {}), "provider_status": str(provider_status)}
        tx.save(update_fields=["status", "failure_code", "failure_message", "metadata", "updated_at"])
        _log_audit(tx.created_by, "update", tx, {"status": {"to": "failed"}, "code": tx.failure_code})


class DocumentDeliveryViewSet(viewsets.ModelViewSet):
    serializer_class = DocumentDeliverySerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = OptionalPageNumberPagination

    def _is_retryable_error(self, e: Exception, err_message: str | None) -> bool:
        msg = (err_message or "").lower()
        if isinstance(e, (PermissionDenied, NotFound)):
            return False
        if isinstance(e, (TypeError, ValueError)):
            return False
        if isinstance(e, ValidationError):
            if "not configured" in msg:
                return False
            if "is required" in msg:
                return False
            if "invalid" in msg:
                return False
            return True
        return True

    def _report(self, delivery: DocumentDelivery) -> dict:
        now = timezone.now()
        max_attempts = int(getattr(settings, "DOCUMENT_DELIVERY_MAX_ATTEMPTS", 6) or 6)
        retry_recommended = bool(delivery.status == "failed" and (delivery.attempt_count or 0) < max_attempts and delivery.next_retry_at)
        retry_after_seconds = None
        if delivery.next_retry_at:
            retry_after_seconds = max(0, int((delivery.next_retry_at - now).total_seconds()))
        recipient = {
            "email": delivery.to_email,
            "phone": delivery.to_phone,
            "printer_name": str((delivery.metadata or {}).get("printer_name") or "").strip() or None,
        }
        document = {
            "type": delivery.document_type,
            "invoice_id": delivery.invoice_id,
            "receipt_id": delivery.receipt_id,
        }
        if delivery.document_type == "invoice" and delivery.invoice_id:
            try:
                document["invoice_number"] = delivery.invoice.invoice_number if delivery.invoice else None
            except Exception:
                document["invoice_number"] = None
        recommendation = "No action required."
        if delivery.status == "failed":
            if retry_recommended and delivery.next_retry_at:
                recommendation = f"Retry recommended. Next retry at {delivery.next_retry_at.isoformat()}."
            elif (delivery.attempt_count or 0) >= max_attempts:
                recommendation = "No more retries. Update recipient details and resend."
            else:
                recommendation = "Update recipient details and resend."
        return {
            "ok": delivery.status == "sent",
            "delivery_id": delivery.id,
            "status": delivery.status,
            "channel": delivery.channel,
            "format": delivery.format,
            "recipient": recipient,
            "document": document,
            "provider_message_id": delivery.provider_message_id,
            "attempt_count": delivery.attempt_count,
            "last_attempt_at": delivery.last_attempt_at.isoformat() if delivery.last_attempt_at else None,
            "created_at": delivery.created_at.isoformat() if delivery.created_at else None,
            "updated_at": delivery.updated_at.isoformat() if delivery.updated_at else None,
            "recommendation": recommendation,
            "error": {
                "code": delivery.last_error_code,
                "message": delivery.last_error_message,
            }
            if delivery.status == "failed"
            else None,
            "retry": {
                "recommended": retry_recommended,
                "next_retry_at": delivery.next_retry_at.isoformat() if delivery.next_retry_at else None,
                "retry_after_seconds": retry_after_seconds,
                "max_attempts": max_attempts,
            },
        }

    def _error_details(self, e: Exception) -> tuple[str | None, str | None]:
        code = e.__class__.__name__
        msg = str(e)
        try:
            if isinstance(e, ValidationError):
                detail = getattr(e, "detail", None)
                if isinstance(detail, dict):
                    if "detail" in detail:
                        msg = str(detail.get("detail"))
                    else:
                        msg = json.dumps(detail, default=str)
                elif isinstance(detail, list) and detail:
                    msg = str(detail[0])
                elif detail is not None:
                    msg = str(detail)
        except Exception:
            msg = str(e)
        return (code[:80] if code else None), (msg[:255] if msg else None)

    def get_queryset(self):
        qs = DocumentDelivery.objects.all().select_related("invoice", "receipt", "receipt__invoice", "user").order_by("-id")
        roles = set(user_role_names(self.request.user))
        if "admin" in roles or "staff" in roles or bool(getattr(self.request.user, "is_superuser", False)):
            return qs
        return qs.filter(user=self.request.user)

    def create(self, request, *args, **kwargs):
        from .documents import create_delivery, send_delivery

        user = request.user
        if not getattr(user, "is_authenticated", False):
            raise PermissionDenied()

        document_type = str(request.data.get("document_type") or "").strip()
        document_id = request.data.get("document_id")
        channel = str(request.data.get("channel") or "").strip()
        fmt = str(request.data.get("format") or "pdf").strip()
        send_now = request.data.get("send_now")
        if send_now in (None, ""):
            send_now = True
        ttl_minutes = request.data.get("ttl_minutes")
        try:
            ttl_minutes_int = int(ttl_minutes) if ttl_minutes not in (None, "") else 60
        except (TypeError, ValueError):
            raise ValidationError({"ttl_minutes": "ttl_minutes must be an integer"})

        try:
            document_id_int = int(document_id)
        except (TypeError, ValueError):
            raise ValidationError({"document_id": "document_id must be an integer"})

        delivery_metadata: dict[str, Any] = {}
        email_subject_template = str(request.data.get("email_subject_template") or "").strip()
        email_message_template = str(request.data.get("email_message_template") or "").strip()
        if email_subject_template:
            delivery_metadata["email_subject_template"] = email_subject_template
        if email_message_template:
            delivery_metadata["email_message_template"] = email_message_template

        delivery, token = create_delivery(
            user=user,
            document_type=document_type,
            document_id=document_id_int,
            channel=channel,
            fmt=fmt,
            to_email=request.data.get("to_email"),
            to_phone=request.data.get("to_phone"),
            ttl_minutes=ttl_minutes_int,
            metadata=delivery_metadata,
        )
        if channel == "print":
            printer_name = str(request.data.get("printer_name") or "").strip()
            if printer_name:
                delivery.metadata = {**(delivery.metadata or {}), "printer_name": printer_name}
                delivery.save(update_fields=["metadata", "updated_at"])

        _log_audit(user, "create", delivery, {"channel": channel, "format": fmt, "document_type": document_type, "document_id": document_id_int})

        if bool(send_now):
            try:
                delivery.status = "sending"
                delivery.save(update_fields=["status", "updated_at"])
                delivery = send_delivery(request, delivery, token)
                _log_audit(user, "update", delivery, {"status": {"to": "sent"}})
                logger.info("document_delivery.sent delivery_id=%s channel=%s", delivery.id, delivery.channel)
            except Exception as e:
                err_code, err_msg = self._error_details(e)
                retryable = self._is_retryable_error(e, err_msg)
                delivery.status = "failed"
                delivery.attempt_count = (delivery.attempt_count or 0) + 1
                delivery.last_attempt_at = timezone.now()
                delivery.last_error_code = err_code
                delivery.last_error_message = err_msg
                delivery.next_retry_at = (
                    timezone.now() + timedelta(seconds=_delivery_backoff_seconds(delivery.attempt_count)) if retryable else None
                )
                delivery.save(update_fields=["status", "attempt_count", "last_attempt_at", "last_error_code", "last_error_message", "next_retry_at", "updated_at"])
                logger.exception("document_delivery.send_failed delivery_id=%s", delivery.id)
                _log_audit(user, "update", delivery, {"status": {"to": "failed"}, "error": {"code": delivery.last_error_code, "message": delivery.last_error_message}})

        payload = {
            "delivery": self.get_serializer(delivery).data,
            "report": self._report(delivery),
        }
        return Response(payload, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"])
    def send(self, request, pk=None):
        from .documents import send_delivery

        delivery = self.get_object()
        if delivery.status == "sent":
            return Response({"delivery": self.get_serializer(delivery).data, "report": self._report(delivery)}, status=status.HTTP_200_OK)
        try:
            delivery.status = "sending"
            delivery.save(update_fields=["status", "updated_at"])
            delivery = send_delivery(request, delivery, token=None)
            _log_audit(request.user, "update", delivery, {"status": {"to": "sent"}})
            logger.info("document_delivery.sent delivery_id=%s channel=%s", delivery.id, delivery.channel)
        except Exception as e:
            err_code, err_msg = self._error_details(e)
            retryable = self._is_retryable_error(e, err_msg)
            delivery.status = "failed"
            delivery.attempt_count = (delivery.attempt_count or 0) + 1
            delivery.last_attempt_at = timezone.now()
            delivery.last_error_code = err_code
            delivery.last_error_message = err_msg
            delivery.next_retry_at = timezone.now() + timedelta(seconds=_delivery_backoff_seconds(delivery.attempt_count)) if retryable else None
            delivery.save(update_fields=["status", "attempt_count", "last_attempt_at", "last_error_code", "last_error_message", "next_retry_at", "updated_at"])
            logger.exception("document_delivery.resend_failed delivery_id=%s", delivery.id)
            _log_audit(request.user, "update", delivery, {"status": {"to": "failed"}, "error": {"code": delivery.last_error_code, "message": delivery.last_error_message}})
        return Response({"delivery": self.get_serializer(delivery).data, "report": self._report(delivery)}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="download", permission_classes=[permissions.AllowAny])
    def download(self, request, pk=None):
        from .documents import render_invoice, render_receipt, verify_download_token

        try:
            delivery = DocumentDelivery.objects.select_related("invoice", "receipt", "receipt__invoice", "user").get(pk=int(pk))
        except (ValueError, TypeError, DocumentDelivery.DoesNotExist):
            raise NotFound("Delivery not found")

        token = str(request.query_params.get("token") or "").strip()
        now = timezone.now()
        token_ok = False
        if token and delivery.download_token_hash and (delivery.download_expires_at is None or delivery.download_expires_at > now):
            token_ok = bool(verify_download_token(token, delivery.download_token_hash))

        user_ok = False
        if getattr(request.user, "is_authenticated", False):
            roles = set(user_role_names(request.user))
            if "admin" in roles or "staff" in roles or bool(getattr(request.user, "is_superuser", False)):
                user_ok = True
            elif delivery.user_id == request.user.id:
                user_ok = True

        if not token_ok and not user_ok:
            raise PermissionDenied("Invalid or expired token")

        if delivery.document_type == "invoice":
            doc = render_invoice(request, delivery.invoice, delivery.format)
        else:
            doc = render_receipt(request, delivery.receipt, delivery.format)

        resp = HttpResponse(doc.content, content_type=doc.content_type)
        resp["Content-Disposition"] = f'attachment; filename="{doc.filename}"'
        resp["X-Content-Type-Options"] = "nosniff"
        return resp


class SavedDocumentViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = SavedDocumentSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = OptionalPageNumberPagination

    def get_queryset(self):
        qs = SavedDocument.objects.all().select_related("invoice", "receipt", "receipt__invoice", "user").order_by("-id")
        roles = set(user_role_names(self.request.user))
        if "admin" in roles or "staff" in roles or bool(getattr(self.request.user, "is_superuser", False)):
            return qs
        return qs.filter(user=self.request.user)

    def _report(self, saved_document: SavedDocument) -> dict[str, Any]:
        return {
            "ok": True,
            "saved_document_id": saved_document.id,
            "document_type": saved_document.document_type,
            "format": saved_document.format,
            "storage_backend": saved_document.storage_backend,
            "size_bytes": saved_document.size_bytes,
            "created_at": saved_document.created_at.isoformat() if saved_document.created_at else None,
        }

    def create(self, request, *args, **kwargs):
        from .documents import save_document_backup

        user = request.user
        if not getattr(user, "is_authenticated", False):
            raise PermissionDenied()

        document_type = str(request.data.get("document_type") or "").strip()
        document_id = request.data.get("document_id")
        try:
            document_id_int = int(document_id)
        except (TypeError, ValueError):
            raise ValidationError({"document_id": "document_id must be an integer"})

        save_label = str(request.data.get("label") or "").strip()
        try:
            saved_document = save_document_backup(
                request,
                user=user,
                document_type=document_type,
                document_id=document_id_int,
                metadata={"label": save_label} if save_label else {},
            )
        except ValidationError:
            raise
        except Exception as exc:
            logger.exception("saved_document.create_failed document_type=%s document_id=%s", document_type, document_id_int)
            raise APIException(f"Failed to save document backup: {exc}") from exc

        _log_audit(user, "create", saved_document, {"document_type": document_type, "document_id": document_id_int})
        payload = {
            "saved_document": self.get_serializer(saved_document).data,
            "download_url": f"/api/documents/saved/{saved_document.id}/download/",
            "report": self._report(saved_document),
        }
        return Response(payload, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"])
    def download(self, request, pk=None):
        saved_document = self.get_object()
        try:
            with saved_document.file.storage.open(saved_document.file.name, "rb") as fh:
                content = fh.read()
        except Exception as exc:
            logger.exception("saved_document.download_failed saved_document_id=%s", saved_document.id)
            raise APIException(f"Failed to open saved document: {exc}") from exc

        response = HttpResponse(content, content_type=saved_document.content_type or "application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{saved_document.original_filename}"'
        response["X-Content-Type-Options"] = "nosniff"
        return response


class PrinterListApi(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        import subprocess

        roles = set(user_role_names(request.user))
        if not ("admin" in roles or "staff" in roles or bool(getattr(request.user, "is_superuser", False))):
            raise PermissionDenied()

        printers: list[dict] = []
        try:
            proc = subprocess.run(["lpstat", "-p"], capture_output=True, text=True, timeout=3, check=False)
            for line in (proc.stdout or "").splitlines():
                s = line.strip()
                if not s.startswith("printer "):
                    continue
                parts = s.split()
                if len(parts) >= 2:
                    printers.append({"name": parts[1], "raw": s})
        except Exception:
            printers = []
        return Response({"printers": printers}, status=status.HTTP_200_OK)


class PaymentTransactionViewSet(viewsets.ModelViewSet):
    serializer_class = PaymentTransactionSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = OptionalPageNumberPagination

    def get_queryset(self):
        qs = PaymentTransaction.objects.all().select_related("invoice", "created_by").order_by("-id")
        roles = set(user_role_names(self.request.user))
        if "admin" in roles or "staff" in roles or bool(getattr(self.request.user, "is_superuser", False)):
            return qs
        return qs.filter(created_by=self.request.user)

    def create(self, request, *args, **kwargs):
        user = request.user
        if not getattr(user, "is_authenticated", False):
            raise PermissionDenied()

        idempotency_key = request.META.get("HTTP_IDEMPOTENCY_KEY") or request.headers.get("Idempotency-Key")
        if not idempotency_key or not isinstance(idempotency_key, str):
            raise ValidationError({"detail": "Idempotency-Key header is required"})
        idem_hash = _tx_idempotency_hash(idempotency_key)

        provider = str(request.data.get("provider") or "").strip()
        if provider not in ("bank_transfer", "opay", "flutterwave", "paystack"):
            raise ValidationError({"provider": "Invalid provider"})
        invoice_id = request.data.get("invoice")
        try:
            invoice_id_int = int(invoice_id)
        except (TypeError, ValueError):
            raise ValidationError({"invoice": "invoice must be an integer"})

        with transaction.atomic():
            try:
                invoice = Invoice.objects.select_for_update().get(pk=invoice_id_int, is_deleted=False)
            except Invoice.DoesNotExist:
                raise NotFound("Invoice not found")

            existing = PaymentTransaction.objects.filter(invoice=invoice, provider=provider, idempotency_key_hash=idem_hash).first()
            if existing:
                return Response(self.get_serializer(existing).data, status=status.HTTP_200_OK)

            raw_amount = request.data.get("amount")
            if raw_amount in (None, ""):
                total_paid = invoice.receipts.filter(is_deleted=False).aggregate(total=Sum("amount_paid"))["total"] or Decimal("0.00")
                outstanding = outstanding_invoice_amount(invoice.total_amount, total_paid)
                amount = outstanding
            else:
                try:
                    amount = Decimal(str(raw_amount)).quantize(CENTS, rounding=ROUND_HALF_UP)
                except (InvalidOperation, TypeError):
                    raise ValidationError({"amount": "amount must be a valid number"})
            if amount <= 0:
                raise ValidationError({"amount": "amount must be > 0"})

            currency_code = str(request.data.get("currency_code") or "").strip().upper()
            if not currency_code:
                region = _effective_region_settings_for_user(request, request.user)
                currency_code = str(region.get("currency_code") or "NGN").strip().upper()

            reference = uuid.uuid4().hex[:18].upper()
            tx = PaymentTransaction.objects.create(
                invoice=invoice,
                created_by=user,
                provider=provider,
                status="initiated",
                amount=amount,
                currency_code=currency_code,
                reference=reference,
                idempotency_key_hash=idem_hash,
                metadata={"client_ip": _client_ip(request)},
            )

        redirect_url = str(request.data.get("redirect_url") or "").strip()
        if not redirect_url:
            redirect_url = str(getattr(settings, "FRONTEND_BASE_URL", "http://127.0.0.1:3000")).rstrip("/") + "/receipts"

        try:
            if provider == "paystack":
                secret = str(getattr(settings, "PAYSTACK_SECRET_KEY", "") or "").strip()
                if not secret:
                    raise ValidationError({"detail": "Paystack is not configured"})
                email = str(request.data.get("email") or getattr(invoice.customer, "email", "") or "").strip()
                if not email:
                    raise ValidationError({"email": "Customer email is required"})
                amount_minor = _to_minor_units(amount, currency_code)
                status_code, parsed, _raw = _http_json(
                    "POST",
                    "https://api.paystack.co/transaction/initialize",
                    {"Authorization": f"Bearer {secret}"},
                    {"email": email, "amount": amount_minor, "currency": currency_code, "reference": reference, "callback_url": redirect_url, "metadata": {"invoice_id": invoice.id}},
                    timeout=20,
                )
                if status_code >= 400 or not parsed.get("status"):
                    raise PaymentGatewayUnavailable("Paystack initialize failed")
                data = parsed.get("data") or {}
                payment_url = str(data.get("authorization_url") or "").strip() or None
                provider_ref = str(data.get("reference") or reference).strip() or reference
                tx.payment_url = payment_url
                tx.provider_reference = provider_ref
                tx.status = "pending"
                tx.save(update_fields=["payment_url", "provider_reference", "status", "updated_at"])
            elif provider == "flutterwave":
                secret = str(getattr(settings, "FLUTTERWAVE_SECRET_KEY", "") or "").strip()
                if not secret:
                    raise ValidationError({"detail": "Flutterwave is not configured"})
                email = str(request.data.get("email") or getattr(invoice.customer, "email", "") or "").strip()
                if not email:
                    raise ValidationError({"email": "Customer email is required"})
                phone = str(request.data.get("phone") or getattr(invoice.customer, "phone", "") or "").strip()
                customer_name = str(getattr(invoice.customer, "name", "") or "Customer")
                status_code, parsed, _raw = _http_json(
                    "POST",
                    "https://api.flutterwave.com/v3/payments",
                    {"Authorization": f"Bearer {secret}"},
                    {
                        "tx_ref": reference,
                        "amount": str(amount),
                        "currency": currency_code,
                        "redirect_url": redirect_url,
                        "customer": {"email": email, "phonenumber": phone, "name": customer_name},
                        "customizations": {"title": f"Invoice {invoice.invoice_number}", "description": "Invoice payment"},
                        "meta": {"invoice_id": invoice.id},
                    },
                    timeout=20,
                )
                if status_code >= 400 or str(parsed.get("status") or "").lower() not in ("success", "successful"):
                    raise PaymentGatewayUnavailable("Flutterwave initialize failed")
                data = parsed.get("data") or {}
                payment_url = str(data.get("link") or "").strip() or None
                tx.payment_url = payment_url
                tx.provider_reference = reference
                tx.status = "pending"
                tx.save(update_fields=["payment_url", "provider_reference", "status", "updated_at"])
            elif provider == "opay":
                merchant_id = str(getattr(settings, "OPAY_MERCHANT_ID", "") or "").strip()
                private_key = str(getattr(settings, "OPAY_PRIVATE_KEY", "") or "").strip()
                env = str(getattr(settings, "OPAY_ENV", "test") or "test").strip().lower()
                if not merchant_id or not private_key:
                    raise ValidationError({"detail": "OPay is not configured"})
                base = "https://testapi.opaycheckout.com" if env != "live" else "https://liveapi.opaycheckout.com"
                url = base + "/api/v1/international/payment/create"
                country = str(request.data.get("country") or "NG").strip().upper()
                amount_minor = _to_minor_units(amount, currency_code)
                payload = {
                    "reference": reference,
                    "amount": {"total": amount_minor, "currency": currency_code},
                    "country": country,
                    "payMethod": "BankCard",
                    "product": {"name": f"Invoice {invoice.invoice_number}", "description": "Invoice payment"},
                    "returnUrl": redirect_url,
                    "callbackUrl": request.build_absolute_uri("/api/payments/webhooks/opay/"),
                }
                ordered = json.dumps(payload, sort_keys=True, separators=(",", ":"))
                sig = hmac.new(private_key.encode("utf-8"), ordered.encode("utf-8"), hashlib.sha512).hexdigest()
                status_code, parsed, _raw = _http_json(
                    "POST",
                    url,
                    {"Authorization": f"Bearer {sig}", "MerchantId": merchant_id},
                    payload,
                    timeout=20,
                )
                data = parsed.get("data") or {}
                cashier = str(data.get("cashierUrl") or data.get("cashier_url") or data.get("redirectUrl") or "").strip() or None
                if status_code >= 400 or not cashier:
                    raise PaymentGatewayUnavailable("OPay initialize failed")
                tx.payment_url = cashier
                tx.provider_reference = reference
                tx.provider_transaction_id = str(data.get("orderNo") or data.get("order_no") or "").strip() or None
                tx.status = "pending"
                tx.save(update_fields=["payment_url", "provider_reference", "provider_transaction_id", "status", "updated_at"])
            elif provider == "bank_transfer":
                bank_code = str(request.data.get("bank_code") or "").strip()
                bank_name = str(request.data.get("bank_name") or "").strip()
                account_number = str(request.data.get("account_number") or "").strip()
                account_name = str(request.data.get("account_name") or "").strip()
                tx.metadata = {
                    **(tx.metadata or {}),
                    "bank_code": bank_code or None,
                    "bank_name": bank_name or None,
                    "account_number": account_number or None,
                    "account_name": account_name or None,
                    "instructions": "Complete a bank transfer using the reference. Funds will be confirmed automatically if a reconciliation webhook is configured, otherwise manually.",
                }
                tx.status = "pending"
                tx.save(update_fields=["metadata", "status", "updated_at"])
                reconcile_url = str(getattr(settings, "BANK_TRANSFER_RECONCILIATION_URL", "") or "").strip()
                if reconcile_url:
                    headers = {}
                    secret = str(getattr(settings, "BANK_TRANSFER_RECONCILIATION_SECRET", "") or "").strip()
                    if secret:
                        headers["x-bank-transfer-secret"] = secret
                    try:
                        status_code, parsed, _raw = _http_json(
                            "POST",
                            reconcile_url,
                            headers,
                            {"reference": tx.reference, "amount": str(tx.amount), "currency": tx.currency_code, "invoice_id": invoice.id},
                            timeout=10,
                        )
                        if status_code >= 400:
                            tx.metadata = {**(tx.metadata or {}), "reconciliation_error": "reconciliation_http_error", "reconciliation_status_code": status_code}
                            tx.save(update_fields=["metadata", "updated_at"])
                        else:
                            tx.metadata = {**(tx.metadata or {}), "reconciliation_accepted": True, "reconciliation_response": parsed}
                            tx.save(update_fields=["metadata", "updated_at"])
                    except Exception as e:
                        tx.metadata = {**(tx.metadata or {}), "reconciliation_error": e.__class__.__name__, "reconciliation_error_message": str(e)[:255]}
                        tx.save(update_fields=["metadata", "updated_at"])
            else:
                raise ValidationError({"provider": "Invalid provider"})
        except ValidationError:
            raise
        except PaymentGatewayTimeout:
            raise
        except Exception as e:
            _fail_payment_transaction(tx=tx, failure_code=e.__class__.__name__, failure_message=str(e), provider_status=None)
            raise PaymentGatewayUnavailable(str(e))

        _log_audit(user, "create", tx, {"provider": provider, "invoice": invoice_id_int, "amount": str(tx.amount), "currency": tx.currency_code})
        return Response(self.get_serializer(tx).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"])
    def verify(self, request, pk=None):
        tx = self.get_object()
        if tx.status == "succeeded":
            return Response(self.get_serializer(tx).data, status=status.HTTP_200_OK)

        try:
            if tx.provider == "paystack":
                secret = str(getattr(settings, "PAYSTACK_SECRET_KEY", "") or "").strip()
                if not secret:
                    raise ValidationError({"detail": "Paystack is not configured"})
                status_code, parsed, _raw = _http_json(
                    "GET",
                    f"https://api.paystack.co/transaction/verify/{urllib.parse.quote(tx.reference)}",
                    {"Authorization": f"Bearer {secret}"},
                    None,
                    timeout=20,
                )
                data = (parsed.get("data") or {}) if isinstance(parsed, dict) else {}
                if status_code < 400 and parsed.get("status") and str(data.get("status") or "").lower() == "success":
                    _settle_payment_transaction(
                        tx=tx,
                        provider_paid_at=str(data.get("paid_at") or ""),
                        provider_status=str(data.get("status") or ""),
                        provider_tx_id=str(data.get("id") or ""),
                        provider_ref=str(data.get("reference") or tx.reference),
                    )
                elif status_code >= 400:
                    raise PaymentGatewayUnavailable("Paystack verify failed")
            elif tx.provider == "flutterwave":
                secret = str(getattr(settings, "FLUTTERWAVE_SECRET_KEY", "") or "").strip()
                if not secret:
                    raise ValidationError({"detail": "Flutterwave is not configured"})
                flw_id = str(tx.provider_transaction_id or "").strip()
                if not flw_id:
                    raise ValidationError({"detail": "Missing Flutterwave transaction id"})
                status_code, parsed, _raw = _http_json(
                    "GET",
                    f"https://api.flutterwave.com/v3/transactions/{urllib.parse.quote(flw_id)}/verify",
                    {"Authorization": f"Bearer {secret}"},
                    None,
                    timeout=20,
                )
                data = (parsed.get("data") or {}) if isinstance(parsed, dict) else {}
                if status_code < 400 and str(parsed.get("status") or "").lower() in ("success", "successful") and str(data.get("status") or "").lower() == "successful":
                    _settle_payment_transaction(
                        tx=tx,
                        provider_paid_at=str(data.get("created_at") or ""),
                        provider_status=str(data.get("status") or ""),
                        provider_tx_id=str(data.get("id") or ""),
                        provider_ref=str(data.get("tx_ref") or tx.reference),
                    )
                elif status_code >= 400:
                    raise PaymentGatewayUnavailable("Flutterwave verify failed")
            return Response(self.get_serializer(PaymentTransaction.objects.get(pk=tx.pk)).data, status=status.HTTP_200_OK)
        except ValidationError:
            raise
        except Exception as e:
            raise PaymentGatewayUnavailable(str(e))


class PaystackWebhookApi(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        secret = str(getattr(settings, "PAYSTACK_SECRET_KEY", "") or "").strip()
        raw_body = (request.body or b"").decode("utf-8", errors="replace")
        signature = str(request.headers.get("x-paystack-signature") or "").strip()
        calc = hmac.new(secret.encode("utf-8"), (request.body or b""), hashlib.sha512).hexdigest() if secret else ""
        signature_valid = bool(secret) and bool(signature) and hmac.compare_digest(calc, signature)

        try:
            payload = json.loads(raw_body or "{}")
            if not isinstance(payload, dict):
                payload = {}
        except json.JSONDecodeError:
            payload = {}

        data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
        event_id = str(data.get("id") or "").strip() or None
        reference = str(data.get("reference") or "").strip() or None

        row, created = PaymentWebhookEvent.objects.get_or_create(
            provider="paystack",
            event_id=event_id,
            defaults={"reference": reference, "signature_valid": signature_valid, "headers": _safe_webhook_headers(request), "payload": payload, "raw_body": raw_body},
        )
        if not created:
            if row.status == "processed":
                return Response({"ok": True}, status=status.HTTP_200_OK)
            row.reference = row.reference or reference
            row.signature_valid = bool(row.signature_valid or signature_valid)
            row.headers = row.headers or _safe_webhook_headers(request)
            row.payload = row.payload or payload
            row.raw_body = row.raw_body or raw_body
            row.save(update_fields=["reference", "signature_valid", "headers", "payload", "raw_body"])

        if not signature_valid:
            row.status = "ignored"
            row.processed_at = timezone.now()
            row.save(update_fields=["status", "processed_at"])
            return Response({"ok": True}, status=status.HTTP_200_OK)

        event = str(payload.get("event") or "").strip().lower()
        try:
            tx = PaymentTransaction.objects.filter(provider="paystack", reference=reference).select_related("invoice", "created_by").first() if reference else None
            if tx is None:
                row.status = "ignored"
                row.processed_at = timezone.now()
                row.save(update_fields=["status", "processed_at"])
                return Response({"ok": True}, status=status.HTTP_200_OK)

            if event == "charge.success" or str(data.get("status") or "").lower() == "success":
                _settle_payment_transaction(
                    tx=tx,
                    provider_paid_at=str(data.get("paid_at") or ""),
                    provider_status=str(data.get("status") or ""),
                    provider_tx_id=str(data.get("id") or ""),
                    provider_ref=str(data.get("reference") or tx.reference),
                )
                row.status = "processed"
            else:
                row.status = "ignored"
            row.processed_at = timezone.now()
            row.save(update_fields=["status", "processed_at"])
            return Response({"ok": True}, status=status.HTTP_200_OK)
        except Exception as e:
            row.status = "failed"
            row.error_message = str(e)[:255]
            row.processed_at = timezone.now()
            row.save(update_fields=["status", "error_message", "processed_at"])
            logger.exception("paystack.webhook_failed event_id=%s reference=%s", event_id, reference)
            return Response({"ok": True}, status=status.HTTP_200_OK)


class FlutterwaveWebhookApi(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        secret_hash = str(getattr(settings, "FLUTTERWAVE_WEBHOOK_SECRET_HASH", "") or "").strip()
        raw_body_bytes = request.body or b""
        raw_body = raw_body_bytes.decode("utf-8", errors="replace")
        header_sig = str(request.headers.get("flutterwave-signature") or request.headers.get("verif-hash") or "").strip()
        calc = hmac.new(secret_hash.encode("utf-8"), raw_body_bytes, hashlib.sha256).hexdigest() if secret_hash else ""
        signature_valid = bool(secret_hash) and bool(header_sig) and hmac.compare_digest(calc, header_sig)

        try:
            payload = json.loads(raw_body or "{}")
            if not isinstance(payload, dict):
                payload = {}
        except json.JSONDecodeError:
            payload = {}

        data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
        event_id = str(data.get("id") or "").strip() or None
        reference = str(data.get("tx_ref") or data.get("reference") or "").strip() or None

        row, created = PaymentWebhookEvent.objects.get_or_create(
            provider="flutterwave",
            event_id=event_id,
            defaults={"reference": reference, "signature_valid": signature_valid, "headers": _safe_webhook_headers(request), "payload": payload, "raw_body": raw_body},
        )
        if not created:
            if row.status == "processed":
                return Response({"ok": True}, status=status.HTTP_200_OK)
            row.reference = row.reference or reference
            row.signature_valid = bool(row.signature_valid or signature_valid)
            row.headers = row.headers or _safe_webhook_headers(request)
            row.payload = row.payload or payload
            row.raw_body = row.raw_body or raw_body
            row.save(update_fields=["reference", "signature_valid", "headers", "payload", "raw_body"])

        if not signature_valid:
            row.status = "ignored"
            row.processed_at = timezone.now()
            row.save(update_fields=["status", "processed_at"])
            return Response({"ok": True}, status=status.HTTP_200_OK)

        event = str(payload.get("event") or "").strip().lower()
        status_str = str(data.get("status") or "").strip().lower()

        try:
            tx = PaymentTransaction.objects.filter(provider="flutterwave", reference=reference).select_related("invoice", "created_by").first() if reference else None
            if tx is None:
                row.status = "ignored"
                row.processed_at = timezone.now()
                row.save(update_fields=["status", "processed_at"])
                return Response({"ok": True}, status=status.HTTP_200_OK)

            if event == "charge.completed" and status_str in ("successful", "success"):
                _settle_payment_transaction(
                    tx=tx,
                    provider_paid_at=str(data.get("created_at") or ""),
                    provider_status=str(data.get("status") or ""),
                    provider_tx_id=str(data.get("id") or ""),
                    provider_ref=str(data.get("tx_ref") or tx.reference),
                )
                row.status = "processed"
            elif status_str in ("failed", "cancelled"):
                _fail_payment_transaction(tx=tx, failure_code=str(data.get("processor_response") or "FAILED"), failure_message=str(data.get("processor_response") or "Failed"), provider_status=status_str)
                row.status = "processed"
            else:
                row.status = "ignored"
            row.processed_at = timezone.now()
            row.save(update_fields=["status", "processed_at"])
            return Response({"ok": True}, status=status.HTTP_200_OK)
        except Exception as e:
            row.status = "failed"
            row.error_message = str(e)[:255]
            row.processed_at = timezone.now()
            row.save(update_fields=["status", "error_message", "processed_at"])
            logger.exception("flutterwave.webhook_failed event_id=%s reference=%s", event_id, reference)
            return Response({"ok": True}, status=status.HTTP_200_OK)


class OPayWebhookApi(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        merchant_id = str(getattr(settings, "OPAY_MERCHANT_ID", "") or "").strip()
        private_key = str(getattr(settings, "OPAY_PRIVATE_KEY", "") or "").strip()
        raw_body_bytes = request.body or b""
        raw_body = raw_body_bytes.decode("utf-8", errors="replace")
        header_sig = str(request.headers.get("Authorization") or request.headers.get("authorization") or "").strip()
        header_mid = str(request.headers.get("MerchantId") or request.headers.get("merchantid") or "").strip()
        req_ts = str(request.headers.get("RequestTimestamp") or request.headers.get("requesttimestamp") or "").strip()
        provided = header_sig.replace("Bearer", "").strip()
        calc = hmac.new(private_key.encode("utf-8"), (req_ts + raw_body).encode("utf-8"), hashlib.sha512).hexdigest() if private_key and req_ts else ""
        signature_valid = bool(private_key) and bool(merchant_id) and header_mid == merchant_id and bool(provided) and bool(calc) and hmac.compare_digest(calc, provided)

        try:
            payload = json.loads(raw_body or "{}")
            if not isinstance(payload, dict):
                payload = {}
        except json.JSONDecodeError:
            payload = {}

        event_id = str(payload.get("notifyId") or payload.get("notify_id") or "").strip() or None
        reference = str(payload.get("reference") or "").strip() or None

        row, created = PaymentWebhookEvent.objects.get_or_create(
            provider="opay",
            event_id=event_id,
            defaults={"reference": reference, "signature_valid": signature_valid, "headers": _safe_webhook_headers(request), "payload": payload, "raw_body": raw_body},
        )
        if not created:
            if row.status == "processed":
                return Response({"ok": True}, status=status.HTTP_200_OK)
            row.reference = row.reference or reference
            row.signature_valid = bool(row.signature_valid or signature_valid)
            row.headers = row.headers or _safe_webhook_headers(request)
            row.payload = row.payload or payload
            row.raw_body = row.raw_body or raw_body
            row.save(update_fields=["reference", "signature_valid", "headers", "payload", "raw_body"])

        if not signature_valid:
            row.status = "ignored"
            row.processed_at = timezone.now()
            row.save(update_fields=["status", "processed_at"])
            return Response({"ok": True}, status=status.HTTP_200_OK)

        try:
            tx = PaymentTransaction.objects.filter(provider="opay", reference=reference).select_related("invoice", "created_by").first() if reference else None
            if tx is None:
                row.status = "ignored"
                row.processed_at = timezone.now()
                row.save(update_fields=["status", "processed_at"])
                return Response({"ok": True}, status=status.HTTP_200_OK)

            status_str = str(payload.get("status") or "").strip().upper()
            if status_str in ("SUCCESS",):
                _settle_payment_transaction(
                    tx=tx,
                    provider_paid_at=str(payload.get("timestamp") or ""),
                    provider_status=status_str,
                    provider_tx_id=str(payload.get("orderNo") or payload.get("order_no") or ""),
                    provider_ref=reference,
                )
                row.status = "processed"
            elif status_str in ("FAIL", "CLOSE"):
                _fail_payment_transaction(tx=tx, failure_code=str(payload.get("failureCode") or ""), failure_message=str(payload.get("failureReason") or ""), provider_status=status_str)
                row.status = "processed"
            else:
                row.status = "ignored"
            row.processed_at = timezone.now()
            row.save(update_fields=["status", "processed_at"])
            return Response({"ok": True}, status=status.HTTP_200_OK)
        except Exception as e:
            row.status = "failed"
            row.error_message = str(e)[:255]
            row.processed_at = timezone.now()
            row.save(update_fields=["status", "error_message", "processed_at"])
            logger.exception("opay.webhook_failed event_id=%s reference=%s", event_id, reference)
            return Response({"ok": True}, status=status.HTTP_200_OK)


class BankTransferWebhookApi(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        secret = str(getattr(settings, "BANK_TRANSFER_WEBHOOK_SECRET", "") or "").strip()
        provided = str(request.headers.get("x-bank-transfer-secret") or "").strip()
        if not secret or not provided or not hmac.compare_digest(secret, provided):
            return Response({"ok": True}, status=status.HTTP_200_OK)

        reference = str(request.data.get("reference") or "").strip()
        if not reference:
            return Response({"ok": True}, status=status.HTTP_200_OK)

        tx = PaymentTransaction.objects.filter(provider="bank_transfer", reference=reference).select_related("invoice", "created_by").first()
        if tx is None:
            return Response({"ok": True}, status=status.HTTP_200_OK)

        status_str = str(request.data.get("status") or "SUCCESS").strip().upper()
        if status_str == "SUCCESS":
            _settle_payment_transaction(
                tx=tx,
                provider_paid_at=str(request.data.get("paid_at") or ""),
                provider_status=status_str,
                provider_tx_id=str(request.data.get("bank_tx_id") or ""),
                provider_ref=reference,
            )
        elif status_str in ("FAIL", "FAILED", "CANCELLED"):
            _fail_payment_transaction(tx=tx, failure_code=str(request.data.get("failure_code") or "FAILED"), failure_message=str(request.data.get("failure_message") or "Failed"), provider_status=status_str)

        return Response({"ok": True}, status=status.HTTP_200_OK)


BUSINESS_ROLE_LEVEL = {"viewer": 1, "member": 2, "admin": 3, "owner": 4}


def _is_app_admin(user) -> bool:
    if not getattr(user, "is_authenticated", False):
        return False
    roles = set(user_role_names(user))
    return "admin" in roles or "staff" in roles or bool(getattr(user, "is_superuser", False))


def _require_app_admin(request) -> None:
    if not _is_app_admin(request.user):
        raise PermissionDenied()


def _business_role_for_user(user, business_id: int) -> str | None:
    if not getattr(user, "is_authenticated", False):
        return None
    if _is_app_admin(user):
        return "owner"
    row = BusinessMembership.objects.filter(business_id=business_id, user=user).only("role").first()
    return row.role if row else None


def _require_business_role(request, business_id: int, minimum_role: str) -> str:
    role = _business_role_for_user(request.user, business_id)
    if role is None:
        raise PermissionDenied()
    if BUSINESS_ROLE_LEVEL.get(role, 0) < BUSINESS_ROLE_LEVEL.get(minimum_role, 999):
        raise PermissionDenied()
    return role


class BusinessAccountViewSet(viewsets.ModelViewSet):
    serializer_class = BusinessAccountSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = OptionalPageNumberPagination

    def get_queryset(self):
        qs = BusinessAccount.objects.all().select_related("owner").order_by("-id")
        if _is_app_admin(self.request.user):
            return qs
        biz_ids = BusinessMembership.objects.filter(user=self.request.user).values_list("business_id", flat=True)
        return qs.filter(id__in=biz_ids)

    def create(self, request, *args, **kwargs):
        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        with transaction.atomic():
            biz = BusinessAccount.objects.create(name=str(ser.validated_data["name"]).strip(), owner=request.user)
            BusinessMembership.objects.get_or_create(business=biz, user=request.user, defaults={"role": "owner"})
        _log_audit(request.user, "create", biz, {"name": biz.name})
        return Response(self.get_serializer(biz).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        biz = self.get_object()
        _require_business_role(request, biz.id, "admin")
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        biz = self.get_object()
        _require_business_role(request, biz.id, "admin")
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        biz = self.get_object()
        _require_business_role(request, biz.id, "owner")
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["get"], url_path="members")
    def members(self, request, pk=None):
        biz = self.get_object()
        _require_business_role(request, biz.id, "member")
        qs = BusinessMembership.objects.filter(business=biz).select_related("user").order_by("id")
        return Response({"members": BusinessMembershipSerializer(qs, many=True).data}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="members/add")
    def add_member(self, request, pk=None):
        biz = self.get_object()
        _require_business_role(request, biz.id, "admin")
        username = str(request.data.get("username") or "").strip()
        role = str(request.data.get("role") or "").strip()
        if not username:
            raise ValidationError({"username": "username is required"})
        if role not in dict(BusinessMembership.ROLE_CHOICES):
            raise ValidationError({"role": "Invalid role"})
        if role == "owner":
            raise ValidationError({"role": "owner role cannot be assigned"})
        User = get_user_model()
        user = User.objects.filter(username=username).first()
        if user is None:
            raise ValidationError({"username": "User not found"})
        row, created = BusinessMembership.objects.get_or_create(business=biz, user=user, defaults={"role": role})
        if not created and row.role != role:
            row.role = role
            row.save(update_fields=["role"])
        _log_audit(request.user, "update", biz, {"member": {"user_id": str(user.id), "role": role}})
        return Response(BusinessMembershipSerializer(row).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="members/remove")
    def remove_member(self, request, pk=None):
        biz = self.get_object()
        _require_business_role(request, biz.id, "admin")
        user_id = request.data.get("user_id")
        try:
            user_id_int = int(user_id)
        except (TypeError, ValueError):
            raise ValidationError({"user_id": "user_id must be an integer"})
        if user_id_int == biz.owner_id:
            raise ValidationError({"user_id": "Cannot remove the business owner"})
        deleted, _ = BusinessMembership.objects.filter(business=biz, user_id=user_id_int).delete()
        return Response({"removed": bool(deleted)}, status=status.HTTP_200_OK)
